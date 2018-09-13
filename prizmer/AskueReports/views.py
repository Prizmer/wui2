# coding -*- coding: utf-8 -*-

#from django.shortcuts import render
from __future__ import unicode_literals
from django.shortcuts import render_to_response, HttpResponse
from django.db import connection
import StringIO
from openpyxl import Workbook
from openpyxl.compat import range
import datetime
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Font
from openpyxl.cell import get_column_letter
import common_sql
import re
from excel_response import ExcelResponse
from django.shortcuts import redirect
#from datetime import datetime, date, time
import decimal
# для работы с xml
from lxml import etree


def zagotovka(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
    ws['B5'] = 'Заготовка'
    ws['B5'].style = ali_grey
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'zagotovka' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

# Стили
ali_grey   = Style(fill=PatternFill(fill_type='solid', start_color='DCDCDC'), border=Border(left=Side(border_style='thin',color='FF000000'), bottom=Side(border_style='thin',color='FF000000'), right=Side(border_style='thin',color='FF000000'), top=Side(border_style='thin',color='FF000000')), alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True))
ali_white  = Style(border=Border(left=Side(border_style='thin',color='FF000000'), bottom=Side(border_style='thin',color='FF000000'), right=Side(border_style='thin',color='FF000000'), top=Side(border_style='thin',color='FF000000')), alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True))
ali_blue   = Style(fill=PatternFill(fill_type='solid', start_color='E6E6FA'), border=Border(left=Side(border_style='thin',color='FF000000'), bottom=Side(border_style='thin',color='FF000000'), right=Side(border_style='thin',color='FF000000'), top=Side(border_style='thin',color='FF000000')), alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True))
ali_pink   = Style(fill=PatternFill(fill_type='solid', start_color='FFF0F5'), border=Border(left=Side(border_style='thin',color='FF000000'), bottom=Side(border_style='thin',color='FF000000'), right=Side(border_style='thin',color='FF000000'), top=Side(border_style='thin',color='FF000000')), alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True))

ali_yellow = Style(fill=PatternFill(fill_type='solid', start_color='EEEE00'), border=Border(left=Side(border_style='thin',color='FF000000'), bottom=Side(border_style='thin',color='FF000000'), right=Side(border_style='thin',color='FF000000'), top=Side(border_style='thin',color='FF000000')), alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True))
ali_white_size_18  = Style(font=Font(size=18))
# Конец описания стилей

def translate(name):
 
    #Заменяем пробелы и преобразуем строку к нижнему регистру
    name = name.replace(' ','-').lower()
 
    #
    transtable = (
        ## Большие буквы
        (u"Щ", u"Sch"),
        (u"Щ", u"SCH"),
        # two-symbol
        (u"Ё", u"Yo"),
        (u"Ё", u"YO"),
        (u"Ж", u"Zh"),
        (u"Ж", u"ZH"),
        (u"Ц", u"Ts"),
        (u"Ц", u"TS"),
        (u"Ч", u"Ch"),
        (u"Ч", u"CH"),
        (u"Ш", u"Sh"),
        (u"Ш", u"SH"),
        (u"Ы", u"Yi"),
        (u"Ы", u"YI"),
        (u"Ю", u"Yu"),
        (u"Ю", u"YU"),
        (u"Я", u"Ya"),
        (u"Я", u"YA"),
        # one-symbol
        (u"А", u"A"),
        (u"Б", u"B"),
        (u"В", u"V"),
        (u"Г", u"G"),
        (u"Д", u"D"),
        (u"Е", u"E"),
        (u"З", u"Z"),
        (u"И", u"I"),
        (u"Й", u"J"),
        (u"К", u"K"),
        (u"Л", u"L"),
        (u"М", u"M"),
        (u"Н", u"N"),
        (u"О", u"O"),
        (u"П", u"P"),
        (u"Р", u"R"),
        (u"С", u"S"),
        (u"Т", u"T"),
        (u"У", u"U"),
        (u"Ф", u"F"),
        (u"Х", u"H"),
        (u"Э", u"E"),
        (u"Ъ", u"`"),
        (u"Ь", u"'"),
        ## Маленькие буквы
        # three-symbols
        (u"щ", u"sch"),
        # two-symbols
        (u"ё", u"yo"),
        (u"ж", u"zh"),
        (u"ц", u"ts"),
        (u"ч", u"ch"),
        (u"ш", u"sh"),
        (u"ы", u"yi"),
        (u"ю", u"yu"),
        (u"я", u"ya"),
        # one-symbol
        (u"а", u"a"),
        (u"б", u"b"),
        (u"в", u"v"),
        (u"г", u"g"),
        (u"д", u"d"),
        (u"е", u"e"),
        (u"з", u"z"),
        (u"и", u"i"),
        (u"й", u"j"),
        (u"к", u"k"),
        (u"л", u"l"),
        (u"м", u"m"),
        (u"н", u"n"),
        (u"о", u"o"),
        (u"п", u"p"),
        (u"р", u"r"),
        (u"с", u"s"),
        (u"т", u"t"),
        (u"у", u"u"),
        (u"ф", u"f"),
        (u"х", u"h"),
        (u"ъ", u"`"),
        (u"ь", u"'"),
        (u"э", u"e"),
    )
    #перебираем символы в таблице и заменяем
    for symb_in, symb_out in transtable:
        name = name.replace(symb_in, symb_out)
    #возвращаем переменную
    return name

def get_k_t_n_by_serial_number(serial_number):
    """Получаем Ктн по серийному номеру счтчика"""
    simpleq = connection.cursor()
    simpleq.execute("""SELECT 
                          link_abonents_taken_params.coefficient_2
                        FROM 
                          public.meters, 
                          public.link_abonents_taken_params, 
                          public.taken_params
                        WHERE 
                          link_abonents_taken_params.guid_taken_params = taken_params.guid AND
                          taken_params.guid_meters = meters.guid AND
                          meters.factory_number_manual = %s
                        ORDER BY
                          link_abonents_taken_params.coefficient_2 DESC
                        LIMIT 1;""", [serial_number])
    simpleq = simpleq.fetchall()
    return simpleq[0][0]
    

    
def get_k_t_t_by_serial_number(serial_number):
    """Получаем Ктт по серийному номеру счтчика"""
    simpleq = connection.cursor()
    simpleq.execute("""SELECT 
                          link_abonents_taken_params.coefficient
                        FROM 
                          public.meters, 
                          public.link_abonents_taken_params, 
                          public.taken_params
                        WHERE 
                          link_abonents_taken_params.guid_taken_params = taken_params.guid AND
                          taken_params.guid_meters = meters.guid AND
                          meters.factory_number_manual = %s
                        ORDER BY
                          link_abonents_taken_params.coefficient DESC
                        LIMIT 1;""", [serial_number])
    simpleq = simpleq.fetchall()
    return simpleq[0][0]
    
def get_k_a_by_serial_number(serial_number):
    """Получаем Коэффициент А по серийному номеру счтчика"""
    simpleq = connection.cursor()
    simpleq.execute("""SELECT 
                          link_abonents_taken_params.coefficient_3
                        FROM 
                          public.meters, 
                          public.link_abonents_taken_params, 
                          public.taken_params
                        WHERE 
                          link_abonents_taken_params.guid_taken_params = taken_params.guid AND
                          taken_params.guid_meters = meters.guid AND
                          meters.factory_number_manual = %s
                        ORDER BY
                          link_abonents_taken_params.coefficient DESC
                        LIMIT 1;""", [serial_number])
    simpleq = simpleq.fetchall()
    return simpleq[0][0]


def report_3_tarifa_k(request): # Отчет по А+ и R+ с коэффициентами

    response = StringIO.StringIO()    
    wb = Workbook()    
    ws = wb.active                   

# Шапка отчета   
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Срез показаний счетчика с коэффициентами за период' + ' ' + str(request.session["electric_data_start"]) + " - " + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктт'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктн'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма A+
    ws.merge_cells('F4:I4')
    ws['F4'] = 'Суммарные показания/энергия A+'
    ws['F4'].style = ali_grey
    ws['G4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['F5'] = 'Показания на ' + str(request.session["electric_data_start"])
    ws['F5'].style = ali_grey
    ws['G5'] = 'Показания на ' + str(request.session["electric_data_end"])
    ws['G5'].style = ali_grey
    ws['H5'] = 'Разность показаний'
    ws['H5'].style = ali_grey
    ws['I5'] = 'Энергия кВт*ч'
    ws['I5'].style = ali_grey
    
    # Сумма R+
    ws.merge_cells('J4:M4')
    ws['J4'] = 'Суммарные показания/энергия R+'
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['L4'].style = ali_grey
    ws['M4'].style = ali_grey
    ws['J5'] = 'Показания на ' + str(request.session["electric_data_start"])
    ws['J5'].style = ali_grey
    ws['K5'] = 'Показания на ' + str(request.session["electric_data_end"])
    ws['K5'].style = ali_grey
    ws['L5'] = 'Разность показаний'
    ws['L5'].style = ali_grey
    ws['M5'] = 'Энергия кВт*ч'
    ws['M5'].style = ali_grey
    
   
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17
    
# Заполняем таблицу данными ----------------------------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------------------------------------------------------
    is_abonent_level    = re.compile(r'abonent')
    is_object_level     = re.compile(r'level')    
    obj_title           = request.session["obj_title"]
    electric_data_start = request.session["electric_data_start"]
    electric_data_end   = request.session["electric_data_end"]
    obj_key             = request.session["obj_key"]
    data_table = []
    
#--------------------------------------------------------------------------------------------------------------------------------------------------------------
    if bool(is_object_level.search(obj_key)): # Если это объект, то формируем список абонентов
        cursor_abonents_list = connection.cursor()
        cursor_abonents_list.execute("""
                                  SELECT 
                                   abonents.name
                                  FROM 
                                   public.objects, 
                                   public.abonents
                                  WHERE 
                                   objects.guid = abonents.guid_objects AND
                                   objects.name = %s
                                  ORDER BY
                                   abonents.name ASC;""",[obj_title])
        abonents_list = cursor_abonents_list.fetchall()

    elif bool(is_abonent_level.search(obj_key)):   # Если это отдельный абонент, то делаем выборку для одного абонента, а за имя родительсого объекта берем завод.
        abonents_list = [(obj_title,)]
        obj_title = u"Завод"
     
    for x in range(len(abonents_list)):
        # delta for groups abonents 'start date' A+
        cursor_t0_aplus_delta_start_temp = connection.cursor()
        cursor_t0_aplus_delta_start_temp.execute("""
                    SELECT 
                      daily_values.date, 
                      daily_values.value, 
                      abonents.name, 
                      daily_values.id_taken_params, 
                      objects.name, 
                      names_params.name, 
                      meters.factory_number_manual, 
                      resources.name,
                      link_abonents_taken_params.coefficient,
                      link_abonents_taken_params.coefficient_2,
                      link_abonents_taken_params.coefficient_3
                    FROM 
                      public.daily_values, 
                      public.link_abonents_taken_params, 
                      public.taken_params, 
                      public.abonents, 
                      public.objects, 
                      public.names_params, 
                      public.params, 
                      public.meters, 
                      public.resources
                    WHERE 
                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                      taken_params.id = daily_values.id_taken_params AND
                      taken_params.guid_params = params.guid AND
                      taken_params.guid_meters = meters.guid AND
                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                      objects.guid = abonents.guid_objects AND
                      names_params.guid = params.guid_names_params AND
                      resources.guid = names_params.guid_resources AND
                      abonents.name = %s AND 
                      objects.name = %s AND 
                      names_params.name = 'T0 A+' AND 
                      daily_values.date = %s AND 
                      resources.name = 'Электричество'
                      ORDER BY
                      objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
        data_table_t0_aplus_delta_start_temp = cursor_t0_aplus_delta_start_temp.fetchall()
    
	# delta for groups abonents 'end date' A+
        cursor_t0_aplus_delta_end_temp = connection.cursor()
        cursor_t0_aplus_delta_end_temp.execute("""
                    SELECT 
                      daily_values.date, 
                      daily_values.value, 
                      abonents.name, 
                      daily_values.id_taken_params, 
                      objects.name, 
                      names_params.name, 
                      meters.factory_number_manual, 
                      resources.name,
                      link_abonents_taken_params.coefficient,
                      link_abonents_taken_params.coefficient_2,
                      link_abonents_taken_params.coefficient_3
                    FROM 
                      public.daily_values, 
                      public.link_abonents_taken_params, 
                      public.taken_params, 
                      public.abonents, 
                      public.objects, 
                      public.names_params, 
                      public.params, 
                      public.meters, 
                      public.resources
                    WHERE 
                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                      taken_params.id = daily_values.id_taken_params AND
                      taken_params.guid_params = params.guid AND
                      taken_params.guid_meters = meters.guid AND
                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                      objects.guid = abonents.guid_objects AND
                      names_params.guid = params.guid_names_params AND
                      resources.guid = names_params.guid_resources AND
                      abonents.name = %s AND 
                      objects.name = %s AND 
                      names_params.name = 'T0 A+' AND 
                      daily_values.date = %s AND 
                      resources.name = 'Электричество'
                      ORDER BY
                      objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
        data_table_t0_aplus_delta_end_temp = cursor_t0_aplus_delta_end_temp.fetchall()
        
    # delta for groups abonents 'start date' R+
        cursor_t0_rplus_delta_start_temp = connection.cursor()
        cursor_t0_rplus_delta_start_temp.execute("""
                    SELECT 
                      daily_values.date, 
                      daily_values.value, 
                      abonents.name, 
                      daily_values.id_taken_params, 
                      objects.name, 
                      names_params.name, 
                      meters.factory_number_manual, 
                      resources.name,
                      link_abonents_taken_params.coefficient,
                      link_abonents_taken_params.coefficient_2,
                      link_abonents_taken_params.coefficient_3
                    FROM 
                      public.daily_values, 
                      public.link_abonents_taken_params, 
                      public.taken_params, 
                      public.abonents, 
                      public.objects, 
                      public.names_params, 
                      public.params, 
                      public.meters, 
                      public.resources
                    WHERE 
                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                      taken_params.id = daily_values.id_taken_params AND
                      taken_params.guid_params = params.guid AND
                      taken_params.guid_meters = meters.guid AND
                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                      objects.guid = abonents.guid_objects AND
                      names_params.guid = params.guid_names_params AND
                      resources.guid = names_params.guid_resources AND
                      abonents.name = %s AND 
                      objects.name = %s AND 
                      names_params.name = 'T0 R+' AND 
                      daily_values.date = %s AND 
                      resources.name = 'Электричество'
                      ORDER BY
                      objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_start])
        data_table_t0_rplus_delta_start_temp = cursor_t0_rplus_delta_start_temp.fetchall()
    
	# delta for groups abonents 'end date' R+
        cursor_t0_rplus_delta_end_temp = connection.cursor()
        cursor_t0_rplus_delta_end_temp.execute("""
                    SELECT 
                      daily_values.date, 
                      daily_values.value, 
                      abonents.name, 
                      daily_values.id_taken_params, 
                      objects.name, 
                      names_params.name, 
                      meters.factory_number_manual, 
                      resources.name,
                      link_abonents_taken_params.coefficient,
                      link_abonents_taken_params.coefficient_2,
                      link_abonents_taken_params.coefficient_3
                    FROM 
                      public.daily_values, 
                      public.link_abonents_taken_params, 
                      public.taken_params, 
                      public.abonents, 
                      public.objects, 
                      public.names_params, 
                      public.params, 
                      public.meters, 
                      public.resources
                    WHERE 
                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                      taken_params.id = daily_values.id_taken_params AND
                      taken_params.guid_params = params.guid AND
                      taken_params.guid_meters = meters.guid AND
                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                      objects.guid = abonents.guid_objects AND
                      names_params.guid = params.guid_names_params AND
                      resources.guid = names_params.guid_resources AND
                      abonents.name = %s AND 
                      objects.name = %s AND 
                      names_params.name = 'T0 R+' AND 
                      daily_values.date = %s AND 
                      resources.name = 'Электричество'
                      ORDER BY
                      objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
        data_table_t0_rplus_delta_end_temp = cursor_t0_rplus_delta_end_temp.fetchall()
    
       
        data_table_temp = []
        data_table_temp.append(abonents_list[x][0]) # наименование канала
        try:# заводской номер
            data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][6])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:# T0 A+ нач
            data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")        
        try:# T0 A+ кон
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:# расход T0 A+
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1] - data_table_t0_aplus_delta_start_temp[0][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try: #k_1
            data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][8])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try: #k_2
            data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][9])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:#k_3
            data_table_temp.append(data_table_t0_aplus_delta_start_temp[0][10])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")

        try:# T0 R+ нач
            data_table_temp.append(data_table_t0_rplus_delta_start_temp[0][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")        
        try:# T0 R+ кон
            data_table_temp.append(data_table_t0_rplus_delta_end_temp[0][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:# расход T0 R+
            data_table_temp.append(data_table_t0_rplus_delta_end_temp[0][1] - data_table_t0_rplus_delta_start_temp[0][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
                      
        data_table.append(data_table_temp)
        
    for row in range(6, len(abonents_list)+6):
        ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0]) # наименование канала
        ws.cell('A%s'%(row)).style = ali_grey
        
        ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
        ws.cell('B%s'%(row)).style = ali_white
        
        ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][2])  # T0 A+ нач
        ws.cell('F%s'%(row)).style = ali_white
               
        ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][3])  # T0 A+ кон
        ws.cell('G%s'%(row)).style = ali_white
        
        ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][4]) # расход T0 A+
        ws.cell('H%s'%(row)).style = ali_white
        
        ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][5]) # коэффициент 1  Ктт
        ws.cell('C%s'%(row)).style = ali_white
        
        ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][6]) # коэффициент 2  Ктн
        ws.cell('D%s'%(row)).style = ali_white
                
        ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][7]) # коэффициент 3  Постоянная счётчика
        ws.cell('E%s'%(row)).style = ali_white
        
        
        ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][8])  # T0 R+ нач
        ws.cell('J%s'%(row)).style = ali_white
               
        ws.cell('K%s'%(row)).value = '%s' % (data_table[row-6][9])  # T0 R+ кон
        ws.cell('K%s'%(row)).style = ali_white
        
        ws.cell('L%s'%(row)).value = '%s' % (data_table[row-6][10]) # расход T0 R+
        ws.cell('L%s'%(row)).style = ali_white
        
        try:
            ws.cell('I%s'%(row)).value = '%s' % (float(data_table[row-6][4])*float(data_table[row-6][5])*float(data_table[row-6][6])) # T0 R+ энергия с учёток коэффициентов
            ws.cell('I%s'%(row)).style = ali_yellow
        except UnicodeEncodeError:
            ws.cell('I%s'%(row)).value = '%s' % '-'
            ws.cell('I%s'%(row)).style = ali_yellow
        except TypeError:
            ws.cell('I%s'%(row)).value = '%s' % '-'
            ws.cell('I%s'%(row)).style = ali_yellow
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % (float(data_table[row-6][10])*float(data_table[row-6][5])*float(data_table[row-6][6])) # T0 R+ энергия с учёток коэффициентов
            ws.cell('M%s'%(row)).style = ali_yellow
        except UnicodeEncodeError:
            ws.cell('M%s'%(row)).value = '%s' % '-'
            ws.cell('M%s'%(row)).style = ali_yellow
        except TypeError:
            ws.cell('M%s'%(row)).value = '%s' % '-'
            ws.cell('M%s'%(row)).style = ali_yellow
                    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
#    response['Content-Disposition'] = "attachment; filename=electric.xlsx"
    output_name = u'otchet za period ' + electric_data_start + '-' + electric_data_end+ translate(obj_title)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   

    return response
    
    
def pokazania(request): # Показания по общему тарифу по А+ и R+
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

# Шапка отчета   
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Срез показаний с коэффициентами на дату' + ' ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктт'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктн'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F4:I4')
    ws['F4'] = 'Суммарные показания/энергия'
    ws['F4'].style = ali_grey
    ws['G4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['F5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['G5'].style = ali_yellow
    
    ws['H5'] = 'Показания R+ на ' + str(request.session["electric_data_end"])
    ws['H5'].style = ali_grey
    
    ws['I5'] = 'Энергия R+ на ' + str(request.session["electric_data_end"])
    ws['I5'].style = ali_yellow
    
   
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
# Заполняем таблицу данными ----------------------------------------------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------------------------------------------------------------------------
    is_abonent_level    = re.compile(r'abonent')
    is_object_level     = re.compile(r'level')    
    obj_title           = request.session["obj_title"]
    electric_data_end   = request.session["electric_data_end"]
    obj_key             = request.session["obj_key"]
    data_table = []
    
#--------------------------------------------------------------------------------------------------------------------------------------------------------------
    if bool(is_object_level.search(obj_key)): # Если это объект, то формируем список абонентов
        cursor_abonents_list = connection.cursor()
        cursor_abonents_list.execute("""
                                  SELECT 
                                   abonents.name
                                  FROM 
                                   public.objects, 
                                   public.abonents
                                  WHERE 
                                   objects.guid = abonents.guid_objects AND
                                   objects.name = %s
                                  ORDER BY
                                   abonents.name ASC;""",[obj_title])
        abonents_list = cursor_abonents_list.fetchall()

    elif bool(is_abonent_level.search(obj_key)):   # Если это отдельный абонент, то делаем выборку для одного абонента, а за имя родительсого объекта берем завод.
        abonents_list = [(obj_title,)]
        obj_title = u"Завод"

    for x in range(len(abonents_list)):    
	# delta A+ for groups abonents 'end date'
        cursor_t0_aplus_delta_end_temp = connection.cursor()
        cursor_t0_aplus_delta_end_temp.execute("""
                    SELECT 
                      daily_values.date, 
                      daily_values.value, 
                      abonents.name, 
                      daily_values.id_taken_params, 
                      objects.name, 
                      names_params.name, 
                      meters.factory_number_manual, 
                      resources.name,
                      link_abonents_taken_params.coefficient,
                      link_abonents_taken_params.coefficient_2,
                      link_abonents_taken_params.coefficient_3
                    FROM 
                      public.daily_values, 
                      public.link_abonents_taken_params, 
                      public.taken_params, 
                      public.abonents, 
                      public.objects, 
                      public.names_params, 
                      public.params, 
                      public.meters, 
                      public.resources
                    WHERE 
                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                      taken_params.id = daily_values.id_taken_params AND
                      taken_params.guid_params = params.guid AND
                      taken_params.guid_meters = meters.guid AND
                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                      objects.guid = abonents.guid_objects AND
                      names_params.guid = params.guid_names_params AND
                      resources.guid = names_params.guid_resources AND
                      abonents.name = %s AND 
                      objects.name = %s AND 
                      names_params.name = 'T0 A+' AND 
                      daily_values.date = %s AND 
                      resources.name = 'Электричество'
                      ORDER BY
                      objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
        data_table_t0_aplus_delta_end_temp = cursor_t0_aplus_delta_end_temp.fetchall()
        
        	# delta R+ for groups abonents 'end date'
        cursor_t0_rplus_delta_end_temp = connection.cursor()
        cursor_t0_rplus_delta_end_temp.execute("""
                    SELECT 
                      daily_values.date, 
                      daily_values.value, 
                      abonents.name, 
                      daily_values.id_taken_params, 
                      objects.name, 
                      names_params.name, 
                      meters.factory_number_manual, 
                      resources.name,
                      link_abonents_taken_params.coefficient,
                      link_abonents_taken_params.coefficient_2,
                      link_abonents_taken_params.coefficient_3
                    FROM 
                      public.daily_values, 
                      public.link_abonents_taken_params, 
                      public.taken_params, 
                      public.abonents, 
                      public.objects, 
                      public.names_params, 
                      public.params, 
                      public.meters, 
                      public.resources
                    WHERE 
                      taken_params.guid = link_abonents_taken_params.guid_taken_params AND
                      taken_params.id = daily_values.id_taken_params AND
                      taken_params.guid_params = params.guid AND
                      taken_params.guid_meters = meters.guid AND
                      abonents.guid = link_abonents_taken_params.guid_abonents AND
                      objects.guid = abonents.guid_objects AND
                      names_params.guid = params.guid_names_params AND
                      resources.guid = names_params.guid_resources AND
                      abonents.name = %s AND 
                      objects.name = %s AND 
                      names_params.name = 'T0 R+' AND 
                      daily_values.date = %s AND 
                      resources.name = 'Электричество'
                      ORDER BY
                      objects.name ASC;""",[abonents_list[x][0], obj_title, electric_data_end])
        data_table_t0_rplus_delta_end_temp = cursor_t0_rplus_delta_end_temp.fetchall()
           
        data_table_temp = []
        data_table_temp.append(abonents_list[x][0]) # наименование канала
        try:# заводской номер
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][6])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:# T0 A+ кон
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")        
        try: #k_1
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][8])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try: #k_2
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][9])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:#k_3
            data_table_temp.append(data_table_t0_aplus_delta_end_temp[0][10])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:# T0 R+ кон
            data_table_temp.append(data_table_t0_rplus_delta_end_temp[0][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
                       
        data_table.append(data_table_temp)
        
    for row in range(6, len(abonents_list)+6):
        ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0]) # наименование канала
        ws.cell('A%s'%(row)).style = ali_grey
        
        ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
        ws.cell('B%s'%(row)).style = ali_white
        
        ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][2])  # T0 A+ кон
        ws.cell('F%s'%(row)).style = ali_white
        
        ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3]) # коэффициент 1  Ктт
        ws.cell('C%s'%(row)).style = ali_white
        
        ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4]) # коэффициент 2  Ктн
        ws.cell('D%s'%(row)).style = ali_white
                
        ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5]) # коэффициент 3  Постоянная счётчика
        ws.cell('E%s'%(row)).style = ali_white

        ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][6]) # T0 R+ кон
        ws.cell('H%s'%(row)).style = ali_white
                
        try:
            ws.cell('G%s'%(row)).value = '%s' % (float(data_table[row-6][2])*float(data_table[row-6][3])*float(data_table[row-6][4])) # T0 A+ энергия с учёток коэффициентов
            ws.cell('G%s'%(row)).style = ali_yellow
        except UnicodeEncodeError:
            ws.cell('G%s'%(row)).value = '%s' % '-'
            ws.cell('G%s'%(row)).style = ali_yellow
        except TypeError:
            ws.cell('G%s'%(row)).value = '%s' % '-'
            ws.cell('G%s'%(row)).style = ali_yellow
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (float(data_table[row-6][6])*float(data_table[row-6][3])*float(data_table[row-6][4])) # T0 R+ энергия с учёток коэффициентов
            ws.cell('I%s'%(row)).style = ali_yellow
        except UnicodeEncodeError:
            ws.cell('I%s'%(row)).value = '%s' % '-'
            ws.cell('I%s'%(row)).style = ali_yellow
        except TypeError:
            ws.cell('I%s'%(row)).value = '%s' % '-'
            ws.cell('I%s'%(row)).style = ali_yellow

    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel") 
#    response['Content-Disposition'] = 'attachment; filename=eclectric.xlsx'
    output_name = u'otchet po pokazaniyam za ' + electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    
    return response


def profil_30_min(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
    meters_name         = request.session["obj_title"]
    electric_data_end   = request.session["electric_data_end"]
    
    ws.merge_cells('A2:F2')
    ws['A2'] = u'Профиль мощности по абоненту ' + unicode(request.session['obj_title']) + u' за ' + str(request.session["electric_data_end"])
    ws['A3'] = u'Ктн = ' + str(common_sql.get_k_t_n(meters_name))
    ws['B3'] = u'Ктт = ' + str(common_sql.get_k_t_t(meters_name))
    ws['B5'] = 'Дата'
    ws['B5'].style = ali_grey
    ws['C5'] = 'Время'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Наименование'
    ws['D5'].style = ali_grey
    ws['E5'] = 'A+ кВт'
    ws['E5'].style = ali_grey
    ws['F5'] = 'R+ кВАр'
    ws['F5'].style = ali_grey
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['D'].width = 30

    
    a_plus = connection.cursor()
    a_plus.execute("""SELECT 
                          various_values.date, 
                          various_values."time", 
                          various_values.value, 
                          meters.name, 
                          meters.address, 
                          names_params.name
                        FROM 
                          public.various_values, 
                          public.meters, 
                          public.params, 
                          public.taken_params, 
                          public.names_params
                        WHERE 
                          params.guid_names_params = names_params.guid AND
                          taken_params.guid_params = params.guid AND
                          taken_params.guid_meters = meters.guid AND
                          taken_params.id = various_values.id_taken_params AND
                          various_values.date = %s AND 
                          meters.name = %s AND 
                          names_params.name = 'A+ Профиль';""",[electric_data_end, meters_name])
    a_plus = a_plus.fetchall()
   
        
    r_plus = connection.cursor()
    r_plus.execute("""SELECT 
                          various_values.date, 
                          various_values."time", 
                          various_values.value, 
                          meters.name, 
                          meters.address, 
                          names_params.name
                        FROM 
                          public.various_values, 
                          public.meters, 
                          public.params, 
                          public.taken_params, 
                          public.names_params
                        WHERE 
                          params.guid_names_params = names_params.guid AND
                          taken_params.guid_params = params.guid AND
                          taken_params.guid_meters = meters.guid AND
                          taken_params.id = various_values.id_taken_params AND
                          various_values.date = %s AND 
                          meters.name = %s AND 
                          names_params.name = 'R+ Профиль';""",[electric_data_end, meters_name])
    r_plus = r_plus.fetchall()
          
    data_table = []
    for x in range(len(a_plus)):
        data_table_temp = []
        try:
            data_table_temp.append(a_plus[x][0])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(a_plus[x][1])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(a_plus[x][3])
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(a_plus[x][2]*2*common_sql.get_k_t_n(meters_name)*common_sql.get_k_t_t(meters_name))
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        try:
            data_table_temp.append(r_plus[x][2]*2*common_sql.get_k_t_n(meters_name)*common_sql.get_k_t_t(meters_name))
        except IndexError:
            data_table_temp.append(u"Н/Д")
        except TypeError:
            data_table_temp.append(u"Н/Д")
        data_table.append(data_table_temp)
            
    for row in range(6, len(data_table)+6):
        ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][0]) # дата
        ws.cell('B%s'%(row)).style = ali_grey
        
        ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][1])  # время
        ws.cell('C%s'%(row)).style = ali_white
        
        ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][2])  # канал
        ws.cell('D%s'%(row)).style = ali_white
        
        ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][3]) # значение A+
        ws.cell('E%s'%(row)).style = ali_white
        
        ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][4]) # значение R+
        ws.cell('F%s'%(row)).style = ali_white
           
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'profil 30 min za ' + electric_data_end # формируем имя для excel отчета 
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_hour_increment(request): # Выгрузка excel по часовым приращениям
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
    meters_name         = request.session["obj_title"]
    electric_data_end   = request.session["electric_data_end"]    

# Шапка очета      
    ws.merge_cells('B2:F2')
    ws['B2'] = 'Почасовой учет электроэнергии за ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('B3:F3')
    ws['B3'] = u'Абонент: ' + unicode(request.session['obj_title'])
    
    ws['B5'] = 'Дата'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Время'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Абонент'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Серийный номер'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'A+ кВт*ч'
    ws['F5'].style = ali_grey
    
    ws['G5'] = '+ А+ кВт*ч'
    ws['G5'].style = ali_grey
    
    ws['H5'] = 'R+ кВар*ч'
    ws['H5'].style = ali_grey
    
    ws['I5'] = '+ R+ кВар*ч'
    ws['I5'].style = ali_grey
    
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20  
#-------------
    
#Запрашиваем данные для отчета
    time_list = ['00:00', '00:30','01:00', '01:30', '02:00', '02:30', '03:00', '03:30', '04:00', '04:30', '05:00', '05:30', '06:00', '06:30', '07:00', '07:30', '08:00', '08:30', '09:00', '09:30', '10:00', '10:30', '11:00', '11:30', '12:00', '12:30', '13:00', '13:30', '14:00', '14:30', '15:00', '15:30', '16:00', '16:30', '17:00', '17:30', '18:00', '18:30', '19:00', '19:30', '20:00', '20:30', '21:00', '21:30', '22:00', '22:30', '23:00', '23:30']
    meters_name         = request.session["obj_title"]
    electric_data_end   = request.session["electric_data_end"]
    
    serial_number = common_sql.get_serial_number_by_meter_name(meters_name)
        
    data_table = []
    # Добавляем первую строку в таблицу данных. Делаем запрос показаний на начало суток.
    data_table.append([electric_data_end,u'00:00', meters_name, serial_number, common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 A+' ),common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 R+' ), u'0', u'0'])
    
    if common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 A+' ) != u'Нет данных':  # Если есть показания на начало суток выполняем почасовое приращение  
        for x in range(24):
            data_table_temp = []
            data_table_temp.append(electric_data_end)
            data_table_temp.append(time_list[(2*x)])
            data_table_temp.append(meters_name)
            data_table_temp.append(serial_number)
            data_table_temp.append(data_table[len(data_table)-1][4] + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'A+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'A+ Профиль'))
            data_table_temp.append(common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 R+' ) + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'R+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'R+ Профиль'))
            data_table_temp.append(common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'A+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'A+ Профиль'))
            data_table_temp.append(common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'R+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'R+ Профиль'))    
            if x == 0: # Убираем первую строку. Так как показания на 00:00 берем отдельным запросом
                next
            else:
                data_table.append(data_table_temp)    
#-----------------------------

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][0]) # дата
        ws.cell('B%s'%(row)).style = ali_grey
        
        ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][1])  # время
        ws.cell('C%s'%(row)).style = ali_white
        
        ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][2])  # канал
        ws.cell('D%s'%(row)).style = ali_white
        
        ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][3]) # Заводской номер
        ws.cell('E%s'%(row)).style = ali_white
        
        ws.cell('F%s'%(row)).value = '%s' % (round(data_table[row-6][4],2)) # значение A+
        ws.cell('F%s'%(row)).style = ali_white
        
        ws.cell('G%s'%(row)).value = '%s' % (round(float(data_table[row-6][6]),2)) # значение + A+
        ws.cell('G%s'%(row)).style = ali_white
        
        ws.cell('H%s'%(row)).value = '%s' % (round(data_table[row-6][5],2)) # значение R+
        ws.cell('H%s'%(row)).style = ali_white

        ws.cell('I%s'%(row)).value = '%s' % (round(float(data_table[row-6][7]),2)) # значение + R+
        ws.cell('I%s'%(row)).style = ali_white
#---------------------------
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'chasovie prirasheniya'
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
    
def pokazania_period(request): # Показания по абоненту за период
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

    meters_name         = request.session["obj_title"]
    parent_name         = request.session['obj_parent_title']
    electric_data_end   = request.session["electric_data_end"]
    electric_data_start = request.session['electric_data_start']
    data_table = []
# Шапка отчета
    ws.merge_cells('B2:F2')
    ws['B2'] = u'Ежедневные показания за период с ' + str(request.session["electric_data_start"]) + u' по ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('B3:F3')
    ws['B3'] = u'Абонент: ' + unicode(request.session['obj_title'])
   
    ws['B5'] = 'Дата'
    ws['B5'].style = ali_grey
       
    ws['C5'] = 'Абонент'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Серийный номер'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'A+ кВт*ч'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'R+ кВар*ч'
    ws['F5'].style = ali_grey
       
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20          
# Конец шапки

#Запрашиваем данные для отчета
    end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
    start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
    dates = [x for x in common_sql.daterange(start_date,
                  end_date,
                  step=datetime.timedelta(days=1),
                  inclusive=True)]

    for x in range(len(dates)):
        data_table_temp = get_data_table_by_date_daily(meters_name, parent_name, datetime.datetime.strftime(dates[x], "%d.%m.%Y"))
        if data_table_temp:
            data_table.extend(data_table_temp)
        else:
            data_table.append([datetime.datetime.strftime(dates[x], "%d.%m.%Y"),meters_name,common_sql.get_serial_number_by_meter_name(meters_name), u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д'])
   
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][0]) # дата
        ws.cell('B%s'%(row)).style = ali_grey
        
        ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][1])  # абонент
        ws.cell('C%s'%(row)).style = ali_white
        
        ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
        ws.cell('D%s'%(row)).style = ali_white
        
        ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][3]) # значение A+
        ws.cell('E%s'%(row)).style = ali_white
        
        ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][8]) # значение R+
        ws.cell('F%s'%(row)).style = ali_white
        
#---------------------------
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'pokazania za period' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_rejim_day(request): # Отчет по режимному дню
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
    meters_name         = request.session["obj_title"]
#    parent_name         = request.session['obj_parent_title']
    electric_data_end   = request.session["electric_data_end"]
#    electric_data_start = request.session['electric_data_start']
    data_table = []
    #print meters_name
    
    general_k = common_sql.get_k_t_t(meters_name) * common_sql.get_k_t_n(meters_name)#поменяла функцию: в запросе нужен factory namber manual, а не name в meters
# Шапка отчета
    ws['A1'] = u"ЗАО 'Кировская керамика'"
    ws['H1'] = u'Шифр'
    ws['H2'] = u'Питающий центр'
    ws['H3'] = u'№ фидера'
    
    ws['D5'] = u'Протокол (первичный)'
    ws['D5'].style = ali_white_size_18
    
    ws['E7'] = u'трансформаторного напряжения _____ вольт'
    
    ws['B6'] = u'записей показаний электросчетчиков и вольтметров, а также определения нагрузок'
    ws['B7'] = u"и тангенса 'фи' за " + str(electric_data_end) + u'г'
    ws['B9'] = u'Акт. Счетчик № '# + str(common_sql.get_serial_number_by_meter_name(meters_name)) 
    ws['E9'] = u'Реакт. Счетчик № '# + str(common_sql.get_serial_number_by_meter_name(meters_name))
    ws['B10'] = u'Расч. Коэфициент' 
    ws['E10'] = u'Расч. Коэфициент'

    ws['A11'] = u'Время' 
    ws['A12'] = u'записи, часы'
    ws['A39'] = u'суточный расход' 
    ws['A40'] = u'активной и реактивной'
    ws['A41'] = u'энергии' 
    ws['A42'] = u'Контрольная сумма'
    ws['A44'] = u'Запись показаний счетчиков производили'
    ws['A44'] = u'Запись показаний счетчиков производили'
    ws['A45'] = u'фамилия ______________ подпись ______________'
    ws['A46'] = u'фамилия ______________ подпись ______________'
    ws['A47'] = u'фамилия ______________ подпись ______________'

    ws['G44'] = u'Расчеты произвел:'
    ws['G46'] = u'фамилия ______________ подпись ______________'

    ws['D10'] = ws['G10'] = general_k # Общий коэффициент
    ws['D9'] = ws['G9'] = str(common_sql.get_serial_number_by_meter_name(meters_name)) #Серийный номер прибора #поменяла функцию: в запросе нужен factory namber manual, а не name в meters
        
    ws['A13'] = u'0' 
    ws['A14'] = u'1' 
    ws['A15'] = u'2'
    ws['A16'] = u'3' 
    ws['A17'] = u'4'
    ws['A18'] = u'5' 
    ws['A19'] = u'6'
    ws['A20'] = u'7' 
    ws['A21'] = u'8'
    ws['A22'] = u'9' 
    ws['A23'] = u'10'
    ws['A24'] = u'11' 
    ws['A25'] = u'12'
    ws['A26'] = u'13' 
    ws['A27'] = u'14'
    ws['A28'] = u'15' 
    ws['A29'] = u'16'
    ws['A30'] = u'17' 
    ws['A31'] = u'18'
    ws['A32'] = u'19' 
    ws['A33'] = u'20'
    ws['A34'] = u'21' 
    ws['A35'] = u'22'
    ws['A36'] = u'23' 
    ws['A37'] = u'24'

    ws['B11'] = u'Показания' 
    ws['B12'] = u'счетчика'
    
    ws['C11'] = u'Разность' 
    ws['C12'] = u'показаний'
    
    ws['D11'] = u'расход за' 
    ws['D12'] = u'час(квт)'
    
    ws['E11'] = u'Показания' 
    ws['E12'] = u'счетчика'
    
    ws['F11'] = u'Разность' 
    ws['F12'] = u'показаний'
    
    ws['G11'] = u'расход за' 
    ws['G12'] = u'час(квт)'

    ws['H10'] = u'тангенс'
    ws['H11'] = u'фи'
    
    ws['I9'] = u'Показания'
    ws['I10'] = u'вольтметров на'
    ws['I11'] = u'стороне'
    ws['I12'] = u'в/н'
    ws['J12'] = u'н/н'
    
    ws['K9'] = u'мощность '
    ws['K10'] = u'включенных'
    ws['K11'] = u'компен.устр'
    ws['K12'] = u'кВар'
    
    for col_idx in range(1, 12):
        col = get_column_letter(col_idx)
        for row in range(13, 38):
            ws.cell('%s%s'%(col, row)).style = ali_white

    ws.column_dimensions['A'].width = 12            
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['I'].width = 7
    ws.column_dimensions['J'].width = 7    
    ws.row_dimensions[5].height = 30
    # Конец шапки

#Запрашиваем данные для отчета
    time_list = ['00:00', '00:30','01:00', '01:30', '02:00', '02:30', '03:00', '03:30', '04:00', '04:30', '05:00', '05:30', '06:00', '06:30', '07:00', '07:30', '08:00', '08:30', '09:00', '09:30', '10:00', '10:30', '11:00', '11:30', '12:00', '12:30', '13:00', '13:30', '14:00', '14:30', '15:00', '15:30', '16:00', '16:30', '17:00', '17:30', '18:00', '18:30', '19:00', '19:30', '20:00', '20:30', '21:00', '21:30', '22:00', '22:30', '23:00', '23:30']
    
    serial_number = common_sql.get_serial_number_by_meter_name(meters_name)
    # Добавляем первую строку в таблицу данных. Делаем запрос показаний на начало суток.
    data_table.append([electric_data_end,u'00:00', meters_name, serial_number, common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 A+' ),common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 R+' ), u'0', u'0'])
    
    if common_sql.get_daily_value_by_meter_name(meters_name, electric_data_end, 'T0 A+' ) != u'Нет данных':  # Если есть показания на начало суток выполняем почасовое приращение  
        for x in range(24):
            data_table_temp = []
            data_table_temp.append(electric_data_end) # Дата
            data_table_temp.append(time_list[(2*x)])  # Отчетный час
            data_table_temp.append(meters_name)       # Имя абонента
            data_table_temp.append(serial_number)     # Серийный номер
            data_table_temp.append(data_table[len(data_table)-1][4] + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'A+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'A+ Профиль'))     # Показиние счётчика за предыдущий час + две получасовки А+           
            data_table_temp.append(data_table[len(data_table)-1][5] + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'R+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'R+ Профиль'))     # Показиние счётчика за предыдущий час + две получасовки R+

            data_table_temp.append(common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'A+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'A+ Профиль'))                                        # Сумма двух получасовок: потребленная энергия за час А+
            data_table_temp.append(common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)-1], 'R+ Профиль') + common_sql.get_30_min_by_meter_name(meters_name, electric_data_end, time_list[(2*x)], 'R+ Профиль'))                                        # Сумма двух получасовок: потребленная энергия за час R+  
            if x == 0: # Убираем первую строку. Так как показания на 00:00 берем отдельным запросом 
                next
            else:
                data_table.append(data_table_temp)
    if data_table[23][4] and data_table[23][5]:
        data_table.append([(datetime.datetime.strptime(electric_data_end, u'%d.%m.%Y') + datetime.timedelta(days=1)).strftime(u'%d.%m.%Y'),u'00:00', meters_name, serial_number, common_sql.get_daily_value_by_meter_name(meters_name, (datetime.datetime.strptime(electric_data_end, u'%d.%m.%Y') + datetime.timedelta(days=1)).strftime(u'%d.%m.%Y'), 'T0 A+' ),common_sql.get_daily_value_by_meter_name(meters_name, (datetime.datetime.strptime(electric_data_end, u'%d.%m.%Y') + datetime.timedelta(days=1)).strftime(u'%d.%m.%Y'), 'T0 R+' ),common_sql.get_daily_value_by_meter_name(meters_name, (datetime.datetime.strptime(electric_data_end, u'%d.%m.%Y') + datetime.timedelta(days=1)).strftime(u'%d.%m.%Y'), 'T0 A+' ) - data_table[23][4],common_sql.get_daily_value_by_meter_name(meters_name, (datetime.datetime.strptime(electric_data_end, u'%d.%m.%Y') + datetime.timedelta(days=1)).strftime(u'%d.%m.%Y'), 'T0 R+' ) - data_table[23][5]])

    #------------

# Заполняем отчет значениями
    for row in range(13, len(data_table)+13):
        
        ws.cell('B%s'%(row)).value = '%s' % (round(data_table[row-13][4],4)) # значение A+
        ws.cell('B%s'%(row)).style = ali_white
        
        ws.cell('C%s'%(row)).value = '%s' % (round(float(data_table[row-13][6]),4)) # значение + A+
        ws.cell('C%s'%(row)).style = ali_white
        
        ws.cell('D%s'%(row)).value = '%s' % (round(float(data_table[row-13][6]),4)*general_k) # значение + A+
        ws.cell('D%s'%(row)).style = ali_white
        
        ws.cell('E%s'%(row)).value = '%s' % (round(data_table[row-13][5],4)) # значение R+
        ws.cell('E%s'%(row)).style = ali_white

        ws.cell('F%s'%(row)).value = '%s' % (round(float(data_table[row-13][7]),4)) # значение + R+
        ws.cell('F%s'%(row)).style = ali_white
        
        ws.cell('G%s'%(row)).value = '%s' % (round(float(data_table[row-13][7]),4)*general_k) # значение + R+
        ws.cell('G%s'%(row)).style = ali_white
        
        fi=0
        try:
            fi = round((float(data_table[row-13][7])*general_k)/(float(data_table[row-13][6])*general_k),2)
        except:
            fi = u''
        
        ws.cell('H%s'%(row)).value = '%s' % fi # значение + R+/+ A+
        ws.cell('H%s'%(row)).style = ali_white  
    #---------------------------      
    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'rejimniy den '+ str(electric_data_end) 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    

def report_economic_electric(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    #--------------------------------------------------------------------------------------------------------------------      
# Шапка отчета
    ws.merge_cells('B2:F2')
    ws['B2'] = u'Таблица расчета удельного коэффициента период с ' + str(request.session["electric_data_start"]) + u' по ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('B3:F3')
    ws['B3'] = u'Абонент: Литейный цех'
   
    ws['B5'] = 'Дата'
    ws['B5'].style = ali_grey
       
    ws['C5'] = 'Изготовленная продукция, кг'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Затраченная A+, кВт*ч'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Удельный расход A+, кВт*ч/кг'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'Затраченная R+, кВар*ч'
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Удельный расход R+, кВт*ч/кг'
    ws['G5'].style = ali_grey

    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18    
    ws.row_dimensions[5].height = 30    
# Конец шапки
    
#Запрашиваем данные для отчета---
    data_table = []
    
    electric_data_end   = request.session["electric_data_end"]
    electric_data_start   = request.session["electric_data_start"]

    if request.is_ajax():
        if request.method == 'GET':
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']    

    end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
    start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
    dates = [x for x in common_sql.daterange(start_date,
                  end_date,
                  step=datetime.timedelta(days=1),
                  inclusive=True)]
                  
    for x in range(len(dates)):
        try:
            data_table_temp = []
            delta_a_plus = 1
            delta_r_plus = 1

            try:
                delta_a_plus = common_sql.delta_sum_a_plus(dates[x+1])-common_sql.delta_sum_a_plus(dates[x])
                if delta_a_plus > 0:
				    delta_a_plus = delta_a_plus
                else:
				    delta_a_plus = u'Н/Д'
                delta_r_plus = common_sql.delta_sum_r_plus(dates[x+1])-common_sql.delta_sum_r_plus(dates[x])
                if delta_r_plus > 0:
				    delta_r_plus = delta_r_plus
                else:
                    delta_r_plus = u'Н/Д'

            except:
                delta_a_plus = u'Н/Д'
                delta_r_plus = u'Н/Д'

            data_table_temp.append(dates[x])
            data_table_temp.append(common_sql.product_sum(dates[x]))
            data_table_temp.append(delta_a_plus)
            data_table_temp.append(delta_a_plus/(common_sql.product_sum(dates[x])))
            data_table_temp.append(delta_r_plus)
            data_table_temp.append(delta_r_plus/(common_sql.product_sum(dates[x])))
        except:
            next
        data_table.append(data_table_temp)

#Конец запроса данных------------
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        ws.cell('B%s'%(row)).value = '%s' % ((data_table[row-6][0]).strftime("%d-%m-%Y")) # дата
        ws.cell('B%s'%(row)).style = ali_grey
        
        ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][1])  # изготовленная продукция, кг
        ws.cell('C%s'%(row)).style = ali_white
        
        ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][2])  # затраченная А+
        ws.cell('D%s'%(row)).style = ali_white
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][3])  # удельных расход А+/кг
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][4])  # затраченная R+
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][5])  # удельный расход R+/кг
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
        
        
        
#---------------------------    
    #--------------------------------------------------------------------------------------------------------------------    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'udelniy_coefficient_liteiniy_ceh' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response


def report_pokazaniya_water_identificators(request): # Показания по воде за дату с идентификаторами
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

    meters_name         = request.session["obj_title"]
    parent_name         = request.session['obj_parent_title']
    electric_data_end   = request.session["electric_data_end"]
#    electric_data_start = request.session['electric_data_start']
    data_table = []

# Шапка отчета
    ws.merge_cells('B2:F2')
    ws['B2'] = u'Показания по воде с идентификаторами за ' + str(request.session["electric_data_end"])
       
    ws['B5'] = 'Абонент'
    ws['B5'].style = ali_grey
       
    ws['C5'] = 'Идентификатор'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания'
    ws['D5'].style = ali_grey
    
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20        
# Конец шапки

#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_channel(meters_name, electric_data_end)
    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []        
        for x in range(len(list_of_abonents_2)):
            data_table_temp = common_sql.get_daily_water_channel(list_of_abonents_2[x], electric_data_end)
            data_table.extend(data_table_temp)
    elif(bool(is_object_level_1.search(obj_key))):
        
        list_of_objects_2 = common_sql.list_of_objects(common_sql.return_parent_guid_by_abonent_name(meters_name)) #Список квартир для объекта с пульсарами
        data_table = []
        for x in range(len(list_of_objects_2)):
            data_table_temp = [(list_of_objects_2[x][0],)]
            list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(meters_name), list_of_objects_2[x][0])
            for y in range(len(list_of_abonents_2)):
                data_table_temp2 = common_sql.get_daily_water_channel(list_of_abonents_2[y], electric_data_end)

                data_table_temp.extend(data_table_temp2)                                
                      
            data_table.extend(data_table_temp)
              
    else:
        data_table = [1,1,1,1]

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][0]) # абонент
        ws.cell('B%s'%(row)).style = ali_grey
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][4])  # идентификатор
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][2])  # показания м3
            ws.cell('D%s'%(row)).style = ali_white

        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
        
#---------------------------

    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'pokazania po vode' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_electric_simple_2_zones(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Срез показаний с коэффициентами на дату' + ' ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктт'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F4:G4')
    ws['F4'] = 'Сумма'
    ws['F4'].style = ali_grey
    ws['G4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['F5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['G5'].style = ali_yellow
    
    # Тариф 1
    ws.merge_cells('H4:I4')
    ws['H4'] = 'Тариф 1'
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['H5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['H5'].style = ali_grey
    
    ws['I5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['I5'].style = ali_yellow
    
    # Тариф 2
    ws.merge_cells('J4:K4')
    ws['J4'] = 'Тариф 2'
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['J5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['J5'].style = ali_grey
    
    ws['K5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['K5'].style = ali_yellow
    
       
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    is_electric_monthly = request.session['is_electric_monthly']
    is_electric_daily   = request.session['is_electric_daily']
    data_table = []
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
                        
#Запрашиваем данные для отчета конец
                  
    if (bool(is_abonent_level.search(obj_key))):
        if (is_electric_monthly == "1"):
            data_table = common_sql.get_data_table_by_date_monthly_2_zones(meters_name, parent_name, electric_data_end)
        elif (is_electric_daily == "1"):
            data_table = common_sql.get_data_table_by_date_daily_2_zones(meters_name, parent_name, electric_data_end)

    elif (bool(is_object_level_2.search(obj_key))):
            list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
            data_table = []
            for x in range(len(list_of_abonents_2)):
                if (is_electric_monthly == "1"):                
                    data_table_temp = common_sql.get_data_table_by_date_monthly_2_zones(list_of_abonents_2[x], meters_name, electric_data_end)
                elif (is_electric_daily == "1"):                
                    data_table_temp = common_sql.get_data_table_by_date_daily_2_zones(list_of_abonents_2[x], meters_name, electric_data_end)
                if data_table_temp:
                    data_table.extend(data_table_temp)
                else:
                    data_table.extend([[u'', list_of_abonents_2[x],u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д']])
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (get_k_t_n_by_serial_number(data_table[row-6][2]))  # Ктн
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (get_k_t_t_by_serial_number(data_table[row-6][2]))  # Ктт
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (get_k_a_by_serial_number(data_table[row-6][2]))  # Ка
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
                   
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][3])  # Сумма А+
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][3]*get_k_t_n_by_serial_number(data_table[row-6][2])*get_k_t_t_by_serial_number(data_table[row-6][2]))  # "Энергия Сумма А+
            ws.cell('G%s'%(row)).style = ali_yellow
        except:
            ws.cell('G%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][4])  # Тариф 1 А+
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][4]*get_k_t_n_by_serial_number(data_table[row-6][2])*get_k_t_t_by_serial_number(data_table[row-6][2]))  # "Энергия Тариф 1 А+
            ws.cell('I%s'%(row)).style = ali_yellow
        except:
            ws.cell('I%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][5])  # Тариф 2 А+
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table[row-6][5]*get_k_t_n_by_serial_number(data_table[row-6][2])*get_k_t_t_by_serial_number(data_table[row-6][2]))  # "Энергия Тариф 2 А+
            ws.cell('K%s'%(row)).style = ali_yellow
        except:
            ws.cell('K%s'%(row)).style = ali_yellow
            next           
# Сохраняем в ecxel    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'2_tariffa' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response
    

def report_electric_simple_2_zones_v2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Срез показаний с коэффициентами на дату' + ' ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктт'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F4:G4')
    ws['F4'] = 'Сумма'
    ws['F4'].style = ali_grey
    ws['G4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['F5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['G5'].style = ali_yellow
    
    # Тариф 1
    ws.merge_cells('H4:I4')
    ws['H4'] = 'Тариф 1'
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['H5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['H5'].style = ali_grey
    
    ws['I5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['I5'].style = ali_yellow
    
    # Тариф 2
    ws.merge_cells('J4:K4')
    ws['J4'] = 'Тариф 2'
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['J5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['J5'].style = ali_grey
    
    ws['K5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['K5'].style = ali_yellow
    
       
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    
    obj_parent_title    = request.session['obj_parent_title']
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    is_electric_monthly = request.session['is_electric_monthly']
    is_electric_daily   = request.session['is_electric_daily']
    is_electric_period  = request.session['is_electric_period']
    data_table = []
    if True:
        if True:            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'monthly')
                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'daily')

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # показания на начало месяца для объекта
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]        

#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for abonents group
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]
            elif (is_electric_daily == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]
              
            elif (is_electric_monthly == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][9])  # Ктн
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][8])  # Ктт
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][10])  # Ка
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
                   
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][3])  # Сумма А+
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % round((data_table[row-6][3]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Сумма А+
            ws.cell('G%s'%(row)).style = ali_yellow
        except:
            ws.cell('G%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][4])  # Тариф 1 А+
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % round((data_table[row-6][4]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Тариф 1 А+
            ws.cell('I%s'%(row)).style = ali_yellow
        except:
            ws.cell('I%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][5])  # Тариф 2 А+
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % round((data_table[row-6][5]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Тариф 2 А+
            ws.cell('K%s'%(row)).style = ali_yellow
        except:
            ws.cell('K%s'%(row)).style = ali_yellow
            next           
# Сохраняем в ecxel    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'2_tariffa' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response
    
def report_electric_simple_3_zones(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Срез показаний с коэффициентами на дату' + ' ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктт'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F4:G4')
    ws['F4'] = 'Сумма'
    ws['F4'].style = ali_grey
    ws['G4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['F5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['G5'].style = ali_yellow
    
    # Тариф 1
    ws.merge_cells('H4:I4')
    ws['H4'] = 'Тариф 1'
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['H5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['H5'].style = ali_grey
    
    ws['I5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['I5'].style = ali_yellow
    
    # Тариф 2
    ws.merge_cells('J4:K4')
    ws['J4'] = 'Тариф 2'
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['J5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['J5'].style = ali_grey
    
    ws['K5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['K5'].style = ali_yellow
    
    # Тариф 3
    ws.merge_cells('L4:M4')
    ws['L4'] = 'Тариф 3'
    ws['L4'].style = ali_grey
    ws['M4'].style = ali_grey
    ws['L4'].style = ali_grey
    ws['M4'].style = ali_grey
    ws['L5'] = 'Показания A+ на ' + str(request.session["electric_data_end"])
    ws['L5'].style = ali_grey
    
    ws['M5'] = 'Энергия A+ на ' + str(request.session["electric_data_end"])
    ws['M5'].style = ali_yellow
         
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    is_electric_monthly = request.session['is_electric_monthly']
    is_electric_daily   = request.session['is_electric_daily']
    data_table = []
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
                     
#Запрашиваем данные для отчета конец                  
    if (bool(is_abonent_level.search(obj_key))):
        if (is_electric_monthly == "1"):
            data_table = common_sql.get_data_table_by_date_monthly_3_zones(meters_name, parent_name, electric_data_end)
        elif (is_electric_daily == "1"):
            data_table = common_sql.get_data_table_by_date_daily_3_zones(meters_name, parent_name, electric_data_end)

    elif (bool(is_object_level_2.search(obj_key))):
            list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
            data_table = []
            for x in range(len(list_of_abonents_2)):
                if (is_electric_monthly == "1"):                
                    data_table_temp = common_sql.get_data_table_by_date_monthly_3_zones(list_of_abonents_2[x], meters_name, electric_data_end)
                elif (is_electric_daily == "1"):                
                    data_table_temp = common_sql.get_data_table_by_date_daily_3_zones(list_of_abonents_2[x], meters_name, electric_data_end)
                if data_table_temp:
                    data_table.extend(data_table_temp)
                else:
                    data_table.extend([[u'', list_of_abonents_2[x],u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д']])
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (get_k_t_n_by_serial_number(data_table[row-6][2]))  # Ктн
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (get_k_t_t_by_serial_number(data_table[row-6][2]))  # Ктт
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (get_k_a_by_serial_number(data_table[row-6][2]))  # Ка
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
                   
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][3])  # Сумма А+
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][3]*get_k_t_n_by_serial_number(data_table[row-6][2])*get_k_t_t_by_serial_number(data_table[row-6][2]))  # "Энергия Сумма А+
            ws.cell('G%s'%(row)).style = ali_yellow
        except:
            ws.cell('G%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][4])  # Тариф 1 А+
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][4]*get_k_t_n_by_serial_number(data_table[row-6][2])*get_k_t_t_by_serial_number(data_table[row-6][2]))  # "Энергия Тариф 1 А+
            ws.cell('I%s'%(row)).style = ali_yellow
        except:
            ws.cell('I%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][5])  # Тариф 2 А+
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table[row-6][5]*get_k_t_n_by_serial_number(data_table[row-6][2])*get_k_t_t_by_serial_number(data_table[row-6][2])) # "Энергия Тариф 2 А+
            ws.cell('K%s'%(row)).style = ali_yellow
        except:
            ws.cell('K%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table[row-6][6])  # Тариф 3 А+
            ws.cell('L%s'%(row)).style = ali_white

        except:
            ws.cell('L%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % (data_table[row-6][6]*get_k_t_n_by_serial_number(data_table[row-6][2])*get_k_t_t_by_serial_number(data_table[row-6][2]))  # "Энергия Тариф 3 А+
            ws.cell('M%s'%(row)).style = ali_yellow
        except:
            ws.cell('M%s'%(row)).style = ali_yellow
            next

# Сохраняем в ecxel  
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'3_tariffa' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def report_electric_simple_3_zones_v2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']    
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+' .Срез показаний с коэффициентами на дату' + ' ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктт'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F4:G4')
    ws['F4'] = 'Сумма'
    ws['F4'].style = ali_grey
    ws['G4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['F5'] = 'Показания A+ на ' + electric_data_end
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Энергия A+ на ' + electric_data_end
    ws['G5'].style = ali_yellow
    
    # Тариф 1
    ws.merge_cells('H4:I4')
    ws['H4'] = 'Тариф 1'
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['H5'] = 'Показания A+ на ' +  electric_data_end
    ws['H5'].style = ali_grey
    
    ws['I5'] = 'Энергия A+ на ' +  electric_data_end
    ws['I5'].style = ali_yellow
    
    # Тариф 2
    ws.merge_cells('J4:K4')
    ws['J4'] = 'Тариф 2'
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['J5'] = 'Показания A+ на ' +  electric_data_end
    ws['J5'].style = ali_grey
    
    ws['K5'] = 'Энергия A+ на ' +  electric_data_end
    ws['K5'].style = ali_yellow
    
    # Тариф 3
    ws.merge_cells('L4:M4')
    ws['L4'] = 'Тариф 3'
    ws['L4'].style = ali_grey
    ws['M4'].style = ali_grey
    ws['L4'].style = ali_grey
    ws['M4'].style = ali_grey
    ws['L5'] = 'Показания A+ на ' +  electric_data_end
    ws['L5'].style = ali_grey
    
    ws['M5'] = 'Энергия A+ на ' +  electric_data_end
    ws['M5'].style = ali_yellow
         
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
    #выборка данных из БД
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    
    obj_parent_title    = request.session['obj_parent_title']
            
    obj_key             = request.session['obj_key']
    is_electric_monthly = request.session['is_electric_monthly']
    is_electric_daily   = request.session['is_electric_daily']
    is_electric_period  = request.session['is_electric_period']
    data_table = []
    if True:
        if True:            
            if (is_electric_monthly == '1') & (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'monthly')
                
            elif (is_electric_daily == '1') & (is_electric_period == "0") & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'daily')

#*********************************************************************************************************************************************************************      
            elif (is_electric_monthly == '1') & (bool(is_object_level.search(obj_key))): # показания на начало месяца для объекта
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]        

#*********************************************************************************************************************************************************************
            elif (is_electric_daily == '1') & (bool(is_object_level.search(obj_key))): # daily for abonents group
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]
            elif (is_electric_daily == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'daily')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]
              
            elif (is_electric_monthly == '1') & (bool(is_group_level.search(obj_key))): # поиск по баланскной группе
                    data_table= common_sql.get_data_table_by_date_for_group_3_zones_v2(obj_title, electric_data_end, 'monthly')
                    if not data_table:
                        data_table = [[electric_data_end, obj_title, u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д', u'Н/Д']]

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][9])  # Ктн
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][8])  # Ктт
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][10])  # Ка
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
                   
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][3])  # Сумма А+
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % round((data_table[row-6][3]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Сумма А+
            ws.cell('G%s'%(row)).style = ali_yellow
        except:
            ws.cell('G%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][4])  # Тариф 1 А+
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % round((data_table[row-6][4]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Тариф 1 А+
            ws.cell('I%s'%(row)).style = ali_yellow
        except:
            ws.cell('I%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][5])  # Тариф 2 А+
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % round((data_table[row-6][5]*data_table[row-6][8]*data_table[row-6][9]),3) # "Энергия Тариф 2 А+
            ws.cell('K%s'%(row)).style = ali_yellow
        except:
            ws.cell('K%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table[row-6][6])  # Тариф 3 А+
            ws.cell('L%s'%(row)).style = ali_white

        except:
            ws.cell('L%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % round((data_table[row-6][6]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Тариф 3 А+
            ws.cell('M%s'%(row)).style = ali_yellow
        except:
            ws.cell('M%s'%(row)).style = ali_yellow
            next

# Сохраняем в ecxel  
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'3_tariffa_'+translate(obj_title)+'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def report_current_3_zones_v2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']    
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+' .Срез показаний с коэффициентами на дату' + ' ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктт'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F4:G4')
    ws['F4'] = 'Сумма'
    ws['F4'].style = ali_grey
    ws['G4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['F5'] = 'Показания A+ на ' + electric_data_end
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Энергия A+ на ' + electric_data_end
    ws['G5'].style = ali_yellow
    
    # Тариф 1
    ws.merge_cells('H4:I4')
    ws['H4'] = 'Тариф 1'
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['H4'].style = ali_grey
    ws['I4'].style = ali_grey
    ws['H5'] = 'Показания A+ на ' +  electric_data_end
    ws['H5'].style = ali_grey
    
    ws['I5'] = 'Энергия A+ на ' +  electric_data_end
    ws['I5'].style = ali_yellow
    
    # Тариф 2
    ws.merge_cells('J4:K4')
    ws['J4'] = 'Тариф 2'
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['J4'].style = ali_grey
    ws['K4'].style = ali_grey
    ws['J5'] = 'Показания A+ на ' +  electric_data_end
    ws['J5'].style = ali_grey
    
    ws['K5'] = 'Энергия A+ на ' +  electric_data_end
    ws['K5'].style = ali_yellow
    
    # Тариф 3
    ws.merge_cells('L4:M4')
    ws['L4'] = 'Тариф 3'
    ws['L4'].style = ali_grey
    ws['M4'].style = ali_grey
    ws['L4'].style = ali_grey
    ws['M4'].style = ali_grey
    ws['L5'] = 'Показания A+ на ' +  electric_data_end
    ws['L5'].style = ali_grey
    
    ws['M5'] = 'Энергия A+ на ' +  electric_data_end
    ws['M5'].style = ali_yellow
         
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
    #выборка данных из БД
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    
    obj_parent_title    = request.session['obj_parent_title']
            
    obj_key             = request.session['obj_key']
    is_electric_monthly = request.session['is_electric_monthly']
    is_electric_daily   = request.session['is_electric_daily']
    is_electric_period  = request.session['is_electric_period']
    data_table = []
    if True:
        if True:            
            if (bool(is_abonent_level.search(obj_key))):   # monthly for abonents
                data_table = common_sql.get_data_table_by_date_monthly_3_zones_v2(obj_title, obj_parent_title, electric_data_end, 'current')         
                            
            elif (bool(is_object_level.search(obj_key))): # показания на начало месяца для объекта
                    data_table= common_sql.get_data_table_by_date_for_object_3_zones_v2(obj_title, electric_data_end, 'current')

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][9])  # Ктн
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][8])  # Ктт
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][10])  # Ка
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
                   
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][3])  # Сумма А+
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % round((data_table[row-6][3]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Сумма А+
            ws.cell('G%s'%(row)).style = ali_yellow
        except:
            ws.cell('G%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][4])  # Тариф 1 А+
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % round((data_table[row-6][4]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Тариф 1 А+
            ws.cell('I%s'%(row)).style = ali_yellow
        except:
            ws.cell('I%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][5])  # Тариф 2 А+
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % round((data_table[row-6][5]*data_table[row-6][8]*data_table[row-6][9]),3) # "Энергия Тариф 2 А+
            ws.cell('K%s'%(row)).style = ali_yellow
        except:
            ws.cell('K%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table[row-6][6])  # Тариф 3 А+
            ws.cell('L%s'%(row)).style = ali_white

        except:
            ws.cell('L%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % round((data_table[row-6][6]*data_table[row-6][8]*data_table[row-6][9]),3)  # "Энергия Тариф 3 А+
            ws.cell('M%s'%(row)).style = ali_yellow
        except:
            ws.cell('M%s'%(row)).style = ali_yellow
            next

# Сохраняем в ecxel  
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'3_tariffa_'+translate(obj_title)+'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def electric_between_report(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Значения профиля показаний за период с' + ' '+str(request.session["electric_data_start"]) +' по '+ str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:C5')
    ws['C4'] = 'Дата'
    ws['C4'].style = ali_grey
    ws['C5'].style = ali_grey
    
    # Сумма
    ws.merge_cells('D4:D5')
    ws['D4'] = 'Сумма - Показания A+ '
    ws['D4'].style = ali_grey
    ws['D5'].style = ali_grey
 
    # Дельта
    ws.merge_cells('E4:E5')
    ws['E4'] = 'Сумма - Расход за прошедшие сутки'
    ws['E4'].style = ali_grey
    ws['E5'].style = ali_grey
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_electric_daily    = request.session['is_electric_daily']
    obj_parent_title    = request.session['obj_parent_title']
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
    obj_key             = request.session['obj_key']

    data_table = []
    if True:
        if True:            
            if (is_electric_daily == '1') & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end)
            else:
                pass
            
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][3])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][4])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # дата
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][5])  # сумма-показания
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][12])  # Расход за прошедшие сутки
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next

# Сохраняем в ecxel    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_electric' + str(electric_data_start) + u' - ' + str(electric_data_end)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response
    
def electric_between_2_zones_report(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Значения профиля показаний за период с' + ' '+str(request.session["electric_data_start"]) +' по '+ str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:C5')
    ws['C4'] = 'Дата'
    ws['C4'].style = ali_grey
    ws['C5'].style = ali_grey
    
    # Сумма
    ws.merge_cells('D4:D5')
    ws['D4'] = 'Сумма - Показания T0 A+ '
    ws['D4'].style = ali_grey
    ws['D5'].style = ali_grey
 
    # Дельта
    ws.merge_cells('E4:E5')
    ws['E4'] = 'Сумма - Расход за прошедшие сутки T0'
    ws['E4'].style = ali_grey
    ws['E5'].style = ali_grey
    
        # Сумма
    ws.merge_cells('F4:F5')
    ws['F4'] = 'Сумма - Показания T1 A+ '
    ws['F4'].style = ali_grey
    ws['F5'].style = ali_grey
 
    # Дельта
    ws.merge_cells('G4:G5')
    ws['G4'] = 'Сумма - Расход за прошедшие сутки T1'
    ws['G4'].style = ali_grey
    ws['G5'].style = ali_grey
    
        # Сумма
    ws.merge_cells('H4:H5')
    ws['H4'] = 'Сумма - Показания T2 A+ '
    ws['H4'].style = ali_grey
    ws['H5'].style = ali_grey
 
    # Дельта
    ws.merge_cells('I4:I5')
    ws['I4'] = 'Сумма - Расход за прошедшие сутки T2'
    ws['I4'].style = ali_grey
    ws['I5'].style = ali_grey
    
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_electric_daily    = request.session['is_electric_daily']
    obj_parent_title    = request.session['obj_parent_title']
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
    obj_key             = request.session['obj_key']

    data_table = []
    if True:
        if True:            
            if (is_electric_daily == '1') & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end)
            else:
                pass
            
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][3])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][4])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # дата
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][5])  # сумма-показания t0
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][12])  # Расход за прошедшие сутки t0
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # сумма-показанияt1
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][13])  # Расход за прошедшие суткиt1
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # сумма-показанияt2
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][14])  # Расход за прошедшие суткиt2
            ws.cell('I%s'%(row)).style = ali_white
        except:
            ws.cell('I%s'%(row)).style = ali_white
            next

# Сохраняем в ecxel    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_electric_2_zones_' + str(electric_data_start) + u' - ' + str(electric_data_end)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response
    
def electric_between_3_zones_report(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+'. Значения профиля показаний за период с' + ' '+electric_data_start +' по '+ electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:C5')
    ws['C4'] = 'Дата'
    ws['C4'].style = ali_grey
    ws['C5'].style = ali_grey
    
    # Сумма
    ws.merge_cells('D4:D5')
    ws['D4'] = 'Показания T0 A+ '
    ws['D4'].style = ali_grey
    ws['D5'].style = ali_grey
 
    # Дельта
    ws.merge_cells('E4:E5')
    ws['E4'] = 'Расход за прошедшие сутки T0'
    ws['E4'].style = ali_grey
    ws['E5'].style = ali_grey
    
        # Сумма
    ws.merge_cells('F4:F5')
    ws['F4'] = 'Показания T1 A+ '
    ws['F4'].style = ali_grey
    ws['F5'].style = ali_grey
 
    # Дельта
    ws.merge_cells('G4:G5')
    ws['G4'] = 'Расход за прошедшие сутки T1'
    ws['G4'].style = ali_grey
    ws['G5'].style = ali_grey
    
        # Сумма
    ws.merge_cells('H4:H5')
    ws['H4'] = 'Показания T2 A+ '
    ws['H4'].style = ali_grey
    ws['H5'].style = ali_grey
 
    # Дельта
    ws.merge_cells('I4:I5')
    ws['I4'] = 'Расход за прошедшие сутки T2'
    ws['I4'].style = ali_grey
    ws['I5'].style = ali_grey
    
        # Сумма
    ws.merge_cells('J4:J5')
    ws['J4'] = 'Показания T3 A+ '
    ws['J4'].style = ali_grey
    ws['J5'].style = ali_grey
 
    # Дельта
    ws.merge_cells('K4:K5')
    ws['K4'] = 'Расход за прошедшие сутки T3'
    ws['K4'].style = ali_grey
    ws['K5'].style = ali_grey
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_electric_daily    = request.session['is_electric_daily']
    obj_parent_title    = request.session['obj_parent_title']
    
    obj_key             = request.session['obj_key']

    data_table = []
    if True:
        if True:            
            if (is_electric_daily == '1') & (bool(is_abonent_level.search(obj_key))):   # daily for abonents
                data_table = common_sql.get_data_table_electric_between(obj_title, obj_parent_title,electric_data_start, electric_data_end)
            else:
                pass
            
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][3])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][4])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # дата
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][5])  # сумма-показания t0
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][12])  # Расход за прошедшие сутки t0
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # сумма-показанияt1
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][13])  # Расход за прошедшие суткиt1
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # сумма-показанияt2
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][14])  # Расход за прошедшие суткиt2
            ws.cell('I%s'%(row)).style = ali_white
        except:
            ws.cell('I%s'%(row)).style = ali_white
            next
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][8])  # сумма-показанияt3
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table[row-6][15])  # Расход за прошедшие суткиt3
            ws.cell('K%s'%(row)).style = ali_white
        except:
            ws.cell('K%s'%(row)).style = ali_white
            next

# Сохраняем в ecxel    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_electric_3_zones_'+translate(obj_title)+u'_' + electric_data_start + u' - ' +electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)    
    return response

def report_electric_potreblenie_2_zones(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление электроэнергии в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктт'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F3:I3')
    ws['F3'] = 'Сумма A+, кВт*ч'
    ws['F3'].style = ali_grey
    ws['G3'].style = ali_grey
    ws['H3'].style = ali_grey
    ws['I3'].style = ali_grey
    
    ws.merge_cells('F4:G4')
    ws['F4'] = 'На ' + str(request.session["electric_data_start"])
    ws['F4'].style = ali_grey
    
    ws.merge_cells('H4:I4')
    ws['H4'] = 'На ' + str(request.session["electric_data_end"])
    ws['H4'].style = ali_grey

    ws['F5'] = 'Показания'
    ws['F5'].style = ali_grey     
    ws['G5'] = 'Энергия'
    ws['G5'].style = ali_yellow
    
    ws['H5'] = 'Показания'
    ws['H5'].style = ali_grey     
    ws['I5'] = 'Энергия'
    ws['I5'].style = ali_yellow
    
    # Тариф 1
    ws.merge_cells('J3:M3')
    ws['J3'] = 'Тариф 1 A+, кВт*ч'
    ws['J3'].style = ali_grey
    ws['K3'].style = ali_grey
    ws['L3'].style = ali_grey
    ws['M3'].style = ali_grey
    
    ws.merge_cells('J4:K4')
    ws['J4'] = 'На ' + str(request.session["electric_data_start"])
    ws['J4'].style = ali_grey
    
    ws.merge_cells('L4:M4')
    ws['L4'] = 'На ' + str(request.session["electric_data_end"])
    ws['L4'].style = ali_grey

    ws['J5'] = 'Показания'
    ws['J5'].style = ali_grey     
    ws['K5'] = 'Энергия'
    ws['K5'].style = ali_yellow
    
    ws['L5'] = 'Показания'
    ws['L5'].style = ali_grey     
    ws['M5'] = 'Энергия'
    ws['M5'].style = ali_yellow
    
    # Тариф 2
    ws.merge_cells('N3:Q3')
    ws['N3'] = 'Тариф 2 A+, кВт*ч'
    ws['N3'].style = ali_grey
    ws['O3'].style = ali_grey
    ws['O3'].style = ali_grey
    ws['Q3'].style = ali_grey
    
    ws.merge_cells('N4:O4')
    ws['N4'] = 'На ' + str(request.session["electric_data_start"])
    ws['N4'].style = ali_grey
    
    ws.merge_cells('P4:Q4')
    ws['P4'] = 'На ' + str(request.session["electric_data_end"])
    ws['P4'].style = ali_grey

    ws['N5'] = 'Показания'
    ws['N5'].style = ali_grey     
    ws['O5'] = 'Энергия'
    ws['O5'].style = ali_yellow
    
    ws['P5'] = 'Показания'
    ws['P5'].style = ali_grey     
    ws['Q5'] = 'Энергия'
    ws['Q5'].style = ali_yellow
         
    # Расход
    ws.merge_cells('R3:W3')
    ws['R3'] = 'Расход А+, кВт*ч'
    ws['R3'].style = ali_grey
    ws['W3'].style = ali_grey
        # Расход Т0
    ws.merge_cells('R4:S4')
    ws['R4'] = 'Сумма'
    ws['R4'].style = ali_grey
    ws['R5'] = 'Показания'
    ws['R5'].style = ali_grey
    ws['S5'] = 'Энергия'
    ws['S5'].style = ali_yellow
        # Расход Т1
    ws.merge_cells('T4:U4')
    ws['T4'] = 'Tариф 1'
    ws['T4'].style = ali_grey
    ws['T5'] = 'Показания'
    ws['T5'].style = ali_grey
    ws['U5'] = 'Энергия'
    ws['U5'].style = ali_yellow
        # Расход Т2
    ws.merge_cells('V4:W4')
    ws['V4'] = 'Tариф 2'
    ws['V4'].style = ali_grey
    ws['V5'] = 'Показания'
    ws['V5'].style = ali_grey
    ws['W5'] = 'Энергия'
    ws['W5'].style = ali_yellow
  
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']                        
    obj_key             = request.session['obj_key']
    is_electric_monthly = request.session['is_electric_monthly']
    is_electric_daily   = request.session['is_electric_daily']
    data_table_end   = []
    data_table_start = []
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
                     
                 
    if (bool(is_abonent_level.search(obj_key))):
        if (is_electric_monthly == "1"):
            data_table_end   = common_sql.get_data_table_by_date_monthly_2_zones(meters_name, parent_name, electric_data_end)
            data_table_start = common_sql.get_data_table_by_date_monthly_2_zones(meters_name, parent_name, electric_data_start)
        elif (is_electric_daily == "1"):
            data_table_end   = common_sql.get_data_table_by_date_daily_2_zones(meters_name, parent_name, electric_data_end)
            data_table_start = common_sql.get_data_table_by_date_daily_2_zones(meters_name, parent_name, electric_data_start)


    elif (bool(is_object_level_2.search(obj_key))):
            list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
            data_table_end   = []
            data_table_start = []
            for x in range(len(list_of_abonents_2)):
                if (is_electric_monthly == "1"):                
                    data_table_temp_end = common_sql.get_data_table_by_date_monthly_2_zones(list_of_abonents_2[x], meters_name, electric_data_end)
                    data_table_temp_start = common_sql.get_data_table_by_date_monthly_2_zones(list_of_abonents_2[x], meters_name, electric_data_start)

                elif (is_electric_daily == "1"):                
                    data_table_temp_end = common_sql.get_data_table_by_date_daily_2_zones(list_of_abonents_2[x], meters_name, electric_data_end)
                    data_table_temp_start = common_sql.get_data_table_by_date_daily_2_zones(list_of_abonents_2[x], meters_name, electric_data_start)

                if data_table_temp_end and data_table_start:
                    data_table_end.extend(data_table_temp_end)
                    data_table_start.extend(data_table_temp_start)
                    
                else:
                    data_table_start.extend([[u'', list_of_abonents_2[x],u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д']])
                    data_table_end.extend([[u'', list_of_abonents_2[x],u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д']])
#Запрашиваем данные для отчета конец
                    
# Заполняем отчет значениями
    for row in range(6, len(data_table_end)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table_end[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table_end[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (get_k_t_n_by_serial_number(data_table_end[row-6][2]))  # Ктн
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Ктт
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (get_k_a_by_serial_number(data_table_end[row-6][2]))  # Ка
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
                   
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table_end[row-6][3])  # Сумма А+ на конец интервала
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table_end[row-6][3]*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Энергия Сумма А+ на конец интервала
            ws.cell('I%s'%(row)).style = ali_yellow
        except:
            ws.cell('I%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table_start[row-6][3])  # Сумма А+ на начало интервала
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table_start[row-6][3]*get_k_t_n_by_serial_number(data_table_start[row-6][2])*get_k_t_t_by_serial_number(data_table_start[row-6][2]))  # Энергия Сумма А+ на начало интервала
            ws.cell('G%s'%(row)).style = ali_yellow
        except:
            ws.cell('G%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table_end[row-6][4])  # Тариф 1 А+ на конец интервала
            ws.cell('L%s'%(row)).style = ali_white
        except:
            ws.cell('L%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % (data_table_end[row-6][4]*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Энергия Тариф 1 А+ на конец интервала
            ws.cell('M%s'%(row)).style = ali_yellow
        except:
            ws.cell('M%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table_start[row-6][4])  # Тариф 1 А+ на начало интервала
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table_start[row-6][4]*get_k_t_n_by_serial_number(data_table_start[row-6][2])*get_k_t_t_by_serial_number(data_table_start[row-6][2]))  # Энергия Тариф 1 А+ на начало интервала
            ws.cell('K%s'%(row)).style = ali_yellow
        except:
            ws.cell('K%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table_end[row-6][5])  # Тариф 2 А+ на конец интервала
            ws.cell('P%s'%(row)).style = ali_white
        except:
            ws.cell('P%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('Q%s'%(row)).value = '%s' % (data_table_end[row-6][5]*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Энергия Тариф 2 А+ на конец интервала
            ws.cell('Q%s'%(row)).style = ali_yellow
        except:
            ws.cell('Q%s'%(row)).style = ali_yellow
            next

        try:
            ws.cell('N%s'%(row)).value = '%s' % (data_table_start[row-6][5])  # Тариф 2 А+ на начало интервала
            ws.cell('N%s'%(row)).style = ali_white
        except:
            ws.cell('N%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table_start[row-6][5]*get_k_t_n_by_serial_number(data_table_start[row-6][2])*get_k_t_t_by_serial_number(data_table_start[row-6][2]))  # Энергия Тариф 2 А+ на начало интервала
            ws.cell('O%s'%(row)).style = ali_yellow
        except:
            ws.cell('O%s'%(row)).style = ali_yellow
            next
            
        # Расход
        try:
            ws.cell('R%s'%(row)).value = '%s' % (data_table_end[row-6][3] - data_table_start[row-6][3] )  # Расход Сумма А+
            ws.cell('R%s'%(row)).style = ali_white
        except:
            ws.cell('R%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('S%s'%(row)).value = '%s' % ((data_table_end[row-6][3] - data_table_start[row-6][3])*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Расход Сумма Энергия А+
            ws.cell('S%s'%(row)).style = ali_yellow
        except:
            ws.cell('S%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('T%s'%(row)).value = '%s' % (data_table_end[row-6][4] - data_table_start[row-6][4] )  # Расход Тариф 1 А+
            ws.cell('T%s'%(row)).style = ali_white
        except:
            ws.cell('T%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('U%s'%(row)).value = '%s' % ((data_table_end[row-6][4] - data_table_start[row-6][4])*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Расход Тариф 1 Энергия А+
            ws.cell('U%s'%(row)).style = ali_yellow
        except:
            ws.cell('U%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('V%s'%(row)).value = '%s' % (data_table_end[row-6][5] - data_table_start[row-6][5] )  # Расход Тариф 2 А+
            ws.cell('V%s'%(row)).style = ali_white
        except:
            ws.cell('V%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('W%s'%(row)).value = '%s' % ((data_table_end[row-6][5] - data_table_start[row-6][5])*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Расход Тариф 2 Энергия А+
            ws.cell('W%s'%(row)).style = ali_yellow
        except:
            ws.cell('W%s'%(row)).style = ali_yellow
            next            
# Конец наполнения отчёта
            
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'rashod_2_zones ' + str(electric_data_start) + u' - ' + str(electric_data_end)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_electric_potreblenie_3_zones(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление электроэнергии в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктт'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F3:I3')
    ws['F3'] = 'Сумма A+, кВт*ч'
    ws['F3'].style = ali_grey
    ws['G3'].style = ali_grey
    ws['H3'].style = ali_grey
    ws['I3'].style = ali_grey
    
    ws.merge_cells('F4:G4')
    ws['F4'] = 'На ' + str(request.session["electric_data_start"])
    ws['F4'].style = ali_grey
    
    ws.merge_cells('H4:I4')
    ws['H4'] = 'На ' + str(request.session["electric_data_end"])
    ws['H4'].style = ali_grey

    ws['F5'] = 'Показания'
    ws['F5'].style = ali_grey     
    ws['G5'] = 'Энергия'
    ws['G5'].style = ali_yellow
    
    ws['H5'] = 'Показания'
    ws['H5'].style = ali_grey     
    ws['I5'] = 'Энергия'
    ws['I5'].style = ali_yellow
    
    # Тариф 1
    ws.merge_cells('J3:M3')
    ws['J3'] = 'Тариф 1 A+, кВт*ч'
    ws['J3'].style = ali_grey
    ws['K3'].style = ali_grey
    ws['L3'].style = ali_grey
    ws['M3'].style = ali_grey
    
    ws.merge_cells('J4:K4')
    ws['J4'] = 'На ' + str(request.session["electric_data_start"])
    ws['J4'].style = ali_grey
    
    ws.merge_cells('L4:M4')
    ws['L4'] = 'На ' + str(request.session["electric_data_end"])
    ws['L4'].style = ali_grey

    ws['J5'] = 'Показания'
    ws['J5'].style = ali_grey     
    ws['K5'] = 'Энергия'
    ws['K5'].style = ali_yellow
    
    ws['L5'] = 'Показания'
    ws['L5'].style = ali_grey     
    ws['M5'] = 'Энергия'
    ws['M5'].style = ali_yellow
    
    # Тариф 2
    ws.merge_cells('N3:Q3')
    ws['N3'] = 'Тариф 2 A+, кВт*ч'
    ws['N3'].style = ali_grey
    ws['O3'].style = ali_grey
    ws['O3'].style = ali_grey
    ws['Q3'].style = ali_grey
    
    ws.merge_cells('N4:O4')
    ws['N4'] = 'На ' + str(request.session["electric_data_start"])
    ws['N4'].style = ali_grey
    
    ws.merge_cells('P4:Q4')
    ws['P4'] = 'На ' + str(request.session["electric_data_end"])
    ws['P4'].style = ali_grey

    ws['N5'] = 'Показания'
    ws['N5'].style = ali_grey     
    ws['O5'] = 'Энергия'
    ws['O5'].style = ali_yellow
    
    ws['P5'] = 'Показания'
    ws['P5'].style = ali_grey     
    ws['Q5'] = 'Энергия'
    ws['Q5'].style = ali_yellow
    
    # Тариф 3
    ws.merge_cells('R3:U3')
    ws['R3'] = 'Тариф 3 A+, кВт*ч'
    ws['R3'].style = ali_grey
    ws['S3'].style = ali_grey
    ws['T3'].style = ali_grey
    ws['U3'].style = ali_grey
    
    ws.merge_cells('R4:S4')
    ws['R4'] = 'На ' + str(request.session["electric_data_start"])
    ws['R4'].style = ali_grey
    
    ws.merge_cells('T4:U4')
    ws['T4'] = 'На ' + str(request.session["electric_data_end"])
    ws['T4'].style = ali_grey

    ws['R5'] = 'Показания'
    ws['R5'].style = ali_grey     
    ws['S5'] = 'Энергия'
    ws['S5'].style = ali_yellow
    
    ws['T5'] = 'Показания'
    ws['T5'].style = ali_grey     
    ws['U5'] = 'Энергия'
    ws['U5'].style = ali_yellow
         
    # Расход
    ws.merge_cells('V3:AC3')
    ws['V3'] = 'Расход А+, кВт*ч'
    ws['V3'].style = ali_grey
    ws['AC3'].style = ali_grey
        # Расход Т0
    ws.merge_cells('V4:W4')
    ws['V4'] = 'Сумма'
    ws['V4'].style = ali_grey
    ws['V5'] = 'Показания'
    ws['V5'].style = ali_grey
    ws['W5'] = 'Энергия'
    ws['W5'].style = ali_yellow
        # Расход Т1
    ws.merge_cells('X4:Y4')
    ws['X4'] = 'Tариф 1'
    ws['X4'].style = ali_grey
    ws['X5'] = 'Показания'
    ws['X5'].style = ali_grey
    ws['Y5'] = 'Энергия'
    ws['Y5'].style = ali_yellow
        # Расход Т2
    ws.merge_cells('Z4:AA4')
    ws['Z4'] = 'Tариф 2'
    ws['Z4'].style = ali_grey
    ws['Z5'] = 'Показания'
    ws['Z5'].style = ali_grey
    ws['AA5'] = 'Энергия'
    ws['AA5'].style = ali_yellow
        # Расход Т3
    ws.merge_cells('AB4:AC4')
    ws['AB4'] = 'Tариф 3'
    ws['AB4'].style = ali_grey
    ws['AC4'].style = ali_grey
    ws['AB5'] = 'Показания'
    ws['AB5'].style = ali_grey
    ws['AC5'] = 'Энергия'
    ws['AC5'].style = ali_yellow
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']                        
    obj_key             = request.session['obj_key']
    is_electric_monthly = request.session['is_electric_monthly']
    is_electric_daily   = request.session['is_electric_daily']
    data_table_end   = []
    data_table_start = []
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']
                     
                 
    if (bool(is_abonent_level.search(obj_key))):
        if (is_electric_monthly == "1"):
            data_table_end   = common_sql.get_data_table_by_date_monthly_3_zones(meters_name, parent_name, electric_data_end)
            data_table_start = common_sql.get_data_table_by_date_monthly_3_zones(meters_name, parent_name, electric_data_start)
        elif (is_electric_daily == "1"):
            data_table_end   = common_sql.get_data_table_by_date_daily_3_zones(meters_name, parent_name, electric_data_end)
            data_table_start = common_sql.get_data_table_by_date_daily_3_zones(meters_name, parent_name, electric_data_start)


    elif (bool(is_object_level_2.search(obj_key))):
            list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
            data_table_end   = []
            data_table_start = []
            for x in range(len(list_of_abonents_2)):
                if (is_electric_monthly == "1"):                
                    data_table_temp_end = common_sql.get_data_table_by_date_monthly_3_zones(list_of_abonents_2[x], meters_name, electric_data_end)
                    data_table_temp_start = common_sql.get_data_table_by_date_monthly_3_zones(list_of_abonents_2[x], meters_name, electric_data_start)

                elif (is_electric_daily == "1"):                
                    data_table_temp_end = common_sql.get_data_table_by_date_daily_3_zones(list_of_abonents_2[x], meters_name, electric_data_end)
                    data_table_temp_start = common_sql.get_data_table_by_date_daily_3_zones(list_of_abonents_2[x], meters_name, electric_data_start)

                if data_table_temp_end and data_table_start:
                    data_table_end.extend(data_table_temp_end)
                    data_table_start.extend(data_table_temp_start)
                    
                else:
                    data_table_start.extend([[u'', list_of_abonents_2[x],u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д']])
                    data_table_end.extend([[u'', list_of_abonents_2[x],u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д']])
#Запрашиваем данные для отчета конец
                    
# Заполняем отчет значениями
    for row in range(6, len(data_table_end)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table_end[row-6][1])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table_end[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (get_k_t_n_by_serial_number(data_table_end[row-6][2]))  # Ктн
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Ктт
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (get_k_a_by_serial_number(data_table_end[row-6][2]))  # Ка
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
                   
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table_end[row-6][3])  # Сумма А+ на конец интервала
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table_end[row-6][3]*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Энергия Сумма А+ на конец интервала
            ws.cell('I%s'%(row)).style = ali_yellow
        except:
            ws.cell('I%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table_start[row-6][3])  # Сумма А+ на начало интервала
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table_start[row-6][3]*get_k_t_n_by_serial_number(data_table_start[row-6][2])*get_k_t_t_by_serial_number(data_table_start[row-6][2]))  # Энергия Сумма А+ на начало интервала
            ws.cell('G%s'%(row)).style = ali_yellow
        except:
            ws.cell('G%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table_end[row-6][4])  # Тариф 1 А+ на конец интервала
            ws.cell('L%s'%(row)).style = ali_white
        except:
            ws.cell('L%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % (data_table_end[row-6][4]*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Энергия Тариф 1 А+ на конец интервала
            ws.cell('M%s'%(row)).style = ali_yellow
        except:
            ws.cell('M%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table_start[row-6][4])  # Тариф 1 А+ на начало интервала
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table_start[row-6][4]*get_k_t_n_by_serial_number(data_table_start[row-6][2])*get_k_t_t_by_serial_number(data_table_start[row-6][2]))  # Энергия Тариф 1 А+ на начало интервала
            ws.cell('K%s'%(row)).style = ali_yellow
        except:
            ws.cell('K%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table_end[row-6][5])  # Тариф 2 А+ на конец интервала
            ws.cell('P%s'%(row)).style = ali_white
        except:
            ws.cell('P%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('Q%s'%(row)).value = '%s' % (data_table_end[row-6][5]*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Энергия Тариф 2 А+ на конец интервала
            ws.cell('Q%s'%(row)).style = ali_yellow
        except:
            ws.cell('Q%s'%(row)).style = ali_yellow
            next

        try:
            ws.cell('N%s'%(row)).value = '%s' % (data_table_start[row-6][5])  # Тариф 2 А+ на начало интервала
            ws.cell('N%s'%(row)).style = ali_white
        except:
            ws.cell('N%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table_start[row-6][5]*get_k_t_n_by_serial_number(data_table_start[row-6][2])*get_k_t_t_by_serial_number(data_table_start[row-6][2]))  # Энергия Тариф 2 А+ на начало интервала
            ws.cell('O%s'%(row)).style = ali_yellow
        except:
            ws.cell('O%s'%(row)).style = ali_yellow
            next
            

            
        try:
            ws.cell('T%s'%(row)).value = '%s' % (data_table_end[row-6][6])  # Тариф 3 А+ на конец интервала
            ws.cell('T%s'%(row)).style = ali_white
        except:
            ws.cell('T%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('U%s'%(row)).value = '%s' % (data_table_end[row-6][6]*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # "Энергия Тариф 3 А+ на конец интервала
            ws.cell('U%s'%(row)).style = ali_yellow
        except:
            ws.cell('U%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('R%s'%(row)).value = '%s' % (data_table_start[row-6][6])  # Тариф 3 А+ на начало интервала
            ws.cell('R%s'%(row)).style = ali_white
        except:
            ws.cell('R%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('S%s'%(row)).value = '%s' % (data_table_start[row-6][6]*get_k_t_n_by_serial_number(data_table_start[row-6][2])*get_k_t_t_by_serial_number(data_table_start[row-6][2]))  # "Энергия Тариф 3 А+ на начало интервала
            ws.cell('S%s'%(row)).style = ali_yellow
        except:
            ws.cell('S%s'%(row)).style = ali_yellow
            next
        # Расход
        try:
            ws.cell('V%s'%(row)).value = '%s' % (data_table_end[row-6][3] - data_table_start[row-6][3] )  # Расход Сумма А+
            ws.cell('V%s'%(row)).style = ali_white
        except:
            ws.cell('V%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('W%s'%(row)).value = '%s' % ((data_table_end[row-6][3] - data_table_start[row-6][3])*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Расход Сумма Энергия А+
            ws.cell('W%s'%(row)).style = ali_yellow
        except:
            ws.cell('W%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('X%s'%(row)).value = '%s' % (data_table_end[row-6][4] - data_table_start[row-6][4] )  # Расход Тариф 1 А+
            ws.cell('X%s'%(row)).style = ali_white
        except:
            ws.cell('X%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('Y%s'%(row)).value = '%s' % ((data_table_end[row-6][4] - data_table_start[row-6][4])*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Расход Тариф 1 Энергия А+
            ws.cell('Y%s'%(row)).style = ali_yellow
        except:
            ws.cell('Y%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('Z%s'%(row)).value = '%s' % (data_table_end[row-6][5] - data_table_start[row-6][5] )  # Расход Тариф 2 А+
            ws.cell('Z%s'%(row)).style = ali_white
        except:
            ws.cell('Z%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('AA%s'%(row)).value = '%s' % ((data_table_end[row-6][5] - data_table_start[row-6][5])*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Расход Тариф 2 Энергия А+
            ws.cell('AA%s'%(row)).style = ali_yellow
        except:
            ws.cell('AA%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('AB%s'%(row)).value = '%s' % (data_table_end[row-6][6] - data_table_start[row-6][6] )  # Расход Тариф 3 А+
            ws.cell('AB%s'%(row)).style = ali_white
        except:
            ws.cell('AB%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('AC%s'%(row)).value = '%s' % ((data_table_end[row-6][6] - data_table_start[row-6][6])*get_k_t_n_by_serial_number(data_table_end[row-6][2])*get_k_t_t_by_serial_number(data_table_end[row-6][2]))  # Расход Тариф 3 Энергия А+
            ws.cell('AC%s'%(row)).style = ali_yellow
        except:
            ws.cell('AC%s'%(row)).style = ali_yellow
            next
# Конец наполнения отчёта
            
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'rashod_3_zones ' + str(electric_data_start) + u' - ' + str(electric_data_end)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_electric_potreblenie_3_zones_v2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = obj_title+'. Потребление электроэнергии в период с ' + electric_data_start + ' по ' + electric_data_end
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктт'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F3:I3')
    ws['F3'] = 'Сумма A+, кВт*ч'
    ws['F3'].style = ali_grey
    ws['G3'].style = ali_grey
    ws['H3'].style = ali_grey
    ws['I3'].style = ali_grey
    
    ws.merge_cells('F4:G4')
    ws['F4'] = 'На ' + str(request.session["electric_data_start"])
    ws['F4'].style = ali_grey
    
    ws.merge_cells('H4:I4')
    ws['H4'] = 'На ' + str(request.session["electric_data_end"])
    ws['H4'].style = ali_grey

    ws['F5'] = 'Показания'
    ws['F5'].style = ali_grey     
    ws['G5'] = 'Энергия'
    ws['G5'].style = ali_yellow
    
    ws['H5'] = 'Показания'
    ws['H5'].style = ali_grey     
    ws['I5'] = 'Энергия'
    ws['I5'].style = ali_yellow
    
    # Тариф 1
    ws.merge_cells('J3:M3')
    ws['J3'] = 'Тариф 1 A+, кВт*ч'
    ws['J3'].style = ali_grey
    ws['K3'].style = ali_grey
    ws['L3'].style = ali_grey
    ws['M3'].style = ali_grey
    
    ws.merge_cells('J4:K4')
    ws['J4'] = 'На ' + str(request.session["electric_data_start"])
    ws['J4'].style = ali_grey
    
    ws.merge_cells('L4:M4')
    ws['L4'] = 'На ' + str(request.session["electric_data_end"])
    ws['L4'].style = ali_grey

    ws['J5'] = 'Показания'
    ws['J5'].style = ali_grey     
    ws['K5'] = 'Энергия'
    ws['K5'].style = ali_yellow
    
    ws['L5'] = 'Показания'
    ws['L5'].style = ali_grey     
    ws['M5'] = 'Энергия'
    ws['M5'].style = ali_yellow
    
    # Тариф 2
    ws.merge_cells('N3:Q3')
    ws['N3'] = 'Тариф 2 A+, кВт*ч'
    ws['N3'].style = ali_grey
    ws['O3'].style = ali_grey
    ws['O3'].style = ali_grey
    ws['Q3'].style = ali_grey
    
    ws.merge_cells('N4:O4')
    ws['N4'] = 'На ' + str(request.session["electric_data_start"])
    ws['N4'].style = ali_grey
    
    ws.merge_cells('P4:Q4')
    ws['P4'] = 'На ' + str(request.session["electric_data_end"])
    ws['P4'].style = ali_grey

    ws['N5'] = 'Показания'
    ws['N5'].style = ali_grey     
    ws['O5'] = 'Энергия'
    ws['O5'].style = ali_yellow
    
    ws['P5'] = 'Показания'
    ws['P5'].style = ali_grey     
    ws['Q5'] = 'Энергия'
    ws['Q5'].style = ali_yellow
    
    # Тариф 3
    ws.merge_cells('R3:U3')
    ws['R3'] = 'Тариф 3 A+, кВт*ч'
    ws['R3'].style = ali_grey
    ws['S3'].style = ali_grey
    ws['T3'].style = ali_grey
    ws['U3'].style = ali_grey
    
    ws.merge_cells('R4:S4')
    ws['R4'] = 'На ' + str(request.session["electric_data_start"])
    ws['R4'].style = ali_grey
    
    ws.merge_cells('T4:U4')
    ws['T4'] = 'На ' + str(request.session["electric_data_end"])
    ws['T4'].style = ali_grey

    ws['R5'] = 'Показания'
    ws['R5'].style = ali_grey     
    ws['S5'] = 'Энергия'
    ws['S5'].style = ali_yellow
    
    ws['T5'] = 'Показания'
    ws['T5'].style = ali_grey     
    ws['U5'] = 'Энергия'
    ws['U5'].style = ali_yellow
         
    # Расход
    ws.merge_cells('V3:AC3')
    ws['V3'] = 'Расход А+, кВт*ч'
    ws['V3'].style = ali_grey
    ws['AC3'].style = ali_grey
        # Расход Т0
    ws.merge_cells('V4:W4')
    ws['V4'] = 'Сумма'
    ws['V4'].style = ali_grey
    ws['V5'] = 'Показания'
    ws['V5'].style = ali_grey
    ws['W5'] = 'Энергия'
    ws['W5'].style = ali_yellow
        # Расход Т1
    ws.merge_cells('X4:Y4')
    ws['X4'] = 'Tариф 1'
    ws['X4'].style = ali_grey
    ws['X5'] = 'Показания'
    ws['X5'].style = ali_grey
    ws['Y5'] = 'Энергия'
    ws['Y5'].style = ali_yellow
        # Расход Т2
    ws.merge_cells('Z4:AA4')
    ws['Z4'] = 'Tариф 2'
    ws['Z4'].style = ali_grey
    ws['Z5'] = 'Показания'
    ws['Z5'].style = ali_grey
    ws['AA5'] = 'Энергия'
    ws['AA5'].style = ali_yellow
        # Расход Т3
    ws.merge_cells('AB4:AC4')
    ws['AB4'] = 'Tариф 3'
    ws['AB4'].style = ali_grey
    ws['AC4'].style = ali_grey
    ws['AB5'] = 'Показания'
    ws['AB5'].style = ali_grey
    ws['AC5'] = 'Энергия'
    ws['AC5'].style = ali_yellow
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    
    obj_parent_title    = request.session['obj_parent_title']
    
    obj_key             = request.session['obj_key']
    is_electric_delta  = request.session['is_electric_delta']
    is_electric_monthly=request.session['is_electric_monthly']
    data_table = []
    if True:
        if True:                        
            res=u'Электричество'
            
            if (is_electric_monthly=="1"):
                dm='monthly'
            else:
                dm='daily'
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                    isAbon=True                    
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
                
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                    isAbon=False
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
            #*********************************************************************************************************************************************************************
            elif (is_electric_delta == '1') &(bool(is_group_level.search(obj_key))):
                    data_table=common_sql.get_data_table_electric_period_for_group(obj_title,obj_parent_title, electric_data_start, electric_data_end, res)
                    request.session["data_table_export"] = data_table
#Запрашиваем данные для отчета конец
                    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][23])  # Ктн
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][20])  # Ктт
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][24])  # Ка
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
                   
        try:
            ws.cell('H%s'%(row)).value = '%s' % (format(data_table[row-6][7],'.3f'))  # Сумма А+ на конец интервала
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (format(data_table[row-6][7]*data_table[row-6][20]*data_table[row-6][23],'.3f'))  # Энергия Сумма А+ на конец интервала
            ws.cell('I%s'%(row)).style = ali_yellow
        except:
            ws.cell('I%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s'%(format(data_table[row-6][2], '.3f'))  # '%s' % (data_table[row-6][2])  # Сумма А+ на начало интервала
            ws.cell('F%s'%(row)).style = ali_white
            #ws.cell('F%s'%(row)).number_format = '0.000'
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (format(data_table[row-6][2]*data_table[row-6][20]*data_table[row-6][23],'.3f'))  # Энергия Сумма А+ на начало интервала
            ws.cell('G%s'%(row)).style = ali_yellow
        except:
            ws.cell('G%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (format(data_table[row-6][8],'.3f'))  # Тариф 1 А+ на конец интервала
            ws.cell('L%s'%(row)).style = ali_white
        except:
            ws.cell('L%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % (format(data_table[row-6][8]*data_table[row-6][20]*data_table[row-6][23],'.3f'))  # Энергия Тариф 1 А+ на конец интервала
            ws.cell('M%s'%(row)).style = ali_yellow
        except:
            ws.cell('M%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (format(data_table[row-6][3],'.3f'))  # Тариф 1 А+ на начало интервала
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (format(data_table[row-6][3]*data_table[row-6][20]*data_table[row-6][23],'.3f'))  # Энергия Тариф 1 А+ на начало интервала
            ws.cell('K%s'%(row)).style = ali_yellow
        except:
            ws.cell('K%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (format(data_table[row-6][9],'.3f'))  # Тариф 2 А+ на конец интервала
            ws.cell('P%s'%(row)).style = ali_white
        except:
            ws.cell('P%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('Q%s'%(row)).value = '%s' % (format(data_table[row-6][9]*data_table[row-6][20]*data_table[row-6][23],'.3f'))  # Энергия Тариф 2 А+ на конец интервала
            ws.cell('Q%s'%(row)).style = ali_yellow
        except:
            ws.cell('Q%s'%(row)).style = ali_yellow
            next

        try:
            ws.cell('N%s'%(row)).value = '%s' % (format(data_table[row-6][4],'.3f'))  # Тариф 2 А+ на начало интервала
            ws.cell('N%s'%(row)).style = ali_white
        except:
            ws.cell('N%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (format(data_table[row-6][4]*data_table[row-6][20]*data_table[row-6][23],'.3f'))  # Энергия Тариф 2 А+ на начало интервала
            ws.cell('O%s'%(row)).style = ali_yellow
        except:
            ws.cell('O%s'%(row)).style = ali_yellow
            next
            

            
        try:
            ws.cell('T%s'%(row)).value = '%s' % (format(data_table[row-6][10],'.3f'))  # Тариф 3 А+ на конец интервала
            ws.cell('T%s'%(row)).style = ali_white
        except:
            ws.cell('T%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('U%s'%(row)).value = '%s' % (format(data_table[row-6][10]*data_table[row-6][20]*data_table[row-6][23],'.3f'))  # "Энергия Тариф 3 А+ на конец интервала
            ws.cell('U%s'%(row)).style = ali_yellow
        except:
            ws.cell('U%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('R%s'%(row)).value = '%s' % (format(data_table[row-6][5],'.3f'))  # Тариф 3 А+ на начало интервала
            ws.cell('R%s'%(row)).style = ali_white
        except:
            ws.cell('R%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('S%s'%(row)).value = '%s' % (format(data_table[row-6][5]*data_table[row-6][20]*data_table[row-6][23],'.3f'))  # "Энергия Тариф 3 А+ на начало интервала
            ws.cell('S%s'%(row)).style = ali_yellow
        except:
            ws.cell('S%s'%(row)).style = ali_yellow
            next
        # Расход
        try:
            ws.cell('V%s'%(row)).value = '%s' % (format(data_table[row-6][12],'.3f'))  # Расход Сумма А+
            ws.cell('V%s'%(row)).style = ali_white
            #ws.cell('V%s'%(row)).number_format = '0.000'
        except:
            ws.cell('V%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('W%s'%(row)).value = '%s' % (format(data_table[row-6][12]*data_table[row-6][20]*data_table[row-6][23],'.3f'))  # Расход Сумма Энергия А+
            ws.cell('W%s'%(row)).style = ali_yellow
            #ws.cell('W%s'%(row)).number_format = '0.000'
        except:
            ws.cell('W%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('X%s'%(row)).value = '%s' % (format(data_table[row-6][13],'.3f'))  # Расход Тариф 1 А+
            ws.cell('X%s'%(row)).style = ali_white
        except:
            ws.cell('X%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('Y%s'%(row)).value = '%s' % (format(data_table[row-6][13]*data_table[row-6][20]*data_table[row-6][23],'.3f'))  # Расход Тариф 1 Энергия А+
            ws.cell('Y%s'%(row)).style = ali_yellow
        except:
            ws.cell('Y%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('Z%s'%(row)).value = '%s' % (format(data_table[row-6][14],'.3f'))  # Расход Тариф 2 А+
            ws.cell('Z%s'%(row)).style = ali_white
        except:
            ws.cell('Z%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('AA%s'%(row)).value = '%s' % (format(data_table[row-6][14]*data_table[row-6][20]*data_table[row-6][23],'.3f'))  # Расход Тариф 2 Энергия А+
            ws.cell('AA%s'%(row)).style = ali_yellow
        except:
            ws.cell('AA%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('AB%s'%(row)).value = '%s' % (format(data_table[row-6][15],'.3f'))  # Расход Тариф 3 А+
            ws.cell('AB%s'%(row)).style = ali_white
        except:
            ws.cell('AB%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('AC%s'%(row)).value = '%s' % (format(data_table[row-6][15]*data_table[row-6][20]*data_table[row-6][23],'.3f'))  # Расход Тариф 3 Энергия А+
            ws.cell('AC%s'%(row)).style = ali_yellow
        except:
            ws.cell('AC%s'%(row)).style = ali_yellow
            next
# Конец наполнения отчёта
            
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'rashod_3_zones_'+translate(obj_title)+'_' + str(electric_data_start) + u'-' + str(electric_data_end)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_electric_potreblenie_2_zones_v2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
# Шапка отчета    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление электроэнергии в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    
    ws.merge_cells('A4:A5')
    ws['A4'] = 'Наименование канала'
    ws['A4'].style = ali_grey
    ws['A5'].style = ali_grey
    
    ws.merge_cells('B4:B5')
    ws['B4'] = 'Заводской номер'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = 'Коэффициенты'
    ws['C4'].style = ali_grey
    ws['D4'].style = ali_grey
    ws['E4'].style = ali_grey
    
    ws['C5'] = 'Ктн'
    ws['C5'].style = ali_grey
    ws['D5'] = 'Ктт'
    ws['D5'].style = ali_grey
    ws['E5'] = 'А'
    ws['E5'].style = ali_grey 
    
    # Сумма
    ws.merge_cells('F3:I3')
    ws['F3'] = 'Сумма A+, кВт*ч'
    ws['F3'].style = ali_grey
    ws['G3'].style = ali_grey
    ws['H3'].style = ali_grey
    ws['I3'].style = ali_grey
    
    ws.merge_cells('F4:G4')
    ws['F4'] = 'На ' + str(request.session["electric_data_start"])
    ws['F4'].style = ali_grey
    
    ws.merge_cells('H4:I4')
    ws['H4'] = 'На ' + str(request.session["electric_data_end"])
    ws['H4'].style = ali_grey

    ws['F5'] = 'Показания'
    ws['F5'].style = ali_grey     
    ws['G5'] = 'Энергия'
    ws['G5'].style = ali_yellow
    
    ws['H5'] = 'Показания'
    ws['H5'].style = ali_grey     
    ws['I5'] = 'Энергия'
    ws['I5'].style = ali_yellow
    
    # Тариф 1
    ws.merge_cells('J3:M3')
    ws['J3'] = 'Тариф 1 A+, кВт*ч'
    ws['J3'].style = ali_grey
    ws['K3'].style = ali_grey
    ws['L3'].style = ali_grey
    ws['M3'].style = ali_grey
    
    ws.merge_cells('J4:K4')
    ws['J4'] = 'На ' + str(request.session["electric_data_start"])
    ws['J4'].style = ali_grey
    
    ws.merge_cells('L4:M4')
    ws['L4'] = 'На ' + str(request.session["electric_data_end"])
    ws['L4'].style = ali_grey

    ws['J5'] = 'Показания'
    ws['J5'].style = ali_grey     
    ws['K5'] = 'Энергия'
    ws['K5'].style = ali_yellow
    
    ws['L5'] = 'Показания'
    ws['L5'].style = ali_grey     
    ws['M5'] = 'Энергия'
    ws['M5'].style = ali_yellow
    
    # Тариф 2
    ws.merge_cells('N3:Q3')
    ws['N3'] = 'Тариф 2 A+, кВт*ч'
    ws['N3'].style = ali_grey
    ws['O3'].style = ali_grey
    ws['O3'].style = ali_grey
    ws['Q3'].style = ali_grey
    
    ws.merge_cells('N4:O4')
    ws['N4'] = 'На ' + str(request.session["electric_data_start"])
    ws['N4'].style = ali_grey
    
    ws.merge_cells('P4:Q4')
    ws['P4'] = 'На ' + str(request.session["electric_data_end"])
    ws['P4'].style = ali_grey

    ws['N5'] = 'Показания'
    ws['N5'].style = ali_grey     
    ws['O5'] = 'Энергия'
    ws['O5'].style = ali_yellow
    
    ws['P5'] = 'Показания'
    ws['P5'].style = ali_grey     
    ws['Q5'] = 'Энергия'
    ws['Q5'].style = ali_yellow
    

         
    # Расход
    ws.merge_cells('R3:W3')
    ws['R3'] = 'Расход А+, кВт*ч'
    ws['R3'].style = ali_grey
    ws['W3'].style = ali_grey
        # Расход Т0
    ws.merge_cells('R4:S4')
    ws['R4'] = 'Сумма'
    ws['R4'].style = ali_grey
    ws['R5'] = 'Показания'
    ws['R5'].style = ali_grey
    ws['S5'] = 'Энергия'
    ws['S5'].style = ali_yellow
        # Расход Т1
    ws.merge_cells('T4:U4')
    ws['T4'] = 'Tариф 1'
    ws['T4'].style = ali_grey
    ws['T5'] = 'Показания'
    ws['T5'].style = ali_grey
    ws['U5'] = 'Энергия'
    ws['U5'].style = ali_yellow
        # Расход Т2
    ws.merge_cells('V4:W4')
    ws['V4'] = 'Tариф 2'
    ws['V4'].style = ali_grey
    ws['V5'] = 'Показания'
    ws['V5'].style = ali_grey
    ws['W5'] = 'Энергия'
    ws['W5'].style = ali_yellow
    
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17    
# Шапка отчета конец
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')
    is_group_level = re.compile(r'group')
    
    obj_parent_title    = request.session['obj_parent_title']
    obj_title           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
    obj_key             = request.session['obj_key']
    is_electric_delta  = request.session['is_electric_delta']
    is_electric_monthly=request.session['is_electric_monthly']
    data_table = []
    if True:
        if True:                        
            res=u'Электричество'
            
            if (is_electric_monthly=="1"):
                dm='monthly'
            else:
                dm='daily'
            if (is_electric_delta == "1") & (bool(is_abonent_level.search(obj_key))): # delta for abonents
                    isAbon=True                    
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
                
            elif (is_electric_delta == '1') & (bool(is_object_level.search(obj_key))): # daily delta for abonents group
                    isAbon=False
                    data_table=common_sql.get_data_table_electric_period(isAbon,obj_title,obj_parent_title, electric_data_start, electric_data_end, res, dm)
                    request.session["data_table_export"] = data_table
            #*********************************************************************************************************************************************************************
            elif (is_electric_delta == '1') &(bool(is_group_level.search(obj_key))):
                    data_table=common_sql.get_data_table_electric_period_for_group(obj_title,obj_parent_title, electric_data_start, electric_data_end, res)
                    request.session["data_table_export"] = data_table
#Запрашиваем данные для отчета конец
                    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Наименование канала
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][23])  # Ктн
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][20])  # Ктт
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][24])  # Ка
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
                   
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # Сумма А+ на конец интервала
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][7]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Сумма А+ на конец интервала
            ws.cell('I%s'%(row)).style = ali_yellow
        except:
            ws.cell('I%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][2])  # Сумма А+ на начало интервала
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][2]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Сумма А+ на начало интервала
            ws.cell('G%s'%(row)).style = ali_yellow
        except:
            ws.cell('G%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table[row-6][8])  # Тариф 1 А+ на конец интервала
            ws.cell('L%s'%(row)).style = ali_white
        except:
            ws.cell('L%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % (data_table[row-6][8]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Тариф 1 А+ на конец интервала
            ws.cell('M%s'%(row)).style = ali_yellow
        except:
            ws.cell('M%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][3])  # Тариф 1 А+ на начало интервала
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table[row-6][3]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Тариф 1 А+ на начало интервала
            ws.cell('K%s'%(row)).style = ali_yellow
        except:
            ws.cell('K%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table[row-6][9])  # Тариф 2 А+ на конец интервала
            ws.cell('P%s'%(row)).style = ali_white
        except:
            ws.cell('P%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('Q%s'%(row)).value = '%s' % (data_table[row-6][9]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Тариф 2 А+ на конец интервала
            ws.cell('Q%s'%(row)).style = ali_yellow
        except:
            ws.cell('Q%s'%(row)).style = ali_yellow
            next

        try:
            ws.cell('N%s'%(row)).value = '%s' % (data_table[row-6][4])  # Тариф 2 А+ на начало интервала
            ws.cell('N%s'%(row)).style = ali_white
        except:
            ws.cell('N%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table[row-6][4]*data_table[row-6][20]*data_table[row-6][23])  # Энергия Тариф 2 А+ на начало интервала
            ws.cell('O%s'%(row)).style = ali_yellow
        except:
            ws.cell('O%s'%(row)).style = ali_yellow
            next
            

        # Расход
        try:
            ws.cell('R%s'%(row)).value = '%s' % (data_table[row-6][12])  # Расход Сумма А+
            ws.cell('R%s'%(row)).style = ali_white
        except:
            ws.cell('R%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('S%s'%(row)).value = '%s' % (data_table[row-6][12]*data_table[row-6][20]*data_table[row-6][23])  # Расход Сумма Энергия А+
            ws.cell('S%s'%(row)).style = ali_yellow
        except:
            ws.cell('S%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('T%s'%(row)).value = '%s' % (data_table[row-6][13])  # Расход Тариф 1 А+
            ws.cell('T%s'%(row)).style = ali_white
        except:
            ws.cell('T%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('U%s'%(row)).value = '%s' % (data_table[row-6][13]*data_table[row-6][20]*data_table[row-6][23])  # Расход Тариф 1 Энергия А+
            ws.cell('U%s'%(row)).style = ali_yellow
        except:
            ws.cell('U%s'%(row)).style = ali_yellow
            next
            
        try:
            ws.cell('V%s'%(row)).value = '%s' % (data_table[row-6][14])  # Расход Тариф 2 А+
            ws.cell('V%s'%(row)).style = ali_white
        except:
            ws.cell('V%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('W%s'%(row)).value = '%s' % (data_table[row-6][14]*data_table[row-6][20]*data_table[row-6][23])  # Расход Тариф 2 Энергия А+
            ws.cell('W%s'%(row)).style = ali_yellow
        except:
            ws.cell('W%s'%(row)).style = ali_yellow
            next

# Конец наполнения отчёта
            
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'rashod_2_zones ' + str(electric_data_start) + u' - ' + str(electric_data_end)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response


def pokazaniya_heat_report(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Показания теплосчётчиков на ' + str(request.session["electric_data_end"])
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Заводской номер'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания, Гкал'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания, м3'
    ws['D5'].style = ali_grey

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']                        
    obj_key             = request.session['obj_key']

    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            #request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            #request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']  

    list_except = [u'ВРУ Счётчик01',u'ВРУ Счётчик02',u'ВРУ Счётчик03',u'ВРУ Счётчик04',u'ВРУ Счётчик05',u'ВРУ Счётчик06',u'ВРУ Счётчик07',u'ВРУ Счётчик08',u'ВРУ Счётчик09',u'ВРУ Счётчик10',u'ВРУ Счётчик11',u'ВРУ Счётчик12',u'ВРУ Счётчик13',u'ВРУ Счётчик14',u'ВРУ Счётчик15',u'ВРУ Счётчик16',u'ВРУ Счётчик17',u'ВРУ Счётчик18',u'ВРУ Счётчик19',u'ВРУ Счётчик20',u'ВРУ Счётчик21',u'ВРУ Счётчик22',u'ВРУ Счётчик23',u'Гараж Счётчик 1',u'Гараж Счётчик 2']
                     
    if (bool(is_abonent_level.search(obj_key))):     
        data_table = common_sql.get_data_table_by_date_heat(meters_name, parent_name, electric_data_end)

    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []
        for x in range(len(list_of_abonents_2)):
            data_table_temp = common_sql.get_data_table_by_date_heat(list_of_abonents_2[x], meters_name, electric_data_end)

            if list_of_abonents_2[x][0] in list_except:
                next
            elif data_table_temp:            
                data_table.extend(data_table_temp)
            else:
                data_table.extend([[0,list_of_abonents_2[x][0],u'Н/Д',u'Н/Д',u'Н/Д']])
                             
    else:
        data_table = []

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по расходу воды
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17 
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'heat_report' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def pokazaniya_heat_report_v2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Показания теплосчётчиков на ' + str(request.session["electric_data_end"])
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Заводской номер'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания, Гкал'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания, м3'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Время работы, ч'
    ws['E5'].style = ali_grey

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']

    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            #request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            #request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']  

    list_except = []
                     
    data_table = []
    list_except = []
    if (bool(is_abonent_level.search(obj_key))):     
        data_table = common_sql.get_data_table_by_date_heat_v2(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_heat_v2(meters_name, parent_name, electric_data_end, False)
        for row in data_table:
            for x in list_except:
                if x==row[2]:
                    data_table.remove(x)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по расходу воды
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5])  # время работы
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17 
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'heat_report_'+translate(parent_name)+'_'+translate(meters_name)+'-'+electric_data_end

    #output_name = u'heat_report' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pokazaniya_sayany(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    meters_name           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = meters_name+'. Показания теплосчётчиков Саяны на ' + electric_data_end 
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Номер счётчика'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Дата'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания Q1, Гкал'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания M1, т'
    ws['E5'].style = ali_grey
    
    ws['F5'] = u't1, C˚'
    ws['F5'].style = ali_grey
    
    ws['G5'] = u't2, C˚'
    ws['G5'].style = ali_grey

# ниже не переделывала
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')

    
    parent_name    = request.session['obj_parent_title']
    
    obj_key             = request.session['obj_key']

    
    #print parent_name,meters_name,electric_data_end, obj_key
    
    data_table = []
#    if request.is_ajax():
#        #if request.method == 'GET':
#            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
#            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
#            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
#            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # Дата
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу Q1
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу Q2
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания по теплу t1
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания по теплу t2
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17 
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'heat_sayany_report_'+translate(meters_name)+'_'+electric_data_end 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pokazaniya_sayany_archive(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Показания теплосчётчиков Саяны на ' + str(request.session["electric_data_end"])
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Номер счётчика'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Дата'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания Q1'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания Q2'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 't1'
    ws['F5'].style = ali_grey
    
    ws['G5'] = 't2'
    ws['G5'].style = ali_grey

# ниже не переделывала
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')

    
    parent_name    = request.session['obj_parent_title']
    meters_name           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']

    
    data_table = []

    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # Дата
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу Q1
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу Q2
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания по теплу t1
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания по теплу t2
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17 

    import zipfile

    o=StringIO.StringIO()
    
    zf = zipfile.ZipFile(response, mode='w', compression=zipfile.ZIP_DEFLATED)
    zf.writestr('README.txt', 'test msg')
    wb.save(o)
    zf.writestr('test.xlsx',o.getvalue())
    zf.close()
    response=HttpResponse(response.getvalue())
    response['Content-Type'] = 'application/x-zip-compressed'
    response['Content-Disposition'] = "attachment; filename=\"sayani_test.zip\""
    
    return response
    
def report_sayany_last(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

    meters_name           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = meters_name+'. Показания теплосчётчиков Саяны на ' + electric_data_end
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Номер счётчика'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Дата'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания Q1, Гкал'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания М1, т'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 't1, C˚'
    ws['F5'].style = ali_grey
    
    ws['G5'] = 't2, C˚'
    ws['G5'].style = ali_grey

# ниже не переделывала
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')

    
    parent_name    = request.session['obj_parent_title']
    meters_name           = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']

    
    data_table = []
#    if request.is_ajax():
#        #if request.method == 'GET':
#            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
#            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
#            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
#            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_heat_sayany_v2(meters_name, parent_name, electric_data_end, False)
    
    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        if (data_table[i][3] is None):
            data_table[i][0]=electric_data_end
            dt=common_sql.get_data_table_by_date_heat_sayany_v2(data_table[i][1], meters_name, None, True)
            if (len(dt)>0):
                data_table[i]=dt[0]
        data_table[i]=tuple(data_table[i])
    
    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][0])  # Дата
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу Q1
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу Q2
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания по теплу t1
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания по теплу t2
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 17 
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'heat_sayany_report_'+translate(meters_name)+'_' +electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_heat_potreblenie_sayany(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
#Шапка
    ws.merge_cells('A2:G2')
    ws['A2'] = meters_name+'. Потребление по теплосчётчикам Sayany в период с ' + electric_data_start + ' по ' +electric_data_end
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания Q1 на '  + electric_data_start+u', Гкал'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания Q1 на '  + electric_data_end+u', Гкал'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Потребление Q1, Гкал'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'Показания M1 на '  + electric_data_start+u', т'
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Показания M1 на '  + electric_data_end+u', т'
    ws['G5'].style = ali_grey
    
    ws['H5'] = 'Потребление M1, т'
    ws['H5'].style = ali_grey

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
                        
    obj_key             = request.session['obj_key']

    data_table = []
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_period_heat_sayany(meters_name, parent_name,electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_period_heat_sayany(meters_name, parent_name,electric_data_start, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания Q1 на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания Q1 на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление Q1
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания M1 на начало
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания M1 на конец
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # Потребление M1
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
#    ws.column_dimensions['C'].width = 35
#    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['H'].width = 18
#    ws.column_dimensions['F'].width = 18
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'potreblenie_heat_report_'+translate(meters_name)+u'_'+electric_data_start+u'-'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_water_potreblenie_pulsar(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
    meters_name         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')
    electric_data_start   = request.GET.get('electric_data_start')
    
#Шапка
    ws.merge_cells('A2:G2')
    ws['A2'] = meters_name+'. Потребление по импульсным водосчётчикам в период с ' + electric_data_start + ' по ' +electric_data_end
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Тип ресурса '
    ws['C5'].style = ali_grey
    
    ws['d5'] = 'Показания на '  + electric_data_start+u', м3'
    ws['d5'].style = ali_grey
    
    ws['e5'] = 'Показания на '  + electric_data_end+u', м3'
    ws['e5'].style = ali_grey
    
    ws['f5'] = 'Потребление, м3'
    ws['f5'].style = ali_grey
    
    ws['g5'] = 'Лицевой номер '
    ws['g5'].style = ali_grey    
    



    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'level2')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level1')

    parent_name         = request.GET.get('obj_parent_title')
    obj_key             = request.GET.get('obj_key')
    data_table=[]
    
    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_water_period_pulsar(meters_name, parent_name,electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_water_period_pulsar(meters_name, parent_name,electric_data_start, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        num=data_table[i][3]
        if ('ХВС, №' in num) or ('ГВС, №' in num):
            num=num.replace(u'ХВС, №', ' ')
            num=num.replace(u'ГВС, №', ' ')
            data_table[i][3]=num
        data_table[i]=tuple(data_table[i])

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][3])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][4])  # Тип ресурса
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания на начало
            ws.cell('d%s'%(row)).style = ali_white
        except:
            ws.cell('d%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания  на конец
            ws.cell('e%s'%(row)).style = ali_white
        except:
            ws.cell('e%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('f%s'%(row)).value = '%s' % (data_table[row-6][7])  # Потребление
            ws.cell('f%s'%(row)).style = ali_white
        except:
            ws.cell('f%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('g%s'%(row)).value = '%s' % (data_table[row-6][1])  # Лицевой номер
            ws.cell('g%s'%(row)).style = ali_white
        except:
            ws.cell('g%s'%(row)).style = ali_white
            next


    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    #ws.column_dimensions['H'].width = 25
    ws.column_dimensions['F'].width = 18
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'potreblenie_water_report_'+translate(meters_name)+u'_'+electric_data_start+u'-'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def pokazaniya_water_current_report(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Последние считанные показания по ХВС и ГВС на ' + str(datetime.now())
    
    ws['A5'] = 'Дата'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Время'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Абонент'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Заводской номер счётчика'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания ХВС, м3'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'Показания ГВС, м3'
    ws['F5'].style = ali_grey

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']

    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
    
    data_table=[]
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_current_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end,  True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table_temp=common_sql.get_current_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end,  False)
        for row in data_table_temp:
            if row[4]==u'Н/Д' and row[5]==u'Н/Д':
                row2=common_sql.get_current_water_gvs_hvs(unicode(row[2]), unicode(row[6]) , electric_data_end, True)
                if len(row2)==0:
                    r=[unicode(electric_data_end), u'Н/Д', unicode(row[2]),unicode(row[3]), u'Н/Д', u'Н/Д']
                    data_table.append(r)
                else:
                    data_table.append(row2[0])
            else:
                data_table.append(row)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Дата
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # время
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # абонент
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # заводской номер
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # хвс
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # гвс
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 17 
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'water_report_current' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def pokazaniya_water_daily_report(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Показания импульсные по ХВС и ГВС за ' + str(request.session["electric_data_end"])
    
    ws['A5'] = 'Дата'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Абонент'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Заводской номер счётчика'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания ХВС, м3'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания ГВС, м3'
    ws['E5'].style = ali_grey

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = obj_parent_title         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end, 'daily', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table=common_sql.get_daily_water_gvs_hvs(obj_title, obj_parent_title , electric_data_end, 'daily', False)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Дата
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  #  абонент
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  #хвс
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # гвс
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next

    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25 
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'water_report' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

#_________________________________-

def pokazaniya_heat_current_report(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Показания теплосчётчиков. Последние считанные данные'
    

    ws['A5'] = 'Дата'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Время'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Абонент'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Заводской номер'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания, Гкал'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'Показания, м3'
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Время работы, ч'
    ws['G5'].style = ali_grey
    
    ws['H5'] = 'Твхода, С'
    ws['H5'].style = ali_grey

    ws['I5'] = 'Твыхода, С'
    ws['I5'].style = ali_grey
    
    ws['J5'] = 'Разница Т, С'
    ws['J5'].style = ali_grey
    
    ws['K5'] = 'Код ошибки'
    ws['K5'].style = ali_grey
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    #electric_data_end   = request.session['electric_data_end']
    #electric_data_start   = request.session['electric_data_start']                        
    obj_key             = request.session['obj_key']
    #is_electric_monthly = request.session['is_electric_monthly']
    #is_electric_daily   = request.session['is_electric_daily']
    #data_table_end   = []
    #data_table_start = []
    list_except = [u'ВРУ Счётчик01',u'ВРУ Счётчик02',u'ВРУ Счётчик03',u'ВРУ Счётчик04',u'ВРУ Счётчик05',u'ВРУ Счётчик06',u'ВРУ Счётчик07',u'ВРУ Счётчик08',u'ВРУ Счётчик09',u'ВРУ Счётчик10',u'ВРУ Счётчик11',u'ВРУ Счётчик12',u'ВРУ Счётчик13',u'ВРУ Счётчик14',u'ВРУ Счётчик15',u'ВРУ Счётчик16',u'ВРУ Счётчик17',u'ВРУ Счётчик18',u'ВРУ Счётчик19',u'ВРУ Счётчик20',u'ВРУ Счётчик21',u'ВРУ Счётчик22',u'ВРУ Счётчик23',u'Гараж Счётчик 1',u'Гараж Счётчик 2']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            #request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']
            #request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
            #request.session["is_electric_monthly"] = is_electric_monthly = request.GET['is_electric_monthly']
            #request.session["is_electric_daily"]   = is_electric_daily   = request.GET['is_electric_daily']  

    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_current_heat(meters_name, parent_name)
    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []
        for x in range(len(list_of_abonents_2)):
            data_table_temp = common_sql.get_data_table_current_heat(list_of_abonents_2[x], parent_name)
            
            if list_of_abonents_2[x][0] in list_except:
                next
            elif data_table_temp:            
                data_table.extend(data_table_temp)
            else:
                data_table.extend([[u'Н/Д',u'Н/Д',list_of_abonents_2[x][0],u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д']])
                             
    else:
        data_table = []

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по расходу воды
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Абонент
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # заводской номер
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания по теплу
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # Показания по расходу воды
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][8])  # заводской номер
            ws.cell('I%s'%(row)).style = ali_white
        except:
            ws.cell('I%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][9])  # Показания по теплу
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table[row-6][10])  # Показания по расходу воды
            ws.cell('K%s'%(row)).style = ali_white
        except:
            ws.cell('K%s'%(row)).style = ali_white

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'heat_current_report' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

#_________________________________________________________________________
def report_all_res_by_date(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]

#Шапка
    ws['A1'] = 'Лицевой счет'
    ws['A1'].style = ali_grey
    
    ws['B1'] = 'Дата начала работы'
    ws['B1'].style = ali_grey
    
    ws['C1'] = 'Номер прибора'
    ws['C1'].style = ali_grey
    
    ws['D1'] = 'тип прибора'
    ws['D1'].style = ali_grey
    
    ws['E1'] = 'Наименование ПУ'
    ws['E1'].style = ali_grey
    
    ws['F1'] = 'Показания'
    ws['F1'].style = ali_grey
    
    ws['G1'] = 'Дата Окончания' 
    ws['G1'].style = ali_grey
    
    ws['H1'] = 'Дата'
    ws['H1'].style = ali_grey

    ws['I1'] = 'Квартирный' 
    ws['I1'].style = ali_grey
    
    ws['J1'] = 'Адрес'
    ws['J1'].style = ali_grey
    
    ws['K1'] = 'Операция'
    ws['K1'].style = ali_grey
    
    ws['L1'] = 'Заменяемый счетчик'
    ws['L1'].style = ali_grey
    
    ws['M1'] = 'Показания заменяемого счетчика'
    ws['M1'].style = ali_grey

    ws['N1'] = 'Улица'
    ws['N1'].style = ali_grey

    ws['O1'] = 'Дом'
    ws['O1'].style = ali_grey

    ws['P1'] = 'Квартира'
    ws['P1'].style = ali_grey
#Запрашиваем данные для отчета

           
    
    data_table = []
    data_table = common_sql.get_data_table_report_all_res_by_date(electric_data_end)

    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        for j in range(1,len(data_table[i])):
            if (data_table[i][j] == None) or (data_table[i][j] is None):
                data_table[i][j]=u''


# Заполняем отчет значениями
    for row in range(2, len(data_table)+2):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-2][0])  # лицевой счёт
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            a=str(data_table[row-2][1])
            b=datetime.datetime.strptime(a,'%Y-%m-%d')
            d=datetime.datetime.strftime(b,"%d.%m.%Y")
            ws.cell('B%s'%(row)).value = '%s' % d  # начальная дата
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-2][2])  # номер счётчика
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-2][3])  # тип энергоресурса
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-2][4])  # наименование ПУ
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-2][5])  # показания на начальную дату
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % electric_data_end  # дата снятия показаний
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
                        
        try:
            ws.cell('K%s'%(row)).value = 'ВводПоказаний'
            ws.cell('K%s'%(row)).style = ali_white
        except:
            ws.cell('K%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table[row-2][8])  # Дом-корпус
            ws.cell('O%s'%(row)).style = ali_white
        except:
            ws.cell('O%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table[row-2][7])  # № квартиры
            ws.cell('P%s'%(row)).style = ali_white
        except:
            ws.cell('P%s'%(row)).style = ali_white
            next
            

#    ws.row_dimensions[5].height = 41
#    ws.column_dimensions['A'].width = 10 
#    ws.column_dimensions['B'].width = 10 
#    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['j'].width = 15
    ws.column_dimensions['k'].width = 15
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_all_resources_'+str(electric_data_end)
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_electric_all_by_date(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]

#Шапка
    
    ws['A1'] = 'Лицевой счет'
    ws['A1'].style = ali_grey
    
    ws['B1'] = 'Дата начала работы'
    ws['B1'].style = ali_grey
    
    ws['C1'] = 'Номер прибора'
    ws['C1'].style = ali_grey
    
    ws['D1'] = 'тип прибора'
    ws['D1'].style = ali_grey
    
    ws['E1'] = 'Наименование ПУ'
    ws['E1'].style = ali_grey
    
    ws['F1'] = 'Показания'
    ws['F1'].style = ali_grey
    
    ws['G1'] = 'Дата Окончания' 
    ws['G1'].style = ali_grey
    
    ws['H1'] = 'Дата'
    ws['H1'].style = ali_grey

    ws['I1'] = 'Квартирный' 
    ws['I1'].style = ali_grey
    
    ws['J1'] = 'Адрес'
    ws['J1'].style = ali_grey
    
    ws['K1'] = 'Операция'
    ws['K1'].style = ali_grey
    
    ws['L1'] = 'Заменяемый счетчик'
    ws['L1'].style = ali_grey
    
    ws['M1'] = 'Показания заменяемого счетчика'
    ws['M1'].style = ali_grey

    ws['N1'] = 'Улица'
    ws['N1'].style = ali_grey

    ws['O1'] = 'Дом'
    ws['O1'].style = ali_grey

    ws['P1'] = 'Квартира'
    ws['P1'].style = ali_grey
#Запрашиваем данные для отчета

           
    
    data_table = []
    data_table = common_sql.get_data_table_report_electric_res_by_date(electric_data_end)

    #zamenyem None na N/D vezde
#    if len(data_table)>0: 
#        data_table=common_sql.ChangeNull(data_table, None)
    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        for j in range(1,len(data_table[i])):
            if (data_table[i][j] == None) or (data_table[i][j] is None):
                data_table[i][j]=u''
        data_table[i]=tuple(data_table[i])


# Заполняем отчет значениями
    for row in range(2, len(data_table)+2):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-2][0])  # лицевой счёт
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            a=str(data_table[row-2][1])
            b=datetime.datetime.strptime(a,'%Y-%m-%d')
            d=datetime.datetime.strftime(b,"%d.%m.%Y")

            ws.cell('B%s'%(row)).value = '%s' %d  # начальная дата
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-2][2])  # номер счётчика
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-2][3])  # тип энергоресурса
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-2][9])  # наименование ПУ
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-2][5])  # показания на начальную дату
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            #a=str(data_table[row-2][6])
            a=str(electric_data_end)
            #b=datetime.datetime.strptime(a,'%Y-%m-%d')
            #d=datetime.datetime.strftime(b,"%d.%m.%Y")
            ws.cell('H%s'%(row)).value = '%s' % a  # дата снятия показаний
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = 'ВводПоказаний'
            ws.cell('K%s'%(row)).style = ali_white
        except:
            ws.cell('K%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table[row-2][8])  # Дом-корпус
            ws.cell('O%s'%(row)).style = ali_white
        except:
            ws.cell('O%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table[row-2][7])  # № квартиры
            ws.cell('P%s'%(row)).style = ali_white
        except:
            ws.cell('P%s'%(row)).style = ali_white
            next
            

#    ws.row_dimensions[5].height = 41
#    ws.column_dimensions['A'].width = 10 
#    ws.column_dimensions['B'].width = 10 
#    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['j'].width = 15
    ws.column_dimensions['k'].width = 15
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_electric_resources_'+str(electric_data_end)
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_water_all_by_date(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]

#Шапка
    
    ws['A1'] = 'Лицевой счет'
    ws['A1'].style = ali_grey
    
    ws['B1'] = 'Дата начала работы'
    ws['B1'].style = ali_grey
    
    ws['C1'] = 'Номер прибора'
    ws['C1'].style = ali_grey
    
    ws['D1'] = 'тип прибора'
    ws['D1'].style = ali_grey
    
    ws['E1'] = 'Наименование ПУ'
    ws['E1'].style = ali_grey
    
    ws['F1'] = 'Показания'
    ws['F1'].style = ali_grey
    
    ws['G1'] = 'Дата Окончания' 
    ws['G1'].style = ali_grey
    
    ws['H1'] = 'Дата'
    ws['H1'].style = ali_grey

    ws['I1'] = 'Квартирный' 
    ws['I1'].style = ali_grey
    
    ws['J1'] = 'Адрес'
    ws['J1'].style = ali_grey
    
    ws['K1'] = 'Операция'
    ws['K1'].style = ali_grey
    
    ws['L1'] = 'Заменяемый счетчик'
    ws['L1'].style = ali_grey
    
    ws['M1'] = 'Показания заменяемого счетчика'
    ws['M1'].style = ali_grey

    ws['N1'] = 'Улица'
    ws['N1'].style = ali_grey

    ws['O1'] = 'Дом'
    ws['O1'].style = ali_grey

    ws['P1'] = 'Квартира'
    ws['P1'].style = ali_grey
#Запрашиваем данные для отчета

           
    
    data_table = []
    data_table = common_sql.get_data_table_report_water_res_by_date(electric_data_end)

    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        for j in range(1,len(data_table[i])):
            if (data_table[i][j] == None) or (data_table[i][j] is None):
                data_table[i][j]=u''

# Заполняем отчет значениями
    for row in range(2, len(data_table)+2):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-2][0])  # лицевой счёт
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            a=str(data_table[row-2][1])
            b=datetime.datetime.strptime(a,'%Y-%m-%d')
            d=datetime.datetime.strftime(b,"%d.%m.%Y")
            ws.cell('B%s'%(row)).value = '%s' % d  # начальная дата
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-2][2])  # номер счётчика
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-2][3])  # тип энергоресурса
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-2][4])  # наименование ПУ
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-2][5])  # показания на начальную дату
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            d=electric_data_end
            ws.cell('H%s'%(row)).value = '%s' % d  # дата снятия показаний
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
                        
        try:
            ws.cell('K%s'%(row)).value = 'ВводПоказаний'
            ws.cell('K%s'%(row)).style = ali_white
        except:
            ws.cell('K%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table[row-2][8])  # Дом-корпус
            ws.cell('O%s'%(row)).style = ali_white
        except:
            ws.cell('O%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table[row-2][7])  # № квартиры
            ws.cell('P%s'%(row)).style = ali_white
        except:
            ws.cell('P%s'%(row)).style = ali_white
            next
            

#    ws.row_dimensions[5].height = 41
#    ws.column_dimensions['A'].width = 10 
#    ws.column_dimensions['D'].width = 30 
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['j'].width = 15
    ws.column_dimensions['k'].width = 15
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_water_resources_'+str(electric_data_end)
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_heat_all_by_date(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]

#Шапка
    
    ws['A1'] = 'Лицевой счет'
    ws['A1'].style = ali_grey
    
    ws['B1'] = 'Дата начала работы'
    ws['B1'].style = ali_grey
    
    ws['C1'] = 'Номер прибора'
    ws['C1'].style = ali_grey
    
    ws['D1'] = 'тип прибора'
    ws['D1'].style = ali_grey
    
    ws['E1'] = 'Наименование ПУ'
    ws['E1'].style = ali_grey
    
    ws['F1'] = 'Показания'
    ws['F1'].style = ali_grey
    
    ws['G1'] = 'Дата Окончания' 
    ws['G1'].style = ali_grey
    
    ws['H1'] = 'Дата'
    ws['H1'].style = ali_grey

    ws['I1'] = 'Квартирный' 
    ws['I1'].style = ali_grey
    
    ws['J1'] = 'Адрес'
    ws['J1'].style = ali_grey
    
    ws['K1'] = 'Операция'
    ws['K1'].style = ali_grey
    
    ws['L1'] = 'Заменяемый счетчик'
    ws['L1'].style = ali_grey
    
    ws['M1'] = 'Показания заменяемого счетчика'
    ws['M1'].style = ali_grey

    ws['N1'] = 'Улица'
    ws['N1'].style = ali_grey

    ws['O1'] = 'Дом'
    ws['O1'].style = ali_grey

    ws['P1'] = 'Квартира'
    ws['P1'].style = ali_grey
#Запрашиваем данные для отчета

    
    data_table = common_sql.get_data_table_report_heat_res_by_date(electric_data_end)
         
    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        if (data_table[i][5] is None):            
            dt=common_sql.get_data_table_by_date_heat_sayany_for_buhgaltery(data_table[i][0], data_table[i][8], electric_data_end)         
            if (len(dt)>0):                
                data_table[i]=dt[0]
        data_table[i]=tuple(data_table[i])
    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

# Заполняем отчет значениями
    for row in range(2, len(data_table)+2):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-2][0])  # лицевой счёт
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            a=str(data_table[row-2][1])
            b=datetime.datetime.strptime(a,'%Y-%m-%d')
            d=datetime.datetime.strftime(b,"%d.%m.%Y")
            ws.cell('B%s'%(row)).value = '%s' % d  # начальная дата
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-2][2])  # номер счётчика
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-2][3])  # тип энергоресурса
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-2][4])  # наименование ПУ
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-2][5])  # показания на начальную дату
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-2][6]) # дата снятия показаний
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
                        
        try:
            ws.cell('K%s'%(row)).value = 'ВводПоказаний'
            ws.cell('K%s'%(row)).style = ali_white
        except:
            ws.cell('K%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table[row-2][8])  # Дом-корпус
            ws.cell('O%s'%(row)).style = ali_white
        except:
            ws.cell('O%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('P%s'%(row)).value = '%s' % (data_table[row-2][7])  # № квартиры
            ws.cell('P%s'%(row)).style = ali_white
        except:
            ws.cell('P%s'%(row)).style = ali_white
            next
            

#    ws.row_dimensions[5].height = 41
#    ws.column_dimensions['A'].width = 10 
#    ws.column_dimensions['B'].width = 10 
#    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['j'].width = 15
    ws.column_dimensions['k'].width = 15
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_heat_resources_'+str(electric_data_end)
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_resources_all(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_start = request.session["electric_data_start"]
    electric_data_end   = request.session["electric_data_end"]
#Шапка
    ws.merge_cells('A2:G2')
    ws['A2'] = 'Филиград: 1, 2, 3 корпуса. Показания по энергоресурсам за период c '+electric_data_start +' по '+electric_data_end
    
    ws['A5'] = 'Лицевой'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Дата начала'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Номер прибора'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Тип прибора'
    ws['D5'].style = ali_grey
    
    #ws['E5'] = 'Показания на '+electric_data_end
    ws['E5'] = 'текущие'
    ws['E5'].style = ali_grey
    
    #ws['F5'] = 'Показания на '+electric_data_start
    ws['F5'] = 'предыдущие'
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'разница' 
    ws['G5'].style = ali_grey
    
    ws['H5'] = 'Дата'#установки
    ws['H5'].style = ali_grey

    ws['I5'] = 'Дата' # съёма
    ws['I5'].style = ali_grey
    
    ws['J5'] = 'Квар'
    ws['J5'].style = ali_grey

    ws['K5'] = 'Адр'
    ws['K5'].style = ali_grey

    ws['L5'] = 'Операция'
    ws['L5'].style = ali_grey
    
    ws['M5'] = 'Квар'
    ws['M5'].style = ali_grey
    
#Запрашиваем данные для отчета

           
    
    data_table = []
    data_table = common_sql.get_data_table_report_all_res_period3(electric_data_start, electric_data_end)

    #zamenyem None na N/D vezde
    for i in range(len(data_table)):
        data_table[i]=list(data_table[i])
        for j in range(1,len(data_table[i])):
            if (data_table[i][j] == None) or (data_table[i][j] is None):
                data_table[i][j]=u''
        data_table[i]=tuple(data_table[i])


# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # лицевой счёт
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            a=str(data_table[row-6][1])
            b=datetime.datetime.strptime(a,'%Y-%m-%d')
            d=datetime.datetime.strftime(b,"%d.%m.%Y")
            ws.cell('B%s'%(row)).value = '%s' %d  # начальная дата
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # номер счётчика
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # тип энергоресурса
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5])  # показания на конченую дату
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # показания на начальную дату
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][7])  # Дельта
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
        try:
            a=str(data_table[row-6][8])
            b=datetime.datetime.strptime(a,'%Y-%m-%d')
            d=datetime.datetime.strftime(b,"%d.%m.%Y")
            ws.cell('H%s'%(row)).value = '%s' %d  # Дата установки приборов
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            a=str(data_table[row-6][9])
            b=datetime.datetime.strptime(a,'%Y-%m-%d')
            d=datetime.datetime.strftime(b,"%d.%m.%Y")
            ws.cell('I%s'%(row)).value = '%s' %d  # Дата конечная
            ws.cell('I%s'%(row)).style = ali_white
        except:
            ws.cell('I%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('L%s'%(row)).value = 'ВводПоказаний'  # Операция
            ws.cell('L%s'%(row)).style = ali_white
        except:
            ws.cell('L%s'%(row)).style = ali_white
            next
        try:
            ws.cell('M%s'%(row)).value = '%s' % (data_table[row-6][10])  # Абонент
            ws.cell('M%s'%(row)).style = ali_white
        except:
            ws.cell('M%s'%(row)).style = ali_white
            next

    ws.row_dimensions[5].height = 41
#    ws.column_dimensions['A'].width = 10 
    ws.column_dimensions['L'].width = 20 
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
#    ws.column_dimensions['j'].width = 15
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_all_resources_'+ str(electric_data_start)+'_'+str(electric_data_end)
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def pokazaniya_heat_current_report_v2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Показания теплосчётчиков. Последние считанные данные'
    
    ws['A5'] = 'Дата'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Время'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Абонент'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Заводской номер'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания, Гкал'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'Показания, м3'
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Время работы, ч'
    ws['G5'].style = ali_grey
    
    ws['H5'] = 'Твхода, С'
    ws['H5'].style = ali_grey

    ws['I5'] = 'Твыхода, С'
    ws['I5'].style = ali_grey
    
    ws['J5'] = 'Разница Т, С'
    ws['J5'].style = ali_grey
    
    ws['K5'] = 'Код ошибки'
    ws['K5'].style = ali_grey
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    obj_key             = request.session['obj_key']

    list_except = []
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = obj_title         = request.GET['obj_title']
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']

    data_table=[]
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_current_heat_v2(obj_title, parent_name, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_current_heat_v2(obj_title, parent_name, False)
        for row in data_table:
            for x in list_except:
                if x==row[2]:
                    data_table.remove(x)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по расходу воды
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Абонент
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # заводской номер
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания по теплу
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # Показания по расходу воды
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][8])  # заводской номер
            ws.cell('I%s'%(row)).style = ali_white
        except:
            ws.cell('I%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][9])  # Показания по теплу
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table[row-6][10])  # Показания по расходу воды
            ws.cell('K%s'%(row)).style = ali_white
        except:
            ws.cell('K%s'%(row)).style = ali_white

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'heat_current_report' 
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_potreblenie_heat_v2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']
#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление тепловой энергии в период с ' + str(electric_data_start)  + ' по ' + str(electric_data_end)
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания на '  + str(electric_data_start)  + u', Гкал'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания на '  + str(electric_data_end) + u', Гкал'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Потребление, Гкал'
    ws['E5'].style = ali_grey
    
    ws['f5'] = 'Объём на '  + str(electric_data_start)  + u', м3'
    ws['f5'].style = ali_grey
    
    ws['g5'] = 'Объём на '  + str(electric_data_end) + u', м3'
    ws['g5'].style = ali_grey
    
    ws['h5'] = 'Потребленный объём, м3'
    ws['h5'].style = ali_grey
    
    ws['i5'] = 'Время работы на '  + str(electric_data_start) + u', ч'
    ws['i5'].style = ali_grey
    
    ws['j5'] = 'Время работы на '  + str(electric_data_end) + u', ч'
    ws['j5'].style = ali_grey
    
    ws['k5'] = 'Время работы с ' + str(electric_data_start) + ' по ' + str(electric_data_end) + ', ч'
    ws['k5'].style = ali_grey
    
#    ws['F5'] = 'Время работы с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"]) + ' ,ч'
#    ws['F5'].style = ali_grey
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
#    electric_data_end   = request.session['electric_data_end']
#    electric_data_start   = request.session['electric_data_start']                        
    obj_key             = request.session['obj_key']
    #is_electric_monthly = request.session['is_electric_monthly']
    #is_electric_daily   = request.session['is_electric_daily']
    #data_table_end   = []
    #data_table_start = []
    list_except = []
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_data_table_for_period_v3(meters_name, parent_name, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_for_period_v3(meters_name, parent_name, electric_data_start, electric_data_end, False)
    else:
        data_table = []

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5])  # Потребление
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # Объём на начало
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][7])  #Объём на конец
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('h%s'%(row)).value = '%s' % (data_table[row-6][8])  # Объём - дельта
            ws.cell('h%s'%(row)).style = ali_white
        except:
            ws.cell('h%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('i%s'%(row)).value = '%s' % (data_table[row-6][9])  # Время работы на начао
            ws.cell('i%s'%(row)).style = ali_white
        except:
            ws.cell('i%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('j%s'%(row)).value = '%s' % (data_table[row-6][10])  # Время работы на конец 
            ws.cell('j%s'%(row)).style = ali_white
        except:
            ws.cell('j%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('k%s'%(row)).value = '%s' % (data_table[row-6][11])  # Время работы - дельта
            ws.cell('k%s'%(row)).style = ali_white
        except:
            ws.cell('k%s'%(row)).style = ali_white
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
#    ws.column_dimensions['F'].width = 18
#____________
   
#------------

                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    #output_name = u'potreblenie_heat_report'
    output_name = u'potreblenie_heat_report_'+translate(parent_name)+'_'+translate(meters_name)+'_'+electric_data_start+'-'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_water_potreblenie_tekon_hvs(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Текон. Потребление воды ХВС в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_start"])
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Потребление, м3'
    ws['E5'].style = ali_grey

#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_start = request.session['electric_data_start']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name, electric_data_start, electric_data_end, u'Канал 1',  u'Tekon_hvs', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name,electric_data_start, electric_data_end, u'Канал 1',  u'Tekon_hvs', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
#        try:
#            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # Время работы
#            ws.cell('F%s'%(row)).style = ali_white
#        except:
#            ws.cell('F%s'%(row)).style = ali_white
#            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
#    ws.column_dimensions['F'].width = 18
#____________
   
#------------

                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_water_tekon_hvs_report'
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_water_tekon_hvs(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Текон. Потребление воды ХВС на ' +str(request.session["electric_data_end"])
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['C5'].style = ali_grey
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']    
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, u'Канал 1',  u'Tekon_hvs', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, u'Канал 1',  u'Tekon_hvs', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35

#    ws.column_dimensions['F'].width = 18
#____________
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'pokazaniya_na_datu_water_tekon_hvs_report'
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_water_potreblenie_tekon_gvs(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Текон. Потребление воды ГВС в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_start"])
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Потребление, м3'
    ws['E5'].style = ali_grey
    
#    ws['F5'] = 'Время работы с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"]) + ' ,ч'
#    ws['F5'].style = ali_grey
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_start = request.session['electric_data_start']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name, electric_data_start, electric_data_end, u'Канал 2',  u'Tekon_gvs', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name,electric_data_start, electric_data_end, u'Канал 2',  u'Tekon_gvs', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
#        try:
#            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # Время работы
#            ws.cell('F%s'%(row)).style = ali_white
#        except:
#            ws.cell('F%s'%(row)).style = ali_white
#            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
#    ws.column_dimensions['F'].width = 18
#____________
   
#------------

                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_water_tekon_gvs_report'
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_tekon_heat_potreblenie(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Текон. Потребление воды ГВС в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_start"])
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Потребление, м3'
    ws['E5'].style = ali_grey
    
#    ws['F5'] = 'Время работы с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"]) + ' ,ч'
#    ws['F5'].style = ali_grey
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_start = request.session['electric_data_start']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name, electric_data_start, electric_data_end, u'Канал 3',  u'Tekon_heat', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_period(meters_name, parent_name,electric_data_start, electric_data_end, u'Канал 3',  u'Tekon_heat', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
#        try:
#            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # Время работы
#            ws.cell('F%s'%(row)).style = ali_white
#        except:
#            ws.cell('F%s'%(row)).style = ali_white
#            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
#    ws.column_dimensions['F'].width = 18
#____________
   
#------------

                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_heat_tekon_report'
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_water_tekon_gvs(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Текон. Потребление воды ГВС на ' +str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['C5'].style = ali_grey

    
#    ws['F5'] = 'Время работы с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"]) + ' ,ч'
#    ws['F5'].style = ali_grey
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']    
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, u'Канал 2',  u'Tekon_gvs', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, u'Канал 2',  u'Tekon_gvs', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35

#    ws.column_dimensions['F'].width = 18
#____________
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'pokazaniya_na_datu_water_tekon_gvs_'+translate(parent_name)+'_'+translate(meters_name)+'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_tekon_heat_by_date(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Текон. Потребление воды ГВС на ' +str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['C5'].style = ali_grey

    
#    ws['F5'] = 'Время работы с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"]) + ' ,ч'
#    ws['F5'].style = ali_grey
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']    
    electric_data_end   = request.session['electric_data_end']
    obj_key             = request.session['obj_key']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        # Edinaya f-ya dliya HVS (kanal 1) i GVS (kanal 2), peredaem imiya kanala
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, u'Канал 3',  u'Tekon_heat', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_tekon_daily(meters_name, parent_name, electric_data_end, u'Канал 3',  u'Tekon_heat', False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, electric_data_end)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35

#    ws.column_dimensions['F'].width = 18
#____________
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'pokazaniya_na_datu_heat_tekon_'+translate(parent_name)+'_'+translate(meters_name)+'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
def report_potreblenie_heat(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление тепловой энергии в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания на '  + str(request.session["electric_data_start"])
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Потребление, Гкал'
    ws['E5'].style = ali_grey
    
#    ws['F5'] = 'Время работы с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"]) + ' ,ч'
#    ws['F5'].style = ali_grey
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
#    is_object_level = re.compile(r'level')
#    is_object_level_1 = re.compile(r'level1')
    is_object_level_2 = re.compile(r'level2')
    
    parent_name         = request.session['obj_parent_title']
    meters_name         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']                        
    obj_key             = request.session['obj_key']
    #is_electric_monthly = request.session['is_electric_monthly']
    #is_electric_daily   = request.session['is_electric_daily']
    #data_table_end   = []
    #data_table_start = []
    list_except = [u'ВРУ Счётчик01',u'ВРУ Счётчик02',u'ВРУ Счётчик03',u'ВРУ Счётчик04',u'ВРУ Счётчик05',u'ВРУ Счётчик06',u'ВРУ Счётчик07',u'ВРУ Счётчик08',u'ВРУ Счётчик09',u'ВРУ Счётчик10',u'ВРУ Счётчик11',u'ВРУ Счётчик12',u'ВРУ Счётчик13',u'ВРУ Счётчик14',u'ВРУ Счётчик15',u'ВРУ Счётчик16',u'ВРУ Счётчик17',u'ВРУ Счётчик18',u'ВРУ Счётчик19',u'ВРУ Счётчик20',u'ВРУ Счётчик21',u'ВРУ Счётчик22',u'ВРУ Счётчик23',u'Гараж Счётчик 1',u'Гараж Счётчик 2']
    
    if request.is_ajax():
        if request.method == 'GET':
            request.session["obj_parent_title"]    = parent_name         = request.GET['obj_parent_title']
            request.session["obj_title"]           = meters_name         = request.GET['obj_title']
            request.session["electric_data_end"]   = electric_data_end   = request.GET['electric_data_end']           
            request.session["electric_data_start"]   = electric_data_start   = request.GET['electric_data_start']           
            request.session["obj_key"]             = obj_key             = request.GET['obj_key']
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table_end = common_sql.get_data_table_by_date_heat(meters_name, parent_name, electric_data_end)
        data_table_start = common_sql.get_data_table_by_date_heat(meters_name, parent_name, electric_data_start)
        data_table = []
        for x in range(len(data_table_end)):
            try:
                data_table_temp=[data_table_end[x][0], data_table_end[x][1], data_table_end[x][2], data_table_start[x][3], data_table_end[x][3], data_table_end[x][3]-data_table_start[x][3], data_table_end[x][5] - data_table_start[x][5]]
                data_table.append(data_table_temp)
            except:
                data_table = []
            

    elif (bool(is_object_level_2.search(obj_key))):
        list_of_abonents_2 = common_sql.list_of_abonents(common_sql.return_parent_guid_by_abonent_name(parent_name), meters_name)
        data_table = []
        for x in range(len(list_of_abonents_2)):
            data_table_end_temp = common_sql.get_data_table_by_date_heat(list_of_abonents_2[x][0], meters_name, electric_data_end)
            data_table_start_temp = common_sql.get_data_table_by_date_heat(list_of_abonents_2[x][0], meters_name, electric_data_start)
            data_table_temp = []
            for x in range(len(data_table_end_temp)):

                data_table_temp_2 = []
                try:
                    data_table_temp_2.append(data_table_end_temp[x][0])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")
                try:
                    data_table_temp_2.append(data_table_end_temp[x][1])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")
                try:
                    data_table_temp_2.append(data_table_end_temp[x][2])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")
                try:
                    data_table_temp_2.append(data_table_start_temp[x][3])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")
                try:
                    data_table_temp_2.append(data_table_end_temp[x][3])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")
                try:
                    data_table_temp_2.append(data_table_end_temp[x][3]-data_table_start_temp[x][3])
                except IndexError:
                    data_table_temp_2.append(u"Н/Д")
                except TypeError:
                    data_table_temp_2.append(u"Н/Д")
                #try:
                #    data_table_temp_2.append(data_table_end_temp[x][5]-data_table_start_temp[x][5])
                #except IndexError:
                #    data_table_temp_2.append(u"Н/Д")
                #except TypeError:
                #    data_table_temp_2.append(u"Н/Д")

                data_table_temp.append(data_table_temp_2)
                
            data_table_end_temp = []
            data_table_start_temp = []
            
            if list_of_abonents_2[x][0] in list_except:
                next
            elif data_table_temp:            
                data_table.extend(data_table_temp)
            else:
                data_table.extend([[0,list_of_abonents_2[x][0],u'Н/Д',u'Н/Д',u'Н/Д',u'Н/Д']])
                
              
    else:
        data_table = []

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5])  # Потребление
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
#        try:
#            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # Время работы
#            ws.cell('F%s'%(row)).style = ali_white
#        except:
#            ws.cell('F%s'%(row)).style = ali_white
#            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 18
#    ws.column_dimensions['F'].width = 18
#____________
   
#------------

                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_heat_report'
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_water_by_date(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    meters_name         = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')
#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = meters_name+'. Показания по воде на ' + electric_data_end
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Номер счётчика'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Пульсар'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Канал'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания'
    ws['E5'].style = ali_grey
    
#Запрашиваем данные для отчета

    is_abonent_level = re.compile(r'level2')
    is_object_level_2 = re.compile(r'level1')
    
    parent_name         = request.GET.get('obj_parent_title')
    #meters_name         = request.session['obj_title']
    electric_data_end   = request.GET.get('electric_data_end')           
    obj_key             = request.GET.get('obj_key')
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_water_by_date(meters_name, parent_name, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_water_by_date(meters_name, parent_name, electric_data_end, False)

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # пульсар
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # канал
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
            
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'water_'+translate(parent_name)+'_'+translate(meters_name)+'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

    
def report_forma_80020(request):
    import zipfile
    response = StringIO.StringIO()
    
    
    #Запрашиваем данные для отчета
    
    group_80020_name    = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']
    electric_data_start   = request.session['electric_data_start']                        
   

    # Формируем список дат на основе начальной и конечной даты полученной от web-календарей
    end_date   = datetime.datetime.strptime(electric_data_end, "%d.%m.%Y")
    start_date = datetime.datetime.strptime(electric_data_start, "%d.%m.%Y")
    list_of_dates = [x for x in common_sql.daterange(start_date,
                  end_date,
                  step=datetime.timedelta(days=1),
                  inclusive=True)]
    
    info_group_80020 = common_sql.get_info_group_80020(group_80020_name)
    inn_sender_from_base      = info_group_80020[0][0]
    name_sender_from_base     = info_group_80020[0][1]
    inn_postavshik_from_base  = info_group_80020[0][2]
    name_postavshik_from_base = info_group_80020[0][3]
    dogovor_number_from_base  = info_group_80020[0][4]
    
    #Узнаем GUID счётчиков, которые входят в группу 80020
    meters_guid_list = common_sql.get_meters_guid_list_by_group_name(group_80020_name)

   
    #Создаем архив
    zf = zipfile.ZipFile(response, mode='w', compression=zipfile.ZIP_DEFLATED)
    
    for dates in range(len(list_of_dates)):
    #Формируем файл xml по форме Мосэнергосбыт 80020   
        
        # Создание корневого элемента message
        root = etree.Element('message', {'class': '80020', 'version': '2', 'number': '1' })
        
        # Добавление дочерних элементов - <datetime> <sender> <area> в <root>
        datetimeElt = etree.SubElement(root, 'datetime')
        senderElt = etree.SubElement(root, 'sender')
        areaElt = etree.SubElement(root, 'area')
        
        # Присваиваем значения в <day> <timestamp> <daylightsavingtime>
        day = etree.SubElement(datetimeElt, 'day')
        timestamp = etree.SubElement(datetimeElt, 'timestamp')
        daylightsavingtime = etree.SubElement(datetimeElt, 'daylightsavingtime')
        
        day.text = unicode(list_of_dates[dates].strftime('%Y%m%d'))
        timestamp.text = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
        daylightsavingtime.text = u'0'
        
        # Присваиваем значения <name> <inn> для <sender>
        inn_sender = etree.SubElement(senderElt, 'inn')
        name_sender = etree.SubElement(senderElt, 'name')
        
        inn_sender.text  = inn_sender_from_base
        name_sender.text = name_sender_from_base
        
        # Присваиваем значения для <area>
        inn_area = etree.SubElement(areaElt, 'inn')
        name_abonent_area = etree.SubElement(areaElt, 'name_abonent')
        
        inn_area.text = inn_postavshik_from_base
        name_abonent_area.text = name_postavshik_from_base
        
        # Добавление дочерних элементов
        for x in range(len(meters_guid_list)):
            info_measuring_point = common_sql.get_info_measuring_point_in_group_80020(meters_guid_list[x])
            my_measure_code = unicode(info_measuring_point[0][0])
            my_measure_name = unicode(info_measuring_point[0][1])
            measurepointElt = etree.SubElement(areaElt, 'measuringpoint', code = my_measure_code , name = my_measure_name )
        
            list_of_taken_params = []
            #Получаем считываемые параметры по guid счётчика.
             #A+
            name_of_type_meters = common_sql.get_name_of_type_meter_by_guid(meters_guid_list[x])
            if name_of_type_meters[0][0] == u'Меркурий 230-УМ':
                guid_params = u'922ad57c-8f5e-4f00-a78d-e3ba89ef859f'
            elif name_of_type_meters[0][0] == u'Меркурий 230':
                guid_params = u'6af9ddce-437a-4e07-bd70-6cf9dcc10b31'
            else:
                pass
            result = common_sql.get_taken_param_by_guid_meters_and_guid_params(meters_guid_list[x], guid_params)
            result_list = []
            result_list.append(unicode(result[0][0]))
            result_list.append(unicode(result[0][1]))
            list_of_taken_params.append(result_list)
            #R+
            if name_of_type_meters[0][0] == u'Меркурий 230-УМ':
                 guid_params = u'61101fa3-a96a-4934-9482-e32036c12829'
            elif name_of_type_meters[0][0] == u'Меркурий 230':
                 guid_params = u'66e997c0-8128-40a7-ae65-7e8993fbea61'
            else:
                pass
            
            result = common_sql.get_taken_param_by_guid_meters_and_guid_params(meters_guid_list[x], guid_params)
            result_list = []
            result_list.append(unicode(result[0][0]))
            result_list.append(unicode(result[0][1]))
            list_of_taken_params.append(result_list)
            
            for y in range(len(list_of_taken_params)):
                my_dict_of_profil = {0: [u'0000', u'0030', 0, 1],
                                     1: [u'0030', u'0100', 0, 1],
                                     2: [u'0100', u'0130', 0, 1],
                                     3: [u'0130', u'0200', 0, 1],
                                     4: [u'0200', u'0230', 0, 1],
                                     5: [u'0230', u'0300', 0, 1],
                                     6: [u'0300', u'0330', 0, 1],
                                     7: [u'0330', u'0400', 0, 1],
                                     8: [u'0400', u'0430', 0, 1],
                                     9: [u'0430', u'0500', 0, 1],
                                     10: [u'0500', u'0530', 0, 1],
                                     11: [u'0530', u'0600', 0, 1],
                                     12: [u'0600', u'0630', 0, 1],
                                     13: [u'0630', u'0700', 0, 1],
                                     14: [u'0700', u'0730', 0, 1],
                                     15: [u'0730', u'0800', 0, 1],
                                     16: [u'0800', u'0830', 0, 1],
                                     17: [u'0830', u'0900', 0, 1],
                                     18: [u'0900', u'0930', 0, 1],
                                     19: [u'0930', u'1000', 0, 1],
                                     20: [u'1000', u'1030', 0, 1],
                                     21: [u'1030', u'1100', 0, 1],
                                     22: [u'1100', u'1130', 0, 1],                                
                                     23: [u'1130', u'1200', 0, 1],
                                     24: [u'1200', u'1230', 0, 1],
                                     25: [u'1230', u'1300', 0, 1],
                                     26: [u'1300', u'1330', 0, 1],
                                     27: [u'1330', u'1400', 0, 1],
                                     28: [u'1400', u'1430', 0, 1],
                                     29: [u'1430', u'1500', 0, 1],
                                     30: [u'1500', u'1530', 0, 1],
                                     31: [u'1530', u'1600', 0, 1],
                                     32: [u'1600', u'1630', 0, 1],
                                     33: [u'1630', u'1700', 0, 1],
                                     34: [u'1700', u'1730', 0, 1],
                                     35: [u'1730', u'1800', 0, 1],
                                     36: [u'1800', u'1830', 0, 1],
                                     37: [u'1830', u'1900', 0, 1],
                                     38: [u'1900', u'1930', 0, 1],
                                     39: [u'1930', u'2000', 0, 1],
                                     40: [u'2000', u'2030', 0, 1],                                
                                     41: [u'2030', u'2100', 0, 1],
                                     42: [u'2100', u'2130', 0, 1],
                                     43: [u'2130', u'2200', 0, 1],
                                     44: [u'2200', u'2230', 0, 1],
                                     45: [u'2230', u'2300', 0, 1],
                                     46: [u'2300', u'2330', 0, 1],
                                     47: [u'2330', u'0000', 0, 1] }
                time_table = ['00:00:00', '00:30:00',
                              '01:00:00', '01:30:00',
                              '02:00:00', '02:30:00',
                              '03:00:00', '03:30:00',
                              '04:00:00', '04:30:00',
                              '05:00:00', '05:30:00',
                              '06:00:00', '06:30:00',
                              '07:00:00', '07:30:00',
                              '08:00:00', '08:30:00',
                              '09:00:00', '09:30:00',
                              '10:00:00', '10:30:00',
                              '11:00:00', '11:30:00',
                              '12:00:00', '12:30:00',
                              '13:00:00', '13:30:00',
                              '14:00:00', '14:30:00',
                              '15:00:00', '15:30:00',
                              '16:00:00', '16:30:00',
                              '17:00:00', '17:30:00',
                              '18:00:00', '18:30:00',
                              '19:00:00', '19:30:00',
                              '20:00:00', '20:30:00',
                              '21:00:00', '21:30:00',
                              '22:00:00', '22:30:00',
                              '23:00:00', '23:30:00']
                if list_of_taken_params[y][1] == u'A+ Профиль':
                    my_measuring_channel_code = u'01'
                    for time in range(len(time_table)):
                        result_30_min = common_sql.get_30_min_value_by_meters_number_param_names_and_datetime(list_of_taken_params[y][0], list_of_taken_params[y][1], list_of_dates[dates].strftime('%Y-%m-%d'), time_table[time])
                        if result_30_min:
                            my_dict_of_profil[time]=[my_dict_of_profil[time][0], my_dict_of_profil[time][1], float(result_30_min[0][4])*common_sql.get_k_t_t_by_factory_number_manual(list_of_taken_params[y][0]), 0] #Квт.ч
                            #print u'------ Есть значение',list_of_taken_params[y][1] , my_dict_of_profil[time][0], my_dict_of_profil[time][1]
                            
                        else:
                            #print u'Оставляем по умолчанию', list_of_taken_params[y][1], my_dict_of_profil[time][0], my_dict_of_profil[time][1]
                            pass
                    
                elif list_of_taken_params[y][1] == u'R+ Профиль':
                    my_measuring_channel_code = u'03'
                    for time in range(len(time_table)):
                        result_30_min = common_sql.get_30_min_value_by_meters_number_param_names_and_datetime(list_of_taken_params[y][0], list_of_taken_params[y][1], list_of_dates[dates].strftime('%Y-%m-%d'), time_table[time])
                        if result_30_min:
                            my_dict_of_profil[time]=[my_dict_of_profil[time][0], my_dict_of_profil[time][1], float(result_30_min[0][4])*common_sql.get_k_t_t_by_factory_number_manual(list_of_taken_params[y][0]), 0] # Квар.ч
                            #print u'------ Есть значение',list_of_taken_params[y][1], my_dict_of_profil[time][0], my_dict_of_profil[time][1]
                            
                        else:
                           # print u'Оставляем по умолчанию', list_of_taken_params[y][1], my_dict_of_profil[time][0], my_dict_of_profil[time][1]
                           pass
                else:
                    #print u'Нет совпадений параметров'
                    my_measuring_channel_code = u''
                       
                my_measuring_channel_desc = unicode(list_of_taken_params[y][0]) + u' ' + unicode(list_of_taken_params[y][1])            
                measuringchannelElt = etree.SubElement(measurepointElt, 'measuringchannel', code = my_measuring_channel_code, desc = my_measuring_channel_desc )
                
                                                  
                for z in range (len(my_dict_of_profil)):
                    periodElt = etree.SubElement(measuringchannelElt, 'period', start = unicode(my_dict_of_profil[z][0]), end = unicode(my_dict_of_profil[z][1]))
                    value  = etree.SubElement(periodElt, 'value', status = unicode(my_dict_of_profil[z][3]))
                    value.text = unicode(my_dict_of_profil[z][2])
            
        
        # Создание и сохранение документа
        doc = etree.ElementTree(root) 
        myxml_IO=StringIO.StringIO()   
        doc.write(myxml_IO, xml_declaration=True, encoding='UTF-8')
        # Формируем имя документа
        name_of_document = u'80020'
        name_of_file_80020 =name_of_document + u'_'+ inn_sender_from_base + u'_' + unicode(list_of_dates[dates].strftime('%Y%m%d')) + u'_1' + u'.xml'
        zf.writestr(name_of_file_80020, myxml_IO.getvalue())
       
    zf.close()
    
    response=HttpResponse(response.getvalue())
    response['Content-Type'] = 'application/x-zip-compressed'
    
    output_name = u'80020_' + unicode(dogovor_number_from_base) + u'_' + start_date.strftime('%Y%m%d') + u'-'+ end_date.strftime('%Y%m%d')
    file_ext = u'zip'
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
  
    return response

def report_elf_hvs_by_date(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Эльф. Потребление воды ХВС на ' + unicode(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик Эльф'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Счётчик ХВС'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['D5'].style = ali_grey
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         =  request.session['obj_parent_title']
    obj_title         =  request.session['obj_title']
    electric_data_end   =  request.session['electric_data_end']            
    obj_key             =  request.session['obj_key']
    
    data_table = []
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_elf(obj_title, obj_parent_title , electric_data_end, '1','attr1', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_daily_water_elf(obj_title, obj_parent_title , electric_data_end, '1', 'attr1',False)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # счётчик эльф
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # счётчик хвс
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания 
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next


    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17

#------------

                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'elf_hvs_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_end
    print output_name
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    
    return response
    
def report_elf_hvs_potreblenie(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление воды ХВС в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик Эльф'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Счётчик ХВС'
    ws['C5'].style = ali_grey
        
    ws['D5'] = 'Показания на ' + str(request.session["electric_data_start"])
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания на '  + str(request.session["electric_data_end"]) 
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'Потребление'
    ws['F5'].style = ali_grey

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    electric_data_start = request.session['electric_data_start']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_elf_period(obj_title, obj_parent_title , electric_data_end, electric_data_start,'1','attr1', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_daily_water_elf_period(obj_title, obj_parent_title , electric_data_end, electric_data_start,'1','attr1', False)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][3])  # Потребление
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Время работы
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 18
#____________
   
#------------

                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_elf_hvs_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_start+u'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_elf_gvs_potreblenie(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление воды ГВС в период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик Эльф'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Счётчик ГВС'
    ws['C5'].style = ali_grey
        
    ws['D5'] = 'Показания на '  + str(request.session["electric_data_start"])
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания на '  +  str(request.session["electric_data_end"])
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'Потребление'
    ws['F5'].style = ali_grey

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    electric_data_start = request.session['electric_data_start']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_elf_period(obj_title, obj_parent_title , electric_data_end, electric_data_start,'2','attr2', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_daily_water_elf_period(obj_title, obj_parent_title , electric_data_end, electric_data_start,'2','attr2', False)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][3])  # Потребление
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Время работы
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 18
#____________
   
#------------

                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_elf_gvs_report'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_start+u'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_elf_gvs_by_date(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Эльф. Потребление воды ГВС на ' + unicode(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик Эльф'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Счётчик ГВС'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['D5'].style = ali_grey
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         =  request.session['obj_parent_title']
    obj_title         =  request.session['obj_title']
    electric_data_end   =  request.session['electric_data_end']            
    obj_key             =  request.session['obj_key']
    
    data_table = []
                     
    if (bool(is_abonent_level.search(obj_key))):        
        data_table = common_sql.get_daily_water_elf(obj_title, obj_parent_title , electric_data_end, '2','attr2', True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_daily_water_elf(obj_title, obj_parent_title , electric_data_end, '2', 'attr2',False)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # счётчик эльф
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # счётчик гвс
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания 
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next


    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17

#------------

                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'elf_gvs_report_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pulsar_water_period(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Пульсар. Потребление воды на ' + unicode(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Тип счётчика'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Стояк'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Счётчик'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания на '  + str(request.session["electric_data_start"])+', м3'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'Показания на '  + str(request.session["electric_data_end"])+', м3'
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Разница, м3'
    ws['G5'].style = ali_grey

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         =  request.session['obj_parent_title']
    obj_title         =  request.session['obj_title']
    electric_data_end   =  request.session['electric_data_end']   
    electric_data_start  =  request.session['electric_data_start']            
    obj_key             =  request.session['obj_key']
    
    data_table = []
                     
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_for_period(obj_parent_title, obj_title, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_for_period(obj_parent_title, obj_title, electric_data_start,electric_data_end, False)

    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # тип
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # стояк
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # счётчик 
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # значения на начало
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # значения на конец
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # дельта
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next


    ws.row_dimensions[5].height = 63
    ws.column_dimensions['A'].width = 17 
#    ws.column_dimensions['B'].width = 17 
#    ws.column_dimensions['C'].width = 17
#    ws.column_dimensions['D'].width = 17

#------------

                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'pulsar_water_period_report_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_start+'-'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pulsar_water_daily(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Пульсар. Потребление воды на ' + unicode(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Тип счётчика'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Стояк'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Счётчик'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания на '  + str(request.session["electric_data_end"])+', м3'
    ws['E5'].style = ali_grey
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         =  request.session['obj_parent_title']
    obj_title         =  request.session['obj_title']
    electric_data_end   =  request.session['electric_data_end']            
    obj_key             =  request.session['obj_key']
    
    data_table = []
    
    
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_daily(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_daily(obj_parent_title, obj_title, electric_data_end, False)

    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # тип
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # стояк
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # счётчик
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5])  # показаня
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next


    ws.row_dimensions[5].height = 63
    ws.column_dimensions['A'].width = 23 
#    ws.column_dimensions['B'].width = 17 
#    ws.column_dimensions['C'].width = 17
#    ws.column_dimensions['D'].width = 17

#------------
                   
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'pulsar_water_report_'+translate(obj_title)+'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pulsar_water_daily_row(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]

#Шапка
    ws['A1'] = 'Абонент'
    ws['A1'].style = ali_grey
    ws.merge_cells('A1:A2')
    
    ws['B1'] = 'Стояк 1'
    ws['B1'].style = ali_grey
    ws.merge_cells('B1:E1')
    
    ws['B2'] = 'Счётчик ХВС'
    ws['B2'].style = ali_grey
    
    ws['C2'] = 'Значение ХВС, м3'
    ws['C2'].style = ali_grey   
    
    
    ws['D2'] = 'Счётчик ГВС'
    ws['D2'].style = ali_grey
    
    ws['E2'] = 'Значение ГВС, м3'
    ws['E2'].style = ali_grey
    
    
    ws['F1'] = 'Стояк 2'
    ws['F1'].style = ali_grey
    ws.merge_cells('F1:I1')
    
    ws['F2'] = 'Счётчик ХВС'
    ws['F2'].style = ali_grey
    
    ws['G2'] = 'Значение ХВС, м3'
    ws['G2'].style = ali_grey   
    
    
    ws['H2'] = 'Счётчик ГВС'
    ws['H2'].style = ali_grey
    
    ws['I2'] = 'Значение ГВС, м3'
    ws['I2'].style = ali_grey
    
    
    ws['J1'] = 'Стояк 3'
    ws['J1'].style = ali_grey
    ws.merge_cells('J1:M1')
    
    ws['J2'] = 'Счётчик ХВС'
    ws['J2'].style = ali_grey
    
    ws['K2'] = 'Значение ХВС, м3'
    ws['K2'].style = ali_grey   
    
    
    ws['L2'] = 'Счётчик ГВС'
    ws['L2'].style = ali_grey
    
    ws['M2'] = 'Значение ГВС, м3'
    ws['M2'].style = ali_grey
    
    ws['N1'] = 'Сумма ХВС, м3'
    ws['N1'].style = ali_grey
    ws.merge_cells('N1:N2')
    
    ws['O1'] = 'Сумма ГВС, м3'
    ws['O1'].style = ali_grey
    ws.merge_cells('O1:O2')
#Запрашиваем данные для отчета

    
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
             
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_daily_row(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_daily_row(obj_parent_title, obj_title, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull_for_pulsar(data_table)

    
# Заполняем отчет значениями
    for row in range(3, len(data_table)+3):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-3][1])  # абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-3][2])  # nomer hvs
            ws.cell('B%s'%(row)).style = ali_blue
        except:
            ws.cell('B%s'%(row)).style = ali_blue
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-3][3])  # value hvs
            ws.cell('C%s'%(row)).style = ali_blue
        except:
            ws.cell('C%s'%(row)).style = ali_blue
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-3][4])  # gvs num
            ws.cell('D%s'%(row)).style = ali_pink
        except:
            ws.cell('D%s'%(row)).style = ali_pink
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-3][5])  # value gvs
            ws.cell('E%s'%(row)).style = ali_pink
        except:
            ws.cell('E%s'%(row)).style = ali_pink
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-3][6])  # num hvs
            ws.cell('F%s'%(row)).style = ali_blue
        except:
            ws.cell('F%s'%(row)).style = ali_blue
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-3][7]) # val hvs
            ws.cell('G%s'%(row)).style = ali_blue
        except:
            ws.cell('G%s'%(row)).style = ali_blue
            next
                        
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-3][8])  # gvs num
            ws.cell('H%s'%(row)).style = ali_pink
        except:
            ws.cell('H%s'%(row)).style = ali_pink
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-3][9])  # gvs val
            ws.cell('I%s'%(row)).style = ali_pink
        except:
            ws.cell('I%s'%(row)).style = ali_pink
            next
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-3][10])  # hvs num
            ws.cell('J%s'%(row)).style = ali_blue
        except:
            ws.cell('J%s'%(row)).style = ali_blue
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % (data_table[row-3][11])  # hvs val
            ws.cell('K%s'%(row)).style = ali_blue
        except:
            ws.cell('K%s'%(row)).style = ali_blue
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table[row-3][12])  # hvs num
            ws.cell('L%s'%(row)).style = ali_pink
        except:
            ws.cell('L%s'%(row)).style = ali_pink
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % (data_table[row-3][13])  # hvs val
            ws.cell('M%s'%(row)).style = ali_pink
        except:
            ws.cell('M%s'%(row)).style = ali_pink
            next
            
        try:
            ws.cell('N%s'%(row)).value = '%s' % (data_table[row-3][14])  # hvs num
            ws.cell('N%s'%(row)).style = ali_blue
        except:
            ws.cell('N%s'%(row)).style = ali_blue
            next
            
        try:
            ws.cell('O%s'%(row)).value = '%s' % (data_table[row-3][15])  # hvs val
            ws.cell('O%s'%(row)).style = ali_pink
        except:
            ws.cell('O%s'%(row)).style = ali_pink
            next

#    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
        
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_water_pulsar_row_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+str(electric_data_end)
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pulsar_heat_daily(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Пульсар. Показания по теплу на ' + unicode(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Энергия, Гкал'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Объем, м3'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Температура входа, С'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'Температура выхода, С'
    ws['F5'].style = ali_grey
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         =  request.session['obj_parent_title']
    obj_title         =  request.session['obj_title']
    electric_data_end   =  request.session['electric_data_end']            
    obj_key             =  request.session['obj_key']
    
    data_table = []
                     
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_daily_pulsar_teplo(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_daily_pulsar_teplo(obj_parent_title, obj_title, electric_data_end, False)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # тип
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][3])  # стояк
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value =  '%s' % (data_table[row-6][4])  # счётчик
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][5])  # показаня
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][6])  # показаня
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next

    ws.row_dimensions[5].height = 63
    ws.column_dimensions['A'].width = 20 
    ws.column_dimensions['C'].width = 23 
    ws.column_dimensions['E'].width = 17 
    ws.column_dimensions['F'].width = 17
#    ws.column_dimensions['D'].width = 17

#------------
            
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'pulsar_heat_report_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pulsar_heat_period(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]
    electric_data_start   = request.session["electric_data_start"]

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Пульсар. Потребление тепла с ' +unicode(electric_data_start)+' по '+ unicode(electric_data_end)
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания Энергии на ' + str(electric_data_start)+', Гкал '
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания Энергии на ' + str(electric_data_end)+', Гкал '
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Потребление Энергии, Гкал'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'Показания Объёма на ' + str(electric_data_start)+', м3'
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Показания Объёма на ' + str(electric_data_end)+', м3'
    ws['G5'].style = ali_grey
    
    ws['H5'] = 'Потребление Объёма, м3'
    ws['H5'].style = ali_grey
#Запрашиваем данные для отчета

    
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
          
    obj_key             = request.session['obj_key']
             
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title,electric_data_end, electric_data_start, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title, electric_data_end,electric_data_start, False)
               
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (round(float(data_table[row-6][2]),7)) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (round(float(data_table[row-6][3]),7)) # '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value =  '%s' % (round(float(data_table[row-6][4]),7))  # '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value =  '%s' % (round(float(data_table[row-6][5]),7)) # '%s' % (data_table[row-6][5])  # Время работы
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (round(float(data_table[row-6][6]),7)) # '%s' % (data_table[row-6][6])  # Время работы
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('H%s'%(row)).value =  '%s' % (round(float(data_table[row-6][7]),7))  #'%s' % (data_table[row-6][7])  # Время работы
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
     
    ws.row_dimensions[5].height = 51
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['H'].width = 15
        
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_heat_pulsar_period_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+str(electric_data_start)+'-'+str(electric_data_end)
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_pulsar_heat_period_2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]
    electric_data_start   = request.session["electric_data_start"]

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Пульсар. Потребление энергии с ' +unicode(electric_data_start)+' по '+ unicode(electric_data_end)
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания Энергии на ' + str(electric_data_start)+', Гкал '
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания Энергии на ' + str(electric_data_end)+', Гкал '
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Потребление Энергии, Гкал'
    ws['E5'].style = ali_grey
    
#Запрашиваем данные для отчета

    
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
          
    obj_key             = request.session['obj_key']
    data_table=[]      
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title,electric_data_end, electric_data_start, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_teplo_for_period(obj_parent_title, obj_title, electric_data_end,electric_data_start, False)
               
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value =  '%s' % (format(data_table[row-6][2],'.7f'))   # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value =  '%s' % (format(data_table[row-6][3],'.7f'))   # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value =  '%s' % (format(data_table[row-6][4],'.7f'))  # Потребление
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        

            
     
#    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['E'].width = 17 
        
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_heat_pulsar_energy_period_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+str(electric_data_end)+'-'+str(electric_data_start)
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_pulsar_heat_daily_2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Пульсар. Показания по теплу на ' + unicode(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Энергия, Гкал'
    ws['C5'].style = ali_grey
    

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         =  request.session['obj_parent_title']
    obj_title         =  request.session['obj_title']
    electric_data_end   =  request.session['electric_data_end']            
    obj_key             =  request.session['obj_key']
    
    data_table = []
                     
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_daily_pulsar_teplo(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_by_date_daily_pulsar_teplo(obj_parent_title, obj_title, electric_data_end, False)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][1])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][2])  # тип
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value =  '%s' % (format(data_table[row-6][3],'.7f'))   # стояк
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
        

    ws.row_dimensions[5].height = 63
    ws.column_dimensions['A'].width = 20 
    ws.column_dimensions['C'].width = 23 
    ws.column_dimensions['E'].width = 17 
#    ws.column_dimensions['F'].width = 17
#    ws.column_dimensions['D'].width = 17

#------------
      
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'pulsar_heat_energy_report_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_end
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_heat_elf_period_2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    obj_parent_title         = request.GET.get('obj_parent_title')
    obj_title         = request.GET.get('obj_title')          


    electric_data_end   = request.GET.get("electric_data_end")
    electric_data_start   = request.GET.get("electric_data_start")


#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Теплосчётчик Эльф. Потребление тепла в период c ' +unicode(electric_data_start)+' по '+ unicode(electric_data_end)
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания Энергии на ' + str(electric_data_start)+', Гкал '
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания Энергии на ' + str(electric_data_end)+', Гкал '
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Потребление Энергии, Гкал'
    ws['E5'].style = ali_grey
    
    ws['f5'] = 'Показания Объёма на ' + str(electric_data_start)+', м3'
    ws['f5'].style = ali_grey
    
    ws['g5'] = 'Показания Объёма на ' + str(electric_data_end)+', м3'
    ws['g5'].style = ali_grey       
    
    ws['h5'] = 'Расход объёма, м3'
    ws['h5'].style = ali_grey
    
#Запрашиваем данные для отчета

    
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')    
    obj_parent_title         = request.GET.get('obj_parent_title')
    obj_title         = request.GET.get('obj_title')          
    obj_key             = request.GET.get('obj_key')
#    print obj_parent_title
#    print obj_title
    data_table=[]      
    if (bool(is_abonent_level.search(obj_key))):
         data_table = common_sql.get_data_table_elf_period(obj_parent_title, obj_title, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
         data_table = common_sql.get_data_table_elf_period(obj_parent_title, obj_title, electric_data_start,electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)

    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (format(data_table[row-6][3],'.7f'))  # Показания по теплу на конец
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (format(data_table[row-6][2],'.7f'))  # Показания по теплу на начало
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (format(data_table[row-6][4],'.7f'))  # Потребление
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('f%s'%(row)).value = '%s' % (format(data_table[row-6][6],'.7f'))  # Показания по теплу на конец
            ws.cell('f%s'%(row)).style = ali_white
        except:
            ws.cell('f%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('g%s'%(row)).value = '%s' % (format(data_table[row-6][5],'.7f'))  # Показания по теплу на начало
            ws.cell('g%s'%(row)).style = ali_white
        except:
            ws.cell('g%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('h%s'%(row)).value = '%s' % (format(data_table[row-6][7],'.7f'))  # Потребление
            ws.cell('h%s'%(row)).style = ali_white
        except:
            ws.cell('h%s'%(row)).style = ali_white
            next

            
     
#    ws.row_dimensions[5].height = 41
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 18 
        
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'elf_heat_energy_report_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_start+'-'+electric_data_end
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response


def report_heat_elf_period(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_start = request.session['electric_data_start'] #request.session['electric_data_start']
    electric_data_end   = request.session['electric_data_end'] 
#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Теплосчётчик Эльф. Потребление тепла в период с ' + unicode(electric_data_start) + ' по ' + unicode(electric_data_end)
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания Энергии на ' + unicode(electric_data_end)+', Гкал'
    ws['C5'].style = ali_grey
        
    ws['D5'] = 'Показания Энергии на ' + unicode(electric_data_start)+', Гкал'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Потребление Энергии, Гкал'
    ws['E5'].style = ali_grey
    
    ws['f5'] = 'Показания Объёма на ' + unicode(electric_data_end)+', м3'
    ws['f5'].style = ali_grey
        
    ws['g5'] = 'Показания Объёма на ' + unicode(electric_data_start)+', м3'
    ws['g5'].style = ali_grey
    
    ws['h5'] = 'Потребленённый Объём, м3'
    ws['h5'].style = ali_grey

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    #electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    #electric_data_start = request.session['electric_data_start']
    
#    print unicode(request.session.items())
#    print obj_parent_title
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        pass
        # какая-то фигня - в obj_parent_title, obj_title передаются одинаковые значения
        #data_table = common_sql.get_data_table_elf_period(obj_parent_title, obj_title, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_elf_period(obj_parent_title, obj_title, electric_data_start,electric_data_end, False)
    
  
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Время работы
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('g%s'%(row)).value = '%s' % (data_table[row-6][6])  # Время работы
            ws.cell('g%s'%(row)).style = ali_white
        except:
            ws.cell('g%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('h%s'%(row)).value = '%s' % (data_table[row-6][7])  # Время работы
            ws.cell('h%s'%(row)).style = ali_white
        except:
            ws.cell('h%s'%(row)).style = ali_white
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 18
#____________
                   
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_elf_report_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_start+u'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
  

def report_heat_elf_daily(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Теплосчётчик Эльф. Показания по теплу на ' + str(request.session["electric_data_end"])
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания Энергии на ' + str(request.session["electric_data_end"])+', Гкал'
    ws['C5'].style = ali_grey
        
    ws['D5'] = 'Показания Объёма на ' + str(request.session["electric_data_end"])+', м3'
    ws['D5'].style = ali_grey


    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title    = request.GET.get('obj_parent_title')
    obj_title           = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')            
    obj_key             = request.GET.get('obj_key')

    
#    print unicode(request.session.items())
#    print obj_parent_title
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_elf_heat_daily(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_elf_heat_daily(obj_parent_title, obj_title, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)
  
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
       

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
#____________
                   
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'pokazaniya_teplo_elf_report_'+translate(obj_parent_title)+u'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response

def report_heat_water_elf_daily(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
    obj_parent_title    = request.GET.get('obj_parent_title')
    obj_title           = request.GET.get('obj_title')
    electric_data_end   = request.GET.get('electric_data_end')            
    obj_key             = request.GET.get('obj_key')
#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Теплосчётчик Эльф. Показания по теплу и воде на ' + str(electric_data_end)
    
    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Показания Энергии на ' + str(electric_data_end)+', Гкал'
    ws['C5'].style = ali_grey
        
    ws['D5'] = 'Показания Объёма на ' + str(electric_data_end)+', м3'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Счётчик ХВС'
    ws['E5'].style = ali_grey
    
    ws['f5'] = 'Значение на '+ str(electric_data_end)
    ws['f5'].style = ali_grey
        
    ws['g5'] = 'Счётчик ГВС'
    ws['g5'].style = ali_grey
    
    ws['h5'] = 'Значение на '+ str(electric_data_end)
    ws['h5'].style = ali_grey

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
#    obj_parent_title         = request.session['obj_parent_title']
#    obj_title         = request.session['obj_title']
#    electric_data_end   = request.session['electric_data_end']            
#    obj_key             = request.session['obj_key']
    #electric_data_start = request.session['electric_data_start']
    
#    print unicode(request.session.items())
#    print obj_parent_title
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_elf_heat_water_daily(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_elf_heat_water_daily(obj_parent_title, obj_title, electric_data_end, False)
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)
  
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Время работы
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('g%s'%(row)).value = '%s' % (data_table[row-6][6])  # Время работы
            ws.cell('g%s'%(row)).style = ali_white
        except:
            ws.cell('g%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('h%s'%(row)).value = '%s' % (data_table[row-6][7])  # Время работы
            ws.cell('h%s'%(row)).style = ali_white
        except:
            ws.cell('h%s'%(row)).style = ali_white
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 18
#____________
                   
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'pokazaniya_elf_report_'+translate(obj_parent_title)+u'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_water_pulsar_potreblenie_skladochnaya(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    obj_parent_title         = request.GET.get('obj_parent_title')
    obj_title                = request.GET.get('obj_title')
    electric_data_end        = request.GET.get("electric_data_end")
    electric_data_start      = request.GET.get("electric_data_start")


#Шапка
    
#    ws.merge_cells('A2:E2')
#    ws['A2'] = 'Теплосчётчик Пульсар. Потребление воды в период c ' +unicode(electric_data_start)+' по '+ unicode(electric_data_end)
    ws.merge_cells('B2:N2')
    ws['B2'] =u'Шапка отчёта'
    ws['B2'].style = ali_grey
    ws['C2'].style = ali_grey
    ws['D2'].style = ali_grey
    ws['E2'].style = ali_grey
    ws['F2'].style = ali_grey
    ws['G2'].style = ali_grey
    ws['H2'].style = ali_grey
    ws['I2'].style = ali_grey
    ws['J2'].style = ali_grey
    ws['K2'].style = ali_grey
    ws['L2'].style = ali_grey
    ws['M2'].style = ali_grey
    ws['N2'].style = ali_grey
    
    ws.merge_cells('B3:N3')
    ws['B3'] =u'Шапка отчёта'
    ws['B3'].style = ali_grey
    ws['C3'].style = ali_grey
    ws['D3'].style = ali_grey
    ws['E3'].style = ali_grey
    ws['F3'].style = ali_grey
    ws['G3'].style = ali_grey
    ws['H3'].style = ali_grey
    ws['I3'].style = ali_grey
    ws['J3'].style = ali_grey
    ws['K3'].style = ali_grey
    ws['L3'].style = ali_grey
    ws['M3'].style = ali_grey
    ws['N3'].style = ali_grey

    ws.merge_cells('B4:B5')
    ws['B4'] = 'Квартира'
    ws['B4'].style = ali_grey
    ws['B5'].style = ali_grey
    
    ws.merge_cells('C4:E4')
    ws['C4'] = u'ГВС'
    ws['C4'].style = ali_grey
    
    ws['C5'] = 'Счётчик'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Нач, показание Т1 (м^3)'
    ws['D5'].style = ali_grey

    ws['E5'] = 'Дата/Время'
    ws['E5'].style = ali_grey
    
    ws.merge_cells('F4:H4')
    ws['F4'] = 'ГВС'
    ws['F4'].style = ali_grey
    
    ws['F5'] = 'Кон, показание Т1 (м^3)'
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Дата/Время'
    ws['G5'].style = ali_grey

    ws['H5'] = 'Разница ГВС'
    ws['H5'].style = ali_grey
    
    ws.merge_cells('I4:K4')
    ws['I4'] = 'ХВС'
    ws['I4'].style = ali_grey
    
    ws['I5'] = 'Счётчик'
    ws['I5'].style = ali_grey
    
    ws['J5'] = 'Нач, показание Т1 (м^3)'
    ws['J5'].style = ali_grey

    ws['K5'] = 'Дата/Время'
    ws['K5'].style = ali_grey
    
    ws.merge_cells('L4:N4')
    ws['L4'] = 'ХВС'
    ws['L4'].style = ali_grey
    ws['N4'].style = ali_grey
    
    ws['L5'] = 'Кон, показание Т1 (м^3)'
    ws['L5'].style = ali_grey
    
    ws['M5'] = 'Дата/Время'
    ws['M5'].style = ali_grey

    ws['N5'] = 'Разница ХВС'
    ws['N5'].style = ali_grey
    
#Запрашиваем данные для отчета

    
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')   
    obj_parent_title         = request.GET.get('obj_parent_title')
    obj_title                = request.GET.get('obj_title')
    electric_data_end        = request.GET.get("electric_data_end")
    electric_data_start      = request.GET.get("electric_data_start")             
    obj_key                 = request.GET.get('obj_key')

    data_table=[]      
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_for_period_Skladochnaya(obj_parent_title, obj_title, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_pulsar_water_for_period_Skladochnaya(obj_parent_title, obj_title, electric_data_start,electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)

    
 #Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][2])  # заводской номер
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % ((data_table[row-6][3]).strftime("%d-%m-%Y"))  # заводской номер
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][4])  # заводской номер
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % ((data_table[row-6][5]).strftime("%d-%m-%Y"))  # заводской номер
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][6])  # заводской номер
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('I%s'%(row)).value = '%s' % (data_table[row-6][7])  # заводской номер
            ws.cell('I%s'%(row)).style = ali_white
        except:
            ws.cell('I%s'%(row)).style = ali_white
            next
            
            
        try:
            ws.cell('J%s'%(row)).value = '%s' % (data_table[row-6][8])  # заводской номер
            ws.cell('J%s'%(row)).style = ali_white
        except:
            ws.cell('J%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('K%s'%(row)).value = '%s' % ((data_table[row-6][9]).strftime("%d-%m-%Y"))  # заводской номер
            ws.cell('K%s'%(row)).style = ali_white
        except:
            ws.cell('K%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('L%s'%(row)).value = '%s' % (data_table[row-6][10])  # заводской номер
            ws.cell('L%s'%(row)).style = ali_white
        except:
            ws.cell('L%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('M%s'%(row)).value = '%s' % ((data_table[row-6][11]).strftime("%d-%m-%Y")) # заводской номер
            ws.cell('M%s'%(row)).style = ali_white
        except:
            ws.cell('M%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('N%s'%(row)).value = '%s' % (data_table[row-6][12])  # заводской номер
            ws.cell('N%s'%(row)).style = ali_white
        except:
            ws.cell('N%s'%(row)).style = ali_white
            next

     
#    ws.row_dimensions[5].height = 41
    #ws.row_dimensions[5].height = 41
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 17
    #ws.column_dimensions['F'].width = 35
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['H'].width = 18 
    
    ws.column_dimensions['J'].width = 25    
    ws.column_dimensions['K'].width = 17
    #ws.column_dimensions['L'].width = 35
    ws.column_dimensions['L'].width = 25
    ws.column_dimensions['N'].width = 18 
        
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'pulsar_water_report_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_start+'-'+electric_data_end
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
    
def report_rejim_electro(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    obj_parent_title         = request.GET.get('obj_parent_title')
    obj_title                = request.GET.get('obj_title')
    electric_data_end        = request.GET.get("electric_data_end")    


#Шапка
 
    ws['A2'] = u"Адрес"
    #ws['H1'] = u'Шифр'
    ws['G2'] = u'Абонент'
    ws['G3'] = u'Питающий центр'
    #ws['G3'] = u'№ фидера'
    
    ws['D5'] = u'Протокол (первичный)'
    ws['D5'].style = ali_white_size_18
    
    ws['E7'] = u'трансформаторного напряжения _____ вольт'
    
    ws['B6'] = u'записей показаний электросчетчиков и вольтметров, а также определения нагрузок'
    ws['B7'] = u"и тангенса 'фи' за " + str(electric_data_end) + u'г'
    ws['B9'] = u'Счётчик Акт. и Реакт. энергии № '# + str(common_sql.get_serial_number_by_meter_name(meters_name)) 
    #ws['E9'] = u'Реакт. Счетчик № '# + str(common_sql.get_serial_number_by_meter_name(meters_name))
    ws['B10'] = u'КТТ' 
    ws['B11'] = u'КТН'
    ws['B12'] = u'Расч. коэффициент трансформации'

    #ws['A11'] = u'Время' 
    ws['A17'] = u'час'
    ws['B14'] = u'30-минутные срезы мощности (график мощности'
    ws['B15'] = u'              в памяти счетчика)'
    ws['B16'] = u' активная, Вт'
    ws['B17'] = u'отпуск'
    ws['C17'] = u'приём'
    ws['D16'] = u' реактивная, ВАр'
    ws['D17'] = u'отпуск'
    ws['E17'] = u'приём'
    
    ws['F15'] = u' Расход энергии'
    ws['F16'] = u' активная, '
    ws['G16'] = u' реактивная, '
    ws['F17'] = u'   КВтч'
    ws['G17'] = u'   КВАрч'
    
    ws['H15'] = u'тангенс'
    ws['I15'] = u'косинус'
    ws['H17'] = u'    фи'
    ws['I17'] = u'    фи'
    
    ws['J15'] = u' Полная'
    ws['J16'] = u' мощность'  
    ws['J17'] = u'   КВа'
    
    ws['K15'] = u' Показания'
    ws['K16'] = u' вольтметров '    
    ws['K17'] = u'   в/н'
    ws['L17'] = u'   н/н'
    
    ws['M15'] = u' Мощность'
    ws['M16'] = u'компен.устр'  
    ws['M17'] = u'   КВаh'
    
    ws['A18'] = u'0-1' 
    ws['A19'] = u'1-2' 
    ws['A20'] = u'2-3'
    ws['A21'] = u'3-4' 
    ws['A22'] = u'4-5'
    ws['A23'] = u'5-6' 
    ws['A24'] = u'6-7'
    ws['A25'] = u'7-8' 
    ws['A26'] = u'8-9'
    ws['A27'] = u'9-10' 
    ws['A28'] = u'10-11'
    ws['A29'] = u'11-12' 
    ws['A30'] = u'12-13'
    ws['A31'] = u'13-14' 
    ws['A32'] = u'14-15'
    ws['A33'] = u'15-16' 
    ws['A34'] = u'16-17'
    ws['A35'] = u'17-18' 
    ws['A36'] = u'18-19'
    ws['A37'] = u'19-20' 
    ws['A38'] = u'20-21'
    ws['A39'] = u'21-22' 
    ws['A40'] = u'22-23'
    ws['A41'] = u'23-24' 
    #ws['A37'] = u'24'
    
    

    ws.column_dimensions['A'].width = 7            
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 11
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['F'].width = 11
    ws.column_dimensions['G'].width = 11
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 11  
    ws.column_dimensions['K'].width = 7 
    ws.column_dimensions['L'].width = 7 
    ws.row_dimensions[5].height = 30
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')      
    obj_key                 = request.GET.get('obj_key')
    dt_activ=[]
    dt_reactiv=[]
    if (bool(is_abonent_level.search(obj_key))):
        dt_activ=common_sql.get_data_table_rejim( obj_title, electric_data_end, 'A+ Профиль')
        dt_reactiv=common_sql.get_data_table_rejim( obj_title, electric_data_end, 'R+ Профиль')
    
    if len(dt_activ)>0: 
#Заполняем отчет значениями
        try:
            ws.cell('E9').value = '%s' % (obj_title)  # Абонент грщ
            ws.cell('E9').style = ali_white
        except:
            ws.cell('E9').style = ali_white
            next
    
        try:
            ws.cell('C10').value = '%s' % (dt_activ[0][1])  # ктт
            ws.cell('C10').style = ali_white
        except:
            ws.cell('C10').style = ali_white
            next
            
        try:
            ws.cell('C11').value = '%s' % (dt_activ[0][2])  # ктн
            ws.cell('C11').style = ali_white
        except:
            ws.cell('C11').style = ali_white
            next
            
        try:
            ws.cell('E12').value = '%s' % (dt_activ[0][2]*dt_activ[0][1])  # ктн
            ws.cell('E12').style = ali_white
        except:
            ws.cell('E12').style = ali_white
            next
          
        try:
            t_pred=dt_activ[0][0]+dt_activ[0][7]
            ws.cell('F18').value = '%s' % (t_pred) #сумма первой получасовки/2 и значения на полнось на начало месяца
            ws.cell('F18').style = ali_white
        except:
            ws.cell('F18').style = ali_white
                
        for row in range(18, 42): 
            n=row-11
            
            try:
                ws.cell('C%s'%(row)).value = '%s' % (dt_activ[0][n]*2) # сумма получасовок за час
                ws.cell('C%s'%(row)).style = ali_white
            except:
                ws.cell('C%s'%(row)).style = ali_white
                
            try:
                ws.cell('E%s'%(row)).value = '%s' % (dt_reactiv[0][n]*2) # сумма получасовок за час
                ws.cell('E%s'%(row)).style = ali_white
            except:
                ws.cell('E%s'%(row)).style = ali_white
        
        
        try:
            t_pred_act=dt_activ[0][0]+dt_activ[0][7]
            ws.cell('F18').value = '%s' % (t_pred_act) #сумма первой получасовки/2 и значения на полнось на начало месяца
            ws.cell('F18').style = ali_white
        except:
            ws.cell('F18').style = ali_white
            
        try:
            t_pred_react=dt_reactiv[0][0]+dt_reactiv[0][7]
            ws.cell('G18').value = '%s' % (t_pred_react) #сумма первой получасовки/2 и значения на полнось на начало месяца
            ws.cell('G18').style = ali_white
        except:
            ws.cell('G18').style = ali_white
            
        for row in range(19, 42): 
            n=row-11
            try:
                t_pred_act+=dt_activ[0][n]
                ws.cell('F%s'%(row)).value = '%s' % (t_pred_act) # 
                ws.cell('F%s'%(row)).style = ali_white
            except:
                ws.cell('F%s'%(row)).style = ali_white
                
            try:
                t_pred_react+=dt_reactiv[0][n]                
                ws.cell('G%s'%(row)).value = '%s' % (t_pred_react) # 
                ws.cell('G%s'%(row)).style = ali_white
            except:
                ws.cell('G%s'%(row)).style = ali_white
            
        
    
    
    
    
    
    for col_idx in range(1, 14):
        col = get_column_letter(col_idx)
        for row in range(18, 42):
            ws.cell('%s%s'%(col, row)).style = ali_white
            
            
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    
    output_name = u'protokol_rejim_den_'+translate(obj_parent_title)+'_'+'_'+electric_data_end
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_heat_karat_potreblenie(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    obj_parent_title         = request.GET.get('obj_parent_title')
    obj_title                = request.GET.get('obj_title')
    electric_data_end        = request.GET.get("electric_data_end")
    electric_data_start      = request.GET.get("electric_data_start")
    obj_key                  = request.GET.get('obj_key')
#Шапка    
#Шапка
    ws.merge_cells('A2:E2')
    ws['A1'] = unicode(obj_parent_title) + " " + unicode(obj_title)
    ws['A2'] = 'Карат 307. Потребление тепла в период с ' + str(electric_data_start) + ' по ' + str(electric_data_end)
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Тепловая Энергия на '  + str(electric_data_start)
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Тепловая Энергия на '  + str(electric_data_end)
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Потребление Энергии, Гкал'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'Масса на '  + str(electric_data_start)
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Масса на '  + str(electric_data_end)
    ws['G5'].style = ali_grey
    
    ws['H5'] = 'Потребление Массы, кг'
    ws['H5'].style = ali_grey
    
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_karat_potreblenie(obj_parent_title, obj_title, electric_data_start, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_karat_potreblenie(obj_parent_title, obj_title,electric_data_start, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания по теплу на начало
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания по теплу на конец
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][7])  # Потребление
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
    ws.row_dimensions[5].height = 41
    
    ws.column_dimensions['A'].width = 20 
    ws.column_dimensions['B'].width = 17 
#    ws.column_dimensions['C'].width = 20
#    ws.column_dimensions['D'].width = 35
#    ws.column_dimensions['E'].width = 18
#    ws.column_dimensions['F'].width = 18
             
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_heat_karat_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+ str(electric_data_start) + "-" + str(electric_data_end)
    file_ext = u'xlsx'
 
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_heat_karat_daily(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    obj_parent_title         = request.GET.get('obj_parent_title')
    obj_title                = request.GET.get('obj_title')
    electric_data_end        = request.GET.get("electric_data_end")
    #electric_data_start      = request.GET.get("electric_data_start")
    obj_key                  = request.GET.get('obj_key')
#Шапка    
#Шапка
    ws.merge_cells('A2:E2')
    ws['A1'] = unicode(obj_parent_title) + " " + unicode(obj_title)
    ws['A2'] = 'Карат 307. Показания на ' + str(electric_data_end)
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Тепловая Энергия, Гкал'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Масса, кг '
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Температура вход'
    ws['E5'].style = ali_grey
    
    ws['F5'] = 'Температура выход'
    ws['F5'].style = ali_grey
    
    ws['G5'] = 'Время наработки, мин.'
    ws['G5'].style = ali_grey
    
    ws['H5'] = 'Код ошибки (0-ошибок нет)'
    ws['H5'].style = ali_grey
    
    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    data_table=[]
    if (bool(is_abonent_level.search(obj_key))):
        data_table = common_sql.get_data_table_karat_heat_water_daily(obj_parent_title, obj_title, electric_data_end, True)
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_karat_heat_water_daily(obj_parent_title, obj_title, electric_data_end, False)
              
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)

# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][2])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][3])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания по теплу 
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][5])  # Показания по массе
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][6])  # tin
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][7])  # tout
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('G%s'%(row)).value = '%s' % (data_table[row-6][8])  # twork
            ws.cell('G%s'%(row)).style = ali_white
        except:
            ws.cell('G%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('H%s'%(row)).value = '%s' % (data_table[row-6][9])  # ошибки
            ws.cell('H%s'%(row)).style = ali_white
        except:
            ws.cell('H%s'%(row)).style = ali_white
            next
    ws.row_dimensions[5].height = 41
    
    ws.column_dimensions['A'].width = 20 
    ws.column_dimensions['B'].width = 17 
#    ws.column_dimensions['C'].width = 20
#    ws.column_dimensions['D'].width = 35
#    ws.column_dimensions['E'].width = 18
#    ws.column_dimensions['F'].width = 18
             
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'heat_karat_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+ str(electric_data_end)
    file_ext = u'xlsx'
 
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_empty_alert(request):
   
    #response = StringIO.StringIO()
    response = HttpResponse('Отчёт не предусмотрен', content_type="text/plain")
    output_name = u'empty'
    file_ext = u'txt'
 
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)  
    return response


def get_month_russian_name_by_num(m):
    if m == 1:
         M='январь'
    elif m==2:
         M='февраль'
    elif m==3:
         M='март'
    elif m==4:
         M='апрель'
    elif m==5:
         M='май'
    elif m==6:
         M='июнь'
    elif m==7:
         M='июль'
    elif m==8:
         M='август'
    elif m==9:
         M='сентябрь'
    elif m==10:
         M='октябрь'
    elif m==11:
         M='ноябрь'
    elif m==12:
         M='декабрь'
    else:
         M='ошибка'
    return M
    
def add_3columns_to_dt(data_table,data_range,n1,n2,n3):

    for i in range(0,len(data_table)):
        data_table[i]=list(data_table[i])
        data_table[i].append(data_range[i][n1])
        data_table[i].append(data_range[i][n2])
        data_table[i].append(data_range[i][n3])
        data_table[i]=tuple(data_table[i])
    return data_table  
    
def report_water_elf_potreblenie_monthly_with_delta(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    
    # запрашиваем данные
    data_table = []
    electric_data_end        = request.GET.get("electric_data_end")
    electric_data_start      = request.GET.get("electric_data_start")
    
    dt_date = []
    dt_range = []
    dt_date = common_sql.generate_monthly_range(electric_data_start,electric_data_end)
    double_dates = []
    dt_row4 = []
    dt_row5 = []
    rash = 'РАСХОД'
    znach = 'Значения'
    pred = 'предыдущие'
    tek = 'текущие'
   
    for row in range(0,len(dt_date)):
       data_start = dt_date[row][0].strftime("%d.%m.%Y")
       if (row+1)<len(dt_date):
           data_end = dt_date[row+1][0].strftime("%d.%m.%Y")
       else:
           data_end = dt_date[row][0].strftime("%d.%m.%Y")
       
       double_dates.append(data_start[3:])
       double_dates.append(data_end[3:])
       double_dates.append(get_month_russian_name_by_num(datetime.datetime.strptime(data_end, '%d.%m.%Y').month))
       
       dt_row4.append(znach)
       dt_row4.append(znach)
       dt_row4.append(rash)
       
       dt_row5.append(pred)
       dt_row5.append(tek)
       dt_row5.append('___')
       
       dt_range = common_sql.get_data_table_elf_period_monthly(data_start, data_end)       
       if row == 0:         
           data_table=dt_range           
       else:                          
           data_table=add_3columns_to_dt(data_table,dt_range,4,5,6)
                         
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table,None)    
        #val_num= len(data_table[0]) - 6   
        
    #count_month=range(1,len(dt_date)-1)
    if len(data_table)>03:
        double_dates.pop()
        double_dates.pop()
        double_dates.pop()


#Шапка
    ws.merge_cells('A2:J2')
    ws['A2'] = 'Потребление по водосчётчикам Эльф по месяцам с ' + dt_date[0][0].strftime("%d.%m.%Y") + ' по ' + dt_date[-1][0].strftime("%d.%m.%Y")
    
    ws.merge_cells('A4:A6')
    ws['A4'] = '№ помещения'
    ws['A4'].style = ali_grey
    
    ws.merge_cells('B4:B6')
    ws['B4'] = 'Тип прибора'
    ws['B4'].style = ali_grey
    
    ws.merge_cells('C4:C6')
    ws['C4'] = 'Номер ПУ'
    ws['C4'].style = ali_grey    

    
    chrNum=68
    nextLitera=False
    if len(double_dates) < 49: #выход за пределы дат - и выполнятся долго будет и нужна спец.обработка букв в эксель, это будет около 1,5 года        
        for i in range(0,len(double_dates)): 
            if chrNum == 91:            
                nextLitera=True
                chrNum=65
            if not nextLitera:            
                bukva = chr(chrNum)
                if chrNum != 90:
                    bukva_next= chr(chrNum+1)
                else:
                     bukva_next= 'AA'
            else:            
                bukva = 'A'+chr(chrNum)
                bukva_next= 'A'+chr(chrNum+1)
            if not bool(i % 3): 
                ws.merge_cells(bukva+'4:'+bukva_next+'4')
            print bukva, chrNum
            ws[bukva+'4'] = dt_row4[i]               
            ws[bukva+'4'].style = ali_grey
            
            ws[bukva+'5'] = dt_row5[i]               
            ws[bukva+'5'].style = ali_grey        
            
            ws[bukva+'6'] =  double_dates[i]  #получаем букву экселевского столбца и перебираем их черех аццкие коды
            ws[bukva+'6'].style = ali_grey
            
            ws.column_dimensions[bukva].width = 15
            chrNum+=1
            
            
    # Заполняем отчет значениями
        for row in range(7, len(data_table)+7):
            try:
                ws.cell('A%s'%(row)).value = '%s' % (data_table[row-7][0][9:])  # номре помещения
                ws.cell('A%s'%(row)).style = ali_white
            except:
                ws.cell('A%s'%(row)).style = ali_white
                next
            
            try:
                ws.cell('B%s'%(row)).value = '%s' % (data_table[row-7][3])  # тип прибора
                ws.cell('B%s'%(row)).style = ali_white
            except:
                ws.cell('B%s'%(row)).style = ali_white
                next
                
            try:
                ws.cell('C%s'%(row)).value = '%s' % (data_table[row-7][2])  # номер счётчика
                ws.cell('C%s'%(row)).style = ali_white
            except:
                ws.cell('C%s'%(row)).style = ali_white
                next
               
               
            chrNum=68
            nextLitera = False
            for dt_col in range(4,len(data_table[row-7])-3):
                 if chrNum == 91:
                    nextLitera=True
                    chrNum=65
                 if not nextLitera:
                    bukva = chr(chrNum)                
                 else:               
                    bukva = 'A'+chr(chrNum)                
                 try:
                    ws.cell(bukva+'%s'%(row)).value = '%s' % (data_table[row-7][dt_col])  #дельта
                    ws.cell(bukva+'%s'%(row)).style = ali_white
                 except:
                    ws.cell(bukva+'%s'%(row)).style = ali_white
                    next
                 chrNum+=1

#    ws.column_dimensions['D'].width = 17
#    ws.column_dimensions['E'].width = 17
#    ws.column_dimensions['F'].width = 18
#____________

    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_elf_'+electric_data_start+u'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_water_elf_daily(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Показания по водосчётчикам Эльф на' + unicode(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик Эльф'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Счётчик воды'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Тип счётчика'
    ws['D5'].style = ali_grey
    
    ws['E5'] = 'Показания на '  + str(request.session["electric_data_end"])
    ws['E5'].style = ali_grey

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         =  request.session['obj_parent_title']
    obj_title         =  request.session['obj_title']
    electric_data_end   =  request.session['electric_data_end']            
    obj_key             =  request.session['obj_key']
    
    data_table = []
                     
    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_water_daily_elf(obj_title, obj_parent_title,  electric_data_end, True)
      
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_water_daily_elf(obj_title, obj_parent_title,  electric_data_end, False)
       

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)

        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # счётчик эльф
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # счётчик хвс
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][4])  # Показания 
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания 
            ws.cell('e%s'%(row)).style = ali_white
        except:
            ws.cell('e%s'%(row)).style = ali_white
            next


    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17

#------------

                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'elf_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_end
    print output_name
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    
    return response
    
def report_water_elf_potreblenie(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Потребление по водосчётчикам Эльф за период с ' + str(request.session["electric_data_start"]) + ' по ' + str(request.session["electric_data_end"])
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Счётчик Эльф'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Счётчик  воды'
    ws['C5'].style = ali_grey
    
    ws['d5'] = 'Тип счётчика'
    ws['d5'].style = ali_grey
        
    ws['e5'] = 'Показания на '  + str(request.session["electric_data_start"])
    ws['e5'].style = ali_grey
    
    ws['f5'] = 'Показания на '  +  str(request.session["electric_data_end"])
    ws['f5'].style = ali_grey
    
    ws['g5'] = 'Потребление'
    ws['g5'].style = ali_grey

    
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level_2 = re.compile(r'level2')
    
    obj_parent_title         = request.session['obj_parent_title']
    obj_title         = request.session['obj_title']
    electric_data_end   = request.session['electric_data_end']            
    obj_key             = request.session['obj_key']
    electric_data_start = request.session['electric_data_start']
    
    data_table = []
    if (bool(is_abonent_level.search(obj_key))): 
        data_table = common_sql.get_data_table_water_period_elf(obj_title, obj_parent_title, electric_data_start, electric_data_end, True)
      
    elif (bool(is_object_level_2.search(obj_key))):
        data_table = common_sql.get_data_table_water_period_elf(obj_title, obj_parent_title, electric_data_start, electric_data_end, False)
       

    #zamenyem None na N/D vezde
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None)
        
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
                  
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][4])  # Потребление
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][5])  # Время работы
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('g%s'%(row)).value = '%s' % (data_table[row-6][6])  # Показания по теплу на начало
            ws.cell('g%s'%(row)).style = ali_white
        except:
            ws.cell('g%s'%(row)).style = ali_white
            next

    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 17
    ws.column_dimensions['g'].width = 18
#____________
   

    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'potreblenie_elf_report'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_start+u'_'+electric_data_end
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_balance_period_electric(request):
    response = StringIO.StringIO()
    wb = Workbook()
    
    ws = wb.active
    ws.title = "percent_noBalance"
    
    obj_parent_title         =  request.session['obj_parent_title']
    obj_title         =         request.session['obj_title']
    electric_data_start   =     request.session['electric_data_start'] 
    electric_data_end   =       request.session['electric_data_end'] 
    is_abonent_level =          re.compile(r'abonent')
    obj_key             =       request.session['obj_key']
#Шапка
    ws.merge_cells('A1:E1')
    ws['A1'] = 'Балансная группа: ' + obj_title
    
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Баланс по электричеству c ' + str(electric_data_start) + ' по ' + str(electric_data_end)
    
    ws['A5'] = 'Дата'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Небаланс, кВт*ч'
    ws['B5'].style = ali_grey
    
    ws['C5'] = '% '
    ws['C5'].style = ali_grey
    
       
#Запрашиваем данные для отчета
          
    dtAll=[]
    dt_type_abon=common_sql.GetSimpleTable('types_abonents',"","")
    
    for i in range(0,len(dt_type_abon)):
         guid_type_abon=dt_type_abon[i][0]         
         if not(bool(is_abonent_level.search(obj_key))):
             data_table = common_sql.get_data_table_balance_electric_perid(obj_parent_title, obj_title,electric_data_start, electric_data_end,guid_type_abon)
             #type_abon=translate(dt_type_abon[i][1])             
             if len(data_table)>0: data_table=common_sql.ChangeNull(data_table, None)          
             dtAll.append(data_table)
             
    dt_delta=[]   
   
    if len(dtAll)>0:
        for j in range(1,len(dtAll[0])):
            sumD=0
            vv=0
            for i in range(0,len(dtAll)):                
                if (dtAll[i][j][6] == 'Н/Д' or dtAll[i][j][6] == None  or dtAll[i][j][6] == 'None'): 
                    if (j+1)<len(dtAll[0]): j+=1                             
                if dtAll[i][j][1] == True:                   
                    sumD+=decimal.Decimal(dtAll[i][j][6])
                    vv=decimal.Decimal(dtAll[i][j][6])                    
                else:
                    sumD-=decimal.Decimal(dtAll[i][j][6])
            #считаем проценты
            percent=0           
            if (vv > decimal.Decimal(0)):
                percent=sumD*100/vv
            #print data_table[j][5]
            dt_delta.append([data_table[j][5],sumD, decimal.Decimal(percent)])

        
# Заполняем отчет значениями
    for row in range(6, len(dt_delta)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (dt_delta[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (dt_delta[row-6][1])  # счётчик эльф
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (round(dt_delta[row-6][2],2))  # счётчик хвс
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
       
    ws.row_dimensions[5].height = 41
    ws.column_dimensions['A'].width = 17 
    ws.column_dimensions['B'].width = 17 
    ws.column_dimensions['C'].width = 17
   
    #___________________________________________________
   
    
    
    for val in dtAll: 
        #print val[1], val[2], val[3]
        ws2 = wb.create_sheet(title="balance_detail_"+translate(val[0][2]))
        ws2.merge_cells('A1:E1')
        ws2['A1'] = 'Балансная группа: ' + obj_title
        
        ws2.merge_cells('A2:E2')
        ws2['A2'] = 'Баланс по электричеству c' + str(electric_data_start) + ' по ' + str(electric_data_end)
        
        ws2['A5'] = 'Тип'
        ws2['A5'].style = ali_grey
        
        ws2['B5'] = 'Сумма T0, кВт*ч'
        ws2['B5'].style = ali_grey
        
        ws2['C5'] = 'Дата '
        ws2['C5'].style = ali_grey
        
        ws2['d5'] = 'Потребление T0, кВт*ч'
        ws2['d5'].style = ali_grey
        
        ws2['e5'] = 'Опрошено счётчиков'
        ws2['e5'].style = ali_grey
        #print val
        for row in range(6, len(val)+6):
            try:
                ws2.cell('A%s'%(row)).value = '%s' % (val[row-6][2])  # Абонент
                ws2.cell('A%s'%(row)).style = ali_white
            except:
                ws2.cell('A%s'%(row)).style = ali_white
                next
            
            try:
                ws2.cell('B%s'%(row)).value = '%s' % (val[row-6][3])  # счётчик эльф
                ws2.cell('B%s'%(row)).style = ali_white
            except:
                ws2.cell('B%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('C%s'%(row)).value = '%s' % (val[row-6][5])  # счётчик хвс
                ws2.cell('C%s'%(row)).style = ali_white
            except:
                ws2.cell('C%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('d%s'%(row)).value = '%s' % (val[row-6][6])  # счётчик хвс
                ws2.cell('d%s'%(row)).style = ali_white
            except:
                ws2.cell('d%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('e%s'%(row)).value = '%s' % (val[row-6][7])  # счётчик хвс
                ws2.cell('e%s'%(row)).style = ali_white
            except:
                ws2.cell('e%s'%(row)).style = ali_white
                next
       
        ws2.row_dimensions[5].height = 41
        ws2.column_dimensions['A'].width = 17 
        ws2.column_dimensions['B'].width = 17 
        ws2.column_dimensions['C'].width = 17    
    
#------------

                    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'balance_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+electric_data_start+'-'+electric_data_end
    print output_name
    file_ext = u'xlsx'
    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    
    return response
    
def report_all_res_by_date_v2(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    obj_title = u'Не выбран'
    obj_key = u'Не выбран'
    obj_parent_title = u'Не выбран'
    electric_data_end = u''
    obj_title           = request.session['obj_title']
    obj_key             = request.session['obj_key']
    obj_parent_title    = request.session['obj_parent_title']       
    electric_data_end   = request.session['electric_data_end']  
    

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Отчёт по всем ресурсам на ' + str(electric_data_end) + u'. ' + unicode(obj_parent_title) + u' ' + unicode(obj_title)
    

    ws['A5'] = 'Абонент'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Тип показаний'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Счётчик'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Показания Энергии на ' + str(electric_data_end)
    ws['D5'].style = ali_grey
    
  
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')
    is_object_level = re.compile(r'level')   
    
    decimal.getcontext().prec = 3   
               
    data_table=[]
    obj_parent_title_water=u""
    obj_parent_title_electric=u""
    if (bool(is_abonent_level.search(obj_key))):      
            obj_parent_title_water=unicode(obj_parent_title)+u" Вода"
            obj_parent_title_electric=obj_parent_title
            #print  obj_parent_title_water, obj_parent_title_electric, len( obj_parent_title_electric)
            data_table = common_sql.get_data_table_all_res_for_abon(obj_parent_title_water, obj_parent_title_electric, obj_title, electric_data_end)
    if (bool(is_object_level.search(obj_key))):
        #здесь условно на уровне абонента
        n=unicode(obj_parent_title).find(u'Вода')
        #print obj_parent_title, n
        if (n>0):
            obj_parent_title_water=obj_parent_title
            obj_parent_title_electric=obj_parent_title[0:n-1]
            #print  obj_parent_title_water, obj_parent_title_electric, len( obj_parent_title_electric)
        data_table = common_sql.get_data_table_all_res_for_abon(obj_parent_title_water, obj_parent_title_electric, obj_title, electric_data_end)
    if len(data_table)>0: 
        data_table=common_sql.ChangeNull(data_table, None) 

    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][2]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (round(float(data_table[row-6][3]),3)) # '%s' % (data_table[row-6][3])  # Показания по теплу на конец
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
    ws.row_dimensions[5].height = 51
    ws.column_dimensions['A'].width = 30 
    ws.column_dimensions['b'].width = 20 
    ws.column_dimensions['c'].width = 20 
   
    #ws.column_dimensions['H'].width = 15
    
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_all_res_'+translate(obj_parent_title)+'_'+translate(obj_title)+'_'+str(electric_data_end)
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_electric_res_status(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]  

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Статистика опроса по электричеству на ' + str(electric_data_end)
    

    ws['A5'] = 'Объект'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Опрошено счётчиков'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Всего счётчиков'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Процент опроса'
    ws['D5'].style = ali_grey
    
    ws['e5'] = 'Не опрошено счётчиков'
    ws['e5'].style = ali_grey
    
  
#Запрашиваем данные для отчета
    
    obj_title           = request.session['obj_title']
    obj_parent_title    = request.session['obj_parent_title']       
    electric_data_end   = request.session['electric_data_end']  
               
    dt_objects = common_sql.get_res_objects('electric')

    dtAll_statistic=[]
    dtAll_no_data_meters=[]
    for obj in dt_objects:
        dt_statistic= common_sql.get_electric_count(obj[0],  electric_data_end)
        dt_no_data_meters=common_sql.get_electric_no_data(obj[0],  electric_data_end)
        dtAll_statistic.append(dt_statistic)
        if len(dt_no_data_meters)>0: 
            dt_no_data_meters=common_sql.ChangeNull(dt_no_data_meters, None)
            dtAll_no_data_meters.append(dt_no_data_meters)

    
# Заполняем отчет значениями
    for row in range(6, len(dtAll_statistic)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][2]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][3]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('d%s'%(row)).style = ali_white
        except:
            ws.cell('d%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][4]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('e%s'%(row)).style = ali_white
        except:
            ws.cell('e%s'%(row)).style = ali_white
            next
            
    ws.row_dimensions[5].height = 51
    ws.column_dimensions['A'].width = 30 
    ws.column_dimensions['b'].width = 20 
    ws.column_dimensions['c'].width = 20 
   
    #ws.column_dimensions['H'].width = 
    #___________________________________________________
    
    #print val[1], val[2], val[3]
    ws2 = wb.create_sheet(title="electric_detail")
    ws2.merge_cells('A1:E1')
           
    ws2.merge_cells('A2:E2')
    ws2['A2'] = 'Не ответившие счётчики на ' + str(electric_data_end)
    
    ws2['A5'] = 'Объект'
    ws2['A5'].style = ali_grey
    
    ws2['B5'] = 'Абонент'
    ws2['B5'].style = ali_grey
    
    ws2['C5'] = 'Счётчик '
    ws2['C5'].style = ali_grey
    
    ws2['d5'] = 'T0, кВт*ч'
    ws2['d5'].style = ali_grey
    
    ws2['e5'] = 'T1, кВт*ч'
    ws2['e5'].style = ali_grey
    
    ws2['f5'] = 'T2, кВт*ч'
    ws2['f5'].style = ali_grey
    
    ws2['g5'] = 'T3, кВт*ч'
    ws2['g5'].style = ali_grey
    
        #print val
    row = 5
    for value in dtAll_no_data_meters:                 
        for val in value: 
            row+=1
            try:
                ws2.cell('A%s'%(row)).value = '%s' % (val[0])  # Абонент
                ws2.cell('A%s'%(row)).style = ali_white
            except:
                ws2.cell('A%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('b%s'%(row)).value = '%s' % (val[1])  # Абонент
                ws2.cell('b%s'%(row)).style = ali_white
            except:
                ws2.cell('b%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('c%s'%(row)).value = '%s' % (val[2])  # Абонент
                ws2.cell('c%s'%(row)).style = ali_white
            except:
                ws2.cell('c%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('d%s'%(row)).value = '%s' % (val[3])  # Абонент
                ws2.cell('d%s'%(row)).style = ali_white
            except:
                ws2.cell('d%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('e%s'%(row)).value = '%s' % (val[4])  # Абонент
                ws2.cell('e%s'%(row)).style = ali_white
            except:
                ws2.cell('e%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('f%s'%(row)).value = '%s' % (val[5])  # Абонент
                ws2.cell('f%s'%(row)).style = ali_white
            except:
                ws2.cell('f%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('g%s'%(row)).value = '%s' % (val[6])  # Абонент
                ws2.cell('g%s'%(row)).style = ali_white
            except:
                ws2.cell('g%s'%(row)).style = ali_white
                next
            
   
    ws2.row_dimensions[5].height = 51
    ws2.column_dimensions['A'].width = 30 
    ws2.column_dimensions['b'].width = 20 
    #ws2.column_dimensions['c'].width = 20 
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_electric_statistic_'+str(electric_data_end)
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_heat_res_status(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]  

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Статистика опроса по теплу на ' + str(electric_data_end)
    

    ws['A5'] = 'Объект'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Опрошено счётчиков'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Всего счётчиков'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Процент опроса'
    ws['D5'].style = ali_grey
    
    ws['e5'] = 'Не опрошено счётчиков'
    ws['e5'].style = ali_grey
    
  
#Запрашиваем данные для отчета
    
    obj_title           = request.session['obj_title']
    obj_parent_title    = request.session['obj_parent_title']       
    electric_data_end   = request.session['electric_data_end']  
               
    dt_objects = common_sql.get_res_objects('heat')
    #print 'print len(dt_objects) ', len(dt_objects)
    dtAll_statistic=[]
    dtAll_no_data_meters=[]
    for obj in dt_objects:
        #print obj[0]
        dt_statistic= common_sql.get_heat_count(obj[0],  electric_data_end)
        dt_no_data_meters=common_sql.get_heat_no_data(obj[0],  electric_data_end)
        dtAll_statistic.append(dt_statistic)
        if len(dt_no_data_meters)>0: 
            dt_no_data_meters=common_sql.ChangeNull(dt_no_data_meters, None)
            dtAll_no_data_meters.append(dt_no_data_meters)

    
# Заполняем отчет значениями
    for row in range(6, len(dtAll_statistic)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][2]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][3]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('d%s'%(row)).style = ali_white
        except:
            ws.cell('d%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][4]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('e%s'%(row)).style = ali_white
        except:
            ws.cell('e%s'%(row)).style = ali_white
            next
            
    ws.row_dimensions[5].height = 51
    ws.column_dimensions['A'].width = 30 
    ws.column_dimensions['b'].width = 20 
    ws.column_dimensions['c'].width = 20 
   
    #ws.column_dimensions['H'].width = 
    #___________________________________________________
    
    #print val[1], val[2], val[3]
    ws2 = wb.create_sheet(title="heat_detail")
    ws2.merge_cells('A1:E1')
           
    ws2.merge_cells('A2:E2')
    ws2['A2'] = 'Не ответившие счётчики на ' + str(electric_data_end)
    
    ws2['A5'] = 'Объект'
    ws2['A5'].style = ali_grey
    
    ws2['B5'] = 'Абонент'
    ws2['B5'].style = ali_grey
    
    ws2['C5'] = 'Счётчик '
    ws2['C5'].style = ali_grey
    
    ws2['d5'] = 'Энергия, Гкал'
    ws2['d5'].style = ali_grey
    
    ws2['e5'] = 'Объем, м3'
    ws2['e5'].style = ali_grey
    
    ws2['f5'] = 'Температура входа, С'
    ws2['f5'].style = ali_grey
    
    ws2['g5'] = 'Температура выхода, С'
    ws2['g5'].style = ali_grey
    
        #print val
    row = 5
    for value in dtAll_no_data_meters:                 
        for val in value: 
            row+=1
            try:
                ws2.cell('A%s'%(row)).value = '%s' % (val[0])  # Абонент
                ws2.cell('A%s'%(row)).style = ali_white
            except:
                ws2.cell('A%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('b%s'%(row)).value = '%s' % (val[1])  # Абонент
                ws2.cell('b%s'%(row)).style = ali_white
            except:
                ws2.cell('b%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('c%s'%(row)).value = '%s' % (val[2])  # Абонент
                ws2.cell('c%s'%(row)).style = ali_white
            except:
                ws2.cell('c%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('d%s'%(row)).value = '%s' % (val[3])  # Абонент
                ws2.cell('d%s'%(row)).style = ali_white
            except:
                ws2.cell('d%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('e%s'%(row)).value = '%s' % (val[4])  # Абонент
                ws2.cell('e%s'%(row)).style = ali_white
            except:
                ws2.cell('e%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('f%s'%(row)).value = '%s' % (val[5])  # Абонент
                ws2.cell('f%s'%(row)).style = ali_white
            except:
                ws2.cell('f%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('g%s'%(row)).value = '%s' % (val[6])  # Абонент
                ws2.cell('g%s'%(row)).style = ali_white
            except:
                ws2.cell('g%s'%(row)).style = ali_white
                next
            
   
    ws2.row_dimensions[5].height = 51
    ws2.column_dimensions['A'].width = 30 
    ws2.column_dimensions['b'].width = 20 
    #ws2.column_dimensions['c'].width = 20 
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_heat_statistic_'+str(electric_data_end)
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_water_impulse_res_status(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]  

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Статистика опроса по воде(импульсные ПУ) на ' + str(electric_data_end)
    

    ws['A5'] = 'Объект'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Опрошено счётчиков'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Всего счётчиков'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Процент опроса'
    ws['D5'].style = ali_grey
    
    ws['e5'] = 'Не опрошено счётчиков'
    ws['e5'].style = ali_grey
    
  
#Запрашиваем данные для отчета
          
    dt_objects = common_sql.get_water_impulse_objects()
    
    dtAll_statistic=[]
    dtAll_no_data_meters=[]
    for obj in dt_objects:
        #print electric_data_end
        dt_statistic= common_sql.get_water_impulse_count(obj[0],  electric_data_end)
        dt_no_data_meters=common_sql.get_water_impulse_no_data(obj[0],  electric_data_end)
        dtAll_statistic.append(dt_statistic)
        if len(dt_no_data_meters)>0: 
            dt_no_data_meters=common_sql.ChangeNull(dt_no_data_meters, None)
            dtAll_no_data_meters.append(dt_no_data_meters)

    
# Заполняем отчет значениями
    for row in range(6, len(dtAll_statistic)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][2]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][3]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('d%s'%(row)).style = ali_white
        except:
            ws.cell('d%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][4]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('e%s'%(row)).style = ali_white
        except:
            ws.cell('e%s'%(row)).style = ali_white
            next
            
    ws.row_dimensions[5].height = 51
    ws.column_dimensions['A'].width = 30 
    ws.column_dimensions['b'].width = 20 
    ws.column_dimensions['c'].width = 20 
   
    #ws.column_dimensions['H'].width = 
    #___________________________________________________
    
    #print val[1], val[2], val[3]
    ws2 = wb.create_sheet(title="water_detail")
    ws2.merge_cells('A1:E1')
           
    ws2.merge_cells('A2:E2')
    ws2['A2'] = 'Не ответившие счётчики на ' + str(electric_data_end)
    
    ws2['A5'] = 'Объект'
    ws2['A5'].style = ali_grey
    
    ws2['B5'] = 'Абонент'
    ws2['B5'].style = ali_grey
    
    ws2['C5'] = 'Счётчик '
    ws2['C5'].style = ali_grey
    
    ws2['d5'] = 'Регистратор'
    ws2['d5'].style = ali_grey
    
    ws2['e5'] = 'Канал'
    ws2['e5'].style = ali_grey
    
    ws2['f5'] = 'Показания'
    ws2['f5'].style = ali_grey
    
    
        #print val
    row = 5
    for value in dtAll_no_data_meters:                 
        for val in value: 
            row+=1
            try:
                ws2.cell('A%s'%(row)).value = '%s' % (val[0])  # Абонент
                ws2.cell('A%s'%(row)).style = ali_white
            except:
                ws2.cell('A%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('b%s'%(row)).value = '%s' % (val[1])  # Абонент
                ws2.cell('b%s'%(row)).style = ali_white
            except:
                ws2.cell('b%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('c%s'%(row)).value = '%s' % (val[2])  # Абонент
                ws2.cell('c%s'%(row)).style = ali_white
            except:
                ws2.cell('c%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('d%s'%(row)).value = '%s' % (val[3])  # Абонент
                ws2.cell('d%s'%(row)).style = ali_white
            except:
                ws2.cell('d%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('e%s'%(row)).value = '%s' % (val[4])  # Абонент
                ws2.cell('e%s'%(row)).style = ali_white
            except:
                ws2.cell('e%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('f%s'%(row)).value = '%s' % (val[5])  # Абонент
                ws2.cell('f%s'%(row)).style = ali_white
            except:
                ws2.cell('f%s'%(row)).style = ali_white
                next
                            
   
    ws2.row_dimensions[5].height = 51
    ws2.column_dimensions['A'].width = 35 
    ws2.column_dimensions['b'].width = 20 
    ws2.column_dimensions['c'].width = 30 
    ws2.column_dimensions['d'].width = 30 
    #ws2.column_dimensions['c'].width = 20 
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_water_impulse_statistic_'+str(electric_data_end)
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
    
def report_heat_res_status(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]  

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Статистика опроса по теплу на ' + str(electric_data_end)
    

    ws['A5'] = 'Объект'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Опрошено счётчиков'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Всего счётчиков'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Процент опроса'
    ws['D5'].style = ali_grey
    
    ws['e5'] = 'Не опрошено счётчиков'
    ws['e5'].style = ali_grey
    
  
#Запрашиваем данные для отчета
    
    obj_title           = request.session['obj_title']
    obj_parent_title    = request.session['obj_parent_title']       
    electric_data_end   = request.session['electric_data_end']  
               
    dt_objects = common_sql.get_res_objects('heat')
    #print 'print len(dt_objects) ', len(dt_objects)
    dtAll_statistic=[]
    dtAll_no_data_meters=[]
    for obj in dt_objects:
        #print obj[0]
        dt_statistic= common_sql.get_heat_count(obj[0],  electric_data_end)
        dt_no_data_meters=common_sql.get_heat_no_data(obj[0],  electric_data_end)
        dtAll_statistic.append(dt_statistic)
        if len(dt_no_data_meters)>0: 
            dt_no_data_meters=common_sql.ChangeNull(dt_no_data_meters, None)
            dtAll_no_data_meters.append(dt_no_data_meters)

    
# Заполняем отчет значениями
    for row in range(6, len(dtAll_statistic)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][1])  # заводской номер
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][2]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('d%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][3]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('d%s'%(row)).style = ali_white
        except:
            ws.cell('d%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('e%s'%(row)).value = '%s' % (dtAll_statistic[row-6][0][4]) # '%s' % (data_table[row-6][2])  # Показания по теплу на начало
            ws.cell('e%s'%(row)).style = ali_white
        except:
            ws.cell('e%s'%(row)).style = ali_white
            next
            
    ws.row_dimensions[5].height = 51
    ws.column_dimensions['A'].width = 30 
    ws.column_dimensions['b'].width = 20 
    ws.column_dimensions['c'].width = 20 
   
    #ws.column_dimensions['H'].width = 
    #___________________________________________________
    
    #print val[1], val[2], val[3]
    ws2 = wb.create_sheet(title="heat_detail")
    ws2.merge_cells('A1:E1')
           
    ws2.merge_cells('A2:E2')
    ws2['A2'] = 'Не ответившие счётчики на ' + str(electric_data_end)
    
    ws2['A5'] = 'Объект'
    ws2['A5'].style = ali_grey
    
    ws2['B5'] = 'Абонент'
    ws2['B5'].style = ali_grey
    
    ws2['C5'] = 'Счётчик '
    ws2['C5'].style = ali_grey
    
    ws2['d5'] = 'Энергия, Гкал'
    ws2['d5'].style = ali_grey
    
    ws2['e5'] = 'Объем, м3'
    ws2['e5'].style = ali_grey
    
    ws2['f5'] = 'Температура входа, С'
    ws2['f5'].style = ali_grey
    
    ws2['g5'] = 'Температура выхода, С'
    ws2['g5'].style = ali_grey
    
        #print val
    row = 5
    for value in dtAll_no_data_meters:                 
        for val in value: 
            row+=1
            try:
                ws2.cell('A%s'%(row)).value = '%s' % (val[0])  # Абонент
                ws2.cell('A%s'%(row)).style = ali_white
            except:
                ws2.cell('A%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('b%s'%(row)).value = '%s' % (val[1])  # Абонент
                ws2.cell('b%s'%(row)).style = ali_white
            except:
                ws2.cell('b%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('c%s'%(row)).value = '%s' % (val[2])  # Абонент
                ws2.cell('c%s'%(row)).style = ali_white
            except:
                ws2.cell('c%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('d%s'%(row)).value = '%s' % (val[3])  # Абонент
                ws2.cell('d%s'%(row)).style = ali_white
            except:
                ws2.cell('d%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('e%s'%(row)).value = '%s' % (val[4])  # Абонент
                ws2.cell('e%s'%(row)).style = ali_white
            except:
                ws2.cell('e%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('f%s'%(row)).value = '%s' % (val[5])  # Абонент
                ws2.cell('f%s'%(row)).style = ali_white
            except:
                ws2.cell('f%s'%(row)).style = ali_white
                next
                
            try:
                ws2.cell('g%s'%(row)).value = '%s' % (val[6])  # Абонент
                ws2.cell('g%s'%(row)).style = ali_white
            except:
                ws2.cell('g%s'%(row)).style = ali_white
                next
            
   
    ws2.row_dimensions[5].height = 51
    ws2.column_dimensions['A'].width = 30 
    ws2.column_dimensions['b'].width = 20 
    #ws2.column_dimensions['c'].width = 20 
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_heat_statistic_'+str(electric_data_end)
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response
    
def report_balance_period_water_impulse(request):
    response = StringIO.StringIO()
    wb = Workbook()
    ws = wb.active
    electric_data_end   = request.session["electric_data_end"]  
    electric_data_start   = request.session["electric_data_start"] 

#Шапка
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Баланс по воде (импульсные приборы) c ' + str(electric_data_start) + u' по ' + str(electric_data_end)
    

    ws['A5'] = 'Дата'
    ws['A5'].style = ali_grey
    
    ws['B5'] = 'Потребление за сутки, м3'
    ws['B5'].style = ali_grey
    
    ws['C5'] = 'Подача за сутки, м3 (вводный приборы)'
    ws['C5'].style = ali_grey
    
    ws['D5'] = 'Небаланс, м3'
    ws['D5'].style = ali_grey
    
    ws['e5'] = 'Процент, %'
    ws['e5'].style = ali_grey
    
    ws['f5'] = 'Опрошено счётчиков'
    ws['f5'].style = ali_grey
    
  
#Запрашиваем данные для отчета
    is_abonent_level = re.compile(r'abonent')      
    obj_title           = request.session['obj_title']
    obj_key             = request.session['obj_key']
    obj_parent_title    = request.session['obj_parent_title']    
            
   
    if not(bool(is_abonent_level.search(obj_key))):
         data_table = common_sql.get_data_table_balance_water_impulse_perid(obj_parent_title, obj_title,electric_data_start, electric_data_end)
    if (len( data_table) >0):                
        data_table=common_sql.ChangeNull(data_table, None)

    
# Заполняем отчет значениями
    for row in range(6, len(data_table)+6):
        try:
            ws.cell('A%s'%(row)).value = '%s' % (data_table[row-6][0])  # Абонент
            ws.cell('A%s'%(row)).style = ali_white
        except:
            ws.cell('A%s'%(row)).style = ali_white
            next
        
        try:
            ws.cell('B%s'%(row)).value = '%s' % (data_table[row-6][3])  # тип
            ws.cell('B%s'%(row)).style = ali_white
        except:
            ws.cell('B%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('C%s'%(row)).value = '%s' % (data_table[row-6][6])  # стояк
            ws.cell('C%s'%(row)).style = ali_white
        except:
            ws.cell('C%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('D%s'%(row)).value = '%s' % (data_table[row-6][7])  # счётчик 
            ws.cell('D%s'%(row)).style = ali_white
        except:
            ws.cell('D%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('E%s'%(row)).value = '%s' % (data_table[row-6][8])  # значения на начало
            ws.cell('E%s'%(row)).style = ali_white
        except:
            ws.cell('E%s'%(row)).style = ali_white
            next
            
        try:
            ws.cell('F%s'%(row)).value = '%s' % (data_table[row-6][9])  # значения на конец
            ws.cell('F%s'%(row)).style = ali_white
        except:
            ws.cell('F%s'%(row)).style = ali_white
            next
            
    ws.row_dimensions[5].height = 51
    ws.column_dimensions['A'].width = 20 
    ws.column_dimensions['b'].width = 20 
    ws.column_dimensions['c'].width = 20 
   
    
    wb.save(response)
    response.seek(0)
    response = HttpResponse(response.read(), content_type="application/vnd.ms-excel")
    #response['Content-Disposition'] = "attachment; filename=profil.xlsx"
    
    output_name = u'report_water_impulse_balance_'+str(electric_data_end)
    file_ext = u'xlsx'    
    response['Content-Disposition'] = 'attachment;filename="%s.%s"' % (output_name.replace('"', '\"'), file_ext)   
    return response