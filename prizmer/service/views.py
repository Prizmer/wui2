# coding -*- coding: utf-8 -*-

from __future__ import unicode_literals
from django.shortcuts import render
from django.shortcuts import render_to_response, HttpResponse
from django.http import HttpResponseRedirect
from django.shortcuts import redirect
from django import forms
from django.core.context_processors import csrf
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from django.db import connection
#from general.models import Objects, Abonents, TypesAbonents, Meters, MonthlyValues, DailyValues, CurrentValues, VariousValues, TypesParams, Params, TakenParams, LinkAbonentsTakenParams, Resources, TypesMeters, Measurement, NamesParams, BalanceGroups, LinkMetersComportSettings, LinkMetersTcpipSettings, ComportSettings, TcpipSettings, LinkBalanceGroupsMeters, Groups80020, LinkGroups80020Meters
from general.models import  Objects, Abonents, TcpipSettings, TypesAbonents, Meters, TypesMeters,LinkAbonentsTakenParams,LinkMetersComportSettings, LinkMetersTcpipSettings, ComportSettings,  TakenParams,Params
from django.db.models.signals import pre_save
from django.db.models.signals import post_save
from django.db.models import signals
import datetime
from django.db.models import Max 
import uuid

cfg_excel_name=""
cfg_sheet_name=""

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
# Create your views here.

isService=False
    
    
class UploadFileForm(forms.Form):
    #title = forms.CharField(max_length=150)
    path  = forms.FileField()

def MakeSheet(request):
    args={}
    fileName=""
    sheets=""
    print request.GET['choice_file']
    print '___________'
    print request.GET.get('choice_file')
    if request.is_ajax():
        if request.method == 'GET':
            request.session["choice_file"]    = fileName    = request.GET['choice_file']
            #print fileName
            directory=os.path.join(BASE_DIR,'static/cfg/')
            wb=load_workbook(directory+fileName)
            sheets=wb.sheetnames

    args['sheets']=sheets
    return render_to_response("service/service_sheets_excel.html", args)

def writeToLog(msg):
    #msg=unicode(msg)
#    directory=os.path.join(BASE_DIR,'static\\log\\')
#    if  not(os.path.exists(directory)):
#        os.mkdir(directory)
#    dir_date=datetime.datetime.now().strftime("%d-%m-%Y")        
#    if  not(os.path.exists(directory+dir_date)):
#        os.mkdir(directory+dir_date)  
#        
#    path=directory+dir_date+'\log.txt'
#     
#    f = open(path, 'w')
#    f.write(msg)
#    f.close()
    pass

def choose_service(request):
    args={}
    directory=os.path.join(BASE_DIR,'static\\cfg\\')
    
    if  not(os.path.exists(directory)):
        os.mkdir(directory)
    print directory
    files = os.listdir(directory) 
    print files
    args['filesFF']= files
    return render_to_response("choose_service.html", args)

@csrf_exempt
def service_electric(request):
    args={}
    return render_to_response("service/service_electric.html", args)


def service_file(request):
    args={}
    args.update(csrf(request))    
    data_table=[]
    status='file not loaded'
    args['data_table'] = data_table
    args['status']=status

    return render_to_response("service/service_file.html", args)
    
def service_file_loading(request):
    args={}
    data_table=[]
    status='file not loaded'
    sPath=""
    if request.method == 'POST':        
        form = UploadFileForm(request.POST, request.FILES)
        #print form.as_table()
        #print form.is_valid()
        
        #print sPath
        if form.is_valid():
            sPath=os.path.join(BASE_DIR,'static/cfg/'+request.FILES['path'].name)
            handle_uploaded_file(request.FILES['path'])
            status= u'Файл загружен'
    else:
        form = UploadFileForm()
        
    args['data_table'] = data_table
    args['status']=status
    args['sPath']=sPath
    #print status
    return render_to_response("choose_service.html", args)

    
def service_electric_load(request):
    args={}
    data_table=[]
    status='file not loaded'

    if request.method == 'POST':

        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():
            handle_uploaded_file(request.FILES['path'])
            status='file loaded'
    else:
        form = UploadFileForm()
        
    args['data_table'] = data_table
    args['status']=status
    return render_to_response("service/service_electric.html", args)
    #return render_to_response("service/service_electric_load.html", args)
    
def handle_uploaded_file(f):

    destination = open(os.path.join(BASE_DIR,'static/cfg/'+f.name), 'wb+')
    for chunk in f.chunks():
        destination.write(chunk)
    #print 'file load'
    destination.close()
    
def load_port(request):
    args={}
    #print '!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!'
    fileName=""
    sheet    = ""
    tcp_ip_status    = ""
    object_status    = ""
    counter_status    = ""
    result=""
    if request.is_ajax():
        if request.method == 'GET':
            request.session["choice_file"]    = fileName    = request.GET['choice_file']
            request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']
            request.session["tcp_ip_status"]    = tcp_ip_status    = request.GET['tcp_ip_status']
            request.session["object_status"]    = object_status    = request.GET['object_status']
            request.session["counter_status"]    = counter_status    = request.GET['counter_status']
            
            #print fileName
            directory=os.path.join(BASE_DIR,'static/cfg/')
            sPath=directory+fileName
            #print sPath, sheet
            result=load_tcp_ip_or_com_ports_from_excel(sPath, sheet)
    writeToLog(result)
    if result:
        tcp_ip_status=u"Порт/ы был успешно добавлен"
    else:
        tcp_ip_status=u"Порт не был загружен, он уже существует в БД"
    
    
    #print fileName
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["tcp_ip_status"]=tcp_ip_status
    args["object_status"]=object_status
    args["counter_status"]=counter_status
    return render_to_response("service/service_electric.html", args)

def checkPortIsExist(ip_adr,ip_port):
    dt_ports=[]
    cursor = connection.cursor()
    sQuery="""
    SELECT guid, ip_address, ip_port, write_timeout, read_timeout, attempts, 
       delay_between_sending
  FROM tcpip_settings
  where ip_address='%s' and  ip_port='%s'"""%(unicode(ip_adr).rstrip(),unicode(ip_port).rstrip())
    #print sQuery
    cursor.execute(sQuery)
    dt_ports = cursor.fetchall()
    #print dt_ports
    if len(dt_ports):  
        return False
    else: 
        return True

def load_tcp_ip_or_com_ports_from_excel(sPath, sSheet):
    #Добавление tcp_ip портов
    global cfg_excel_name
    cfg_excel_name=sPath
    global cfg_sheet_name
    cfg_sheet_name=sSheet
    wb = load_workbook(filename = sPath)
    sheet_ranges = wb[sSheet]
    row = 2
    result=""
    IsAdded=False
    portType=sheet_ranges[u'M1'].value
    while (bool(sheet_ranges[u'G%s'%(row)].value)):
        if sheet_ranges[u'G%s'%(row)].value is not None:
            writeToLog(u'Обрабатываем строку ' + str(u'G%s '%(row)) + str(sheet_ranges[u'G%s'%(row)].value))
            ip_adr=sheet_ranges[u'K%s'%(row)].value
            ip_port=sheet_ranges[u'L%s'%(row)].value
            com_port=sheet_ranges[u'M%s'%(row)].value
            #print ip_adr, ip_port
            if portType==u'Com-port': #добавление com-порта
                writeToLog(com_port)
                if not com_port or com_port==None: 
                    result+="Отсутствует значение для com-порта в строке"+str(row)+". Заполните все ячейки excel таблицы."
                    break
                if not (SimpleCheckIfExist('comport_settings','name', com_port, "", "", "")):
                    add_port=ComportSettings(name=unicode(com_port).rstrip(),baudrate=9600,data_bits=8,parity=0,stop_bits=1, write_timeout=100, read_timeout=100, attempts=2, delay_between_sending=100)
                    add_port.save()
                    result+=u"Новый com-порт добавлен"
                    IsAdded=True
                else: result= u'Порт '+str(com_port)+u" уже существует"
            else:
                # проверка есть ли уже такой порт, запрос в БД с адресом и портом, если ответ пустой-добавляем, в противном случае continue
                if not ip_adr or not ip_port or ip_adr==None or ip_port==None: 
                    result+="Отсутствует значение/я для tcp/ip-порта в строке"+str(row)+". Заполните все ячейки excel таблицы."
                    break
                else:
                    if (checkPortIsExist(ip_adr,ip_port)):
                        add_port=TcpipSettings(ip_address = unicode(ip_adr).rstrip(), ip_port =int(ip_port), write_timeout =300 , read_timeout =700 , attempts =3 , delay_between_sending =400)
                        add_port.save()
                        result =u'Новый tcp/ip порт добавлен'
                        IsAdded=True
    #                add_meter = Meters(name = unicode(sheet_ranges[u'F%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'E%s'%(row)].value), address = unicode(sheet_ranges[u'E%s'%(row)].value),  factory_number_manual = unicode(sheet_ranges[u'E%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"7cd88751-d232-410c-a0ef-6354a79112f1") )
    #                add_meter.save()
                    else: result+= u'Порт '+str(ip_adr)+": "+str(ip_port)+u" уже существует"
        writeToLog( result)
        row+=1
    return IsAdded

def SimpleCheckIfExist(table1,fieldName1, value1, table2, fieldName2, value2):
    dt=[]
    cursor = connection.cursor()
    if len(table2)==0: #проверка для одной таблицы
        sQuery="""
        Select *
        from %s
        where %s.%s='%s'"""%(table1, table1, fieldName1, value1)
    else:#проверка для двух сводных таблиц
        sQuery="""
        Select *
        from %s, %s
        where %s.guid_%s=%s.guid and
        %s.%s='%s' and
        %s.%s='%s'
        """%(table1,table2, table2, table1,table1, table1, fieldName1, value1,table2, fieldName2, value2)
    #print sQuery
    #print bool(dt)
    cursor.execute(sQuery)
    dt = cursor.fetchall()

    if not dt:  
        return False
    else: 
        return True
    
def GetSimpleTable(table,fieldName,value):
    dt=[]
    cursor = connection.cursor()
    sQuery="""
        Select *
        from %s
        where %s.%s='%s'"""%(table, table, fieldName, value)
    #print sQuery
    cursor.execute(sQuery)
    dt = cursor.fetchall()
    return dt
    

def GetTableFromExcel(sPath,sSheet):
    wb = load_workbook(filename = sPath)
    ws = wb[sSheet]
    row = 1
    dt=[]
    while (bool(ws[u'A%s'%(row)].value)):
        A=ws[u'A%s'%(row)].value
        B=ws[u'b%s'%(row)].value
        C=ws[u'c%s'%(row)].value
        D=ws[u'd%s'%(row)].value
        E=ws[u'e%s'%(row)].value
        F=ws[u'f%s'%(row)].value
        G=ws[u'g%s'%(row)].value
        H=ws[u'h%s'%(row)].value
        I=ws[u'i%s'%(row)].value
        J=ws[u'j%s'%(row)].value
        K=ws[u'k%s'%(row)].value
        L=ws[u'l%s'%(row)].value
        M=ws[u'm%s'%(row)].value
        N=ws[u'n%s'%(row)].value
        
        vals =[A,B,C,D,E,F,G,H,I,J,K,L,M,N]
        dt.append(vals)
        row+=1
    return dt
    
def LoadObjectsAndAbons(sPath, sSheet):
    #Добавление объектов
    global cfg_excel_name
    cfg_excel_name=sPath
    global cfg_sheet_name
    cfg_sheet_name=sSheet
    result="Объекты не загружены"

    dtAll=GetTableFromExcel(sPath,sSheet) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    
    for i in range(1,len(dtAll)):
        #print  dtAll[i][2],dtAll[i][3]
        print u'Обрабатываем строку ' + dtAll[i][2]+' - '+dtAll[i][3]
        obj_l0=dtAll[i][0]
        writeToLog( obj_l0)
        obj_l1=dtAll[i][1]
        writeToLog(obj_l1)
        obj_l2=dtAll[i][2]
        writeToLog(obj_l2)
        abon=dtAll[i][3]
        writeToLog(abon)
        account_1=dtAll[i][4]
        writeToLog(account_1)
        account_2=dtAll[i][5]
        writeToLog(account_2)
        isNewObj_l0=SimpleCheckIfExist('objects','name',obj_l0,"","","")
        isNewObj_l1=SimpleCheckIfExist('objects','name',obj_l1,"","","")
        isNewObj_l2=SimpleCheckIfExist('objects','name',obj_l2,"","","")
        isNewAbon=SimpleCheckIfExist('objects','name', obj_l2,'abonents', 'name', abon)
        kv=0
        if not (isNewObj_l0):
            writeToLog('create object '+obj_l0)
            add_parent_object = Objects( name=obj_l0, level=0)
            add_parent_object.save()
            writeToLog('create object '+obj_l1)
            #print add_parent_object
            add_object1=Objects(name=obj_l1, level=1, guid_parent = add_parent_object)
            add_object1.save()
            writeToLog('create object '+obj_l2)
            add_object2=Objects(name=obj_l2, level=2, guid_parent = add_object1)
            add_object2.save()
            
            writeToLog('create abonent '+abon)
            add_abonent = Abonents(name = abon, account_1 =unicode(account_1), account_2 =unicode(account_2), guid_objects =add_object2, guid_types_abonents = TypesAbonents.objects.get(guid= u"e4d813ca-e264-4579-ae15-385cdbf5d28c"))
            add_abonent.save()
            result=u"Объекты: "+obj_l0+", "+obj_l1+u", "+obj_l2+u","+abon+u" созданы"
            continue
        if not (isNewObj_l1):
            writeToLog('create object '+obj_l1)
            dtParent=GetSimpleTable('objects','name',obj_l0)
            if dtParent: #родительский объект есть
                guid_parent=dtParent[0][0]
                add_object1=Objects(name=obj_l1, level=1, guid_parent = Objects.objects.get(guid=guid_parent))
                add_object1.save()                
                add_object2=Objects(name=obj_l2, level=2, guid_parent = add_object1)
                add_object2.save()
                writeToLog('create abonent '+abon)
                add_abonent = Abonents(name = abon, account_1 =unicode(account_1), account_2 =unicode(account_2), guid_objects =add_object2, guid_types_abonents = TypesAbonents.objects.get(guid= u"e4d813ca-e264-4579-ae15-385cdbf5d28c"))
                add_abonent.save()
                result+=u"Объекты: "+obj_l1+u", "+obj_l2+u","+abon+u" созданы"
                continue
        if not (isNewObj_l2):
            writeToLog('create object '+obj_l2)
            dtParent=GetSimpleTable('objects','name',obj_l1)
            if dtParent: #родительский объект есть
                guid_parent=dtParent[0][0]                
                add_object = Objects(name=obj_l2, level=2, guid_parent = Objects.objects.get(guid=guid_parent))
                add_object.save()
                result+=u"Объект: "+obj_l2+u" создан"
        if not (isNewAbon):
            writeToLog('create abonent '+ abon)
            dtObj=GetSimpleTable('objects','name',obj_l2)
            if dtObj: #родительский объект есть
                guid_object=dtObj[0][0]
                add_abonent = Abonents(name = abon, account_1 =unicode(account_1), account_2 =unicode(account_2), guid_objects = Objects.objects.get(guid=guid_object), guid_types_abonents = TypesAbonents.objects.get(guid= u"e4d813ca-e264-4579-ae15-385cdbf5d28c"))
                add_abonent.save()
                kv+=1

    result+=u" Прогружено "+str(kv)+u" абонентов"

    return result

def load_electric_objects(request):
    args={}
    fileName=""
    sheet    = ""
    tcp_ip_status    = ""
    object_status    = ""
    counter_status    = ""
    result="Не загружено"
    writeToLog('test1')
    
    if request.is_ajax():
        if request.method == 'GET':
            
            request.session["choice_file"]    = fileName    = request.GET['choice_file']
            request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']
            request.session["tcp_ip_status"]    = tcp_ip_status    = request.GET['tcp_ip_status']
            request.session["object_status"]    = object_status    = request.GET['object_status']
            request.session["counter_status"]    = counter_status    = request.GET['counter_status']
            
            directory=os.path.join(BASE_DIR,'static/cfg/')
            sPath=directory+fileName
            writeToLog(sPath)
                        
            print 'Path:_____',sPath, sheet
            result=LoadObjectsAndAbons(sPath, sheet)
    
    object_status=result#"Загрузка объектов условно прошла"

    #print fileName
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["tcp_ip_status"]=tcp_ip_status
    args["object_status"]=object_status
    args["counter_status"]=counter_status
    return render_to_response("service/service_electric.html", args)
    
def LoadElectricMeters(sPath, sSheet):
    global cfg_excel_name
    cfg_excel_name=sPath
    global cfg_sheet_name
    cfg_sheet_name=sSheet
    result=u"Счётчики не загружены"
    print type(sPath), sPath, type(sSheet), sSheet
    dtAll=GetTableFromExcel(sPath,sSheet) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    met=0
    print 'load dt - ok'
    for i in range(1,len(dtAll)):
        writeToLog(u'Обрабатываем строку ' + unicode(dtAll[i][3])+' - '+unicode(dtAll[i][6]))
        print unicode(dtAll[i][3]), unicode(dtAll[i][6])
        obj_l2=unicode(dtAll[i][2]) #корпус
        abon=unicode(dtAll[i][3]) #квартира
        meter=unicode(dtAll[i][6]) #номер счётчика
        adr=unicode(dtAll[i][7]) #номер в сети
        type_meter=unicode(dtAll[i][8]) #тип счётчика
        NumLic=unicode(dtAll[i][5]) #номер лицевого счёта, тут используется как пароль для м-230-ум
        Group=unicode(dtAll[i][12])
        attr1=unicode(dtAll[i][13])
#        print obj_l2
#        print abon
#        print meter
#        print adr
#        print type_meter
        isNewMeter=SimpleCheckIfExist('meters','factory_number_manual',meter,"","","")
        isNewAbon=SimpleCheckIfExist('objects','name', obj_l2,'abonents', 'name', abon)        
        
        #writeToLog( u'счётчик существует ', isNewMeter)
        if not (isNewAbon):
            return u"Сначала создайте стурктуру объектов и абонентов"
        if not (isNewMeter):
            
            #writeToLog('create meter '+meter +" adress: "+adr)
            
            if unicode(type_meter) == u'М-200':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"6224d20b-1781-4c39-8799-b1446b60774d") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'М-200')
                
                
            elif unicode(type_meter) == u'М-230':
                writeToLog('m-230')
#                print unicode(type_meter)
#                print unicode(meter)
#                print unicode(adr)
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), password = 111111 , factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"423b33a7-2d68-47b6-b4f6-5b470aedc4f4") )
#                print add_meter
#                print 'bryak'
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'М-230')
                
            elif unicode(type_meter) == u'М-230-УМ':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), password = unicode(NumLic) , factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"20e4767a-49e5-4f84-890c-25e311339c28") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'М-230-УМ')
                
            elif unicode(type_meter) == u'Эльф 1.08':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"1c5a8a80-1c51-4733-8332-4ed8d510a650") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'Эльф 1.08')
            elif unicode(type_meter) == u'СПГ762-1':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"c3ec5c22-d184-41c5-b6bf-66fa30215a41") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'СПГ762-1')
                
            elif unicode(type_meter) == u'СПГ762-2':
                add_meter = Meters(name=unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"5eb7dd59-faf9-4ead-8654-4f3de74de2b0") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'СПГ762-2')
            elif unicode(type_meter) == u'СПГ762-3':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"e4fb7950-a44f-41f0-a6ff-af5e30d9d562") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'СПГ762-3')
            elif unicode(type_meter) == u'Sayany':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"5429b439-233e-4944-b91b-4b521a10f77b") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'Sayany')
            elif unicode(type_meter) == u'Tekon_hvs':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), password = unicode(Group), guid_types_meters = TypesMeters.objects.get(guid = u"8398e7d6-39f7-45d2-9c45-a1c48e751b61") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'Tekon_gvs')
            elif unicode(type_meter) == u'Tekon_hvs':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), password = unicode(Group), guid_types_meters = TypesMeters.objects.get(guid = u"64f02a2c-41e1-48b2-bc72-7873ea9b6431") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'Tekon_gvs')

            elif unicode(type_meter) == u'Tekon_heat':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), password = unicode(Group), guid_types_meters = TypesMeters.objects.get(guid = u"b53173f2-2307-4b70-b84c-61b634521e87") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'Tekon_heat')
            elif unicode(type_meter) == u'Пульсар ХВС':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), attr1 = unicode(attr1), guid_types_meters = TypesMeters.objects.get(guid = u"f1789bb7-7fcd-4124-8432-40320559890f") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'Пульсар ХВС')
            
            elif unicode(type_meter) == u'Пульсар ГВС':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), attr1 = unicode(attr1), guid_types_meters = TypesMeters.objects.get(guid = u"a1a349ba-e070-4ec9-975d-9f39e61c34da") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'Пульсар ГВС')

            elif unicode(type_meter) == u'Пульсар Теплосчётчик':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"82b96b1c-31cf-4753-9d64-d22e2f4d036e") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'Пульсар Теплосчётчик')
            elif unicode(type_meter) == u'Карат 307':
                add_meter = Meters(name = unicode(type_meter) + u' ' + unicode(meter), address = unicode(adr), factory_number_manual = unicode(meter), guid_types_meters = TypesMeters.objects.get(guid = u"84fb7a85-ab91-4e93-9154-76ddee35a316") )
                add_meter.save()
                writeToLog(u'Прибор добавлен' + ' --->   ' + u'Карат 307')
            else:
                writeToLog(u'Не найдено совпадение с существующим типом прибора')
                met-=1
            met+=1
            
    result=u" Загружено счётчиков "+str(met)
    
    return result


def load_electric_counters(request):
    global isService
    isService=True
    OnOffSignals()
    args={}
    fileName=""
    sheet    = ""
    tcp_ip_status    = ""
    object_status    = ""
    counter_status    = ""
    result=""
    if request.is_ajax():
        if request.method == 'GET':
            request.session["choice_file"]    = fileName    = request.GET['choice_file']
            request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']
            request.session["tcp_ip_status"]    = tcp_ip_status    = request.GET['tcp_ip_status']
            request.session["object_status"]    = object_status    = request.GET['object_status']
            request.session["counter_status"]    = counter_status    = request.GET['counter_status']
            directory=os.path.join(BASE_DIR,'static/cfg/')
            sPath=directory+fileName
            result=LoadElectricMeters(sPath, sheet)
    counter_status=result#"Загрузка счётчиков условно прошла"
        
    #print fileName
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["tcp_ip_status"]=tcp_ip_status
    args["object_status"]=object_status
    args["counter_status"]=counter_status
    isService=False
    OnOffSignals()
    return render_to_response("service/service_electric.html", args)


@csrf_exempt
def service_water(request):
    args={}
    return render_to_response("service/service_water.html", args)
    
def add_link_meter(sender, instance, created, **kwargs):
    print u'Vi v f-ii add_link_meter'
    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    writeToLog( unicode(dtAll[1][1]))
    if (dtAll[1][1] == u'Объект'): #вода
        print(u'Добавляем связь портов по воде')
        add_link_meter_port_from_excel_cfg_water_v2(sender, instance, created, **kwargs)
    else:# электрика
        print(u'Добавляем связь портов по электрике')
        add_link_meter_port_from_excel_cfg_electric(sender, instance, created, **kwargs)

def add_link_meter_port_from_excel_cfg_water_v2(sender, instance, created, **kwargs):
    """Делаем привязку счётчика к порту по excel файлу ведомости"""
    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    print u'test'
    for i in range(1,len(dtAll)):
        #print u'Обрабатываем строку ' + unicode(dtAll[i][6])+' - '+unicode(dtAll[i][7])
        #print dtAll[i]
        meter=dtAll[i][5] #счётчик
        #print meter
        #print instance.factory_number_manual
        #print dtAll[0][5], dtAll[0][4]       
        ip_adr=unicode(dtAll[i][7]).strip()
        ip_port=unicode(dtAll[i][8]).strip()
        # Привязка к tpc порту
        if meter is not None:
            if unicode(meter) == instance.factory_number_manual :
                 guid_ip_port_from_excel = connection.cursor()
                 sQuery="""SELECT 
                                      tcpip_settings.guid
                                    FROM 
                                      public.tcpip_settings
                                    WHERE 
                                      tcpip_settings.ip_address = '%s' AND 
                                      tcpip_settings.ip_port = '%s';"""%(unicode(ip_adr), unicode(ip_port))
    #print sQuery
                 guid_ip_port_from_excel.execute(sQuery)
                 guid_ip_port_from_excel = guid_ip_port_from_excel.fetchall()
                 #print guid_ip_port_from_excel
                 
                 IsExistLink=SimpleCheckIfExist("Link_Meters_Tcpip_Settings","guid_meters",instance.guid,"","guid_tcpip_settings", guid_ip_port_from_excel)
                 #print IsExistLink
                 if IsExistLink: break
                 if guid_ip_port_from_excel:
                     guid_ip_port = TcpipSettings.objects.get(guid=guid_ip_port_from_excel[0][0])
                     add_ip_port_link = LinkMetersTcpipSettings(guid_meters = instance, guid_tcpip_settings = guid_ip_port)            
                     add_ip_port_link.save()
                     print u'Связь добавлена ', meter, ip_adr, ip_port
                 else: writeToLog(u'Не прогружен порт')
                 
           
           
#def add_link_meter_port_from_excel_cfg_water(sender, instance, created, **kwargs):
#    """Делаем привязку счётчика к порту по excel файлу ведомости"""
#    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
#    i=3
#    #здесь ошибка-привязки по одному и тому же порту
#    ip_adr=unicode(dtAll[i][7]).strip()
#    ip_port=unicode(dtAll[i][8]).strip()
## Привязка к tpc порту
#    guid_ip_port_from_excel = connection.cursor()
#    sQuery="""SELECT 
#                                      tcpip_settings.guid
#                                    FROM 
#                                      public.tcpip_settings
#                                    WHERE 
#                                      tcpip_settings.ip_address = '%s' AND 
#                                      tcpip_settings.ip_port = '%s';"""%(unicode(ip_adr), unicode(ip_port))
#    #print sQuery
#    guid_ip_port_from_excel.execute(sQuery)
#    guid_ip_port_from_excel = guid_ip_port_from_excel.fetchall()
#
#    if guid_ip_port_from_excel:
#        guid_ip_port = TcpipSettings.objects.get(guid=guid_ip_port_from_excel[0][0])
#        add_ip_port_link = LinkMetersTcpipSettings(guid_meters = instance, guid_tcpip_settings = guid_ip_port)            
#        add_ip_port_link.save()
#    else: writeToLog(u'Нет tcp-ip порта, создайте его!')

def add_link_meter_port_from_excel_cfg_electric(sender, instance, created, **kwargs):
    """Делаем привязку счётчика к порту по excel файлу ведомости"""    
    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    
    for i in range(1,len(dtAll)):
        #print u'Обрабатываем строку ' + unicode(dtAll[i][6])+' - '+unicode(dtAll[i][7])
        meter=dtAll[i][6] #счётчик
        #print dtAll[0][11], dtAll[0][12]
        PortType=dtAll[0][11] # com или tcp-ip
        #print 'i=',i,' len=', len(dtAll)
        ip_adr=unicode(dtAll[i][10]).strip()
        ip_port=unicode(dtAll[i][11]).strip()
        # Привязка к tpc порту
        if meter is not None:
            if unicode(meter) == instance.factory_number_manual :
                if unicode(PortType) == u'Com-port':
                    guid_com_port_from_excel = connection.cursor()
                    guid_com_port_from_excel.execute("""SELECT 
                                                      comport_settings.guid
                                                    FROM 
                                                      public.comport_settings
                                                    WHERE 
                                                      comport_settings.name = '%s';"""%(unicode(dtAll[i][12])))
                    guid_com_port_from_excel = guid_com_port_from_excel.fetchall()
            
                    guid_com_port = ComportSettings.objects.get(guid=guid_com_port_from_excel[0][0])
                    add_com_port_link = LinkMetersComportSettings(guid_meters = instance, guid_comport_settings = guid_com_port)
                    add_com_port_link.save()
                
                else:
                    guid_ip_port_from_excel = connection.cursor()
                    sQuery="""SELECT tcpip_settings.guid
                                                    FROM 
                                                      public.tcpip_settings
                                                    WHERE 
                                                      tcpip_settings.ip_address = '%s' AND 
                                                      tcpip_settings.ip_port = '%s';"""%(ip_adr, ip_port)
                    #print sQuery
                    guid_ip_port_from_excel.execute(sQuery)
                    guid_ip_port_from_excel = guid_ip_port_from_excel.fetchall()
            
                    print guid_ip_port_from_excel
                    if (len(guid_ip_port_from_excel)>0):
                        guid_ip_port = TcpipSettings.objects.get(guid=guid_ip_port_from_excel[0][0])
                        add_ip_port_link = LinkMetersTcpipSettings(guid_meters = instance, guid_tcpip_settings = guid_ip_port)            
                        add_ip_port_link.save()
                    else: print u'Привязки по портам не добавлены'
            else:
                pass
            
def add_link_taken_params(sender, instance, created, **kwargs):
    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    if (dtAll[1][1] == u'Объект'): #вода
        add_link_abonents_taken_params2(sender, instance, created, **kwargs)
    else:# электрика
        add_link_abonent_taken_params_from_excel_cfg_electric(sender, instance, created, **kwargs)


def add_taken_param(sender, instance, created, **kwargs): # Добавляем считываемые параметры при создании счётчика
    if instance.guid_types_meters.name == u'Меркурий 230':
        #Добавляем параметры для Меркурия 230
    # T0 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"bdcd1268-37f3-4579-a9d9-5becb2ba8aa3")) # A+ T0 месячные
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"99cd6002-f81c-4ad6-9cb0-53a92a498519")) # A+ T0 суточные
        add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e8c20ce7-bdb6-4ea6-8401-cee28049a7d7")) # A+ T0 текущие
        #add_param.save()
    # T0 R+
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"2ebc02e6-65c6-40ab-b717-0d98d66b5701")) # R+ T0 месячные
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"345a24a4-95b7-4f67-b004-716706ed2560")) # R+ T0 суточные
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4c93dd55-1ec2-48c7-9865-9ceab580b7b3")) # R+ T0 текущие
        #add_param.save()
        
    # T1 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"17789c36-4593-4ff2-94eb-1d0cebdb5366")) # A+ T1 месячные
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"d262c71a-6da4-4ec0-a9c3-b9ea659c246d")) # A+ T1 суточные
        add_param.save()
    # T2 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c31297be-220b-4971-8642-6b614aa7ecee")) # A+ T2 месячные
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"37011b85-c8af-4f6c-857d-4b93a95d31e1")) # A+ T2 суточные
        add_param.save()
    # T3 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"79741ba9-e8b8-4352-862e-17a9c4d928ce")) # A+ T3 месячные
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c3bb9033-ffcb-4a28-91e2-6b45924b8858")) # A+ T3 суточные
        add_param.save()
    
    # Ток
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"aee312b0-adb1-4be9-9879-b3a3598f9b29")) # Ia текущее
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"7ed0d364-e790-4325-a927-9ef86a685f00")) # Ib текущее
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"474b0809-482a-4851-9a96-4587f8c59152")) # Ic текущее
        #add_param.save()
    # Напряжение
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c06f7315-abc6-4889-97ad-201a936c8f2c")) # Ua текущее
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"632f76fb-4dd9-4e7d-86a0-a57a27fc648a")) # Ub текущее
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1a3ca6ca-8866-4aad-8712-d9df003fe692")) # Uc текущее
        #add_param.save()
    # Мощность
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"3077b3ac-fde2-4435-9e6f-17464310c090")) # P Активная мощность
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e7617c95-7e42-4cfa-9acd-5bc119261c6d")) # Q Реактивная мощность
        #add_param.save()
    #Получасовки
#        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6af9ddce-437a-4e07-bd70-6cf9dcc10b31")) # A+ 30-мин. срез мощности
#        add_param.save()
#        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"66e997c0-8128-40a7-ae65-7e8993fbea61")) # R+ 30-мин. срез мощности
#        add_param.save()
    elif instance.guid_types_meters.name == u'Меркурий 233':
        #Добавляем параметры для Меркурия 233
        pass
    elif instance.guid_types_meters.name == u'Пульсар16':
        #Добавляем параметры для Пульсар16
    # Суточные
      # Канал 1
      # Канал 2
      # Канал 3
      # Канал 4
      # Канал 5
      # Канал 6
      # Канал 7
      # Канал 8
      # Канал 9
      # Канал 10
      # Канал 11
      # Канал 12
      # Канал 13
      # Канал 14
      # Канал 15
      # Канал 16
   
    # Текущие
      # Канал 1
      # Канал 2
      # Канал 3
      # Канал 4
      # Канал 5
      # Канал 6
      # Канал 7
      # Канал 8
      # Канал 9
      # Канал 10
      # Канал 11
      # Канал 12
      # Канал 13
      # Канал 14
      # Канал 15
      # Канал 16
       pass
    elif instance.guid_types_meters.name == u'Пульсар10':
        #Добавляем параметры для Пульсар10
    # Суточные
      # Канал 1
      # Канал 2
      # Канал 3
      # Канал 4
      # Канал 5
      # Канал 6
      # Канал 7
      # Канал 8
      # Канал 9
      # Канал 10
   
    # Текущие
      # Канал 1
      # Канал 2
      # Канал 3
      # Канал 4
      # Канал 5
      # Канал 6
      # Канал 7
      # Канал 8
      # Канал 9
      # Канал 10
   
        pass
    elif instance.guid_types_meters.name == u'Пульсар 16M':
        #Добавляем параметры для Пульсар16
    # Месячные
      # Канал 1
      # Канал 2
      # Канал 3
      # Канал 4
      # Канал 5
      # Канал 6
      # Канал 7
      # Канал 8
      # Канал 9
      # Канал 10
      # Канал 11
      # Канал 12
      # Канал 13
      # Канал 14
      # Канал 15
      # Канал 16
    
    # Суточные
      # Канал 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"fc4a9568-4674-4a80-b497-e4f34399acd5"))
        add_param.save()
      # Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"9e6e308f-abec-4b47-9b99-9cb590c55d0c"))
        add_param.save()
      # Канал 3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e6815dd5-fbbc-480f-8b95-025d7f9a0403"))
        add_param.save()
      # Канал 4
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"612d2f20-d454-4e14-910b-1fd89bbb31dd"))
        add_param.save()
      # Канал 5
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"d82c7576-8e5e-4e93-ae10-58459b31e4a0"))
        add_param.save()
      # Канал 6
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6ccc7efb-d9fe-4285-b343-8ed22d2d3625"))
        add_param.save()
      # Канал 7
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"72567365-9a40-4f97-ab25-0911585035bf"))
        add_param.save()
      # Канал 8
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"9203f5ed-d5da-4462-91d1-5aea42e99124"))
        add_param.save()
      # Канал 9
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e4068568-d8c4-42ab-9957-7292753e2891"))
        add_param.save()
      # Канал 10
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"9b5ab67b-40aa-4536-8b7c-340a773ab31b"))
        add_param.save()
      # Канал 11
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4fd440c4-9ec5-4ab9-a073-6c4d3a174777"))
        add_param.save()
      # Канал 12
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"00b7f7c5-37f3-494a-8ceb-5a62f9ebf4e3"))
        add_param.save()
      # Канал 13
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"169a79e0-da6f-4091-9fc7-ab81adc0d7e0"))
        add_param.save()
      # Канал 14
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"17e9c8fe-0d69-4466-b64e-185452c61555"))
        add_param.save()
      # Канал 15
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"25de493d-c680-4ca6-ac02-b778022ee151"))
        add_param.save()
      # Канал 16
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"908e88f0-f9a0-421d-bbe7-9bafdf5d2565"))
        add_param.save() 
  
    # Текущие
      # Канал 1
       # add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e3f1325e-3018-40ba-b94a-ab6d6ac093e9"))
       # add_param.save()
      # Канал 2
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5a6b0338-c15d-4224-a04f-a10fc73c5fc7"))
        #add_param.save()
      # Канал 3
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"48a42afe-d9ac-4180-a733-6dd5f9d9ca80"))
        #add_param.save()
      # Канал 4
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"01a5419c-c701-4185-95b6-457b8c9ca2d0"))
        #add_param.save()
      # Канал 5
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"85c4295e-bc6a-46ec-9866-0bf9f77c6904"))
        #add_param.save()
      # Канал 6
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"68270d0a-5043-4ea2-9b61-4adaa298abad"))
        #add_param.save()
      # Канал 7
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"cd489c4b-6e74-4c65-bfee-c0fa78a853bf"))
        #add_param.save()
      # Канал 8
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"f29062a4-ab60-4117-8f85-0cdec634c797"))
        #add_param.save()
      # Канал 9
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e8521cd7-2f38-4619-935d-8fe86234dbe7"))
        #add_param.save()
      # Канал 10
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1349b747-41ca-4ba8-a690-69c649129f44"))
        #add_param.save()
      # Канал 11
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"99ab1a30-fde8-4b81-9f9e-2f731516ce1b"))
        #add_param.save()
      # Канал 12
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c7f6a397-833d-4020-9d2b-38c19bec272c"))
        #add_param.save()
      # Канал 13
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4413bffb-1832-4900-9351-5ac3666dd8b0"))
        #add_param.save()
      # Канал 14
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6280490b-123d-4e27-bef9-19fd7dc2cf54"))
        #add_param.save()
      # Канал 15
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"93891c5a-1c8f-4906-b7f0-961dc8ad3c9f"))
        #add_param.save()
      # Канал 16
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"22dd3a17-a828-44e0-80d9-db075ba120ae"))
        #add_param.save()

    elif instance.guid_types_meters.name == u'Пульсар 10M':
        #Добавляем параметры для Пульсар10
    # Месячные
      # Канал 1
      # Канал 2
      # Канал 3
      # Канал 4
      # Канал 5
      # Канал 6
      # Канал 7
      # Канал 8
      # Канал 9
      # Канал 10
    
    # Суточные
      # Канал 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"325ec164-9428-4a57-867c-33d4eaf8cc2a"))
        add_param.save()
      # Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"99a99024-65b4-44dd-99fc-6a5cf1d4aaee"))
        add_param.save()
      # Канал 3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"f897f0ca-4e35-4f0d-b345-3379668aa36f"))
        add_param.save()
      # Канал 4
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"034374bd-2dfb-4568-aa16-84255df33c88"))
        add_param.save()
      # Канал 5
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b6bdfae8-4f27-4056-af79-d746b44038ee"))
        add_param.save()
      # Канал 6
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"2c2f7176-8b77-44f4-9678-4773e95e67ce"))
        add_param.save()
      # Канал 7
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"91bb7c43-f802-4ebd-a8fe-75f833acedeb"))
        add_param.save()
      # Канал 8
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"cf24b669-1c5b-4db7-936a-5f9d5c8be928"))
        add_param.save()
      # Канал 9
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"96035c7c-ee7c-41f6-9723-8a75dd9ed573"))
        add_param.save()
      # Канал 10
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"253475ea-614d-4aad-93a8-e81e4c9028e9"))
        add_param.save()   

    # Текущие
      # Канал 1
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"32dad392-ca1e-4110-8f2c-a86b02e26fb3"))
        #add_param.save()
      # Канал 2
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"3e13694b-7cb5-4417-a091-af8a7db34dc7"))
        #add_param.save()
      # Канал 3
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1023b35b-3cbf-4519-aac3-3bf1ebae07c1"))
        #add_param.save()
      # Канал 4
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"eea27ade-44cd-4e66-8298-00a4a6ad915a"))
        #add_param.save()
      # Канал 5
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"25e09d4d-3a48-4381-ad5d-b783c03c4d35"))
        #add_param.save()
      # Канал 6
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"069898ea-9d74-4571-b719-e8e6f1513c12"))
        #add_param.save()
      # Канал 7
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"084aa5f4-75d5-41f6-b0d6-9f2403eacd2c"))
        #add_param.save()
      # Канал 8
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"786ed8b8-aed1-478c-ae75-99caf1358cf0"))
        #add_param.save()
      # Канал 9
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6fc4c39c-9a43-4cb7-a066-c40fd2ca47e5"))
        #add_param.save()
      # Канал 10
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"8b2aa40a-cd91-4e22-b9d1-596e49e5f839"))
        #add_param.save()  

    elif instance.guid_types_meters.name == u'Пульсар 2M':
        #Добавляем параметры для Пульсар10
    
    # Месячные
      # Канал 1
      # Канал 2
    
    # Суточные
      # Канал 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"0239dffb-de88-45e5-b6f6-18bf39f92525"))
        add_param.save()
      # Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"a1cb319d-ac09-466d-894b-91d90aba4239"))
        add_param.save()   
    
    # Текущие
      # Канал 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"fcc28118-66c0-4cdf-aeba-5da1171aae48"))
        add_param.save()
      # Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1faeb517-bd1f-4ba0-96a5-67f00764822f"))
        add_param.save()

    elif instance.guid_types_meters.name == u'ПСЧ-3ТА.04':
        #Добавляем параметры для ПСЧ-3ТА.04
        pass
    elif instance.guid_types_meters.name == u'ТЭМ-104':
        #Добавляем параметры для ТЭМ-104
        pass
    elif instance.guid_types_meters.name == u'СЭТ-4ТМ.03М':
        #Добавляем параметры для СЭТ-4ТМ.03М
        pass
    elif instance.guid_types_meters.name == u'Меркурий 200':
        #Добавляем параметры для Меркурий 200

    # Значения суточные (текущие)
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"9cbc001d-a262-481f-a1aa-47d02bf18af1")) #T0
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b65d4227-69a5-487b-9999-5539ca3fc004")) #T1
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5e312de9-34cd-4ba7-a744-c9b94a77d98b")) #T2
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4260ea05-78f8-4c5c-9172-fa161fa96068")) #T3
        add_param.save()
    # Значения на начало месяца
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"86cd925b-48c2-40b8-b211-f116e0e6dbea")) #T0
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"62a3796a-eaae-445d-9166-2ad517186b78")) #T1
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5f6e1e3d-4128-4cfe-94cf-57ac84a7694a")) #T2
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"0c28c135-58f2-4dff-a222-9f3d9f3c742b")) #T3
        add_param.save()
    # Значения на начало суток
        #Не поддерживается прибором, но текущие переделаны на суточные
        

    elif instance.guid_types_meters.name == u'Эльф 1.08':
        #Добавляем параметры для счётчика тепла Elf 108
    
        #-------------Текущие
        # "Энергия"
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"f2bbf267-456e-477a-95d2-abb94c78ba43"))
        add_param.save()
        # "Объем"       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"dad6e2eb-e978-46f4-b7ec-442834b04e7a"))
        add_param.save()
        # "ElfTon"  Время работы прбора     
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"d3c9563d-51ed-4ca7-922f-ac3731065ead"))
        add_param.save()
        # "ElfErr"  Время работы прибора с ошибкой
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"dade3324-b9b0-41c8-bc76-70f617573e43"))
        add_param.save()
        # "Ti"      Температура входа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"acca627e-f21a-4f8b-be7e-038f534b5d11"))
        add_param.save()
        # "To"      Температура выхода
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"01487323-a28f-419e-9589-2563d785ab2a"))
        add_param.save()
        # "Канал 1"      Импульсный вход 1 текущий
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6e7f0d37-df5c-4850-991e-b5d7cb793924"))
        #add_param.save()
        # "Канал 1"      Импульсный вход 1 суточный
#        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"9af27a62-d6c8-4b67-bd36-da7103e0b1f1"))
#        add_param.save()
        # "Канал 2"      Импульсный вход 2 суточный
#        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"86acc33d-7bea-4977-a5b5-c5858ce9a09d"))
#        add_param.save()
        # "Канал 2"      Импульсный вход 2 текущий
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"de7bfdfd-c17f-4a7c-942d-b28e85db33cb"))
        #add_param.save()
        #-------------Архивные
        # "Энергия"
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"ae439e1f-5c4b-494c-8a53-a61b85c804a0"))
        add_param.save()
        # "Объем"
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b02153a4-00c0-4800-a55a-c7f9dfbb14e7"))
        add_param.save()
        # "ElfTon" Время работы прибора
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"aa611d48-f1fe-462a-8b0a-0a7596792b69"))
        add_param.save()
        # "ElfErr" Время работы прибора с ошибкой
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"af047098-bd45-4579-a60c-b75bed376bbe"))
        add_param.save()
        
        
    elif instance.guid_types_meters.name == u'СПГ762-1':
        #Добавляем параметры для счётчика газа СПГ762 Подсистема 1
    
        #-------------Часовые
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"edfc6dc0-1628-4a7e-bd04-71107882039a"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1e27e72c-79d6-4c68-bc04-1be84d061622"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"0fa5d9ef-4c6c-4f78-bc64-9d9b34002344"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"433d7025-15f3-4ab0-9d73-39bd0e425566"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"446a6bb6-c17f-4478-b1d8-252c7eb454d3"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"003c2fb2-0092-4d7d-a513-3dcc50a255da"))
        add_param.save()
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"9c99da7b-1a73-48b0-a3f5-54438a3ea824"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"832374da-5834-4fe0-abe3-07d48d447af2"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4920e70a-452e-4ecf-918e-14ca288c7a1f"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"825dfd73-82fa-4f1a-9635-77bdcb244997"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"daa4a90e-7993-493c-9ac0-c03241b2ab2c"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"72b5f4ba-d179-419c-8795-e2f86f5ee2ff"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"41774f43-9c9c-4867-bf21-e3d1df4fd2f8"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"f610969e-cb2d-436f-9357-e63da72d162e"))
        add_param.save()
        
        #-------------Суточные
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"ac9fe54a-6f51-4ee9-a849-448f0f10a4b6"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"2eefa93a-be60-4f23-9b09-8d1c6bad0a15"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"78c30686-a5d0-436f-9c56-57b076769774"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"d7900bac-b85a-4b83-ad67-b822b470a698"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"068f4de0-f041-4608-b09a-81dbc8f319ff"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"dad040cd-1047-4060-a7ac-e9a6e7f30fb4"))
        add_param.save()        
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"3be53eb7-c931-4728-a442-44d14c9da44f"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1dff6784-dd8d-4898-ae8b-f1b60fbdc1af"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"af26b128-b634-47a4-96ba-42de6f039fdb"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5bb8af68-588e-470c-9b64-373482f71468"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c32e7e01-106a-49c6-b8b0-6490448548ad"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"8fc09300-4b19-4478-aa44-d2fb1cf792d5"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"53693586-e5b5-4204-922a-a0b0153298ea"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"bf06727c-9cab-4efe-a0f2-6242bb320372"))
        add_param.save()
        
        # Масса газа нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5399a54f-0d2c-47e8-8ffb-882f5dddc239"))
        add_param.save()
        # Объем газа при ст.условиях нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c14eaa34-3264-4fe4-98ab-8da6618fc431"))
        add_param.save()
        # Объем газа при раб.условиях нарастающим итогом              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4ed5ba8d-9ead-4a60-96e4-726f38432d9a"))
        add_param.save()
        
    elif instance.guid_types_meters.name == u'СПГ762-2':
        #Добавляем параметры для счётчика газа СПГ762 Подсистема 1
    
        #-------------Часовые
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"9a17ea4f-a1f8-4fb9-a21e-62f43978535a"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"49728e75-1d62-4b9e-8633-762cb7117b52"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c44af390-ed2a-49d8-9c67-25f543db9935"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"24bd767d-f667-444e-acfa-a935fb8f4699"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1e675e58-2cc9-4103-8410-2d37704a2bcf"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"d14fa3fa-5bd5-4dd9-b740-61049d38e694"))
        add_param.save()
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6de742eb-a391-473a-bb4b-ab780a4642b8"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"04b43217-af06-4f09-8d95-0e2d3dbd0905"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b62c6838-7578-41c9-a94c-d06788cc2d41"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"aad9d0b4-1c12-4165-a1b1-4fac9de00c38"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"205a8c8c-de26-44e5-ab72-efb7fe72040c"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c84c6130-ede7-487e-a414-b384964eb81e"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5907cc5f-1386-4fbc-9e5c-7d3f77dba6d6"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c5638410-44b6-41d7-b501-6e5c0a002f48"))
        add_param.save()
        
        #-------------Суточные
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c1643ab8-1707-4b73-9610-0226b1fb6860"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"dbacff0c-2b3d-40c5-aa03-8d51a64919dd"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"279bdfd5-7b22-4d7e-900c-21e4077506dd"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b74e6743-3996-4af7-8024-da3912d14b45"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c7220fc7-5c01-4bcc-ac2a-7c851276af4d"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"381ccc3f-a9d9-4dcf-a9aa-2e5bd0e4efc8"))
        add_param.save()        
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"7f0f0e09-3bd0-4595-84dd-754f4c21bc5e"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"ee687ef8-36de-4de8-9a05-4ac841c9c144"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"8ee128d0-5c21-4faa-a2fe-7432ff9be684"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c61c87eb-5a4f-4095-ac50-4324e7899340"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"bae3e866-f057-4be5-99a0-7474f6c7cbc1"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"8ce68aef-85fa-4ca6-8f9f-dfa1f9e71cdd"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"ea094191-7ce5-4c42-ad5e-e886d02e73e0"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e4d751b4-ef6c-45ca-b31a-f107f47a97aa"))
        add_param.save()
        
        # Масса газа нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"add68490-64f4-47a2-a801-1fafa48c09a2"))
        add_param.save()
        # Объем газа при ст.условиях нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"525f9439-ce13-43be-a4f2-67f590f4842b"))
        add_param.save()
        # Объем газа при раб.условиях нарастающим итогом              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"a760c696-fda0-47d4-8fb5-899d742957f1"))
        add_param.save()
        
    elif instance.guid_types_meters.name == u'СПГ762-3':
        #Добавляем параметры для счётчика газа СПГ762 Подсистема 1
    
        #-------------Часовые
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"187b787c-1693-4c90-b6df-d868effef692"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"49c86cc3-57c2-4bdf-b4e3-b07f64673d37"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1e92a9a8-1cd9-4252-b9c7-b33357bafce7"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"06cd5dad-7ea9-438e-abbf-043e8918eb3e"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"8064aad9-778a-4902-b0c0-75b23289469a"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b747eeb5-8c69-443e-b74b-2bb89af64206"))
        add_param.save()
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"38c0b41b-0883-4990-bb0c-8b532caed34c"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"7bf3d68d-4344-49f3-8169-370f6142351a"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"fa188255-c1cc-4c2d-844c-3b40a3a7559e"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"06c97066-1e35-4bb1-a96f-fe3c0056cf39"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e530ab26-92e8-4edc-8e6d-5cd6184bfbe7"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5d315c1e-b237-46d6-9273-be4e597ad1c2"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"54b86222-3d1c-440a-b02f-bedbef0e9e28"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"acb12185-dd88-4449-8f03-76b6fd148958"))
        add_param.save()
        
        #-------------Суточные
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"d1f0258f-42ba-4e4f-a66c-74aed4d512ce"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"7ec1a0fb-88f0-497a-8917-01c0b731b88a"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"3465cb7c-57ea-4ad8-afde-74fb2814ddeb"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"a3014fec-73df-4fd4-a68c-9c3ff737d140"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"986f24b0-df76-4c36-9b7f-fbbe05a10c94"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"2a5ec8f1-b6fe-4eff-b91f-42a7712dd663"))
        add_param.save()        
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"cbfc543d-3a19-46c3-8075-ff59492d2620"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"8a24d34e-aee6-4865-bf21-56d9c07dcd1e"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"85de87d5-8e0a-4088-8248-8a64367db47e"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b2bd2c95-ee85-4156-9ef6-7fc25d29a244"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"8674a61c-af88-46c5-b553-fecc9a7d0837"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5880bd0f-699d-407c-a3f0-6cea0ebde423"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"9057618b-445c-4581-86a6-4715469db938"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4991e5c0-0827-4467-b9d2-7613d1b6dd09"))
        add_param.save()
        
        # Масса газа нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"dcd5ed6a-7bd0-41ba-8850-5b88a9831c04"))
        add_param.save()
        # Объем газа при ст.условиях нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"cbeaa4fa-d1fb-4bf5-9688-7084b57fbfe4"))
        add_param.save()
        # Объем газа при раб.условиях нарастающим итогом              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"d16b31ea-87d1-409a-bbf8-4a743b678dbb"))
        add_param.save()

    elif instance.guid_types_meters.name == u'Sayany':
        #Добавляем параметры для счётчика Sayany
    
        #-------------Суточные
        # "Q" Тепловая энергия. Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e7f2ffba-9a40-43e1-80f3-ddd22596cdb8"))
        add_param.save()    
        # "Q" Тепловая энергия. Канал2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6f9cd79e-ca34-447e-8ad1-d54531389fe1"))
        add_param.save() 
        # "M" Расход воды. Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b05de8e2-6176-4fc0-bc44-79ceb4229c80"))
        add_param.save() 
        # "M" ТРасход воды. Канал2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5f256e9b-1cb3-4f27-a53a-d08b446dda58"))
        add_param.save() 
        # "T" Температура. Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"75474616-f3db-4903-91d5-1f22f6593394"))
        add_param.save() 
        # "T" Температура. Канал2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"f3210c5b-afde-4c9a-b201-9c7c403c4cf2"))
        add_param.save() 
        # "T" Температура. Канал3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b12762a0-0a06-49a4-b842-8ad3378f4602"))
        add_param.save() 
        # "T" Температура. Канал4
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"472ba2fd-cc06-4147-a1e7-c1bb66096536"))
        add_param.save() 
        
    elif instance.guid_types_meters.name == u'Tekon_hvs':
        #Добавляем параметры для счётчика Tekon. Читаем один тэк с opcretranslator
    
        #-------------Суточные
        # "Показание". Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"d76796ea-ea63-4317-982f-ffbcde2074dc"))
        add_param.save()  
        
    elif instance.guid_types_meters.name == u'Tekon_gvs':
        #Добавляем параметры для счётчика Tekon. Читаем один тэк с opcretranslator
    
        #-------------Суточные
        # "Показание". Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e71a7206-7a30-45d9-981d-0b7592b96337"))
        add_param.save() 
        
    elif instance.guid_types_meters.name == u'Tekon_heat':
        #Добавляем параметры для счётчика Tekon. Читаем один тэк с opcretranslator
    
        #-------------Суточные
        # "Показание". Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1dca7dab-371a-4429-afa1-8b4877b38b5b"))
        add_param.save()
        
    elif instance.guid_types_meters.name == u'Меркурий 230-УМ':
        #Добавляем параметры для счётчика Меркурий на УСПД УМ-RTU.    
        
        #-------------Суточные
        # "Показание". T0 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b6e89205-3814-463d-86d1-f52cec7d8962"))
        add_param.save()
        # "Показание". T1 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"7f3c42e6-4000-4373-a0e6-37e66ce819a9"))
        add_param.save() 
        # "Показание". T2 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c6512649-56ea-4214-aa33-84516bfe8dc1"))
        add_param.save() 
        # "Показание". T3 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4e20bda9-6e75-4b0f-a99a-0e4c1cd07d3b"))
        add_param.save()
        
        #-------------Мощность        
        #А+ Профиль
#        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"922ad57c-8f5e-4f00-a78d-e3ba89ef859f")) # A+ 30-мин. срез мощности
#        add_param.save()        
#        #R+ Профиль
#        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"61101fa3-a96a-4934-9482-e32036c12829")) # R+ 30-мин. срез мощности
#        add_param.save()
        
    elif instance.guid_types_meters.name == u'Пульсар Теплосчётчик':
        #Добавляем параметры для Теплосчётчика Пульсар.
        #------------Суточные
        # "Показание Энергии" Q, Гкал
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"24ae9f51-40a4-4758-a826-a5f8286e1a2e"))
        add_param.save()
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"a3da78fb-b07b-4d53-a980-54b51e26819a"))
        add_param.save()
        # "Показание Температура подачи" Ti, C0
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"de66ecd2-b243-467c-8d1a-cfcb42377300"))
        add_param.save()
        # "Показание Температура выхода" To, C0
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"d3433b80-cb8c-4038-a682-947e6d05955e"))
        add_param.save()

        
    elif instance.guid_types_meters.name == u'Пульсар ХВС':
        #Добавляем параметры для водосчётчика Пульсар ХВС.
        #------------Суточные
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"209894a8-8d19-4e4d-bad8-1767eec4fedf"))
        add_param.save()

    
    elif instance.guid_types_meters.name == u'Пульсар ГВС':
        #Добавляем параметры для водосчётчика Пульсар ГВС.
        #------------Суточные
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5fc2ff3b-999e-4154-ba49-84d3971369b0"))
        add_param.save()
        
    elif instance.guid_types_meters.name == u'Карат 307':
        print u'Добавляем параметры для счётчика Карат 307'
        #Суточные 
        #Объём     
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"3024fd72-d1e8-4476-a876-4bc09553dde9"))
        add_param.save()
        #Тепло
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"46a63ef5-5761-4e16-a854-1979ddc9668f"))
        add_param.save()
        #Tout
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6dd6ea63-20dc-46d0-b56e-6890a2b83f48"))
        add_param.save()
        #Tin
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"8a5f5921-5b70-410d-83de-8403ec2a4d87"))
        add_param.save()
        #Ton наработка в минутах
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"9c86e183-dd53-4c7f-b728-ffe75a55c633"))
        add_param.save()
        #Terr время работы в ошибке
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"abd41546-02f6-4e2c-8bd2-a60ab80ffe66"))
        add_param.save()
        #Масса 
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"eb617f04-14a3-403c-90e8-286412872232"))
        add_param.save()
    else:
        pass
        #print u'Тип счётчика не определен'
        
signals.post_save.disconnect(add_taken_param, sender=Meters)
signals.post_save.disconnect(add_link_meter, sender=Meters) 
signals.post_save.disconnect(add_link_taken_params, sender=TakenParams)  
        
if (isService):
    print 'signals ON'
    signals.post_save.connect(add_link_taken_params, sender=TakenParams)
    signals.post_save.connect(add_link_meter, sender=Meters)
    signals.post_save.connect(add_taken_param, sender=Meters)
else:
    signals.post_save.disconnect(add_link_meter, sender=Meters)
    signals.post_save.disconnect(add_taken_param, sender=Meters)
    signals.post_save.disconnect(add_link_taken_params, sender=TakenParams)

def OnOffSignals():
    if (isService):
        print 'signals ON'
        signals.post_save.connect(add_link_taken_params, sender=TakenParams)
        signals.post_save.connect(add_link_meter, sender=Meters)
        signals.post_save.connect(add_taken_param, sender=Meters)
    else:
        print 'signals Off'
        signals.post_save.disconnect(add_link_meter, sender=Meters)
        signals.post_save.disconnect(add_taken_param, sender=Meters)
        signals.post_save.disconnect(add_link_taken_params, sender=TakenParams)

def add_link_abonents_taken_params(sender, instance, created, **kwargs):
    def get_taken_param_by_abonent_from_excel_cfg(input_taken_param):
        """Функция, которая читает excel файл. Составляет имя считываемого параметра типа "Пульсар 16M 33555 Пульсар 16M Канал 11". В случае совпадения должна привязать этот параметр к абоненту. Абоненты должны быть предварительно созданы."""    
        dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    
        def shrink_taken_param_name(taken_param_name):
            if taken_param_name.find(u'Текущий') != -1: # Ищем слово "Текущий"
                nn = taken_param_name.find(u'Текущий')  # Если нашли. то Записываем позицию где.        
            elif taken_param_name.find(u'Суточный') != -1:
                nn = taken_param_name.find(u'Суточный')
            else:
                pass
            return taken_param_name[:nn -1]

        for i in range(2,len(dtAll)):
            #taken_param = u'Пульсар' + u' ' + unicode(dtAll[i][3])[17:20] + u' ' + unicode(dtAll[i][3])[2:8] + u' ' + u'Пульсар' + u' ' + unicode(dtAll[i][3])[17:20] + u' ' + u'Канал' + u' ' + unicode(dtAll[i][4])
            taken_param = unicode(dtAll[i][6]) + u' ' + unicode(dtAll[i][5]) + u' '+ unicode(dtAll[i][6]) + u' ' + u'Канал' + u' ' + unicode(dtAll[i][4])
#            print taken_param
#            print shrink_taken_param_name(input_taken_param)
            if taken_param == shrink_taken_param_name(input_taken_param):
                try:
                    return unicode(dtAll[i][2])
                except:
                    return None
            else:
                pass
    
    writeToLog(u'--------')
    writeToLog(instance.name)
    writeToLog(u'==>', get_taken_param_by_abonent_from_excel_cfg(instance.name))
    if get_taken_param_by_abonent_from_excel_cfg(instance.name) is not None:
        writeToLog(u'Совпадение')
        try:
            add_link_abonents_taken_param = LinkAbonentsTakenParams (name = Abonents.objects.get(name= get_taken_param_by_abonent_from_excel_cfg(instance.name)).name + u" " + instance.guid_params.guid_names_params.name + u" " + instance.guid_params.guid_types_params.name ,coefficient=1, coefficient_2 = 1, guid_abonents = Abonents.objects.get(name= get_taken_param_by_abonent_from_excel_cfg(unicode(instance.name))) , guid_taken_params = instance )
            add_link_abonents_taken_param.save()
        except:
            pass
    else:
        pass
    
            
def add_link_abonents_taken_params2(sender, instance, created, **kwargs):
    writeToLog(instance.name)
    isExistTakenParam=SimpleCheckIfExist('taken_params','name',instance.name,"","","")
    if not isExistTakenParam:
        print(u'Параметра не существует!!! Связать невозможно')
        return None
    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    for i in range(2,len(dtAll)):
        abon=unicode(dtAll[i][2])
        type_pulsar=unicode(dtAll[i][6])
        channel=unicode(dtAll[i][4])
        num_pulsar=unicode(dtAll[i][5])
        taken_param = type_pulsar+u' '+num_pulsar+u' '+type_pulsar+u' Канал '+channel+u' Суточный -- adress: '+channel+u'  channel: 0'
        #print taken_param
        if (taken_param==instance.name):
            isExistAbonent=SimpleCheckIfExist('abonents','name',abon,'','','')
            if isExistAbonent:
                writeToLog(u'Совпадение')
                #"ХВС, №47622 Канал 4 Суточный"
                guidAbon=GetSimpleTable('abonents','name',abon)[0][0]
                
                linkName=abon+u' Канал '+channel+' Суточный'
                writeToLog(linkName)
                try:
                    add_link_abonents_taken_param = LinkAbonentsTakenParams (name = linkName,coefficient=1, coefficient_2 = 1, guid_abonents = Abonents.objects.get(guid=guidAbon) , guid_taken_params = instance )
                    add_link_abonents_taken_param.save()
                    writeToLog(u'Связь добавлена: '+abon+u' -- '+taken_param)
                except:
                    writeToLog(u'ошибка')
                else:
                    pass
    
#    
#    
#    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
#    for i in range(2,len(dtAll)):
#            #taken_param = u'Пульсар' + u' ' + unicode(dtAll[i][3])[17:20] + u' ' + unicode(dtAll[i][3])[2:8] + u' ' + u'Пульсар' + u' ' + unicode(dtAll[i][3])[17:20] + u' ' + u'Канал' + u' ' + unicode(dtAll[i][4])
#            # "Пульсар 2M 062726 Пульсар 2M Канал 1 Суточный -- adress: 1  channel: 0"
#            # "Пульсар 10M 203677 Пульсар 10M Канал 7 Суточный -- adress: 7  channel: 0"
#        type_pulsar=unicode(dtAll[i][6])
#        channel=unicode(dtAll[i][4])
#        num_pulsar=unicode(dtAll[i][5])
#        taken_param = type_pulsar+u' '+num_pulsar+u' '+type_pulsar+u' Канал '+channel+u' Суточный -- adress: '+channel+u'  channel: 0'
#        print taken_param
#    
#    print u'--------'
#    print instance.name
#    print u'==>', get_taken_param_by_abonent_from_excel_cfg(instance.name)
#    if get_taken_param_by_abonent_from_excel_cfg(instance.name) is not None:
#        print u'Совпадение'
#        try:
#            add_link_abonents_taken_param = LinkAbonentsTakenParams (name = Abonents.objects.get(name= get_taken_param_by_abonent_from_excel_cfg(instance.name)).name + u" " + instance.guid_params.guid_names_params.name + u" " + instance.guid_params.guid_types_params.name ,coefficient=1, coefficient_2 = 1, guid_abonents = Abonents.objects.get(name= get_taken_param_by_abonent_from_excel_cfg(unicode(instance.name))) , guid_taken_params = instance )
#            add_link_abonents_taken_param.save()
#        except:
#            pass
#    else:
#        pass




def add_link_abonent_taken_params_from_excel_cfg_electric(sender, instance, created, **kwargs):
    dtAll=GetTableFromExcel(cfg_excel_name,cfg_sheet_name) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    #print dtAll[0][0]
    for i in range(1,len(dtAll)):
        meter=dtAll[i][6]
        abon=unicode(dtAll[i][3])
        obj=unicode(dtAll[i][2])
        if meter is not None:
            cursor = connection.cursor()
            sQuery="""SELECT abonents.guid FROM public.objects, public.abonents
                      WHERE objects.guid = abonents.guid_objects 
                      AND abonents.name = '%s' 
                      AND objects.name = '%s';"""%(abon,obj )
            #print sQuery
            cursor.execute(sQuery)
            guid_abonent_by_excel = cursor.fetchall()
            #print guid_abonent_by_excel

            if unicode(meter) == instance.guid_meters.factory_number_manual:
                writeToLog(u'Абонент найден' + u' ' + unicode(instance.name))
                #print guid_abonent_by_excel 
                add_link_abonents_taken_param = LinkAbonentsTakenParams (name = unicode(dtAll[i][3]) + u' - ' +  unicode(instance.guid_meters.name)  ,coefficient=unicode(dtAll[i][9]), coefficient_2 = 1, guid_abonents = Abonents.objects.get(guid =guid_abonent_by_excel[0][0]), guid_taken_params = instance)
                add_link_abonents_taken_param.save()
            else:
                pass
    



def load_water_objects(request):
    args={}
    fileName=""
    sheet    = ""
    tcp_ip_status    = ""
    object_status    = ""
    counter_status    = ""
    result="Не загружено"
    if request.is_ajax():
        if request.method == 'GET':
            request.session["choice_file"]    = fileName    = request.GET['choice_file']
            request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']
            request.session["tcp_ip_status"]    = tcp_ip_status    = request.GET['tcp_ip_status']
            request.session["object_status"]    = object_status    = request.GET['object_status']
            request.session["counter_status"]    = counter_status    = request.GET['counter_status']
            
            directory=os.path.join(BASE_DIR,'static/cfg/')
            sPath=directory+fileName
            result=LoadObjectsAndAbons_water(sPath, sheet)
    
    object_status=result

    #print fileName
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["port_status"]=tcp_ip_status
    args["object_status"]=object_status
    args["counter_status"]=counter_status
    return render_to_response("service/service_water.html", args)
    
def CheckIfExistInObjects(name_parent, name_child):
    dt=[]
    cursor = connection.cursor()
    sQuery="""
    With obj as 
(Select guid as guid_child, objects.name as name_child, objects.level as level_child, guid_parent
 from objects)
Select guid as grand_parent, objects.name as name_parent, objects.level, objects.guid_parent, 
obj.guid_child,obj.name_child, obj.level_child, obj.guid_parent
FROM 
  public.objects, obj
where obj.guid_parent=objects.guid
and objects.name='%s' 
and obj.name_child='%s'
order by name_parent    """%(name_parent, name_child)
    cursor.execute(sQuery)
    dt = cursor.fetchall()

    if not dt:  
        return None
    else: 
        return dt[0][4]# возвращаем guid квариры
    
    
def LoadObjectsAndAbons_water(sPath, sheet):
    result=""
    dtAll=GetTableFromExcel(sPath,sheet) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    kv=0
    for i in range(2,len(dtAll)):
        obj_l0=u'Вода' # всегда будет Вода как объект-родитель
        obj_l1=dtAll[i][0] #корпус
        obj_l2=dtAll[i][1] #квартира
        if not dtAll[i][1] or dtAll[i][1]==None:
            j=i
            while not obj_l2 or obj_l2==None:
                j-=1
                obj_l2=dtAll[j][1]
        abon=dtAll[i][2] #абонент он же счётчик по воде
#        chanel=dtAll[i][4] # канал пульсара
#        numPulsar=dtAll[i][5] #номер пульсара
#        typePulsar=dtAll[i][5] #тип пульсара
        isNewObj_l0=SimpleCheckIfExist('objects','name',obj_l0,"","","")#вода
        isNewObj_l1=SimpleCheckIfExist('objects','name',obj_l1,"","","")#корпус
        
        guid_obj2=CheckIfExistInObjects(obj_l1, obj_l2)#возвращает guid квартиры или None
        
        isNewAbon=SimpleCheckIfExist('objects','name', obj_l2,'abonents', 'name', abon)
        
        #print 'isNewObj_l0 ', not isNewObj_l0,'isNewObj_l1 ', not isNewObj_l1, 'guid_obj2 ', str(guid_obj2), ' IsNewAbon', not isNewAbon 
        #print i, obj_l1, obj_l2, abon
        if not (isNewObj_l0):
            writeToLog('Level 0 create object '+obj_l0)
            add_parent_object = Objects(name=obj_l0, level=0) 
            add_parent_object.save()
            writeToLog( " Ok")
            writeToLog('create object '+obj_l1)
            #print add_parent_object
            add_object1=Objects(name=obj_l1, level=1, guid_parent = add_parent_object)
            add_object1.save()
            writeToLog('create object '+obj_l2)
            add_object2=Objects(name=obj_l2, level=2, guid_parent = add_object1)
            add_object2.save()            
            writeToLog('create abonent '+abon)
            add_abonent = Abonents(name = abon, guid_objects =add_object2, guid_types_abonents = TypesAbonents.objects.get(guid= u"e4d813ca-e264-4579-ae15-385cdbf5d28c"))
            add_abonent.save()
            kv+=1
            result=u"Объекты: "+obj_l0+", "+obj_l1+u", "+obj_l2+u","+abon+u" созданы"
            continue
        if not (isNewObj_l1):#новый корпус
            writeToLog('Level 1 create object '+obj_l1)
            dtParent=GetSimpleTable('objects','name',obj_l0)
            if dtParent: #родительский объект есть - корпус
                guid_parent=dtParent[0][0]
                add_object1=Objects(name=obj_l1, level=1, guid_parent = Objects.objects.get(guid=guid_parent))
                add_object1.save()                
                writeToLog('create object '+obj_l2)
                add_object2=Objects(name=obj_l2, level=2, guid_parent = add_object1)
                add_object2.save()
                writeToLog('create abonent '+abon)
                add_abonent = Abonents(name = abon, guid_objects =add_object2, guid_types_abonents = TypesAbonents.objects.get(guid= u"e4d813ca-e264-4579-ae15-385cdbf5d28c"))
                add_abonent.save()
                kv+=1
                result+=u"Объекты: "+obj_l1+u", "+obj_l2+u","+abon+u" созданы"
                continue
            else: 
                writeToLog(u'Не удалось создать объект '+obj_l1)
                continue
            
        if bool(not guid_obj2): #новая квартира
            #переделать добавление на добавление по гуиду
            writeToLog('Level 2 create object '+obj_l2)
            dtParent=GetSimpleTable('objects','name',obj_l1)
            if dtParent: #родительский объект есть
                guid_parent=dtParent[0][0]
                add_object = Objects(name=obj_l2, level=2, guid_parent = Objects.objects.get(guid=guid_parent))
                add_object.save()
                result+=u"Объект: "+obj_l2+u" создан"
                add_abonent = Abonents(name = abon, guid_objects = add_object, guid_types_abonents = TypesAbonents.objects.get(guid= u"e4d813ca-e264-4579-ae15-385cdbf5d28c"))
                add_abonent.save()
                kv+=1
        if not (isNewAbon):
            writeToLog('Just create abonent '+ abon)
            if bool(guid_obj2): #родительский объект есть
                add_abonent = Abonents(name = abon, guid_objects = Objects.objects.get(guid=guid_obj2), guid_types_abonents = TypesAbonents.objects.get(guid= u"e4d813ca-e264-4579-ae15-385cdbf5d28c"))
                add_abonent.save()
                kv+=1            
#            else: 
#                print u'Не удалось создать объект '+abon
                continue

    result+=u" Прогружено "+str(kv)+u" водо-счётчиков"
    return result
    
def load_water_pulsar(request):
    global isService
    isService=True
    OnOffSignals()
    args={}
    result=""
    if request.is_ajax():
        if request.method == 'GET':
            request.session["choice_file"]    = fileName    = request.GET['choice_file']
            request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']
            request.session["tcp_ip_status"]    = tcp_ip_status    = request.GET['tcp_ip_status']
            request.session["object_status"]    = object_status    = request.GET['object_status']
            request.session["counter_status"]    = counter_status    = request.GET['counter_status']
            directory=os.path.join(BASE_DIR,'static/cfg/')
            sPath=directory+fileName
            result=LoadWaterPulsar(sPath, sheet)
    counter_status=result#"Загрузка счётчиков условно прошла"
        
    #print fileName
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["tcp_ip_status"]=tcp_ip_status
    args["object_status"]=object_status
    args["counter_status"]=counter_status
    isService=False
    OnOffSignals()
    return render_to_response("service/service_water.html", args)
    
def LoadWaterPulsar(sPath, sSheet):
    global cfg_excel_name
    cfg_excel_name=sPath
    global cfg_sheet_name
    cfg_sheet_name=sSheet
    result=u""
    dtAll=GetTableFromExcel(sPath,sSheet) #получили из excel все строки до первой пустой строки (проверка по колонке А)
    met=0
    con=0
    for i in range(2,len(dtAll)):
        obj_l0=u'Вода' # всегда будет Вода как объект-родитель
        obj_l1=dtAll[i][0] #корпус
        obj_l2=dtAll[i][1] #квартира
        if not dtAll[i][1] or dtAll[i][1]==None:
            j=i
            while not obj_l2 or obj_l2==None:
                j-=1
                obj_l2=dtAll[j][1]
        abon=unicode(dtAll[i][2]) #абонент он же счётчик по воде
        numPulsar=unicode(dtAll[i][5]) #номер пульсара
        typePulsar=unicode(dtAll[i][6]) #тип пульсара
        
        isNewAbon=SimpleCheckIfExist('objects','name', obj_l2,'abonents', 'name', abon)
        isNewPulsar=SimpleCheckIfExist('meters','address', numPulsar,'','','')
        #writeToLog(u'пульсар существует '+ unicode(isNewPulsar)+ typePulsar+ numPulsar)
        if not (isNewAbon):
            return u"Сначала создайте стурктуру объектов и счётчиков"
        if not (isNewPulsar):
            print (u'Обрабатываем строку '+unicode(obj_l2) +' '+ unicode(numPulsar))
            if unicode(typePulsar) == u'Пульсар 10M':
                    add_meter = Meters(name = unicode(typePulsar) + u' ' + unicode(numPulsar), address = unicode(numPulsar), factory_number_manual = unicode(numPulsar), guid_types_meters = TypesMeters.objects.get(guid = u"cae994a2-6ab9-4ffa-aac3-f21491a2de0b") )
                    add_meter.save()
                    print (u'OK Прибор добавлен в базу')
                    met+=1
            elif unicode(typePulsar) == u'Пульсар 16M':
                   add_meter = Meters(name = unicode(unicode(typePulsar) + u' ' + unicode(numPulsar)), address = unicode(numPulsar),  factory_number_manual = unicode(numPulsar), guid_types_meters = TypesMeters.objects.get(guid = u"7cd88751-d232-410c-a0ef-6354a79112f1") )
                   add_meter.save()
                   print (u'OK Прибор добавлен в базу')
                   met+=1
            elif unicode(typePulsar) == u'Пульсар 2M':
                   add_meter = Meters(name = unicode(unicode(typePulsar) + u' ' + unicode(numPulsar)), address = unicode(numPulsar),  factory_number_manual = unicode(numPulsar), guid_types_meters = TypesMeters.objects.get(guid = u"6599be9a-1f4d-4a6e-a3d9-fb054b8d44e8") )
                   add_meter.save()
                   print (u'OK Прибор добавлен в базу')
                   met+=1
            else:
                print(u'Такой Пульсар уже есть')
        else:
            # надо проверить каналы и подсоединить их 
            #Пульсар 16M 029571 Пульсар 16M Канал 16 Суточный -- adress: 16  channel: 0
            chanel=unicode(dtAll[i][4])
            pulsarName=unicode(dtAll[i][6])
            abonent_name=unicode(dtAll[i][2])
            taken_param = pulsarName + u' ' + unicode(dtAll[i][5]) + u' '+ pulsarName + u' ' + u'Канал ' + chanel+ u' Суточный -- adress: ' +chanel+u'  channel: 0'
            print(taken_param)
            #Sravnenie(taken_param)
            dtTakenParam=GetSimpleTable('taken_params','name',taken_param)
            #writeToLog(bool(dtTakenParam))
            if dtTakenParam:                
                print(u'taken param найден')
                guid_taken_param=dtTakenParam[0][1]
                dtLink=GetSimpleTable('link_abonents_taken_params','guid_taken_params',guid_taken_param)
                if (dtLink):
                    result+=u"\n Привязка канала "+chanel+u" Пульсара "+pulsarName+u" уже существует. Перезапись НЕ произведена для счётчика "+abonent_name
                    continue
                dtAbon=GetSimpleTable('abonents','name', abonent_name)
                guidAbon=dtAbon[0][0]
                #print guidAbon
                #print guid_taken_param
                #print TakenParams.objects.get(guid=guid_taken_param) 
                #"миномес ГВС, №68208 Канал 5 Суточный"
                add_link_abonents_taken_param = LinkAbonentsTakenParams (name = abonent_name+u' Канал '+chanel+u' Суточный',coefficient=1, coefficient_2 = 1, guid_abonents = Abonents.objects.get(guid =guidAbon), guid_taken_params = TakenParams.objects.get(guid=guid_taken_param) )
                add_link_abonents_taken_param.save()
                writeToLog(u'Abonent connected with taken param')
                con+=1
    result=u'Прогружено новых пульсаров '+unicode(met)
    if con>0:
        result+=u'Созданы новые связи'
    return result

#def Sravnenie(takenParam):
#    str_bd='Пульсар 2М 062726 Пульсар 2M Канал 1 Суточный -- adress: 1 channel: 0'
#    i=0
#    print str_bd
#    while i!=len(takenParam):
#        if ord(takenParam[i])!=ord(str_bd[i]):
#            print i, takenParam[i]
#        i+=1

def load_water_port(request):
    args={}

    fileName=""
    sheet    = ""
    tcp_ip_status    = ""
    result=""
    if request.is_ajax():
        if request.method == 'GET':
            request.session["choice_file"]    = fileName    = request.GET['choice_file']
            request.session["choice_sheet"]    = sheet    = request.GET['choice_sheet']
            request.session["tcp_ip_status"]    = tcp_ip_status    = request.GET['tcp_ip_status']
            
            #print fileName
            directory=os.path.join(BASE_DIR,'static/cfg/')
            sPath=directory+fileName
            #print sPath, sheet
            result=load_tcp_ip_water_ports_from_excel(sPath, sheet)
    #print result
    if result:
        tcp_ip_status=u"Порт/ы был успешно добавлен"
    else:
        tcp_ip_status=u"Порт не был загружен, он уже существует в БД"
    
    
    #print fileName
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["tcp_ip_status"]=tcp_ip_status

    return render_to_response("service/service_water.html", args)
  
def change_electric_meters(request):
    args={}

    old_meter=u''
    new_meter=u''
    change_meter_status=u"Функция в разработке"
    if request.is_ajax():
        if request.method == 'GET':
            
            request.session["old_meter"]    = old_meter    = request.GET.get('old_meter')
            request.session["new_meter"]    = new_meter   = request.GET.get('new_meter')
            if (not old_meter or old_meter==None or new_meter==None or not new_meter):
                change_meter_status=u"Заполните обе ячейки"
            else:
                change_meter_status=ChangeMeters(old_meter, new_meter)

                
    #change_meter_status=unicode(old_meter)+unicode(new_meter)
    args["change_meter_status"]=change_meter_status

    return render_to_response("service/service_change_electric.html", args)
  
def ChangeMeters(old_meter, new_meter):
    result=u""
    isExistOldMeter=SimpleCheckIfExist('meters','factory_number_manual',old_meter,"","","")
    isExistNewMeter=SimpleCheckIfExist('meters','factory_number_manual',new_meter,"","","")
    if not isExistOldMeter:
        return u"Номера старого счётчика нет в базе"
    if isExistNewMeter:
        return u"Новый счётчик уже существует а базе"
    
    dtOldMeter=GetSimpleTable('meters','factory_number_manual', old_meter)
    guidOldMeter=unicode(dtOldMeter[0][0])
    
    dtTakenParams=GetSimpleTable('taken_params','guid_meters', guidOldMeter)
    
    oldName=unicode(dtOldMeter[0][1])
    newName=oldName.replace(old_meter,new_meter) #поменять на срез+ добавление или формировать полность по новой
    old_factory_number_manual=unicode(dtOldMeter[0][5])
    new_factory_number_manual=old_factory_number_manual.replace(old_meter,new_meter)  #поменять на срез+ добавление или формировать полность по новой
    old_address=unicode(dtOldMeter[0][2])
    new_address=old_address.replace(old_meter,new_meter)
    
    if UpdateTable('meters','guid', guidOldMeter, 'name', newName, 'factory_number_manual', new_factory_number_manual,'address', new_address):
        result=u"Счётчик "+unicode(old_meter)+ " успешно заменён на "+unicode(new_meter)
    #print result
    con=0
    for i in range(len(dtTakenParams)):
        dtTakenParams[i]=list(dtTakenParams[i])
        guidTaken=unicode(dtTakenParams[i][1])
        dtLinkAbonentsTakenParams=GetSimpleTable('link_abonents_taken_params','guid_taken_params', guidTaken)
        oldTakenParamName=unicode(dtTakenParams[i][4])
        #newTakenParamName=oldTakenParamName.replace(old_meter,new_meter)
        OldLinkAbonentTakenParamName=unicode(dtLinkAbonentsTakenParams[0][1])
        #newLinkAbonentTakenParamName= OldLinkAbonentTakenParamName.replace(old_meter,new_meter)
        #get_taken_param_by_abonent_from_excel_cfg(instance.name)).name + u" " + instance.guid_params.guid_names_params.name + u" " + instance.guid_params.guid_types_params.name
        #"Квартира 0103 - М-230 21949676"
        if (OldLinkAbonentTakenParamName.find('М-230')):
            typeMeter=u'М-230'
        if (OldLinkAbonentTakenParamName.find('Саяны Комбик')):
            typeMeter=u'Саяны Комбик'  
        newLinkAbonentTakenParamName=OldLinkAbonentTakenParamName.split('-')[0]+ u' - '+ typeMeter +u' ' + unicode(new_meter)
        
        # "М-230 22633939 Меркурий 230 T0 A+ Суточный -- adress: 0  channel: 0"
        #"Саяны Комбик 4443 Саяны Комбик Q Система1 Суточный -- adress: 0  channel: 1"
        n=oldTakenParamName.find(old_meter)
        s=oldTakenParamName[n+len(old_meter):]
        newTakenParamName= typeMeter + u' ' + unicode(new_meter) + s
        #print newTakenParamName
#        print newLinkAbonentTakenParamName
        if UpdateTable('link_abonents_taken_params','guid_taken_params', guidTaken, 'name', newLinkAbonentTakenParamName,"","","","") and UpdateTable('taken_params','guid', guidTaken, 'name',newTakenParamName,"","","",""):
            con+=1
    result+=u"; Изменено связей:"+unicode(con)
    
    return result
    
def UpdateTable(table,whereFieled, whereValue,field1,value1,field2,value2,field3,value3):
    isOk=False
    dt=[]
    cursor = connection.cursor()
    if (field2==""):
        sQuery="""           
     UPDATE %s
     SET  %s='%s'       
     WHERE %s='%s'
     RETURNING * 
   """%(table, field1, value1, whereFieled, whereValue)
    elif (field3==""):
        sQuery="""           
     UPDATE %s
     SET  %s='%s', %s='%s'      
     WHERE %s='%s'
     RETURNING * 
   """%(table, field1, value1,field2,value2,whereFieled, whereValue)
    else:
       sQuery="""           
     UPDATE %s
     SET  %s='%s', %s='%s', %s='%s'       
     WHERE %s='%s'
     RETURNING * 
   """%(table, field1, value1,field2,value2,field3,value3,whereFieled, whereValue)
    print sQuery
    cursor.execute(sQuery)
    dt = cursor.fetchall()
    if len(dt):
        isOk=True   
    return isOk
    
def load_tcp_ip_water_ports_from_excel(sPath, sheet):
    #Добавление tcp_ip портов

    wb = load_workbook(filename = sPath)
    sheet_ranges = wb[sheet]
    row = 3
    IsAdded=False
    result=""
    writeToLog('Load port')
    writeToLog(u'Загрузка портов')
    while (bool(sheet_ranges[u'H%s'%(row)].value)):
        if sheet_ranges[u'H%s'%(row)].value is not None:
            ip_adr=unicode(sheet_ranges[u'H%s'%(row)].value)
            ip_port=unicode(sheet_ranges[u'I%s'%(row)].value)
            #print ip_adr, ip_port
            writeToLog(u'Обрабатываем адрес ' +ip_adr + ip_port)
            
            # проверка есть ли уже такой порт, запрос в БД с адресом и портом, если ответ пустой-добавляем, в противном случае continue
            if not ip_adr or not ip_port or ip_adr==None or ip_port==None: 
                result+=u"Отсутствует значение/я для tcp/ip-порта в строке"+unicode(row)+". Заполните все ячейки excel таблицы."
                break
            else:
                if (checkPortIsExist(ip_adr,ip_port)):
                    add_port=TcpipSettings(ip_address = ip_adr, ip_port =int(ip_port), write_timeout =300 , read_timeout =700 , attempts =3 , delay_between_sending =400)
                    add_port.save()
                    result =u'Новый tcp/ip порт добавлен'
                    IsAdded=True
                else: result+= u'Порт '+unicode(ip_adr)+": "+unicode(ip_port)+u" уже существует"
        writeToLog( result)
        row+=1
    return IsAdded
    
def replace_electric_meters(request):
    args={}

    meter1=u''
    meter2=u''
    change_meter_status=u""
    replace_meter_status=u'НЕ удалось поменять счётчики местами'
    if request.is_ajax():
        if request.method == 'GET':                        
            request.session["meter1"]    = meter1    = request.GET.get('meter1')
            request.session["meter2"]    = meter2   = request.GET.get('meter2')
            
            if (not meter1 or meter1==None or meter2==None or not meter2):
                replace_meter_status=u"Заполните обе ячейки"
            else:                
                replace_meter_status=ReplaceMeters(meter1, meter2)

    args["change_meter_status"]=change_meter_status
    args["replace_meter_status"]=replace_meter_status
    return render_to_response("service/service_change_electric.html", args)
    
def ReplaceMeters(meter1, meter2):
    result=u''
    
    isExistOldMeter=SimpleCheckIfExist('meters','factory_number_manual',meter1,"","","")
    isExistNewMeter=SimpleCheckIfExist('meters','factory_number_manual',meter2,"","","")
    if not isExistOldMeter:
        return u"Номера первого счётчика нет в базе"
    if not isExistNewMeter:
        return u"Номера второго счётчика нет в базе"
        
#  objects.guid as obj_guid,      0 
#  objects.name as obj_name,      1
#  abonents.guid as ab_guid,      2 
#  abonents.name as ab_name,      3
#  link_abonents_taken_params.guid as link_ab_taken_guid,       4
#  link_abonents_taken_params.name as link_ab_taken_name,       5
#  taken_params.guid as taken_guid,       6
#  taken_params.name as taken_name,       7
#  meters.guid as meter_guid,             8
#  meters.name as meter_name,             9
#  meters.address as meter_adr,           10
#  meters.factory_number_manual           11
        
    dtAllTakenMeter1=GetSimpleTable('all_taken_params','factory_number_manual', meter1)
    guidAbonent1=unicode(dtAllTakenMeter1[0][2])
    abName1=unicode(dtAllTakenMeter1[0][3])
    
    dtAllTakenMeter2=GetSimpleTable('all_taken_params','factory_number_manual', meter2)    
    guidAbonent2=unicode(dtAllTakenMeter2[0][2])
    abName2=unicode(dtAllTakenMeter2[0][3])
    
    nameParam1=unicode(dtAllTakenMeter1[0][7])
    nameParam2=unicode(dtAllTakenMeter2[0][7])
        
    typeMeter1=getTypeMeter(nameParam1)
    typeMeter2=getTypeMeter(nameParam2)
    
    if len(typeMeter1)<1 or len(typeMeter2)<1:
        return u'Для этого типа счётчика ещё нет функции обработки'
    if typeMeter1 !=typeMeter2:
        return u'Типы счётчиков не совпадают'
    
    result+=changeConnectionMeterAbonent(dtAllTakenMeter1, typeMeter1, meter1, meter2, guidAbonent2, abName2)
    result+=changeConnectionMeterAbonent(dtAllTakenMeter2, typeMeter1, meter2, meter1, guidAbonent1, abName1) 
           
        
    return result

def getTypeMeter(nameParam1):
    typeMeter1=u''
    if (nameParam1.find('М-230') or nameParam1.find('Меркурий 230')):
        typeMeter1=u'М-230'
    elif (nameParam1.find('Саяны Комбик')):
        typeMeter1=u'Саяны Комбик' 
    elif (nameParam1.find('М-200') or nameParam1.find('Меркурий 200')):
        typeMeter1=u"Меркурий 200"
    return  typeMeter1

def changeConnectionMeterAbonent(dtAllTakenMeter1, typeMeter, meter1, meter2, guidAbonent2, abName2):
    result=u''
    #guidMeter1=unicode(dtAllTakenMeter1[0][8])
    #guidAbonent1=unicode(dtAllTakenMeter1[0][1])
    #meterName1=unicode(dtAllTakenMeter1[0][9])
    #abName1=unicode(dtAllTakenMeter1[0][2])
    con1=0
    for i in range(len(dtAllTakenMeter1)):
        dtAllTakenMeter1[i]=list(dtAllTakenMeter1[i])
        guidParam1=unicode(dtAllTakenMeter1[i][6])   
        nameParam1=unicode(dtAllTakenMeter1[i][7])
        guidLinkAbonentParam1=unicode(dtAllTakenMeter1[i][4])
        #nameLinkAbonentParam1=unicode(dtAllTakenMeter1[i][5])
        
        newTakenParamName1= makeNewTakenParamName(nameParam1, meter1, meter2, typeMeter)
        newLinkAbonentTakenParamName1=makeLinkabonentTakenParamName(abName2,typeMeter,meter2)
        print newTakenParamName1        
        print newLinkAbonentTakenParamName1
        
        isUpdateTakenParam=UpdateTable('taken_params','guid', guidParam1, 'name',newTakenParamName1,"","","","")
        isUpdateLinkAbonTakenParam=UpdateTable('link_abonents_taken_params','guid', guidLinkAbonentParam1, 'guid_abonents', guidAbonent2,'name', newLinkAbonentTakenParamName1,"","")   
        if isUpdateTakenParam and isUpdateLinkAbonTakenParam:
            con1+=1
            print con1
      
    if (con1>0):
        result+=u' Счётчик '+meter1+u' привязан к абоненту '+abName2+'. Изменено привязок: '+unicode(con1)
    else: result+=u' Что-то пошло не так, ни одной привязки не изменено! '+meter1
    return result

def makeLinkabonentTakenParamName(abName,typeMeter,new_meter):
    #"Квартира 0103 - М-230 21949676"   
#LinkAbonentsTakenParams (name = Abonents.objects.get(name= get_taken_param_by_abonent_from_excel_cfg(instance.name)).name + u" " + instance.guid_params.guid_names_params.name + u" " + instance.guid_params.guid_types_params.name 
    newLinkAbonentTakenParamName=abName+ u' - '+ typeMeter +u' ' + unicode(new_meter)
    return newLinkAbonentTakenParamName

def makeNewTakenParamName(nameParam1, old_meter, new_meter, typeMeter):
    newName=u''

        # "М-230 22633939 Меркурий 230 T0 A+ Суточный -- adress: 0  channel: 0"
        #"Саяны Комбик 4443 Саяны Комбик Q Система1 Суточный -- adress: 0  channel: 1"
    n=nameParam1.find(old_meter)
    s=nameParam1[n+len(old_meter):]
    newName= typeMeter + u' ' + unicode(new_meter) + s
    return newName
    
def get_electric_progruz(request):
    pass

def get_water_progruz(request):
    pass

def get_heat_progruz(request):
    pass

def get_info(request):
    args={}
    

    return render_to_response("service/service_get_info.html", args)
    
def load_balance_group(request):
    args={}
    fileName=""
    sheet    = ""
    balance_status    = ""
    result="Не загружено"
    if request.is_ajax():
        if request.method == 'GET':
            request.session["choice_file"]    = fileName    = request.GET['choice_file']
            request.session["choice_sheet"]    = sheet    = request.GET.get('choice_sheet')
            request.session["balance_status"]    = balance_status    = request.GET['balance_status']            
            directory=os.path.join(BASE_DIR,'static/cfg/')
            sPath=directory+fileName
            result=LoadBalance(sPath, sheet)
    
    balance_status=result

    #print fileName
    args["choice_file"]    = fileName
    args["choice_sheet"]    = sheet
    args["balance_status"]=balance_status

    return render_to_response("service/service_balance_load.html", args)

def InsertIntoBalanceGroup(guid,name):
    result=u''
    cursor = connection.cursor()
    sQuery="""
    INSERT INTO balance_groups(
            guid, name)
    VALUES ('%s', '%s');
    """%(guid,name)   
    
    cursor.execute(sQuery)
    cursor.close()
    connection.commit()
    result =u'Создана балансная группа '+unicode(name)
    return result

def InsertIntoTypesAbonents(guid,name):
    result=u''
    cursor = connection.cursor()
    sQuery="""
    INSERT INTO types_abonents(
            guid, name)
    VALUES ('%s', '%s');
    """%(guid,name)   
    
    cursor.execute(sQuery)
    cursor.close()
    connection.commit()
    result =u'Создан тип '+unicode(name)
    return result

def GetGuidFromFirstTableCrossWithSecondTable(table1,table2,field1,val1,field2,val2):
    dt=[]
    cursor = connection.cursor()
    sQuery="""
        SELECT 
  %s.guid
FROM 
  public.%s, 
  public.%s
WHERE 
  %s.guid_%s = %s.guid AND
  %s.%s = '%s' AND 
  %s.%s = '%s'"""%(table1, table1, table2,table1, table2,table2,table1,field1,val1,table2,field2,val2)
    #print sQuery
    cursor.execute(sQuery)
    dt = cursor.fetchall()
    #print sQuery
    return dt
   
def UpdateSimpleTable(table,guid,field,val):
    result=False
    cursor = connection.cursor()
    sQuery="""
    UPDATE %s
    SET %s = '%s'
    WHERE guid = '%s'
    """%(table,field,val,guid)   
    
    cursor.execute(sQuery)
    cursor.close()
    connection.commit()
    result =True
    return result
    
def LoadImpulseWaterBalance(dtAll):
    result = "Баланс по водным импульсным счётчикам"
    count_new_link=0
    for i in range(1,len(dtAll)):        
        balance_group=unicode(dtAll[i][0])
        znak=unicode(dtAll[i][1])        
        meter=unicode(dtAll[i][4])
        type_abonent=unicode(dtAll[i][5])
        print balance_group, znak, meter, type_abonent
        isNewBalanceGroup=not SimpleCheckIfExist('balance_groups','name',balance_group,"","","")
        isNewMeter=not SimpleCheckIfExist('meters','factory_number_manual',meter,"","","")
        isNewTypeAbonent=not SimpleCheckIfExist('types_abonents','name',type_abonent,"","","")
        print u'isNewBalanceGroup: ', isNewBalanceGroup
        print u'isNewTypeAbonent: ', isNewTypeAbonent
        print u'isNewMeter: ', isNewMeter
        if isNewBalanceGroup: #если балансной группы ещё не существует, то создаём её
            balance_group_guid=uuid.uuid4()
            result += InsertIntoBalanceGroup(balance_group_guid, balance_group)
            print u'Создана балансная группа '+balance_group
        if isNewTypeAbonent: #если такого типа абонента не существует, то создаём
            types_abonents_guid=uuid.uuid4()
            result += InsertIntoTypesAbonents(types_abonents_guid, type_abonent)
            print u'Создан тип абонента ' + type_abonent       
        if isNewMeter:#ничего не создаём, добавляем сообщение, что абонента надо создать
           result += u'Счётчика '+meter+u' (в таблице строка '+str(i+1)+u') не существует. В балансную группу не добавлен!'
           continue
              
        guid_meters=GetSimpleTable('meters','factory_number_manual',meter)[0][0]        
        if not isNewBalanceGroup:
           balance_group_guid=GetSimpleTable('balance_groups','name',balance_group)[0][0]
       
        #проверяем нет ли такой связи уже
        dt_link=GetSimpleTable('link_balance_groups_meters',"guid_meters",guid_meters[0][0])
        isNewLink=True
        for j in range(1,len(dt_link)):
            print dt_link[j][3]
            if dt_link[j][3] == balance_group_guid:
                isNewLink=False
                result+= u'Счётчик ' + meter + u' уже принадлежит балансной группе ' + balance_group
                break
        if isNewLink:
            print balance_group, meter
            cursor = connection.cursor()
            isZnak=True        
            if znak=='0' or znak == 0:
                isZnak=False
            sQuery="""
                  INSERT INTO link_balance_groups_meters(
                  guid, type, guid_balance_groups, guid_meters)
                  VALUES ('%s', '%s', '%s', '%s');
                  """%(uuid.uuid4(),isZnak,balance_group_guid,guid_meters)      
            cursor.execute(sQuery)
            cursor.close()
            connection.commit()
            count_new_link+=1
    result+= u'  В балансную группу добавлено счётчиков: '+ unicode(count_new_link) 
    return result       
    
def LoadBalance(sPath, sheet):
    result=u"Баланс по цифровым счётчикам"
    count_new_link=0
    dtAll=GetTableFromExcel(sPath, sheet) #получили из excel все строки до первой пустой строки (проверка по колонке А)    

    if len(dtAll)==0: return u'Таблица пуста!'
    
    if (len(str(dtAll[1][3])) < 0) or (dtAll[1][3] is None):
        result = LoadImpulseWaterBalance(dtAll)
    else:      
    
        for i in range(1,len(dtAll)):        
            balance_group=unicode(dtAll[i][0])
            znak=unicode(dtAll[i][1])
            object_name=unicode(dtAll[i][2])
            abonent_name=unicode(dtAll[i][3])
            meter=unicode(dtAll[i][4])
            type_abonent=unicode(dtAll[i][5])
            print balance_group, znak, abonent_name,meter, type_abonent
            isNewBalanceGroup=not SimpleCheckIfExist('balance_groups','name',balance_group,"","","")
            isNewMeter=not SimpleCheckIfExist('meters','factory_number_manual',meter,"","","")
            isNewTypeAbonent=not SimpleCheckIfExist('types_abonents','name',type_abonent,"","","")
            print u'isNewBalanceGroup: ', isNewBalanceGroup
            print u'isNewTypeAbonent: ', isNewTypeAbonent
            print u'isNewMeter: ', isNewMeter
            if isNewBalanceGroup: #если балансной группы ещё не существует, то создаём её
                balance_group_guid=uuid.uuid4()
                result += InsertIntoBalanceGroup(balance_group_guid, balance_group)
                print u'Создана балансная группа '+balance_group
            if isNewTypeAbonent: #если такого типа абонента не существует, то создаём
                types_abonents_guid=uuid.uuid4()
                result += InsertIntoTypesAbonents(types_abonents_guid, type_abonent)
                print u'Создан тип абонента ' + type_abonent       
            if isNewMeter:#ничего не создаём, добавляем сообщение, что абонента надо создать
               result += u'Счётчика '+meter+u' (в таблице должен принадлежать абоненту '+abonent_name+u') не существует. В балансную группу не добавлен!'
               continue
           
            types_abonents_guid=GetSimpleTable('types_abonents','name',type_abonent)[0][0]
            guid_abonent=GetGuidFromFirstTableCrossWithSecondTable('abonents','objects','name',abonent_name,'name',object_name)[0][0]
            isOk=UpdateSimpleTable('abonents', guid_abonent,'guid_types_abonents',types_abonents_guid)
            print u'type of abonents changed: ', isOk 
            
            
            guid_meters=GetSimpleTable('meters','factory_number_manual',meter)[0][0]        
            if not isNewBalanceGroup:
               balance_group_guid=GetSimpleTable('balance_groups','name',balance_group)[0][0]
           
            #проверяем нет ли такой связи уже
            dt_link=GetSimpleTable('link_balance_groups_meters',"guid_meters",guid_meters[0][0])
            isNewLink=True
            for j in range(1,len(dt_link)):
                print dt_link[j][3]
                if dt_link[j][3] == balance_group_guid:
                    isNewLink=False
                    result+= u'Счётчик ' + meter + u' уже принадлежит балансной группе ' + balance_group
                    break
            if isNewLink:
                print balance_group, meter
                cursor = connection.cursor()
                isZnak=True        
                if znak=='0' or znak == 0:
                    isZnak=False
                sQuery="""
                      INSERT INTO link_balance_groups_meters(
                      guid, type, guid_balance_groups, guid_meters)
                      VALUES ('%s', '%s', '%s', '%s');
                      """%(uuid.uuid4(),isZnak,balance_group_guid,guid_meters)      
                cursor.execute(sQuery)
                cursor.close()
                connection.commit()
                count_new_link+=1
        result+= u'  В балансную группу добавлено счётчиков: '+ unicode(count_new_link)    
         
    return result

def service_balance_load(request):    
    args={}
    
    return render_to_response("service/service_balance_load.html", args)

