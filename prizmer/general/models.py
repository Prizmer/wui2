# -*- coding: utf-8 -*-
from django.db import models
from django.db.models import signals
from django.db.models import Max 
from django_extensions.db.fields import UUIDField

#---------------------- Декораторы для того, чтобы работал метод autoconnect для pre_save b post_save
from functools import wraps
from django.db.models.signals import pre_save
from django.db.models.signals import post_save
import common_sql
import service 

from django.conf import settings

#from jsonfield import JSONField

def autoconnect(cls):
    """ 
    Class decorator that automatically connects pre_save / post_save signals on 
    a model class to its pre_save() / post_save() methods.
    """
    def connect(signal, func):
        cls.func = staticmethod(func)
        @wraps(func)
        def wrapper(sender, **kwargs):
            return func(kwargs.get('instance'))
        signal.connect(wrapper, sender=cls)
        return wrapper

    if hasattr(cls, 'pre_save'):
        cls.pre_save = connect(pre_save, cls.pre_save)

    if hasattr(cls, 'post_save'):
        cls.post_save = connect(post_save, cls.post_save)
    
    return cls 
#------------------------------- 
    

#------- Мои модели    
    
class Objects(models.Model):
    guid = UUIDField(version=4, primary_key=True, max_length=38)
    name = models.CharField(max_length=100)
    level = models.SmallIntegerField()
    guid_parent = models.ForeignKey('Objects', db_column = 'guid_parent', blank=True, null=True)
    class Meta:
        db_table = 'objects'
        verbose_name = u'Объект'
        verbose_name_plural = u'Объекты'
        
    def __unicode__(self):
        return self.name
        
class Abonents(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    name = models.CharField(max_length=50)
    account_1 = models.CharField(max_length=16)
    account_2 = models.CharField(max_length=16, blank=True)
    flat_number = models.CharField(max_length=50, blank=True)
    guid_objects = models.ForeignKey('Objects', db_column='guid_objects')
    guid_types_abonents = models.ForeignKey('TypesAbonents', db_column='guid_types_abonents')
    class Meta:
        db_table = 'abonents'
        verbose_name = u'Абонент'
        verbose_name_plural = u'Абоненты'
        
    def __unicode__(self):
        return self.name

# Для личного кабинета создаём привязку авторизованных пользователей к абонентам, которые им доступны для просмотра
@autoconnect 
class LinkAbonentsAuthUser(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    name = models.CharField(max_length=200)
    guid_abonents = models.ForeignKey('Abonents', db_column = 'guid_abonents')
    guid_auth_user = models.ForeignKey(settings.AUTH_USER_MODEL, to_field = 'id', db_column = 'id_auth_user')
    class Meta:
        db_table = 'link_abonents_auth_user'
        verbose_name = u'Привязка абонента к пользователю'
        verbose_name_plural = u'Привязки абонентов к пользователям'
        
    def pre_save(self):
        self.name =  u'%s - %s' % (self.guid_abonents.name, self.guid_auth_user.last_name)
       
    def __unicode__(self):
        return self.name

@autoconnect 
class Comments(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    name=models.CharField('Имя', max_length=50, blank=True)
    comment = models.TextField('Комментарий')
    date = models.DateTimeField('Дата комментария', auto_now_add=True)
    guid_abonents = models.ForeignKey('Abonents', db_column='guid_abonents')
    class Meta:
        db_table = 'comments'
        verbose_name = u'Комментарии'
        verbose_name_plural = u'Комментарии'
        
    def pre_save(self):
        self.name =  u'%s - %s' % (self.guid_abonents.name, self.guid_abonents.guid_objects.name)
       
    def __unicode__(self):
        return self.name  
#        return u'%s %s %s %s %s' % (self.id_taken_params.guid_meters.name ,self.id_taken_params.guid_meters.factory_number_manual , self.id_taken_params.guid_params.guid_names_params.name, self.date, self.value )

""" class Tree_data(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    name=models.CharField('Имя', max_length=50, blank=True)
    info = JSONField()
    class Meta:
        db_table = 'Tree_data'
        verbose_name = u'Дерево объектов'
        verbose_name_plural = u'Дерево объектов'
           
    def __unicode__(self):
        return self.name     """

class TypesAbonents(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    name = models.CharField(unique=True, max_length=50)
    class Meta:
        db_table = 'types_abonents'
        verbose_name = u'Тип абонента'
        verbose_name_plural = u'Типы абонентов'
        
    def __unicode__(self):
        return self.name
        
        
#-------------------------    
    
class MonthlyValues(models.Model):
#    id = models.AutoField(primary_key=True)
    date = models.DateField()
    value = models.FloatField()
    status = models.BooleanField(default=False)
    id_taken_params = models.ForeignKey('TakenParams', db_column = 'id_taken_params')
    class Meta:
        db_table = 'monthly_values'
        verbose_name = u'Месячное значение'
        verbose_name_plural = u'Месячные значения'
    def __unicode__(self):
        return u'%s %s %s %s %s' % (self.id_taken_params.guid_meters.name ,self.id_taken_params.guid_meters.factory_number_manual , self.id_taken_params.guid_params.guid_names_params.name, self.date, self.value )
        
class DailyValues(models.Model):
#    id = models.AutoField(primary_key=True)
    date = models.DateField()
    value = models.FloatField()
    status = models.BooleanField(default=False)
    id_taken_params = models.ForeignKey('TakenParams', db_column = 'id_taken_params')
    class Meta:
        db_table = 'daily_values'
        verbose_name = u'Суточное значение'
        verbose_name_plural = u'Суточные значения'
    def __unicode__(self):
        return u'%s %s %s %s %s' % (self.id_taken_params.guid_meters.name ,self.id_taken_params.guid_meters.factory_number_manual , self.id_taken_params.guid_params.guid_names_params.name, self.date, self.value )
        
class CurrentValues(models.Model):
#    id = models.AutoField(primary_key=True)
    date = models.DateField()
    time = models.TimeField()
    value = models.FloatField()
    status = models.BooleanField(default=False)
    id_taken_params = models.ForeignKey('TakenParams', db_column = 'id_taken_params')
    class Meta:
        db_table = 'current_values'
        verbose_name = u'Текущее значение'
        verbose_name_plural = u'Текущие значения'
    def __unicode__(self):
        return u'%s %s %s %s %s' % (self.id_taken_params.guid_meters.name ,self.id_taken_params.guid_meters.factory_number_manual , self.id_taken_params.guid_params.guid_names_params.name, self.date, self.value )
        
class CurrentValuesArchive(models.Model):
#    id = models.AutoField(primary_key=True)
    date = models.DateField()
    time = models.TimeField()
    value = models.FloatField()
    status = models.BooleanField(default=False)
    id_taken_params = models.ForeignKey('TakenParams', db_column = 'id_taken_params')
    class Meta:
        db_table = 'current_values_archive'
        verbose_name = u'Архивное текущее значение'
        verbose_name_plural = u'Архив текущих значений'
    def __unicode__(self):
        return u'%s %s %s %s %s' % (self.id_taken_params.guid_meters.name ,self.id_taken_params.guid_meters.factory_number_manual , self.id_taken_params.guid_params.guid_names_params.name, self.date, self.value )
    
        
class VariousValues(models.Model):
#    id = models.AutoField(primary_key=True)
    date = models.DateField()
    time = models.TimeField()
    value = models.FloatField()
    status = models.BooleanField(default=False)
    id_taken_params = models.ForeignKey('TakenParams', db_column = 'id_taken_params')
    class Meta:
        db_table = 'various_values'
        verbose_name = u'Настраиваемое значение'
        verbose_name_plural = u'Настраиваемые значения'
    def __unicode__(self):
        return u'%s %s %s %s %s' % (self.id_taken_params.guid_meters.name ,self.id_taken_params.guid_meters.factory_number_manual , self.id_taken_params.guid_params.guid_names_params.name, self.date, self.value )
        
#----------------------------------
        
class TypesParams(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    name = models.CharField(unique=True, max_length=50)
    period = models.IntegerField(blank=True, null=True, default=0)
    type = models.DecimalField(unique=True, max_digits=3, decimal_places=0)
    class Meta:
        db_table = 'types_params'
        verbose_name = u'Тип считываемого параметра'
        verbose_name_plural = u'Типы считываемых параметров'

    def __unicode__(self):
        return unicode(self.name)
    
@autoconnect
class Params(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    name = models.CharField(max_length=200, blank=True)
    param_address = models.IntegerField()
    channel = models.IntegerField(default=0)
    guid_types_meters = models.ForeignKey('TypesMeters', db_column = 'guid_types_meters')
    guid_names_params = models.ForeignKey('NamesParams', db_column = 'guid_names_params')
    guid_types_params = models.ForeignKey('TypesParams', db_column = 'guid_types_params')
    class Meta:
        db_table = 'params'
        verbose_name = u'Параметр'
        verbose_name_plural = u'Параметры'
        
    def pre_save(self):
        self.name = u'%s %s %s -- adress: %s  channel: %s' % (self.guid_types_meters.name, self.guid_names_params.name, self.guid_types_params.name, self.param_address, self.channel)
    
    def __unicode__(self):
        return self.name
        
@autoconnect     
class TakenParams(models.Model):
    #id = models.AutoField(primary_key=True)
    name = models.CharField(max_length=200, blank=True)
    guid = UUIDField(unique = True, max_length=38)
    guid_params = models.ForeignKey('Params', db_column = 'guid_params')
    guid_meters = models.ForeignKey('Meters', db_column = 'guid_meters')
    class Meta:
        db_table = 'taken_params'
        verbose_name = u'Считываемый параметр'
        verbose_name_plural = u'Считываемые параметры'
                
    def pre_save(self):
        self.name = u'%s %s' % (self.guid_meters.name, self.guid_params.name)
        
    def __unicode__(self):
        return self.name
        
class LinkAbonentsTakenParams(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    name = models.CharField(max_length=200)
    coefficient = models.FloatField(default=1)
    coefficient_2 = models.FloatField(default=1)
    coefficient_3 = models.FloatField(default=1000)
    guid_abonents = models.ForeignKey('Abonents', db_column = 'guid_abonents')
    guid_taken_params = models.ForeignKey('TakenParams', to_field = 'guid', db_column = 'guid_taken_params')
    class Meta:
        db_table = 'link_abonents_taken_params'
        verbose_name = u'Привязка абонента к параметру'
        verbose_name_plural = u'Привязки абонентов к параметрам'
        
    def __unicode__(self):
        return self.name
              
       
class Resources(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    name = models.CharField(unique=True, max_length=50)
    type = models.DecimalField(unique=True, max_digits=3, decimal_places=0)
    class Meta:
        db_table = 'resources'
        verbose_name = u'Ресурс'
        verbose_name_plural = u'Ресурсы'
        
    def __unicode__(self):
        return self.name
        
class Meters(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    name = models.CharField('Имя', unique=True, max_length=50)
    address = models.IntegerField('Сетевой адрес')
    password = models.CharField('Пароль', max_length=100, blank=True)
    attr1 = models.CharField('Атрибут 1', max_length=20, blank=True)
    attr2 = models.CharField('Атрибут 2', max_length=20, blank=True)
    attr3 = models.CharField('Атрибут 3', max_length=20, blank=True)
    attr4 = models.CharField('Атрибут 4', max_length=20, blank=True)
    password_type_hex = models.BooleanField('Использовать HEX для пароля?', default=True)
    factory_number_manual = models.CharField('Заводской номер(вручную)', max_length=16)
    factory_number_readed = models.CharField('Заводской номер(из прибора)', max_length=16, blank=True, null=True)
    is_factory_numbers_equal = models.NullBooleanField('Совпадение номеров', blank=True, null=True)
    dt_install = models.DateTimeField('Дата установки', blank=True, null=True)
    dt_last_read = models.DateTimeField('Дата последнего удачного чтения данных', blank=True, null=True)
    guid_types_meters = models.ForeignKey('TypesMeters', db_column='guid_types_meters')
    guid_meters = models.ForeignKey('Meters', db_column='guid_meters', blank=True, null=True)
    time_delay_current = models.IntegerField(default=10)
    class Meta:
        db_table = 'meters'
        verbose_name = u'Счётчик'
        verbose_name_plural = u'Счётчики'
        
    def __unicode__(self):
        return u'%s - %s' % (self.name, self.factory_number_manual)
        
class TypesMeters(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    name = models.CharField(unique=True, max_length=50)
    driver_name = models.CharField(unique=False, max_length=50)
    class Meta:
        db_table = 'types_meters'
        verbose_name = u'Тип счётчик'
        verbose_name_plural = u'Типы счётчиков'
        
    def __unicode__(self):
        return self.name
        
class LinkMetersComportSettings(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    guid_meters = models.ForeignKey('Meters', db_column='guid_meters')
    guid_comport_settings = models.ForeignKey('ComportSettings', db_column='guid_comport_settings')
    class Meta:
        db_table = 'link_meters_comport_settings'
        verbose_name = u'Привязка счётчика к com порту'
        verbose_name_plural = u'Привязки счётчиков к com портам'
        
    def __unicode__(self):
        return u'Com %s - %s' % (self.guid_comport_settings.name, self.guid_meters.name)


class LinkMetersTcpipSettings(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    guid_meters = models.ForeignKey('Meters', db_column='guid_meters')
    guid_tcpip_settings = models.ForeignKey('TcpipSettings', db_column='guid_tcpip_settings')
    class Meta:
        db_table = 'link_meters_tcpip_settings'
        verbose_name = u'Привязка счётчика к tcp/ip порту'
        verbose_name_plural = u'Привязки счётчиков к tcp/ip портам'
        
    def __unicode__(self):
        return u'%s:%s - %s' % (self.guid_tcpip_settings.ip_address, self.guid_tcpip_settings.ip_port, self.guid_meters.name)
        
class ComportSettings(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    name = models.CharField(max_length=3)
    baudrate = models.IntegerField()
    data_bits = models.DecimalField(max_digits=3, decimal_places=0)
    parity = models.DecimalField(max_digits=3, decimal_places=0)
    stop_bits = models.DecimalField(max_digits=3, decimal_places=0)
    write_timeout = models.SmallIntegerField()
    read_timeout = models.SmallIntegerField()
    attempts = models.DecimalField(max_digits=3, decimal_places=0)
    delay_between_sending = models.IntegerField()
    gsm_on = models.BooleanField('Использовать CSD канал?', default=False)
    gsm_phone_number = models.CharField(unique=False, max_length=15)
    gsm_init_string  = models.CharField(unique=False, max_length=50)
    class Meta:
        db_table = 'comport_settings'
        verbose_name = u'Com порт'
        verbose_name_plural = u'Com порты'
        
    def __unicode__(self):
        return u'Com %s. %s %s%s%s' % (self.name, self.baudrate, self.data_bits, str(self.parity).replace("0","n").replace("2","e"), self.stop_bits)
        
class TcpipSettings(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    ip_address = models.CharField(max_length=15)
    ip_port = models.IntegerField()
    write_timeout = models.SmallIntegerField()
    read_timeout = models.SmallIntegerField()
    attempts = models.DecimalField(max_digits=3, decimal_places=0)
    delay_between_sending = models.IntegerField()
    class Meta:
        db_table = 'tcpip_settings'
        verbose_name = u'TCP/IP порт'
        verbose_name_plural = u'TCP/IP порты'
        
    def __unicode__(self):
        return u'%s:%s' % (self.ip_address, self.ip_port)
        
class Measurement(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    name = models.CharField(unique=True, max_length=50)
    comments = models.CharField(max_length=50, blank=True)
    class Meta:
        db_table = 'measurement'
        verbose_name = u'Единица измерения'
        verbose_name_plural = u'Единицы измерения'
        
    def __unicode__(self):
        return self.name
        

class NamesParams(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    name = models.CharField(unique=True, max_length=50)
    guid_resources = models.ForeignKey('Resources', db_column = 'guid_resources')
    guid_measurement = models.ForeignKey('Measurement', db_column = 'guid_measurement')
    class Meta:
        db_table = 'names_params'
        verbose_name = u'Наименование параметра'
        verbose_name_plural = u'Наименования параметров'
        
    def __unicode__(self):
        return self.name
        
class BalanceGroups(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    name = models.CharField(unique=True, max_length=50)
    #guid_names_params = models.ForeignKey('NamesParams', db_column = 'guid_names_params')
    class Meta:
        db_table = 'balance_groups'
        verbose_name = u'Балансная Группа'
        verbose_name_plural = u'Балансные Группы'
        
    def __unicode__(self):
        return self.name
    
        
class LinkBalanceGroupsMeters(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    guid_balance_groups = models.ForeignKey('BalanceGroups', db_column = 'guid_balance_groups')
    guid_meters = models.ForeignKey('Meters', db_column='guid_meters') 
    type = models.BooleanField(default=True, verbose_name = u"Знак входа в группу '+' ?")
    
    class Meta:
        db_table = 'link_balance_groups_meters'
        verbose_name = u'Привязка Групп к счётчику'
        verbose_name_plural = u'Привязка Групп к счетчикам'
        
    def __unicode__(self):
        return u'%s - %s' % (self.guid_balance_groups.name, self.guid_meters.name )
        
class ProductCoefficientsKilns(models.Model):
    sfid = models.IntegerField()
    coefficient = models.FloatField()
    
    class Meta:
        db_table = 'product_coefficients_kilns'
        verbose_name = u'Удельный коэффициент продукта'
        verbose_name_plural = u'Удельные коэффициенты продуктов'

class ProductInfoKilns(models.Model):
    dt = models.DateField()
    kiln_code = models.IntegerField()
    product_caption = models.CharField(max_length=50)
    product_count = models.IntegerField()
    product_coefficient = models.FloatField()
    product_weight = models.FloatField()
    
    class Meta:
        db_table = 'product_info_kilns'
        verbose_name = u'Информация по продукции'
        
class ProductTypeKilns(models.Model):
    nm = models.CharField(max_length=80)
    kind_id = models.IntegerField()
    
    class Meta:
        db_table = 'product_type_kilns'
        verbose_name = u'Типы продукции'  
        
class Groups80020(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    name = models.CharField(unique=True, max_length=50)
    name_sender = models.CharField(max_length=250) 
    inn_sender  = models.CharField(max_length=250)
    name_postavshik = models.CharField(max_length=250)
    inn_postavshik = models.CharField(max_length=250)
    dogovor_number = models.CharField(default=10, max_length=50)
    class Meta:
        db_table = 'groups_80020'
        verbose_name = u'Группа отчётов 80020'
        verbose_name_plural = u'Группы отчётов 80020'
        
    def __unicode__(self):
        return self.name

class LinkGroups80020Meters(models.Model):
    guid = UUIDField(primary_key=True, max_length=38)
    guid_groups_80020 = models.ForeignKey('Groups80020', db_column = 'guid_groups_80020')
    guid_meters = models.ForeignKey('Meters', db_column='guid_meters')
    measuringpoint_code = models.DecimalField(max_digits=18, decimal_places=0)
    measuringpoint_name = models.CharField(max_length=250)    
    class Meta:
        db_table = 'link_groups_80020_meters'
        verbose_name = u'Связь счётчика и Групп 80020'
        verbose_name_plural = u'Связи счётчиков и Групп 80020'
            
    def __unicode__(self):
        return u'%s - %s' % (self.guid_groups_80020.name, self.guid_meters.name )
        

#-------------- Создаем различный набор считываемых парамтров, в зависимости от типа прибора учёта       
def add_taken_param(sender, instance, created, **kwargs): # Добавляем считываемые параметры при создании счётчика
    if instance.guid_types_meters.name == u'Меркурий 230':
        #Добавляем параметры для Меркурия 230
    # T0 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"bdcd1268-37f3-4579-a9d9-5becb2ba8aa3")) # A+ T0 месячные
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"99cd6002-f81c-4ad6-9cb0-53a92a498519")) # A+ T0 суточные
        add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e8c20ce7-bdb6-4ea6-8401-cee28049a7d7")) # A+ T0 текущие
        #add_param.save()
    # T0 R+
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"2ebc02e6-65c6-40ab-b717-0d98d66b5701")) # R+ T0 месячные
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"345a24a4-95b7-4f67-b004-716706ed2560")) # R+ T0 суточные
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4c93dd55-1ec2-48c7-9865-9ceab580b7b3")) # R+ T0 текущие
        #add_param.save()
        
    # T1 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"17789c36-4593-4ff2-94eb-1d0cebdb5366")) # A+ T1 месячные
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"d262c71a-6da4-4ec0-a9c3-b9ea659c246d")) # A+ T1 суточные
        add_param.save()
    # T2 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c31297be-220b-4971-8642-6b614aa7ecee")) # A+ T2 месячные
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"37011b85-c8af-4f6c-857d-4b93a95d31e1")) # A+ T2 суточные
        add_param.save()
    # T3 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"79741ba9-e8b8-4352-862e-17a9c4d928ce")) # A+ T3 месячные
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c3bb9033-ffcb-4a28-91e2-6b45924b8858")) # A+ T3 суточные
        add_param.save()
    
    # Ток
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"aee312b0-adb1-4be9-9879-b3a3598f9b29")) # Ia текущее
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"7ed0d364-e790-4325-a927-9ef86a685f00")) # Ib текущее
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"474b0809-482a-4851-9a96-4587f8c59152")) # Ic текущее
        #add_param.save()
    # Напряжение
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c06f7315-abc6-4889-97ad-201a936c8f2c")) # Ua текущее
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"632f76fb-4dd9-4e7d-86a0-a57a27fc648a")) # Ub текущее
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1a3ca6ca-8866-4aad-8712-d9df003fe692")) # Uc текущее
        #add_param.save()
    # Мощность
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"3077b3ac-fde2-4435-9e6f-17464310c090")) # P Активная мощность
        #add_param.save()
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e7617c95-7e42-4cfa-9acd-5bc119261c6d")) # Q Реактивная мощность
        #add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6af9ddce-437a-4e07-bd70-6cf9dcc10b31")) # A+ 30-мин. срез мощности
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"66e997c0-8128-40a7-ae65-7e8993fbea61")) # R+ 30-мин. срез мощности
        add_param.save()
    elif instance.guid_types_meters.name == u'Меркурий 233':
        #Добавляем параметры для Меркурия 233
        pass
    elif instance.guid_types_meters.name == u'Пульсар16':
        #Добавляем параметры для Пульсар16
    # Суточные
      # Канал 1
      # Канал 2
      # Канал 3
      # Канал 4
      # Канал 5
      # Канал 6
      # Канал 7
      # Канал 8
      # Канал 9
      # Канал 10
      # Канал 11
      # Канал 12
      # Канал 13
      # Канал 14
      # Канал 15
      # Канал 16
   
    # Текущие
      # Канал 1
      # Канал 2
      # Канал 3
      # Канал 4
      # Канал 5
      # Канал 6
      # Канал 7
      # Канал 8
      # Канал 9
      # Канал 10
      # Канал 11
      # Канал 12
      # Канал 13
      # Канал 14
      # Канал 15
      # Канал 16
       pass
    elif instance.guid_types_meters.name == u'Пульсар10':
        #Добавляем параметры для Пульсар10
    # Суточные
      # Канал 1
      # Канал 2
      # Канал 3
      # Канал 4
      # Канал 5
      # Канал 6
      # Канал 7
      # Канал 8
      # Канал 9
      # Канал 10
   
    # Текущие
      # Канал 1
      # Канал 2
      # Канал 3
      # Канал 4
      # Канал 5
      # Канал 6
      # Канал 7
      # Канал 8
      # Канал 9
      # Канал 10
   
        pass
    elif instance.guid_types_meters.name == u'Пульсар 16M':
        #Добавляем параметры для Пульсар16
    # Месячные
      # Канал 1
      # Канал 2
      # Канал 3
      # Канал 4
      # Канал 5
      # Канал 6
      # Канал 7
      # Канал 8
      # Канал 9
      # Канал 10
      # Канал 11
      # Канал 12
      # Канал 13
      # Канал 14
      # Канал 15
      # Канал 16
    
    # Суточные
      # Канал 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"fc4a9568-4674-4a80-b497-e4f34399acd5"))
        add_param.save()
      # Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"9e6e308f-abec-4b47-9b99-9cb590c55d0c"))
        add_param.save()
      # Канал 3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e6815dd5-fbbc-480f-8b95-025d7f9a0403"))
        add_param.save()
      # Канал 4
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"612d2f20-d454-4e14-910b-1fd89bbb31dd"))
        add_param.save()
      # Канал 5
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"d82c7576-8e5e-4e93-ae10-58459b31e4a0"))
        add_param.save()
      # Канал 6
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6ccc7efb-d9fe-4285-b343-8ed22d2d3625"))
        add_param.save()
      # Канал 7
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"72567365-9a40-4f97-ab25-0911585035bf"))
        add_param.save()
      # Канал 8
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"9203f5ed-d5da-4462-91d1-5aea42e99124"))
        add_param.save()
      # Канал 9
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e4068568-d8c4-42ab-9957-7292753e2891"))
        add_param.save()
      # Канал 10
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"9b5ab67b-40aa-4536-8b7c-340a773ab31b"))
        add_param.save()
      # Канал 11
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4fd440c4-9ec5-4ab9-a073-6c4d3a174777"))
        add_param.save()
      # Канал 12
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"00b7f7c5-37f3-494a-8ceb-5a62f9ebf4e3"))
        add_param.save()
      # Канал 13
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"169a79e0-da6f-4091-9fc7-ab81adc0d7e0"))
        add_param.save()
      # Канал 14
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"17e9c8fe-0d69-4466-b64e-185452c61555"))
        add_param.save()
      # Канал 15
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"25de493d-c680-4ca6-ac02-b778022ee151"))
        add_param.save()
      # Канал 16
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"908e88f0-f9a0-421d-bbe7-9bafdf5d2565"))
        add_param.save() 
  
    # Текущие
      # Канал 1
       # add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e3f1325e-3018-40ba-b94a-ab6d6ac093e9"))
       # add_param.save()
      # Канал 2
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5a6b0338-c15d-4224-a04f-a10fc73c5fc7"))
        #add_param.save()
      # Канал 3
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"48a42afe-d9ac-4180-a733-6dd5f9d9ca80"))
        #add_param.save()
      # Канал 4
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"01a5419c-c701-4185-95b6-457b8c9ca2d0"))
        #add_param.save()
      # Канал 5
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"85c4295e-bc6a-46ec-9866-0bf9f77c6904"))
        #add_param.save()
      # Канал 6
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"68270d0a-5043-4ea2-9b61-4adaa298abad"))
        #add_param.save()
      # Канал 7
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"cd489c4b-6e74-4c65-bfee-c0fa78a853bf"))
        #add_param.save()
      # Канал 8
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"f29062a4-ab60-4117-8f85-0cdec634c797"))
        #add_param.save()
      # Канал 9
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e8521cd7-2f38-4619-935d-8fe86234dbe7"))
        #add_param.save()
      # Канал 10
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1349b747-41ca-4ba8-a690-69c649129f44"))
        #add_param.save()
      # Канал 11
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"99ab1a30-fde8-4b81-9f9e-2f731516ce1b"))
        #add_param.save()
      # Канал 12
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c7f6a397-833d-4020-9d2b-38c19bec272c"))
        #add_param.save()
      # Канал 13
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4413bffb-1832-4900-9351-5ac3666dd8b0"))
        #add_param.save()
      # Канал 14
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6280490b-123d-4e27-bef9-19fd7dc2cf54"))
        #add_param.save()
      # Канал 15
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"93891c5a-1c8f-4906-b7f0-961dc8ad3c9f"))
        #add_param.save()
      # Канал 16
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"22dd3a17-a828-44e0-80d9-db075ba120ae"))
        #add_param.save()

    elif instance.guid_types_meters.name == u'Пульсар 10M':
        #Добавляем параметры для Пульсар10
    # Месячные
      # Канал 1
      # Канал 2
      # Канал 3
      # Канал 4
      # Канал 5
      # Канал 6
      # Канал 7
      # Канал 8
      # Канал 9
      # Канал 10
    
    # Суточные
      # Канал 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"325ec164-9428-4a57-867c-33d4eaf8cc2a"))
        add_param.save()
      # Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"99a99024-65b4-44dd-99fc-6a5cf1d4aaee"))
        add_param.save()
      # Канал 3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"f897f0ca-4e35-4f0d-b345-3379668aa36f"))
        add_param.save()
      # Канал 4
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"034374bd-2dfb-4568-aa16-84255df33c88"))
        add_param.save()
      # Канал 5
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b6bdfae8-4f27-4056-af79-d746b44038ee"))
        add_param.save()
      # Канал 6
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"2c2f7176-8b77-44f4-9678-4773e95e67ce"))
        add_param.save()
      # Канал 7
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"91bb7c43-f802-4ebd-a8fe-75f833acedeb"))
        add_param.save()
      # Канал 8
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"cf24b669-1c5b-4db7-936a-5f9d5c8be928"))
        add_param.save()
      # Канал 9
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"96035c7c-ee7c-41f6-9723-8a75dd9ed573"))
        add_param.save()
      # Канал 10
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"253475ea-614d-4aad-93a8-e81e4c9028e9"))
        add_param.save()   

    # Текущие
      # Канал 1
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"32dad392-ca1e-4110-8f2c-a86b02e26fb3"))
        #add_param.save()
      # Канал 2
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"3e13694b-7cb5-4417-a091-af8a7db34dc7"))
        #add_param.save()
      # Канал 3
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1023b35b-3cbf-4519-aac3-3bf1ebae07c1"))
        #add_param.save()
      # Канал 4
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"eea27ade-44cd-4e66-8298-00a4a6ad915a"))
        #add_param.save()
      # Канал 5
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"25e09d4d-3a48-4381-ad5d-b783c03c4d35"))
        #add_param.save()
      # Канал 6
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"069898ea-9d74-4571-b719-e8e6f1513c12"))
        #add_param.save()
      # Канал 7
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"084aa5f4-75d5-41f6-b0d6-9f2403eacd2c"))
        #add_param.save()
      # Канал 8
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"786ed8b8-aed1-478c-ae75-99caf1358cf0"))
        #add_param.save()
      # Канал 9
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6fc4c39c-9a43-4cb7-a066-c40fd2ca47e5"))
        #add_param.save()
      # Канал 10
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"8b2aa40a-cd91-4e22-b9d1-596e49e5f839"))
        #add_param.save()  

    elif instance.guid_types_meters.name == u'Пульсар 2M':
        #Добавляем параметры для Пульсар10
    
    # Месячные
      # Канал 1
      # Канал 2
    
    # Суточные
      # Канал 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"0239dffb-de88-45e5-b6f6-18bf39f92525"))
        add_param.save()
      # Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"a1cb319d-ac09-466d-894b-91d90aba4239"))
        add_param.save()   
    
    # Текущие
      # Канал 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"fcc28118-66c0-4cdf-aeba-5da1171aae48"))
        add_param.save()
      # Канал 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1faeb517-bd1f-4ba0-96a5-67f00764822f"))
        add_param.save()

    elif instance.guid_types_meters.name == u'ПСЧ-3ТА.04':
        #Добавляем параметры для ПСЧ-3ТА.04
        pass
    elif instance.guid_types_meters.name == u'ТЭМ-104':
        #Добавляем параметры для ТЭМ-104
        pass
    elif instance.guid_types_meters.name == u'СЭТ-4ТМ.03М':
        #Добавляем параметры для СЭТ-4ТМ.03М
        pass
    elif instance.guid_types_meters.name == u'Меркурий 200':
        #Добавляем параметры для Меркурий 200

    # Значения текущие
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"9cbc001d-a262-481f-a1aa-47d02bf18af1")) #T0
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b65d4227-69a5-487b-9999-5539ca3fc004")) #T1
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5e312de9-34cd-4ba7-a744-c9b94a77d98b")) #T2
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4260ea05-78f8-4c5c-9172-fa161fa96068")) #T3
        add_param.save()
    # Значения на начало месяца
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"86cd925b-48c2-40b8-b211-f116e0e6dbea")) #T0
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"62a3796a-eaae-445d-9166-2ad517186b78")) #T1
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5f6e1e3d-4128-4cfe-94cf-57ac84a7694a")) #T2
        add_param.save()
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"0c28c135-58f2-4dff-a222-9f3d9f3c742b")) #T3
        add_param.save()
    # Значения на начало суток
        #Не поддерживается прибором

    elif instance.guid_types_meters.name == u'Эльф 1.08':
        #Добавляем параметры для счётчика тепла Elf 108
    
        #-------------Текущие
        # "Энергия"
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"f2bbf267-456e-477a-95d2-abb94c78ba43"))
        add_param.save()
        # "Объем"       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"dad6e2eb-e978-46f4-b7ec-442834b04e7a"))
        add_param.save()
        # "ElfTon"  Время работы прбора     
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"d3c9563d-51ed-4ca7-922f-ac3731065ead"))
        add_param.save()
        # "ElfErr"  Время работы прибора с ошибкой
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"dade3324-b9b0-41c8-bc76-70f617573e43"))
        add_param.save()
        # "Ti"      Температура входа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"acca627e-f21a-4f8b-be7e-038f534b5d11"))
        add_param.save()
        # "To"      Температура выхода
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"01487323-a28f-419e-9589-2563d785ab2a"))
        add_param.save()
        # "Канал 1"      Импульсный вход 1 текущий
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6e7f0d37-df5c-4850-991e-b5d7cb793924"))
        #add_param.save()
        # "Канал 1"      Импульсный вход 1 суточный
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"9af27a62-d6c8-4b67-bd36-da7103e0b1f1"))
        add_param.save()
        # "Канал 2"      Импульсный вход 2 суточный
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"86acc33d-7bea-4977-a5b5-c5858ce9a09d"))
        add_param.save()
        # "Канал 2"      Импульсный вход 2 текущий
        #add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"de7bfdfd-c17f-4a7c-942d-b28e85db33cb"))
        #add_param.save()
        #-------------Архивные
        # "Энергия"
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"ae439e1f-5c4b-494c-8a53-a61b85c804a0"))
        add_param.save()
        # "Объем"
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b02153a4-00c0-4800-a55a-c7f9dfbb14e7"))
        add_param.save()
        # "ElfTon" Время работы прибора
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"aa611d48-f1fe-462a-8b0a-0a7596792b69"))
        add_param.save()
        # "ElfErr" Время работы прибора с ошибкой
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"af047098-bd45-4579-a60c-b75bed376bbe"))
        add_param.save()
        
        
    elif instance.guid_types_meters.name == u'СПГ762-1':
        #Добавляем параметры для счётчика газа СПГ762 Подсистема 1
    
        #-------------Часовые
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"edfc6dc0-1628-4a7e-bd04-71107882039a"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1e27e72c-79d6-4c68-bc04-1be84d061622"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"0fa5d9ef-4c6c-4f78-bc64-9d9b34002344"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"433d7025-15f3-4ab0-9d73-39bd0e425566"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"446a6bb6-c17f-4478-b1d8-252c7eb454d3"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"003c2fb2-0092-4d7d-a513-3dcc50a255da"))
        add_param.save()
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"9c99da7b-1a73-48b0-a3f5-54438a3ea824"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"832374da-5834-4fe0-abe3-07d48d447af2"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4920e70a-452e-4ecf-918e-14ca288c7a1f"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"825dfd73-82fa-4f1a-9635-77bdcb244997"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"daa4a90e-7993-493c-9ac0-c03241b2ab2c"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"72b5f4ba-d179-419c-8795-e2f86f5ee2ff"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"41774f43-9c9c-4867-bf21-e3d1df4fd2f8"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"f610969e-cb2d-436f-9357-e63da72d162e"))
        add_param.save()
        
        #-------------Суточные
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"ac9fe54a-6f51-4ee9-a849-448f0f10a4b6"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"2eefa93a-be60-4f23-9b09-8d1c6bad0a15"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"78c30686-a5d0-436f-9c56-57b076769774"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"d7900bac-b85a-4b83-ad67-b822b470a698"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"068f4de0-f041-4608-b09a-81dbc8f319ff"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"dad040cd-1047-4060-a7ac-e9a6e7f30fb4"))
        add_param.save()        
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"3be53eb7-c931-4728-a442-44d14c9da44f"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1dff6784-dd8d-4898-ae8b-f1b60fbdc1af"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"af26b128-b634-47a4-96ba-42de6f039fdb"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5bb8af68-588e-470c-9b64-373482f71468"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c32e7e01-106a-49c6-b8b0-6490448548ad"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"8fc09300-4b19-4478-aa44-d2fb1cf792d5"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"53693586-e5b5-4204-922a-a0b0153298ea"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"bf06727c-9cab-4efe-a0f2-6242bb320372"))
        add_param.save()
        
        # Масса газа нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5399a54f-0d2c-47e8-8ffb-882f5dddc239"))
        add_param.save()
        # Объем газа при ст.условиях нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c14eaa34-3264-4fe4-98ab-8da6618fc431"))
        add_param.save()
        # Объем газа при раб.условиях нарастающим итогом              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4ed5ba8d-9ead-4a60-96e4-726f38432d9a"))
        add_param.save()
        
    elif instance.guid_types_meters.name == u'СПГ762-2':
        #Добавляем параметры для счётчика газа СПГ762 Подсистема 1
    
        #-------------Часовые
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"9a17ea4f-a1f8-4fb9-a21e-62f43978535a"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"49728e75-1d62-4b9e-8633-762cb7117b52"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c44af390-ed2a-49d8-9c67-25f543db9935"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"24bd767d-f667-444e-acfa-a935fb8f4699"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1e675e58-2cc9-4103-8410-2d37704a2bcf"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"d14fa3fa-5bd5-4dd9-b740-61049d38e694"))
        add_param.save()
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6de742eb-a391-473a-bb4b-ab780a4642b8"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"04b43217-af06-4f09-8d95-0e2d3dbd0905"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b62c6838-7578-41c9-a94c-d06788cc2d41"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"aad9d0b4-1c12-4165-a1b1-4fac9de00c38"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"205a8c8c-de26-44e5-ab72-efb7fe72040c"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c84c6130-ede7-487e-a414-b384964eb81e"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5907cc5f-1386-4fbc-9e5c-7d3f77dba6d6"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c5638410-44b6-41d7-b501-6e5c0a002f48"))
        add_param.save()
        
        #-------------Суточные
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c1643ab8-1707-4b73-9610-0226b1fb6860"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"dbacff0c-2b3d-40c5-aa03-8d51a64919dd"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"279bdfd5-7b22-4d7e-900c-21e4077506dd"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b74e6743-3996-4af7-8024-da3912d14b45"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c7220fc7-5c01-4bcc-ac2a-7c851276af4d"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"381ccc3f-a9d9-4dcf-a9aa-2e5bd0e4efc8"))
        add_param.save()        
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"7f0f0e09-3bd0-4595-84dd-754f4c21bc5e"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"ee687ef8-36de-4de8-9a05-4ac841c9c144"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"8ee128d0-5c21-4faa-a2fe-7432ff9be684"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c61c87eb-5a4f-4095-ac50-4324e7899340"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"bae3e866-f057-4be5-99a0-7474f6c7cbc1"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"8ce68aef-85fa-4ca6-8f9f-dfa1f9e71cdd"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"ea094191-7ce5-4c42-ad5e-e886d02e73e0"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e4d751b4-ef6c-45ca-b31a-f107f47a97aa"))
        add_param.save()
        
        # Масса газа нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"add68490-64f4-47a2-a801-1fafa48c09a2"))
        add_param.save()
        # Объем газа при ст.условиях нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"525f9439-ce13-43be-a4f2-67f590f4842b"))
        add_param.save()
        # Объем газа при раб.условиях нарастающим итогом              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"a760c696-fda0-47d4-8fb5-899d742957f1"))
        add_param.save()
        
    elif instance.guid_types_meters.name == u'СПГ762-3':
        #Добавляем параметры для счётчика газа СПГ762 Подсистема 1
    
        #-------------Часовые
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"187b787c-1693-4c90-b6df-d868effef692"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"49c86cc3-57c2-4bdf-b4e3-b07f64673d37"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1e92a9a8-1cd9-4252-b9c7-b33357bafce7"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"06cd5dad-7ea9-438e-abbf-043e8918eb3e"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"8064aad9-778a-4902-b0c0-75b23289469a"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b747eeb5-8c69-443e-b74b-2bb89af64206"))
        add_param.save()
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"38c0b41b-0883-4990-bb0c-8b532caed34c"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"7bf3d68d-4344-49f3-8169-370f6142351a"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"fa188255-c1cc-4c2d-844c-3b40a3a7559e"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"06c97066-1e35-4bb1-a96f-fe3c0056cf39"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e530ab26-92e8-4edc-8e6d-5cd6184bfbe7"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5d315c1e-b237-46d6-9273-be4e597ad1c2"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"54b86222-3d1c-440a-b02f-bedbef0e9e28"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"acb12185-dd88-4449-8f03-76b6fd148958"))
        add_param.save()
        
        #-------------Суточные
        # "tи" Время работы узла учета
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"d1f0258f-42ba-4e4f-a66c-74aed4d512ce"))
        add_param.save()
        # "tиo" Время работы при ненулевом расходе       
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"7ec1a0fb-88f0-497a-8917-01c0b731b88a"))
        add_param.save()
        # "Рб"  Атмосферное давление   
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"3465cb7c-57ea-4ad8-afde-74fb2814ddeb"))
        add_param.save()
        # "Тнв" Температура наружного воздуха
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"a3014fec-73df-4fd4-a68c-9c3ff737d140"))
        add_param.save()
        # "toт01"  Значение времени интегрирования
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"986f24b0-df76-4c36-9b7f-fbbe05a10c94"))
        add_param.save()
        # "Qoт01"    Среднее значение расхода газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"2a5ec8f1-b6fe-4eff-b91f-42a7712dd663"))
        add_param.save()        
        # "Тт01" Среднее значение температуры газа
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"cbfc543d-3a19-46c3-8075-ff59492d2620"))
        add_param.save()
        # "Pт01" Среднее значение абсолютного давления
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"8a24d34e-aee6-4865-bf21-56d9c07dcd1e"))
        add_param.save()
        # "Д1т01"  Ср.значение доп.датчика 1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"85de87d5-8e0a-4088-8248-8a64367db47e"))
        add_param.save()
        # "Д1т02" Ср.значение доп.датчика 2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b2bd2c95-ee85-4156-9ef6-7fc25d29a244"))
        add_param.save()
        # "Mт01" Масса газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"8674a61c-af88-46c5-b553-fecc9a7d0837"))
        add_param.save()
        # "Vт01" Объем газа при ст.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5880bd0f-699d-407c-a3f0-6cea0ebde423"))
        add_param.save()
        # "Vрт01" Объем газа при раб.условиях
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"9057618b-445c-4581-86a6-4715469db938"))
        add_param.save()
        # "НСот01" Обобщ.сообщения о нештатных ситуац.              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4991e5c0-0827-4467-b9d2-7613d1b6dd09"))
        add_param.save()
        
        # Масса газа нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"dcd5ed6a-7bd0-41ba-8850-5b88a9831c04"))
        add_param.save()
        # Объем газа при ст.условиях нарастающим итогом
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"cbeaa4fa-d1fb-4bf5-9688-7084b57fbfe4"))
        add_param.save()
        # Объем газа при раб.условиях нарастающим итогом              
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"d16b31ea-87d1-409a-bbf8-4a743b678dbb"))
        add_param.save()

    elif instance.guid_types_meters.name == u'Sayany':
        #Добавляем параметры для счётчика Sayany
    
        #-------------Суточные
        # "Q" Тепловая энергия. Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e7f2ffba-9a40-43e1-80f3-ddd22596cdb8"))
        add_param.save()    
        # "Q" Тепловая энергия. Канал2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"6f9cd79e-ca34-447e-8ad1-d54531389fe1"))
        add_param.save() 
        # "M" Расход воды. Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b05de8e2-6176-4fc0-bc44-79ceb4229c80"))
        add_param.save() 
        # "M" ТРасход воды. Канал2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5f256e9b-1cb3-4f27-a53a-d08b446dda58"))
        add_param.save() 
        # "T" Температура. Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"75474616-f3db-4903-91d5-1f22f6593394"))
        add_param.save() 
        # "T" Температура. Канал2
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"f3210c5b-afde-4c9a-b201-9c7c403c4cf2"))
        add_param.save() 
        # "T" Температура. Канал3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b12762a0-0a06-49a4-b842-8ad3378f4602"))
        add_param.save() 
        # "T" Температура. Канал4
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"472ba2fd-cc06-4147-a1e7-c1bb66096536"))
        add_param.save() 
        
    elif instance.guid_types_meters.name == u'Tekon_hvs':
        #Добавляем параметры для счётчика Tekon. Читаем один тэк с opcretranslator
    
        #-------------Суточные
        # "Показание". Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"d76796ea-ea63-4317-982f-ffbcde2074dc"))
        add_param.save()  
        
    elif instance.guid_types_meters.name == u'Tekon_gvs':
        #Добавляем параметры для счётчика Tekon. Читаем один тэк с opcretranslator
    
        #-------------Суточные
        # "Показание". Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"e71a7206-7a30-45d9-981d-0b7592b96337"))
        add_param.save() 
        
    elif instance.guid_types_meters.name == u'Tekon_heat':
        #Добавляем параметры для счётчика Tekon. Читаем один тэк с opcretranslator
    
        #-------------Суточные
        # "Показание". Канал1
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"1dca7dab-371a-4429-afa1-8b4877b38b5b"))
        add_param.save()
        
    elif instance.guid_types_meters.name == u'Меркурий 230-УМ':
        #Добавляем параметры для счётчика Меркурий на УСПД УМ-RTU.    
        
        #-------------Суточные
        # "Показание". T0 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"b6e89205-3814-463d-86d1-f52cec7d8962"))
        add_param.save()
        # "Показание". T1 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"7f3c42e6-4000-4373-a0e6-37e66ce819a9"))
        add_param.save() 
        # "Показание". T2 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"c6512649-56ea-4214-aa33-84516bfe8dc1"))
        add_param.save() 
        # "Показание". T3 A+
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"4e20bda9-6e75-4b0f-a99a-0e4c1cd07d3b"))
        add_param.save()
        
        #-------------Мощность        
        #А+ Профиль
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"922ad57c-8f5e-4f00-a78d-e3ba89ef859f")) # A+ 30-мин. срез мощности
        add_param.save()        
        #R+ Профиль
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"61101fa3-a96a-4934-9482-e32036c12829")) # R+ 30-мин. срез мощности
        add_param.save()
        
    elif instance.guid_types_meters.name == u'Пульсар Теплосчётчик':
        #Добавляем параметры для Теплосчётчика Пульсар.
        #------------Суточные
        # "Показание Энергии" Q, Гкал
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"24ae9f51-40a4-4758-a826-a5f8286e1a2e"))
        add_param.save()
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"a3da78fb-b07b-4d53-a980-54b51e26819a"))
        add_param.save()
        # "Показание Температура подачи" Ti, C0
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"de66ecd2-b243-467c-8d1a-cfcb42377300"))
        add_param.save()
        # "Показание Температура выхода" To, C0
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"d3433b80-cb8c-4038-a682-947e6d05955e"))
        add_param.save()

        
    elif instance.guid_types_meters.name == u'Пульсар ХВС':
        #Добавляем параметры для водосчётчика Пульсар ХВС.
        #------------Суточные
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"209894a8-8d19-4e4d-bad8-1767eec4fedf"))
        add_param.save()

    
    elif instance.guid_types_meters.name == u'Пульсар ГВС':
        #Добавляем параметры для водосчётчика Пульсар ГВС.
        #------------Суточные
        # "Показание Расход воды" Объем, м3
        add_param = TakenParams(id = TakenParams.objects.aggregate(Max('id'))['id__max']+1, guid_meters = instance, guid_params = Params.objects.get(guid = u"5fc2ff3b-999e-4154-ba49-84d3971369b0"))
        add_param.save()
        
     
    else:
        pass
        #print u'Тип счётчика не определен'
    
           
#signals.post_save.connect(add_taken_param, sender=Meters)    
        

cfg_excel_name = 'D:\\Work\\18092017 GSM\\prizmer\\static\cfg\\omon.xlsx'
cfg_sheet_name = u'vru'
is_electic_cfg = True
is_water_cfg = False
is_heat_cfg = False
        
#--------------------!!!!!!! Для работы с ведомостью по Воде
#---------------------------------------------------------------------------------------------------
#---------------------------------------------------------------------------------------------------
#---------------------------------------------------------------------------------------------------

def add_objects(sender, instance, created, **kwargs): #Добавляем объекты:
    from openpyxl import load_workbook
    wb = load_workbook(filename = cfg_excel_name)
    sheet_ranges = wb[cfg_sheet_name]
    row = 3
    while (bool(sheet_ranges[u'B%s'%(row)].value) ):
        if sheet_ranges[u'A%s'%(row)].value is not None:
            print sheet_ranges[u'A%s'%(row)].value
            add_object = Objects( name=unicode(sheet_ranges[u'A%s'%(row)].value), level=2, guid_parent = Objects.objects.get(guid=u"2136ad23-659c-4e55-9e3e-32f804a14e90"))
            add_object.save()
            print u'OK'
        else:
            print u'Фигня какая-то'
        row = row + 1
#signals.post_save.connect(add_objects, sender=Resources)


def add_abonents(sender, instance, created, **kwargs): # Добавляем абонентов при создании объекта
#***********************
    from openpyxl import load_workbook
    wb = load_workbook(filename = cfg_excel_name)
    sheet_ranges = wb[cfg_sheet_name]

    row = 3
    while (bool(sheet_ranges[u'B%s'%(row)].value)):
        print instance.name
        print sheet_ranges[u'A%s'%(row)].value
        if sheet_ranges[u'A%s'%(row)].value == unicode(instance.name):
            print u'123'
            add_abonent = Abonents(name = unicode(sheet_ranges[u'B%s'%(row)].value), account_1 =1, guid_objects = instance, guid_types_abonents = TypesAbonents.objects.get(guid= u"e4d813ca-e264-4579-ae15-385cdbf5d28c"))
            add_abonent.save()
            print sheet_ranges[u'B%s'%(row)].value
            #time.sleep(1)
            x = 1
            while (sheet_ranges[u'A%s'%(row+x)].value is None):
                if sheet_ranges[u'B%s'%(row+x)].value is None:
                    print u'Конец списка!!!'
                    #time.sleep(3)
                    break
                else:
                    add_abonent = Abonents(name = unicode(sheet_ranges[u'B%s'%(row+x)].value), account_1 =1, guid_objects = instance, guid_types_abonents = TypesAbonents.objects.get(guid= u"e4d813ca-e264-4579-ae15-385cdbf5d28c"))
                    add_abonent.save()
                print sheet_ranges[u'B%s'%(row+x)].value
                x = x + 1

        else:
            next
        row = row + 1
#signals.post_save.connect(add_abonents, sender=Objects)

#def add_abonents_from_excel_cfg_water(sender, instance, created, **kwargs): #Добавляем абонентов из файла excel ведомости по электрике:
#    from openpyxl import load_workbook
#    wb = load_workbook(filename = cfg_excel_name)
#    sheet_ranges = wb[cfg_sheet_name]
#    row = 2
#    while (bool(sheet_ranges[u'C%s'%(row)].value) ):
#        if sheet_ranges[u'A%s'%(row)].value != u'1':
#            print sheet_ranges[u'A%s'%(row)].value
#            print type(sheet_ranges[u'A%s'%(row)].value)
#            print u'OK'
#            add_abonent = Abonents(name = unicode(sheet_ranges[u'A%s'%(row)].value), account_1 =unicode(sheet_ranges[u'E%s'%(row)].value), account_2 =unicode(sheet_ranges[u'F%s'%(row)].value), guid_objects = instance, guid_types_abonents = TypesAbonents.objects.get(guid= u"e4d813ca-e264-4579-ae15-385cdbf5d28c"))
#            add_abonent.save()
#        else:
#            pass
#        row = row + 1
#signals.post_save.connect(add_abonents_from_excel_cfg_water, sender=Objects)

#def add_meters(sender, instance, created, **kwargs): #Добавляем пульсары
#    # Заводские номера 10-канальных пульсаров
#    list_pulsar_10 = [170621]
#    # Заводские номера 16-канальных пульсаров   
#    list_pulsar_16 = [181914]
#    
#   
#    list_pulsar = list_pulsar_10 + list_pulsar_16
#    for x in range(len(list_pulsar)):
#        if list_pulsar[x] in list_pulsar_10:
#            add_meter = Meters(name = u'Пульсар 10M '+ str(list_pulsar[x]), address = str(list_pulsar[x]), factory_number_manual = str(list_pulsar[x]), guid_types_meters = TypesMeters.objects.get(guid = u"cae994a2-6ab9-4ffa-aac3-f21491a2de0b") )
#            add_meter.save()
#        elif list_pulsar[x] in list_pulsar_16:
#            add_meter = Meters(name = u'Пульсар 16M '+ str(list_pulsar[x]), address = str(list_pulsar[x]), factory_number_manual = str(list_pulsar[x]), guid_types_meters = TypesMeters.objects.get(guid = u"7cd88751-d232-410c-a0ef-6354a79112f1") )
#            add_meter.save()            
#signals.post_save.connect(add_meters, sender=BalanceGroups)

def add_meters(sender, instance, created, **kwargs): #Добавляем пульсары
    """Добавляем счётчики из excel файла ведомости"""
    #from django.db import connection
    from openpyxl import load_workbook
    wb = load_workbook(filename = cfg_excel_name)
    sheet_ranges = wb[cfg_sheet_name]
    row = 3
    pulsar_exception_list = [u'181973']
    while (bool(sheet_ranges[u'C%s'%(row)].value) ):
        if sheet_ranges[u'C%s'%(row)].value is not None:
            print u'Обрабатываем строку ' + str(u'C%s '%(row))
            if unicode(sheet_ranges[u'E%s'%(row)].value) not in pulsar_exception_list:
                pulsar_exception_list.append(unicode(sheet_ranges[u'E%s'%(row)].value))
                print  pulsar_exception_list
                if unicode(sheet_ranges[u'F%s'%(row)].value) == u'Пульсар 10M':

                    add_meter = Meters(name = unicode(sheet_ranges[u'F%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'E%s'%(row)].value), address = unicode(sheet_ranges[u'E%s'%(row)].value), factory_number_manual = unicode(sheet_ranges[u'E%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"cae994a2-6ab9-4ffa-aac3-f21491a2de0b") )
                    add_meter.save()
                      
                    print u'OK', u'Прибор добавлен в базу'
                    
                elif unicode(sheet_ranges[u'F%s'%(row)].value) == u'Пульсар 16M':

                   add_meter = Meters(name = unicode(sheet_ranges[u'F%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'E%s'%(row)].value), address = unicode(sheet_ranges[u'E%s'%(row)].value),  factory_number_manual = unicode(sheet_ranges[u'E%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"7cd88751-d232-410c-a0ef-6354a79112f1") )
                   add_meter.save()
                                                                      
                   print u'OK', u'Прибор добавлен в базу'
            else:
                print u'Такой Пульсар уже есть'
        
        else:
            pass
        row = row + 1
       
#signals.post_save.connect(add_meters, sender=BalanceGroups)

#def add_link_meter_port(sender, instance, created, **kwargs): #Создаем привязку пульсара к com-порту
#    guid_com_port = ComportSettings.objects.get(guid=u'2376a4a5-a19a-4a1d-9417-7f65014961e1')
#    add_com_port_link = LinkMetersComportSettings(guid_meters = instance, guid_comport_settings = guid_com_port)            
#    add_com_port_link.save()
#
##signals.post_save.connect(add_link_meter_port, sender=Meters)
#
#def add_link_meter_tcpip_water(sender, instance, created, **kwargs): #Создаем привязку пульсара к tcp-порту
#    guid_tcpip_port = TcpipSettings.objects.get(guid=u'0346aea3-24ac-4776-900e-e1eb690af7bf')
#    add_tcpip_port_link = LinkMetersTcpipSettings(guid_meters = instance, guid_tcpip_settings = guid_tcpip_port)            
#    add_tcpip_port_link.save()
#
#signals.post_save.connect(add_link_meter_tcpip_water, sender=Meters)

#def add_link_meter_port_from_excel_cfg_water(sender, instance, created, **kwargs):
#    """Делаем привязку счётчика к порту по excel файлу ведомости"""
#    from django.db import connection
#    from openpyxl import load_workbook
#    wb = load_workbook(filename = cfg_excel_name)
#    sheet_ranges = wb[cfg_sheet_name]
#    row = 2
#    
#
#    while (bool(sheet_ranges[u'C%s'%(row)].value) ):
#        # Привязка к tpc порту
#        if sheet_ranges[u'C%s'%(row)].value is not None:
#            #print sheet_ranges[u'G%s'%(row)].value
#            if unicode(sheet_ranges[u'E%s'%(row)].value) == instance.factory_number_manual :
#                guid_ip_port_from_excel = connection.cursor()
#                guid_ip_port_from_excel.execute("""SELECT 
#                                                  tcpip_settings.guid
#                                                FROM 
#                                                  public.tcpip_settings
#                                                WHERE 
#                                                  tcpip_settings.ip_address = %s AND 
#                                                  tcpip_settings.ip_port = %s;""",[unicode(sheet_ranges[u'G%s'%(row)].value), unicode(sheet_ranges[u'H%s'%(row)].value)])
#                guid_ip_port_from_excel = guid_ip_port_from_excel.fetchall()
#        
#                guid_ip_port = TcpipSettings.objects.get(guid=guid_ip_port_from_excel[0][0])
#                add_ip_port_link = LinkMetersTcpipSettings(guid_meters = instance, guid_tcpip_settings = guid_ip_port)            
#                add_ip_port_link.save()
#            else:
#                pass
#            row = row + 1
#signals.post_save.connect(add_link_meter_port_from_excel_cfg_water, sender=Meters)
def add_link_meter_port_from_excel_cfg_water(sender, instance, created, **kwargs):
    """Делаем привязку счётчика к порту по excel файлу ведомости"""
    from django.db import connection
    from openpyxl import load_workbook
    wb = load_workbook(filename = cfg_excel_name)
    sheet_ranges = wb[cfg_sheet_name]
    row = 3

# Привязка к tpc порту
    guid_ip_port_from_excel = connection.cursor()
    guid_ip_port_from_excel.execute("""SELECT 
                                      tcpip_settings.guid
                                    FROM 
                                      public.tcpip_settings
                                    WHERE 
                                      tcpip_settings.ip_address = %s AND 
                                      tcpip_settings.ip_port = %s;""",[unicode(sheet_ranges[u'G%s'%(row)].value), unicode(sheet_ranges[u'H%s'%(row)].value)])
    guid_ip_port_from_excel = guid_ip_port_from_excel.fetchall()

    guid_ip_port = TcpipSettings.objects.get(guid=guid_ip_port_from_excel[0][0])
    add_ip_port_link = LinkMetersTcpipSettings(guid_meters = instance, guid_tcpip_settings = guid_ip_port)            
    add_ip_port_link.save()

#signals.post_save.connect(add_link_meter_port_from_excel_cfg_water, sender=Meters)


def add_link_abonents_taken_params(sender, instance, created, **kwargs):
    
    def get_taken_param_by_abonent_from_excel_cfg(input_taken_param):
        """Функция, которая читает excel файл. Составляет имя считываемого параметра типа "Пульсар 16M 33555 Пульсар 16M Канал 11". В случае совпадения должна привязать этот параметр к абоненту. Абоненты должны быть предварительно созданы."""    
        from openpyxl import load_workbook
        wb = load_workbook(filename = cfg_excel_name)
        sheet_ranges = wb[cfg_sheet_name]
    
        def shrink_taken_param_name(taken_param_name):
            
            if taken_param_name.find(u'Текущий') != -1: # Ищем слово "Текущий"
                nn = taken_param_name.find(u'Текущий')  # Если нашли. то Записываем позицию где.
        
            elif taken_param_name.find(u'Суточный') != -1:
                nn = taken_param_name.find(u'Суточный')
            
            else:
                pass
        
            return taken_param_name[:nn -1]
    
        for row in range(1,810):
            taken_param = u'Пульсар' + u' ' + unicode(sheet_ranges[u'C%s'%(row)].value)[17:20] + u' ' + unicode(sheet_ranges[u'C%s'%(row)].value)[2:8] + u' ' + u'Пульсар' + u' ' + unicode(sheet_ranges[u'C%s'%(row)].value)[17:20] + u' ' + u'Канал' + u' ' + unicode(sheet_ranges[u'D%s'%(row)].value)
                       
            if taken_param == shrink_taken_param_name(input_taken_param):
                try:
                    return unicode(sheet_ranges[u'B%s'%(row)].value)
                except:
                    return None
                
            else:
                pass
    
    print u'--------'
    print instance.name
    print u'==>', get_taken_param_by_abonent_from_excel_cfg(instance.name)
    if get_taken_param_by_abonent_from_excel_cfg(instance.name) is not None:
        print u'Совпадение'
        try:
            add_link_abonents_taken_param = LinkAbonentsTakenParams (name = Abonents.objects.get(name= get_taken_param_by_abonent_from_excel_cfg(instance.name)).name + u" " + instance.guid_params.guid_names_params.name + u" " + instance.guid_params.guid_types_params.name ,coefficient=1, coefficient_2 = 1, guid_abonents = Abonents.objects.get(name= get_taken_param_by_abonent_from_excel_cfg(unicode(instance.name))) , guid_taken_params = instance )
            add_link_abonents_taken_param.save()
        except:
            pass
    else:
        pass
                 
#signals.post_save.connect(add_link_abonents_taken_params, sender=TakenParams)



#--------------------!!!!!!! Для работы с ведомостью по электрике
#---------------------------------------------------------------------------------------------------
#---------------------------------------------------------------------------------------------------
#---------------------------------------------------------------------------------------------------

def add_objects_from_excel_cfg_electric(sender, instance, created, **kwargs): #Добавляем объекты из файла excel ведомости по электрике:
    from openpyxl import load_workbook
    wb = load_workbook(filename = cfg_excel_name)
    sheet_ranges = wb[cfg_sheet_name]
    row = 2
    while (bool(sheet_ranges[u'D%s'%(row)].value) ):
        if sheet_ranges[u'D%s'%(row)].value is not None:
            print sheet_ranges[u'D%s'%(row)].value
            add_object = Objects( name=unicode(sheet_ranges[u'D%s'%(row)].value), level=2, guid_parent = Objects.objects.get(guid=u'e3d5393e-20db-4e4e-956c-5fe5c689d64f'))
            add_object.save()
            print u'OK'
        else:
            pass
        row = row + 1
        
#signals.post_save.connect(add_objects_from_excel_cfg_electric, sender=Resources)


def add_abonents_from_excel_cfg_electric(sender, instance, created, **kwargs): #Добавляем абонентов из файла excel ведомости по электрике:
    from openpyxl import load_workbook
    wb = load_workbook(filename = cfg_excel_name)
    sheet_ranges = wb[cfg_sheet_name]
    row = 2
    while (bool(sheet_ranges[u'D%s'%(row)].value) ):
        if sheet_ranges[u'D%s'%(row)].value is not None:
            print sheet_ranges[u'D%s'%(row)].value
            print u'OK'
            add_abonent = Abonents(name = unicode(sheet_ranges[u'D%s'%(row)].value), account_1 =unicode(sheet_ranges[u'E%s'%(row)].value), account_2 =unicode(sheet_ranges[u'F%s'%(row)].value), guid_objects = instance, guid_types_abonents = TypesAbonents.objects.get(guid= u"e4d813ca-e264-4579-ae15-385cdbf5d28c"))
            add_abonent.save()
        else:
            pass
        row = row + 1
#signals.post_save.connect(add_abonents_from_excel_cfg_electric, sender=Objects)

                    

def add_meters_from_excel_cfg_electric(sender, instance, created, **kwargs):
    """Добавляем счётчики из excel файла ведомости"""
    from django.db import connection
    from openpyxl import load_workbook
    wb = load_workbook(filename = cfg_excel_name)
    sheet_ranges = wb[cfg_sheet_name]
    row = 2
    
    while (bool(sheet_ranges[u'G%s'%(row)].value) ):
        if sheet_ranges[u'G%s'%(row)].value is not None:
            print u'Обрабатываем строку ' + str(u'G%s '%(row)) + str(sheet_ranges[u'G%s'%(row)].value)
            if unicode(sheet_ranges[u'I%s'%(row)].value) == u'М-200':
                add_meter = Meters(name = unicode(sheet_ranges[u'I%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'G%s'%(row)].value), address = unicode(sheet_ranges[u'H%s'%(row)].value), factory_number_manual = unicode(sheet_ranges[u'G%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"6224d20b-1781-4c39-8799-b1446b60774d") )
                add_meter.save()
                print u'Прибор добавлен' + ' --->   ' + u'М-200'              
                
            elif unicode(sheet_ranges[u'I%s'%(row)].value) == u'М-230':
                add_meter = Meters(name = unicode(sheet_ranges[u'I%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'G%s'%(row)].value), address = unicode(sheet_ranges[u'H%s'%(row)].value), password = 111111 , factory_number_manual = unicode(sheet_ranges[u'G%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"423b33a7-2d68-47b6-b4f6-5b470aedc4f4") )
                add_meter.save()
                print u'Прибор добавлен' + ' --->   ' + u'М-230'
                
            elif unicode(sheet_ranges[u'I%s'%(row)].value) == u'М-230-УМ':
                add_meter = Meters(name = unicode(sheet_ranges[u'I%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'G%s'%(row)].value), address = unicode(sheet_ranges[u'H%s'%(row)].value), password = unicode(sheet_ranges[u'F%s'%(row)].value) , factory_number_manual = unicode(sheet_ranges[u'G%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"20e4767a-49e5-4f84-890c-25e311339c28") )
                add_meter.save()
                print u'Прибор добавлен' + ' --->   ' + u'М-230-УМ'
                
            elif unicode(sheet_ranges[u'I%s'%(row)].value) == u'Эльф 1.08':
                add_meter = Meters(name = unicode(sheet_ranges[u'I%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'G%s'%(row)].value), address = unicode(sheet_ranges[u'H%s'%(row)].value), attr1 = unicode(sheet_ranges[u'N%s'%(row)].value), attr2 = unicode(sheet_ranges[u'O%s'%(row)].value), factory_number_manual = unicode(sheet_ranges[u'G%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"1c5a8a80-1c51-4733-8332-4ed8d510a650") )
                add_meter.save()
                print u'Прибор добавлен' + ' --->   ' + u'Эльф 1.08'
            elif unicode(sheet_ranges[u'I%s'%(row)].value) == u'СПГ762-1':
                add_meter = Meters(name = unicode(sheet_ranges[u'I%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'G%s'%(row)].value), address = unicode(sheet_ranges[u'H%s'%(row)].value), factory_number_manual = unicode(sheet_ranges[u'G%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"c3ec5c22-d184-41c5-b6bf-66fa30215a41") )
                add_meter.save()
                print u'Прибор добавлен' + ' --->   ' + u'СПГ762-1'
                
            elif unicode(sheet_ranges[u'I%s'%(row)].value) == u'СПГ762-2':
                add_meter = Meters(name = unicode(sheet_ranges[u'I%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'G%s'%(row)].value), address = unicode(sheet_ranges[u'H%s'%(row)].value), factory_number_manual = unicode(sheet_ranges[u'G%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"5eb7dd59-faf9-4ead-8654-4f3de74de2b0") )
                add_meter.save()
                print u'Прибор добавлен' + ' --->   ' + u'СПГ762-2'
            elif unicode(sheet_ranges[u'I%s'%(row)].value) == u'СПГ762-3':
                add_meter = Meters(name = unicode(sheet_ranges[u'I%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'G%s'%(row)].value), address = unicode(sheet_ranges[u'H%s'%(row)].value), factory_number_manual = unicode(sheet_ranges[u'G%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"e4fb7950-a44f-41f0-a6ff-af5e30d9d562") )
                add_meter.save()
                print u'Прибор добавлен' + ' --->   ' + u'СПГ762-3'
            elif unicode(sheet_ranges[u'I%s'%(row)].value) == u'Sayany':
                add_meter = Meters(name = unicode(sheet_ranges[u'I%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'G%s'%(row)].value), address = unicode(sheet_ranges[u'H%s'%(row)].value), factory_number_manual = unicode(sheet_ranges[u'G%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"5429b439-233e-4944-b91b-4b521a10f77b") )
                add_meter.save()
                print u'Прибор добавлен' + ' --->   ' + u'Sayany'
            elif unicode(sheet_ranges[u'I%s'%(row)].value) == u'Tekon_hvs':
                add_meter = Meters(name = unicode(sheet_ranges[u'I%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'G%s'%(row)].value), address = unicode(sheet_ranges[u'H%s'%(row)].value), factory_number_manual = unicode(sheet_ranges[u'G%s'%(row)].value), password = unicode(sheet_ranges[u'M%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"8398e7d6-39f7-45d2-9c45-a1c48e751b61") )
                add_meter.save()
                print u'Прибор добавлен' + ' --->   ' + u'Tekon_gvs'
            elif unicode(sheet_ranges[u'I%s'%(row)].value) == u'Tekon_hvs':
                add_meter = Meters(name = unicode(sheet_ranges[u'I%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'G%s'%(row)].value), address = unicode(sheet_ranges[u'H%s'%(row)].value), factory_number_manual = unicode(sheet_ranges[u'G%s'%(row)].value), password = unicode(sheet_ranges[u'M%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"64f02a2c-41e1-48b2-bc72-7873ea9b6431") )
                add_meter.save()
                print u'Прибор добавлен' + ' --->   ' + u'Tekon_gvs'

            elif unicode(sheet_ranges[u'I%s'%(row)].value) == u'Tekon_heat':
                add_meter = Meters(name = unicode(sheet_ranges[u'I%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'G%s'%(row)].value), address = unicode(sheet_ranges[u'H%s'%(row)].value), factory_number_manual = unicode(sheet_ranges[u'G%s'%(row)].value), password = unicode(sheet_ranges[u'M%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"b53173f2-2307-4b70-b84c-61b634521e87") )
                add_meter.save()
                print u'Прибор добавлен' + ' --->   ' + u'Tekon_heat'

            elif unicode(sheet_ranges[u'I%s'%(row)].value) == u'Пульсар ХВС':
                add_meter = Meters(name = unicode(sheet_ranges[u'I%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'G%s'%(row)].value), address = unicode(sheet_ranges[u'H%s'%(row)].value), factory_number_manual = unicode(sheet_ranges[u'G%s'%(row)].value), attr1 = unicode(sheet_ranges[u'N%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"f1789bb7-7fcd-4124-8432-40320559890f") )
                add_meter.save()
                print u'Прибор добавлен' + ' --->   ' + u'Пульсар ХВС'
            
            elif unicode(sheet_ranges[u'I%s'%(row)].value) == u'Пульсар ГВС':
                add_meter = Meters(name = unicode(sheet_ranges[u'I%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'G%s'%(row)].value), address = unicode(sheet_ranges[u'H%s'%(row)].value), factory_number_manual = unicode(sheet_ranges[u'G%s'%(row)].value), attr1 = unicode(sheet_ranges[u'N%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"a1a349ba-e070-4ec9-975d-9f39e61c34da") )
                add_meter.save()
                print u'Прибор добавлен' + ' --->   ' + u'Пульсар ГВС'

            elif unicode(sheet_ranges[u'I%s'%(row)].value) == u'Пульсар Теплосчётчик':
                add_meter = Meters(name = unicode(sheet_ranges[u'I%s'%(row)].value) + u' ' + unicode(sheet_ranges[u'G%s'%(row)].value), address = unicode(sheet_ranges[u'H%s'%(row)].value), factory_number_manual = unicode(sheet_ranges[u'G%s'%(row)].value), guid_types_meters = TypesMeters.objects.get(guid = u"82b96b1c-31cf-4753-9d64-d22e2f4d036e") )
                add_meter.save()
                print u'Прибор добавлен' + ' --->   ' + u'Пульсар Теплосчётчик'
                
            else:
                print u'Не найдено совпадение с существующим типом прибора'
   
             
            # Добавляем привязку к балансной группе 
            #guid_balance_groups_from_excel = connection.cursor()
            #guid_balance_groups_from_excel.execute("""SELECT balance_groups.guid FROM public.balance_groups WHERE balance_groups.name = %s;""",[unicode(sheet_ranges[u'M%s'%(row)].value)])
            #guid_balance_groups_from_excel = guid_balance_groups_from_excel.fetchall()
            #guid_balance_groups = BalanceGroups.objects.get(guid=guid_balance_groups_from_excel[0][0])
                        
            #guid_meters_from_excel = connection.cursor()
            #guid_meters_from_excel.execute("""SELECT meters.guid FROM public.meters WHERE meters.factory_number_manual = %s;""",[unicode(sheet_ranges[u'G%s'%(row)].value)])
            #guid_meters_from_excel = guid_meters_from_excel.fetchall()
            #guid_meters = Meters.objects.get(guid=guid_meters_from_excel[0][0])
            
       
            #add_link_meter_balance_group = LinkBalanceGroupsMeters(guid_balance_groups = guid_balance_groups, guid_meters = guid_meters)
            #add_link_meter_balance_group.save()
            # Добавляем привязку к балансной группе конец
                               
        else:
            pass
        row = row + 1
#signals.post_save.connect(add_meters_from_excel_cfg_electric, sender=BalanceGroups)



def add_link_meter_port_from_excel_cfg_electric(sender, instance, created, **kwargs):
    """Делаем привязку счётчика к порту по excel файлу ведомости"""
    from django.db import connection
    from openpyxl import load_workbook
    wb = load_workbook(filename = cfg_excel_name)
    sheet_ranges = wb[cfg_sheet_name]
    row = 2
    

    while (bool(sheet_ranges[u'G%s'%(row)].value) ):
        # Привязка к tpc порту
        if sheet_ranges[u'G%s'%(row)].value is not None:
            #print sheet_ranges[u'G%s'%(row)].value
            if unicode(sheet_ranges[u'G%s'%(row)].value) == instance.factory_number_manual :
                if unicode(sheet_ranges[u'M1'].value) == u'Com-port':
                    guid_com_port_from_excel = connection.cursor()
                    guid_com_port_from_excel.execute("""SELECT 
                                                      comport_settings.guid
                                                    FROM 
                                                      public.comport_settings
                                                    WHERE 
                                                      comport_settings.name = %s;""",[unicode(sheet_ranges[u'M%s'%(row)].value)])
                    guid_com_port_from_excel = guid_com_port_from_excel.fetchall()
            
                    guid_com_port = ComportSettings.objects.get(guid=guid_com_port_from_excel[0][0])
                    add_com_port_link = LinkMetersComportSettings(guid_meters = instance, guid_comport_settings = guid_com_port)            
                    add_com_port_link.save()
                
                else:
                    guid_ip_port_from_excel = connection.cursor()
                    
                    guid_ip_port_from_excel.execute("""SELECT 
                                                      tcpip_settings.guid
                                                    FROM 
                                                      public.tcpip_settings
                                                    WHERE 
                                                      tcpip_settings.ip_address = %s AND 
                                                      tcpip_settings.ip_port = %s;""",[unicode(sheet_ranges[u'K%s'%(row)].value), unicode(sheet_ranges[u'L%s'%(row)].value)])
                    guid_ip_port_from_excel = guid_ip_port_from_excel.fetchall()
            
                    
                    guid_ip_port = TcpipSettings.objects.get(guid=guid_ip_port_from_excel[0][0])
                    add_ip_port_link = LinkMetersTcpipSettings(guid_meters = instance, guid_tcpip_settings = guid_ip_port)            
                    add_ip_port_link.save()
            else:
                pass
            row = row + 1
#signals.post_save.connect(add_link_meter_port_from_excel_cfg_electric, sender=Meters)   

def return_id_abonent_by_name_and_parent_name(name, parent_name):
    from django.db import connection
    simpleq = connection.cursor()
    simpleq.execute("""SELECT abonents.guid FROM public.objects, public.abonents WHERE objects.guid = abonents.guid_objects AND abonents.name = %s AND objects.name = %s;""", [name, parent_name])
    simpleq = simpleq.fetchall()
    return simpleq

def add_link_abonent_taken_params_from_excel_cfg_electric(sender, instance, created, **kwargs):
    from openpyxl import load_workbook
    wb = load_workbook(filename = cfg_excel_name)
    sheet_ranges = wb[cfg_sheet_name]
    row = 2
    while (bool(sheet_ranges[u'G%s'%(row)].value) ):
        if sheet_ranges[u'G%s'%(row)].value is not None:
            from django.db import connection
            guid_abonent_by_excel = connection.cursor()
            guid_abonent_by_excel.execute("""SELECT abonents.guid FROM public.objects, public.abonents
                                                WHERE objects.guid = abonents.guid_objects AND abonents.name = %s AND objects.name = %s;""", [unicode(sheet_ranges[u'D%s'%(row)].value), unicode(sheet_ranges[u'C%s'%(row)].value)])
            guid_abonent_by_excel = guid_abonent_by_excel.fetchall()

            
            if unicode(sheet_ranges[u'G%s'%(row)].value) == instance.guid_meters.factory_number_manual:
                print u'Абонент найден' + u' ' + unicode(instance.name)
                add_link_abonents_taken_param = LinkAbonentsTakenParams (name = unicode(sheet_ranges[u'D%s'%(row)].value) + u' - ' +  unicode(instance.guid_meters.name)  ,coefficient=unicode(sheet_ranges[u'J%s'%(row)].value), coefficient_2 = 1, guid_abonents = Abonents.objects.get(guid =guid_abonent_by_excel[0][0]), guid_taken_params = instance)
                add_link_abonents_taken_param.save()
            else:
                pass
            row = row + 1    
    
#signals.post_save.connect(add_link_abonent_taken_params_from_excel_cfg_electric, sender=TakenParams)

def add_link_meter_port_by_type_meter(sender, instance, created, **kwargs):
    """Делаем привязку счётчика к порту. Привязать все счётчики одного типа к порту."""
    from django.db import connection
    list_of_meters = connection.cursor()
    list_of_meters.execute("""SELECT 
  meters.guid
FROM 
  public.meters, 
  public.types_meters
WHERE 
  meters.guid_types_meters = types_meters.guid AND
  types_meters.name = 'Sayany';""")
    list_of_meters = list_of_meters.fetchall()
   
    for x in range(len(list_of_meters)):
        print list_of_meters[x][0]
        instance_meter = Meters.objects.get(guid = list_of_meters[x][0])
        print instance_meter
        instance_ip_port = TcpipSettings.objects.get(guid = u"a4405a7f-7459-44b3-80a3-8d2c48e2d03f" )
        
        add_ip_port_link = LinkMetersTcpipSettings(guid_meters = instance_meter, guid_tcpip_settings = instance_ip_port)            
        add_ip_port_link.save()

#signals.post_save.connect(add_link_meter_port_by_type_meter, sender=Resources)


# #_____________________________________________________________________________________________________________
# #При изменении в админке будут переименовыватсья все линки
# #Не удалять!
# #_____________________________________________________________________________________________________________
# #from general.models import  Meters, TypesMeters, LinkAbonentsTakenParams, TakenParams, Params
# def rename_taken_params(sender, instance, **kwargs):
#     if (service.views.isService): 
#         pass
#     else:    
#         #переименовываем taken_params
#         guid_meter=instance.guid
#         #print guid_meter
#         new_val=instance.name
#         old_val= Meters.objects.get(guid=guid_meter).name    
#         common_sql.update_table_with_replace('taken_params', 'name', 'guid_meters', guid_meter, old_val, new_val)

#         #переименовываем link_abonents_taken_params
#         for row in TakenParams.objects.filter(guid_meters=guid_meter):
#             guid_taken_params= row.guid
#             common_sql.update_table_with_replace('link_abonents_taken_params', 'name', 'guid_taken_params', guid_taken_params, old_val, new_val)



# def rename_link_abonents_taken_params(sender, instance, **kwargs): 
#     print  service.views.isService 
#     if (service.views.isService): 
#         pass
#     else:
#         print 'not work in service'
#         guid_abon = instance.guid
#         new_val = instance.name
#         old_val = Abonents.objects.get(guid=guid_abon).name
#         common_sql.update_table_with_replace('link_abonents_taken_params', 'name', 'guid_abonents', guid_abon, old_val, new_val)

# signals.pre_save.connect(rename_link_abonents_taken_params, sender=Abonents)
# signals.pre_save.connect(rename_taken_params, sender=Meters)






























